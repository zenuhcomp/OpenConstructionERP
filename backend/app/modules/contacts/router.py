"""‌⁠‍Contacts API routes.

Endpoints:
    GET    /                — List contacts with filters
    POST   /                — Create contact (auth required)
    GET    /search          — Text search across name, company, email
    POST   /import/file     — Import contacts from Excel/CSV file (auth required)
    GET    /export          — Export contacts as Excel file
    GET    /template        — Download import template Excel file
    GET    /{contact_id}    — Get single contact
    PATCH  /{contact_id}    — Update contact (auth required)
    DELETE /{contact_id}    — Soft-delete contact (auth required)
"""

import csv
import io
import logging
import re
import uuid
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.upload_guards import reject_if_xlsx_bomb
from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.contacts.models import Contact
from app.modules.contacts.schemas import (
    ContactCreate,
    ContactListResponse,
    ContactResponse,
    ContactStatsResponse,
    ContactUpdate,
)
from app.modules.contacts.service import ContactService

router = APIRouter()
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> ContactService:
    return ContactService(session)


# ── IDOR protection helpers ─────────────────────────────────────────────────
#
# Tenancy gate: v2.3.1 promotes the ``tenant_id`` column to the primary
# access filter. ``created_by`` remains as an audit field and as a
# fallback for rows inserted before the backfill migration ran.


async def _is_admin(session: AsyncSession, user_id: str | None) -> bool:
    """‌⁠‍Return True if the given user has the ``admin`` role."""
    if user_id is None:
        return False
    try:
        from app.modules.users.repository import UserRepository

        user_repo = UserRepository(session)
        try:
            user_uuid = uuid.UUID(str(user_id))
        except (ValueError, TypeError):
            return False
        user = await user_repo.get_by_id(user_uuid)
        return user is not None and getattr(user, "role", "") == "admin"
    except Exception:  # noqa: BLE001 — best-effort admin check
        return False


async def _require_contact_access(
    session: AsyncSession,
    contact_id: uuid.UUID,
    user_id: str | None,
) -> Contact:
    """‌⁠‍Load a contact and verify the caller owns it or is an admin.

    Access is granted when the contact's ``tenant_id`` matches the caller,
    with a fallback to ``created_by`` for rows inserted before the v2.3.1
    migration backfilled tenant_id. Legacy rows with neither field set
    are treated as admin-only.
    """
    if user_id is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required",
        )

    contact = await session.get(Contact, contact_id)
    if contact is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Contact {contact_id} not found",
        )

    if await _is_admin(session, user_id):
        return contact

    caller = str(user_id)
    tenant = getattr(contact, "tenant_id", None)
    created_by = getattr(contact, "created_by", None)
    if (tenant is not None and str(tenant) == caller) or (
        tenant is None and created_by is not None and str(created_by) == caller
    ):
        return contact

    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Access denied: you do not own this contact",
    )


# ── List ──────────────────────────────────────────────────────────────────────


@router.get(
    "/",
    response_model=ContactListResponse,
    summary="List contacts",
    description="Retrieve a paginated list of contacts with optional filters by type, "
    "country, and active status. Returns total count for pagination.",
)
async def list_contacts(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    contact_type: str | None = Query(default=None),
    country_code: str | None = Query(default=None),
    is_active: bool = Query(default=True),
    tag: list[str] | None = Query(
        default=None,
        description="Filter by metadata.tags entries. Repeat to AND-combine "
        "(e.g. ?tag=paid&tag=de). Each tag is a substring match against the "
        "JSON-serialised metadata column.",
    ),
    offset: int = Query(default=0, ge=0),
    # Raised cap from 100 → 500 so the shared ``ContactSearchInput``
    # "Select from contacts" browse dropdown (which requests
    # ``?limit=200``) doesn't 422 and render an empty list. Mirrors the
    # identical cap raise on ``/projects/`` for the Header switcher.
    limit: int = Query(default=50, ge=1, le=500),
    sort_by: str | None = Query(default=None, description="Sort field: name, email, created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _perm: None = Depends(RequirePermission("contacts.read")),
    service: ContactService = Depends(_get_service),
) -> ContactListResponse:
    """List contacts with optional filters.

    Results are scoped to the caller's tenant (``tenant_id`` column,
    with a ``created_by`` fallback for pre-v2.3.1 rows). Admins
    bypass the scope and see every contact in the database.
    """
    # Map friendly sort field names to model column names
    _sort_aliases = {"name": "company_name", "email": "primary_email"}
    resolved_sort = _sort_aliases.get(sort_by, sort_by) if sort_by else None

    owner_filter: str | None = None if await _is_admin(session, user_id) else user_id
    items, total = await service.list_contacts(
        contact_type=contact_type,
        country_code=country_code,
        is_active=is_active,
        owner_id=owner_filter,
        tags=tag,
        offset=offset,
        limit=limit,
        sort_by=resolved_sort,
        sort_order=sort_order,
    )
    return ContactListResponse(
        items=[ContactResponse.model_validate(c) for c in items],
        total=total,
        offset=offset,
        limit=limit,
    )


# ── Search ────────────────────────────────────────────────────────────────────


@router.get(
    "/search/",
    response_model=ContactListResponse,
    summary="Search contacts",
    description="Full-text search across contact name, company, and email fields. "
    "Supports optional type filter and pagination.",
)
async def search_contacts(
    session: SessionDep,
    q: str = Query(..., min_length=1, max_length=200),
    contact_type: str | None = Query(default=None),
    is_active: bool = Query(default=True),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("contacts.read")),
    service: ContactService = Depends(_get_service),
) -> ContactListResponse:
    """Search contacts across name, company, email.

    Scoped to the caller's own contacts; admins see everything.
    """
    owner_filter: str | None = None if await _is_admin(session, user_id) else user_id
    items, total = await service.list_contacts(
        search=q,
        contact_type=contact_type,
        is_active=is_active,
        owner_id=owner_filter,
        offset=offset,
        limit=limit,
    )
    return ContactListResponse(
        items=[ContactResponse.model_validate(c) for c in items],
        total=total,
        offset=offset,
        limit=limit,
    )


# ── Stats ────────────────────────────────────────────────────────────────────


@router.get(
    "/stats/",
    response_model=ContactStatsResponse,
    summary="Get contact statistics",
    description="Aggregate statistics: total contacts, breakdown by type and country (top 10), "
    "and count of contacts with expiring prequalification.",
)
async def contact_stats(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("contacts.read")),
    service: ContactService = Depends(_get_service),
) -> ContactStatsResponse:
    """Aggregate contact statistics.

    Returns total, breakdown by type and country (top 10), and count of
    contacts with approved prequalification that have qualified_until set.
    Stats are scoped to the caller's own contacts; admins see global
    aggregates.
    """
    owner_filter: str | None = None if await _is_admin(session, user_id) else user_id
    raw = await service.get_stats(owner_id=owner_filter)
    return ContactStatsResponse(
        total=raw["total"],
        by_type=raw["by_type"],
        by_country_top10=raw["by_country_top10"],
        with_expiring_prequalification=raw["with_expiring_prequalification"],
    )


# ── Tag facets ───────────────────────────────────────────────────────────────


@router.get(
    "/tags/",
    summary="List tag facets",
    description="Aggregate counts of metadata.tags entries across the caller's "
    "active contacts. Used to drive the contacts list filter chip strip.",
)
async def contact_tag_facets(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    limit: int = Query(default=60, ge=1, le=200),
    _perm: None = Depends(RequirePermission("contacts.read")),
    service: ContactService = Depends(_get_service),
) -> dict[str, Any]:
    """Return ``[{tag, count}, ...]`` sorted by count desc."""
    owner_filter: str | None = None if await _is_admin(session, user_id) else user_id
    facets = await service.tag_facets(owner_id=owner_filter, limit=limit)
    return {"items": [{"tag": t, "count": c} for t, c in facets]}


# ── By Company ───────────────────────────────────────────────────────────────


@router.get(
    "/by-company/",
    response_model=ContactListResponse,
    summary="List contacts by company",
    description="Find all contacts belonging to the same company (case-insensitive match).",
)
async def contacts_by_company(
    session: SessionDep,
    company_name: str = Query(..., min_length=1, max_length=255),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("contacts.read")),
    service: ContactService = Depends(_get_service),
) -> ContactListResponse:
    """List all contacts at the same company (case-insensitive match).

    Scoped to the caller's own contacts; admins see everything.
    """
    owner_filter: str | None = None if await _is_admin(session, user_id) else user_id
    items, total = await service.list_by_company(
        company_name,
        owner_id=owner_filter,
        offset=offset,
        limit=limit,
    )
    return ContactListResponse(
        items=[ContactResponse.model_validate(c) for c in items],
        total=total,
        offset=offset,
        limit=limit,
    )


# ── File import (CSV / Excel) ─────────────────────────────────────────────

_CONTACT_COLUMN_ALIASES: dict[str, list[str]] = {
    "company_name": [
        "company_name",
        "company",
        "company name",
        "firma",
        "unternehmen",
    ],
    "first_name": [
        "first_name",
        "first name",
        "vorname",
        "given name",
    ],
    "last_name": [
        "last_name",
        "last name",
        "nachname",
        "family name",
        "surname",
    ],
    "contact_type": [
        "contact_type",
        "type",
        "typ",
        "category",
    ],
    "primary_email": [
        "primary_email",
        "email",
        "e-mail",
        "mail",
    ],
    "primary_phone": [
        "primary_phone",
        "phone",
        "telefon",
        "tel",
    ],
    "country_code": [
        "country_code",
        "country",
        "land",
        "country code",
    ],
    "vat_number": [
        "vat_number",
        "vat",
        "vat number",
        "ust-id",
        "ust",
    ],
    "prequalification_status": [
        "prequalification_status",
        "prequalification",
        "prequal",
    ],
    "payment_terms_days": [
        "payment_terms_days",
        "payment terms",
        "payment_terms",
        "zahlungsziel",
    ],
    "notes": [
        "notes",
        "note",
        "bemerkung",
        "anmerkung",
    ],
}

_ALLOWED_CONTACT_TYPES = {
    "client",
    "subcontractor",
    "supplier",
    "consultant",
    "internal",
    "lead",
    "customer",
}
_ALLOWED_PREQUAL = {"pending", "approved", "rejected", "expired"}
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Magic-byte signatures for the formats we actually accept on
# ``POST /import/file/``. Trusting ``filename.endswith()`` alone lets a
# caller upload an arbitrary binary renamed to ``payload.xlsx`` and have
# us hand the bytes straight to ``openpyxl`` / ``csv.reader``. The
# downstream parsers then either crash (best case) or accidentally honour
# embedded macros / external entities in some future libxml-backed
# branch. Reject anything whose first bytes don't match the declared
# extension up front.
_XLSX_MAGIC = b"PK\x03\x04"  # .xlsx is a zip
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"  # legacy OLE compound document
_CSV_BANNED_PREFIXES = (b"MZ", b"\x7fELF", b"\xca\xfe\xba\xbe", b"PK\x03\x04")


def _sniff_upload_content(filename: str, content: bytes) -> str:
    """Return the canonical content kind (``xlsx`` / ``xls`` / ``csv``).

    Raises ``HTTPException(400)`` when the declared extension does not
    line up with the file's magic bytes. CSV has no canonical signature
    so we use a denylist of known-binary prefixes instead.
    """
    head = content[:8]
    if filename.endswith(".xlsx"):
        if not head.startswith(_XLSX_MAGIC):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not look like a valid .xlsx (missing ZIP signature).",
            )
        return "xlsx"
    if filename.endswith(".xls"):
        if not head.startswith(_XLS_MAGIC):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not look like a valid .xls (missing OLE signature).",
            )
        return "xls"
    # CSV: reject obvious binaries up front.
    for sig in _CSV_BANNED_PREFIXES:
        if head.startswith(sig):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File does not look like CSV (binary signature detected).",
            )
    return "csv"


def _redact_row_for_log(row: dict[str, Any]) -> dict[str, str]:
    """Drop e-mail / phone / personal-name fields from a row before logging.

    Errors during bulk import need enough context for the operator to
    locate the offending row (its index + which fields were present)
    without spraying PII into the log pipeline.
    """
    safe_keys = {"company_name", "country_code", "contact_type", "vat_number"}
    return {k: str(row.get(k, ""))[:80] for k in safe_keys if row.get(k)}


def _match_contact_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map."""
    normalised = header.strip().lower()
    for canonical, aliases in _CONTACT_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _parse_contact_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for contact import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_contact_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_contact_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for contact import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_contact_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


@router.post(
    "/import/file/",
    summary="Import contacts from file",
    description="Upload an Excel (.xlsx) or CSV (.csv) file to bulk-import contacts. "
    "Column headers are auto-detected using flexible aliases (EN/DE). "
    "Returns a summary with imported, skipped, and error counts per row.",
)
async def import_contacts_file(
    _user_id: CurrentUserId,
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    _perm: None = Depends(RequirePermission("contacts.create")),
    service: ContactService = Depends(_get_service),
) -> dict[str, Any]:
    """Import contacts from an Excel or CSV file upload.

    Accepts a multipart file upload. The file must be .xlsx or .csv.

    Expected columns (flexible auto-detection):
    - **Company / Firma** -- company name
    - **First Name / Vorname** -- first name
    - **Last Name / Nachname** -- last name
    - **Type / Typ** -- contact type (client, subcontractor, supplier, consultant, internal)
    - **Email / E-Mail** -- primary email
    - **Phone / Telefon** -- phone number
    - **Country / Land** -- 2-char country code
    - **VAT / USt-ID** -- VAT number

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # Magic-byte sniff — the extension is hostile-supplied, so verify the
    # declared format before we hand the bytes to openpyxl / csv.reader.
    kind = _sniff_upload_content(filename, content)

    # Zip-bomb guard: reject .xlsx whose uncompressed sheets exceed 50 MB.
    reject_if_xlsx_bomb(content)

    # Parse rows based on the sniffed format (not the extension alone).
    try:
        if kind in ("xlsx", "xls"):
            rows = _parse_contact_rows_from_excel(content)
        else:
            rows = _parse_contact_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception:
        # Don't echo the original exception text — it can include cell
        # contents (including PII) from openpyxl / csv parser errors.
        logger.exception("Unexpected error parsing contact import file")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # Convert rows to ContactCreate objects and import
    imported_count = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            company_name = str(row.get("company_name", "")).strip()
            first_name = str(row.get("first_name", "")).strip() or None
            last_name = str(row.get("last_name", "")).strip() or None

            # Skip rows without any identifying data
            if not company_name and not first_name and not last_name:
                skipped += 1
                continue

            # Parse contact_type
            contact_type = str(row.get("contact_type", "")).strip().lower()
            if contact_type not in _ALLOWED_CONTACT_TYPES:
                contact_type = "supplier"  # default

            # Validate email
            primary_email = str(row.get("primary_email", "")).strip() or None
            if primary_email and not _EMAIL_RE.match(primary_email):
                # Caller already knows the email they uploaded — echoing
                # it back to *their own* error response is fine. Don't
                # widen ``data`` to the full row though: that broadcasts
                # phone / notes / other PII alongside.
                errors.append({
                    "row": row_idx,
                    "error": f"Invalid email format: {primary_email}",
                    "data": _redact_row_for_log(row),
                })
                continue

            # Validate country_code
            country_code = str(row.get("country_code", "")).strip().upper() or None
            if country_code and len(country_code) != 2:
                errors.append({
                    "row": row_idx,
                    "error": f"Country code must be 2 characters, got: {country_code}",
                    "data": _redact_row_for_log(row),
                })
                continue

            primary_phone = str(row.get("primary_phone", "")).strip() or None
            vat_number = str(row.get("vat_number", "")).strip() or None

            # Parse prequalification
            prequal = str(row.get("prequalification_status", "")).strip().lower() or None
            if prequal and prequal not in _ALLOWED_PREQUAL:
                prequal = None

            payment_terms = str(row.get("payment_terms_days", "")).strip() or None
            notes = str(row.get("notes", "")).strip() or None

            data = ContactCreate(
                contact_type=contact_type,
                company_name=company_name or None,
                first_name=first_name,
                last_name=last_name,
                primary_email=primary_email,
                primary_phone=primary_phone,
                country_code=country_code,
                vat_number=vat_number,
                prequalification_status=prequal,
                payment_terms_days=payment_terms,
                notes=notes,
            )
            await service.create_contact(data, user_id=_user_id)
            imported_count += 1

        except Exception as exc:
            # ``str(exc)`` on Pydantic ValidationError embeds the *input*
            # value — for primary_email / primary_phone rows that means
            # PII lands in both the JSON response and the log line.
            # Replace the message with the exception class name and log a
            # PII-stripped row dict instead.
            err_kind = type(exc).__name__
            errors.append({
                "row": row_idx,
                "error": f"Row rejected ({err_kind})",
                "data": _redact_row_for_log(row),
            })
            logger.warning(
                "Contact import error at row %d: %s (fields=%s)",
                row_idx,
                err_kind,
                sorted(_redact_row_for_log(row).keys()),
            )

    logger.info(
        "Contact file import complete: imported=%d, skipped=%d, errors=%d",
        imported_count,
        skipped,
        len(errors),
    )

    return {
        "imported": imported_count,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Export contacts as Excel ─────────────────────────────────────────────────


@router.get(
    "/export/",
    summary="Export contacts as Excel",
    description="Download all active contacts as an Excel (.xlsx) file.",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_contacts(
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("contacts.read")),
) -> StreamingResponse:
    """Export the caller's active contacts as an Excel file.

    Mirrors the tenant-scope filter used by ``list_contacts`` /
    ``search_contacts`` / ``get_stats`` — admins see every row, everyone
    else only sees contacts whose ``tenant_id`` matches their user id (with
    a ``created_by`` fallback for pre-v2.3.1 rows). The earlier
    implementation ran a plain ``select(Contact).where(is_active)`` with no
    owner filter and exported every tenant's data to anyone with the
    ``contacts.read`` permission.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    stmt = select(Contact).where(Contact.is_active.is_(True))
    if not await _is_admin(session, user_id):
        caller = str(user_id)
        stmt = stmt.where(
            or_(
                Contact.tenant_id == caller,
                and_(
                    Contact.tenant_id.is_(None),
                    Contact.created_by == caller,
                ),
            )
        )
    result = await session.execute(stmt.limit(50000))
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = [
        "Company",
        "First Name",
        "Last Name",
        "Type",
        "Email",
        "Phone",
        "Country",
        "VAT",
        "Prequalification",
        "Payment Terms",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.company_name)
        ws.cell(row=row_idx, column=2, value=item.first_name)
        ws.cell(row=row_idx, column=3, value=item.last_name)
        ws.cell(row=row_idx, column=4, value=item.contact_type)
        ws.cell(row=row_idx, column=5, value=item.primary_email)
        ws.cell(row=row_idx, column=6, value=item.primary_phone)
        ws.cell(row=row_idx, column=7, value=item.country_code)
        ws.cell(row=row_idx, column=8, value=item.vat_number)
        ws.cell(row=row_idx, column=9, value=item.prequalification_status)
        ws.cell(row=row_idx, column=10, value=item.payment_terms_days)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="contacts_export.xlsx"'},
    )


# ── Download import template ─────────────────────────────────────────────────


@router.get(
    "/template/",
    summary="Download import template",
    description="Download an empty Excel template with correct column headers and "
    "two example rows. Includes a Notes sheet explaining each column.",
    response_description="Excel file stream with template headers and example data",
)
async def download_contacts_template(
    _user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("contacts.read")),
) -> StreamingResponse:
    """Download an empty Excel template with headers and example rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts Import"

    headers = [
        "Company",
        "First Name",
        "Last Name",
        "Type",
        "Email",
        "Phone",
        "Country",
        "VAT",
        "Prequalification",
        "Payment Terms",
        "Notes",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    # Example row 1
    example1 = [
        "Acme Construction GmbH",
        "Max",
        "Mustermann",
        "subcontractor",
        "max@acme.de",
        "+49 170 1234567",
        "DE",
        "DE123456789",
        "approved",
        "30",
        "Preferred subcontractor for structural work",
    ]
    for i, val in enumerate(example1, 1):
        ws.cell(row=2, column=i, value=val)

    # Example row 2
    example2 = [
        "Smith & Partners Ltd",
        "Jane",
        "Smith",
        "consultant",
        "jane@smithpartners.co.uk",
        "+44 20 7946 0958",
        "GB",
        "GB987654321",
        "pending",
        "45",
        "Engineering consultants",
    ]
    for i, val in enumerate(example2, 1):
        ws.cell(row=3, column=i, value=val)

    # Notes sheet
    notes_ws = wb.create_sheet("Notes")
    note_headers = ["Column", "Description", "Required", "Example Values"]
    for i, h in enumerate(note_headers, 1):
        cell = notes_ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    notes_data = [
        ("Company", "Company or organization name", "Recommended", "Acme Construction GmbH"),
        ("First Name", "Contact person first name", "No", "Max"),
        ("Last Name", "Contact person last name", "No", "Mustermann"),
        (
            "Type",
            "Contact category",
            "No (defaults to supplier)",
            "client, subcontractor, supplier, consultant, internal",
        ),
        ("Email", "Primary email address", "No", "name@company.com"),
        ("Phone", "Phone number with country code", "No", "+49 170 1234567"),
        ("Country", "ISO 3166-1 alpha-2 country code", "No", "DE, US, GB, FR"),
        ("VAT", "VAT / tax identification number", "No", "DE123456789"),
        (
            "Prequalification",
            "Prequalification status",
            "No",
            "pending, approved, rejected, expired",
        ),
        ("Payment Terms", "Payment terms in days", "No", "30, 45, 60"),
        ("Notes", "Additional notes or remarks", "No", "Free text"),
    ]
    for row_idx, (col, desc, req, example) in enumerate(notes_data, 2):
        notes_ws.cell(row=row_idx, column=1, value=col)
        notes_ws.cell(row=row_idx, column=2, value=desc)
        notes_ws.cell(row=row_idx, column=3, value=req)
        notes_ws.cell(row=row_idx, column=4, value=example)

    # Auto-width for notes sheet
    for col_idx in range(1, 5):
        notes_ws.column_dimensions[chr(64 + col_idx)].width = 35
    notes_ws.column_dimensions["A"].width = 20
    notes_ws.column_dimensions["C"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="contacts_import_template.xlsx"'
        },
    )


# ── Create ────────────────────────────────────────────────────────────────────


@router.post(
    "/",
    response_model=ContactResponse,
    status_code=201,
    summary="Create contact",
    description="Create a new contact in the directory. Validates email format "
    "and enforces allowed contact types (client, subcontractor, supplier, consultant, internal).",
)
async def create_contact(
    data: ContactCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("contacts.create")),
    service: ContactService = Depends(_get_service),
) -> ContactResponse:
    """Create a new contact."""
    contact = await service.create_contact(data, user_id=user_id)
    return ContactResponse.model_validate(contact)


# ── Get ───────────────────────────────────────────────────────────────────────


@router.get(
    "/{contact_id}",
    response_model=ContactResponse,
    summary="Get contact",
    description="Retrieve a single contact by its UUID.",
)
async def get_contact(
    contact_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("contacts.read")),
    service: ContactService = Depends(_get_service),
) -> ContactResponse:
    """Get a single contact by ID."""
    await _require_contact_access(session, contact_id, user_id)
    contact = await service.get_contact(contact_id)
    return ContactResponse.model_validate(contact)


# ── Update ────────────────────────────────────────────────────────────────────


@router.patch(
    "/{contact_id}",
    response_model=ContactResponse,
    summary="Update contact",
    description="Partially update a contact. Only provided fields are modified.",
)
async def update_contact(
    contact_id: uuid.UUID,
    data: ContactUpdate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("contacts.update")),
    service: ContactService = Depends(_get_service),
) -> ContactResponse:
    """Update a contact."""
    await _require_contact_access(session, contact_id, user_id)
    contact = await service.update_contact(contact_id, data, user_id=user_id)
    return ContactResponse.model_validate(contact)


# ── Delete (soft) ─────────────────────────────────────────────────────────────


@router.delete(
    "/{contact_id}",
    status_code=204,
    summary="Delete contact",
    description="Soft-delete a contact by setting is_active=False. "
    "The record is retained for audit purposes but excluded from default queries.",
)
async def delete_contact(
    contact_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("contacts.delete")),
    service: ContactService = Depends(_get_service),
) -> None:
    """Soft-delete a contact (set is_active=False)."""
    await _require_contact_access(session, contact_id, user_id)
    await service.deactivate_contact(contact_id, user_id=user_id)
