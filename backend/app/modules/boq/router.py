"""‌⁠‍BOQ API routes.

Endpoints:
    POST   /boqs/                              — Create a new BOQ
    GET    /boqs/?project_id=xxx               — List BOQs for a project
    GET    /boqs/templates                     — List available BOQ templates
    POST   /boqs/from-template                 — Create a BOQ from a template
    GET    /boqs/{boq_id}                      — Get BOQ with all positions
    PATCH  /boqs/{boq_id}                      — Update BOQ metadata
    DELETE /boqs/{boq_id}                      — Delete BOQ and all positions
    GET    /boqs/{boq_id}/structured           — Full BOQ with sections + markups
    GET    /boqs/{boq_id}/activity             — Activity log for a BOQ
    POST   /boqs/{boq_id}/positions            — Add a position to a BOQ
    POST   /boqs/{boq_id}/positions/bulk      — Bulk insert multiple positions
    PATCH  /boqs/{boq_id}/positions/bulk-update — v3.12 Stream A: bulk field/factor update
    POST   /boqs/{boq_id}/positions/{position_id}/restore-field — v3.12 Stream A: restore one field from a log entry
    PATCH  /positions/{position_id}            — Update a position
    PATCH  /positions/{position_id}/resources/{resource_idx}/variant/
                                               — Re-pick variant on a resource row
    DELETE /positions/{position_id}            — Delete a position
    POST   /boqs/{boq_id}/positions/reorder   — Reorder positions via drag-and-drop
    POST   /boqs/{boq_id}/sections             — Create a section header
    POST   /boqs/{boq_id}/markups              — Add a markup line
    PATCH  /boqs/{boq_id}/markups/{markup_id}  — Update a markup
    DELETE /boqs/{boq_id}/markups/{markup_id}  — Delete a markup
    POST   /boqs/{boq_id}/markups/apply-defaults — Apply regional default markups
    POST   /boqs/{boq_id}/duplicate            — Duplicate a BOQ with all data
    POST   /positions/{position_id}/duplicate  — Duplicate a single position
    POST   /boqs/{boq_id}/lock                 — Lock BOQ (prevent edits)
    POST   /boqs/{boq_id}/unlock               — Unlock BOQ (admin/manager only)
    POST   /boqs/{boq_id}/recalculate-rates    — Recalculate unit_rates from resources
    POST   /boqs/{boq_id}/validate             — Validate a BOQ against rule sets
    GET    /boqs/{boq_id}/export/csv           — Export BOQ as CSV
    GET    /boqs/{boq_id}/export/excel         — Export BOQ as Excel (xlsx)
    GET    /boqs/{boq_id}/export/pdf           — Export BOQ as PDF report
    GET    /boqs/{boq_id}/export/gaeb          — Export BOQ as GAEB XML 3.3 (X83)
    POST   /boqs/{boq_id}/import/excel         — Import positions from Excel/CSV
    POST   /boqs/{boq_id}/import/smart         — Smart import: any file via AI (incl. CAD/BIM)
    GET    /boqs/{boq_id}/resource-summary    — Aggregated resource summary across positions
    GET    /boqs/{boq_id}/cost-breakdown     — Cost breakdown by resource category
    GET    /boqs/{boq_id}/sensitivity       — Sensitivity analysis (tornado chart)
    GET    /boqs/{boq_id}/cost-risk        — Monte Carlo cost risk simulation
    GET    /positions/{id}/quantity-links     — List model→position quantity links
    POST   /positions/{id}/quantity-links     — Bind a position quantity to BIM elements
    DELETE /positions/{id}/quantity-links/{lid} — Delete a quantity link
    POST   /boqs/{id}/quantity-links/refresh  — Re-pull bound quantities (review only)
    POST   /boqs/{id}/quantity-links/apply    — Apply re-pulled quantities (confirm)
    GET    /boqs/{id}/compare/{other_id}      — Line-level compare of two BOQs
    GET    /projects/{project_id}/activity     — Activity log for a project
    POST   /boqs/classify                    — AI: suggest classification codes
    POST   /boqs/suggest-rate                — AI: suggest market rate
    POST   /boqs/{boq_id}/check-anomalies   — AI: detect pricing anomalies
"""

import asyncio
import csv
import io
import logging
import random
import re
import tempfile
import uuid
import zipfile
from collections.abc import Iterator
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, Field, field_validator

from app.core.csv_safety import neutralise_formula
from app.core.file_signature import detect as detect_signature
from app.core.upload_guards import reject_if_xlsx_bomb
from app.dependencies import (
    CurrentUserId,
    CurrentUserPayload,
    RequirePermission,
    SessionDep,
    verify_project_access,
)
from app.modules.boq.schemas import (
    ActivityLogList,
    AIChatRequest,
    AIChatResponse,
    AnomalyCheckResponse,
    AnomalyResponse,
    BOQCompareResponse,
    BOQCreate,
    BOQFromTemplateRequest,
    BOQListItem,
    BOQResponse,
    BOQUpdate,
    BOQWithPositions,
    BOQWithSections,
    BulkPositionUpdate,
    BulkUpdateResult,
    CheckScopeRequest,
    CheckScopeResponse,
    ClassificationSuggestion,
    ClassifiedElement,
    ClassifyElementsRequest,
    ClassifyElementsResponse,
    ClassifyRequest,
    ClassifyResponse,
    CO2AssignRequest,
    CO2EnrichResponse,
    CO2MaterialBreakdown,
    CostBreakdownResponse,
    CostItemSearchRequest,
    CostItemSearchResponse,
    CostItemSearchResult,
    CostRiskDriver,
    CostRiskHistogramBin,
    CostRiskPercentiles,
    CostRiskResponse,
    CostRollupItem,
    EnhanceDescriptionRequest,
    EnhanceDescriptionResponse,
    EscalateRateRequest,
    EscalateRateResponse,
    EscalationFactors,
    EstimateClassificationResponse,
    LineItemResponse,
    MarkupCreate,
    MarkupResponse,
    MarkupUpdate,
    PositionCO2Detail,
    PositionCreate,
    PositionLinksResponse,
    PositionResponse,
    PositionUpdate,
    PrerequisiteItem,
    PricingAnomaly,
    QuantityLinkApplyRequest,
    QuantityLinkApplyResponse,
    QuantityLinkCreate,
    QuantityLinkRefreshResponse,
    QuantityLinkResponse,
    RateMatch,
    ResourceCodeLookupResponse,
    ResourcePositionRef,
    ResourceSummaryItem,
    ResourceSummaryResponse,
    ResourceTypeSummary,
    RestoreFieldRequest,
    RestoreFieldResponse,
    ScopeMissingItem,
    SectionCreate,
    SensitivityItem,
    SensitivityResponse,
    SnapshotCreate,
    SnapshotResponse,
    SuggestPrerequisitesRequest,
    SuggestPrerequisitesResponse,
    SuggestRateRequest,
    SuggestRateResponse,
    SustainabilityResponse,
    TemplateInfo,
)
from app.modules.boq.service import MAX_NESTING_DEPTH, BOQService
from app.modules.costs.repository import CostItemRepository

router = APIRouter()
_log = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> BOQService:
    return BOQService(session)


async def _verify_boq_owner(
    session: SessionDep,
    boq_id: uuid.UUID,
    user_id: str,
    payload: dict | None = None,
) -> None:
    """‌⁠‍Load a BOQ, then its project, and verify ownership.

    Admins bypass the check. Raises 403 if the user is not the project owner.
    """
    if payload and payload.get("role") == "admin":
        return
    from app.modules.boq.repository import BOQRepository
    from app.modules.projects.repository import ProjectRepository

    boq_repo = BOQRepository(session)
    boq = await boq_repo.get_by_id(boq_id)
    if boq is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BOQ not found")
    project_repo = ProjectRepository(session)
    project = await project_repo.get_by_id(boq.project_id)
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if str(project.owner_id) != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have access to this BOQ",
        )


async def _verify_project_owner_for_boq(
    session: SessionDep,
    project_id: uuid.UUID,
    user_id: str,
    payload: dict | None = None,
) -> None:
    """‌⁠‍Verify the current user owns the given project. Admins bypass.

    Treats archived (soft-deleted) projects as 404 — no operations on
    archived projects are permitted via this gateway.
    """
    is_admin = bool(payload and payload.get("role") == "admin")
    from app.modules.projects.repository import ProjectRepository

    project_repo = ProjectRepository(session)
    project = await project_repo.get_by_id(project_id)
    if project is None or project.status == "archived":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")
    if is_admin:
        return
    if str(project.owner_id) != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have access to this project",
        )


async def _log_activity(
    service: BOQService,
    *,
    user_id: uuid.UUID,
    action: str,
    target_type: str,
    description: str,
    project_id: uuid.UUID | None = None,
    boq_id: uuid.UUID | None = None,
    target_id: uuid.UUID | None = None,
    changes: dict | None = None,
) -> None:
    """Fire-and-forget activity logging — never fails the request."""
    try:
        # Resolve project_id from boq if not provided
        if project_id is None and boq_id is not None:
            try:
                boq = await service.get_boq(boq_id)
                project_id = boq.project_id
            except Exception:
                _log.debug("Activity log: failed to resolve project_id from boq_id", exc_info=True)
        await service.log_activity(
            user_id=user_id,
            action=action,
            target_type=target_type,
            description=description,
            project_id=project_id,
            boq_id=boq_id,
            target_id=target_id,
            changes=changes,
        )
    except Exception:
        _log.debug("Activity log write failed (non-critical)", exc_info=True)


_CONFIDENCE_LABELS = {"high": 0.9, "medium": 0.6, "med": 0.6, "low": 0.3}


def _coerce_confidence(raw: object) -> float | None:
    """Best-effort coerce a stored confidence value to float (0.0-1.0).

    Some legacy / seed rows persisted ``confidence`` as a label
    (``'high'``/``'medium'``/``'low'``) rather than the numeric 0–1
    contract.  The PATCH endpoint must keep responding 200 for those
    rows or the whole grid stops saving — so we map known labels to
    representative floats and drop anything else to ``None``.
    """
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip().lower()
    if text in _CONFIDENCE_LABELS:
        return _CONFIDENCE_LABELS[text]
    try:
        return float(text)
    except ValueError:
        return None


def _position_to_response(position: object) -> PositionResponse:
    """Build a PositionResponse from a Position ORM object."""
    # Issue #79: read back the CostItem linkage stored under
    # ``metadata.cost_item_id``.  Older rows that pre-date the linkage
    # simply return None.  We tolerate any non-UUID string defensively
    # — bad data should not break the GET response.
    raw_meta = getattr(position, "metadata_", None)  # type: ignore[attr-defined]
    cost_item_id_val: uuid.UUID | None = None
    if isinstance(raw_meta, dict):
        raw_cid = raw_meta.get("cost_item_id")
        if raw_cid:
            try:
                cost_item_id_val = uuid.UUID(str(raw_cid))
            except (ValueError, TypeError):
                cost_item_id_val = None

    return PositionResponse(
        id=position.id,  # type: ignore[attr-defined]
        boq_id=position.boq_id,  # type: ignore[attr-defined]
        parent_id=position.parent_id,  # type: ignore[attr-defined]
        ordinal=position.ordinal,  # type: ignore[attr-defined]
        description=position.description,  # type: ignore[attr-defined]
        unit=position.unit,  # type: ignore[attr-defined]
        # BUG-B-011: forward the exact stored decimal string; the schema
        # now keeps it as Decimal and serialises a plain string, so a
        # 999,999,999.99 × 999,999.99 total no longer loses its tail.
        quantity=position.quantity,  # type: ignore[attr-defined]
        unit_rate=position.unit_rate,  # type: ignore[attr-defined]
        total=position.total,  # type: ignore[attr-defined]
        classification=position.classification,  # type: ignore[attr-defined]
        source=position.source,  # type: ignore[attr-defined]
        confidence=_coerce_confidence(position.confidence),  # type: ignore[attr-defined]
        cad_element_ids=position.cad_element_ids,  # type: ignore[attr-defined]
        validation_status=position.validation_status,  # type: ignore[attr-defined]
        metadata=position.metadata_,  # type: ignore[attr-defined]
        sort_order=position.sort_order,  # type: ignore[attr-defined]
        created_at=position.created_at,  # type: ignore[attr-defined]
        updated_at=position.updated_at,  # type: ignore[attr-defined]
        cost_item_id=cost_item_id_val,
        # BUG-CONCURRENCY01: surface the row's optimistic-concurrency
        # token so clients can echo it on the next PATCH.
        version=int(getattr(position, "version", 0) or 0),  # type: ignore[attr-defined]
        # Issue #127: reuse-group fields (read-only). ``linked_instance_count``
        # needs a project-wide query so it is left None here and populated
        # explicitly by the links endpoint / propagation paths.
        reference_code=getattr(position, "reference_code", None),  # type: ignore[attr-defined]
        link_role=getattr(position, "link_role", None),  # type: ignore[attr-defined]
        link_group_id=getattr(position, "link_group_id", None),  # type: ignore[attr-defined]
    )


async def _position_to_response_with_links(
    service: BOQService,
    position: object,
) -> PositionResponse:
    """Issue #127: like ``_position_to_response`` but, for a master,
    populate ``linked_instance_count`` via a single project-wide query.

    Used only by the single-position endpoints (GET / PATCH / add /
    unlink) — the grid/list endpoints stay on the cheap sync builder so
    they don't pay a per-row link query.
    """
    resp = _position_to_response(position)
    if getattr(position, "link_role", None) == "master":
        try:
            links = await service.list_links(position.id)  # type: ignore[attr-defined]
            resp.linked_instance_count = links.instance_count
        except Exception:  # noqa: BLE001 — count is advisory, never break
            _log.debug("linked_instance_count enrichment failed", exc_info=True)
    # Issue #127: merge the transient propagation/unlink outcome (set by
    # the service on a NON-mapped attribute so it never hits the DB).
    info = getattr(position, "_link_propagation_info", None)
    if isinstance(info, dict):
        merged = dict(resp.metadata) if isinstance(resp.metadata, dict) else {}
        merged["link_propagation"] = info
        resp.metadata = merged
    return resp


def _markup_to_response(markup: object) -> MarkupResponse:
    """Build a MarkupResponse from a BOQMarkup ORM object."""
    return MarkupResponse(
        id=markup.id,  # type: ignore[attr-defined]
        boq_id=markup.boq_id,  # type: ignore[attr-defined]
        name=markup.name,  # type: ignore[attr-defined]
        markup_type=markup.markup_type,  # type: ignore[attr-defined]
        category=markup.category,  # type: ignore[attr-defined]
        percentage=float(markup.percentage),  # type: ignore[attr-defined]
        fixed_amount=float(markup.fixed_amount),  # type: ignore[attr-defined]
        apply_to=markup.apply_to,  # type: ignore[attr-defined]
        sort_order=markup.sort_order,  # type: ignore[attr-defined]
        is_active=markup.is_active,  # type: ignore[attr-defined]
        metadata_=markup.metadata_,  # type: ignore[attr-defined]
        created_at=markup.created_at,  # type: ignore[attr-defined]
        updated_at=markup.updated_at,  # type: ignore[attr-defined]
    )


# ── BOQ CRUD ──────────────────────────────────────────────────────────────────


@router.post(
    "/boqs/",
    response_model=BOQResponse,
    status_code=201,
    summary="Create BOQ",
    description="Create a new Bill of Quantities for a project. Verifies project ownership.",
    dependencies=[Depends(RequirePermission("boq.create"))],
)
async def create_boq(
    data: BOQCreate,
    _user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Create a new Bill of Quantities."""
    await _verify_project_owner_for_boq(session, data.project_id, _user_id, payload)
    boq = await service.create_boq(data)
    await _log_activity(
        service,
        user_id=_user_id,
        action="boq_created",
        target_type="boq",
        description=f"Created BOQ '{boq.name}'",
        project_id=data.project_id,
        boq_id=boq.id,
        target_id=boq.id,
    )
    return BOQResponse.model_validate(boq)


@router.get(
    "/boqs/",
    response_model=list[BOQListItem],
    summary="List BOQs",
    description="List all BOQs for a project with computed grand totals and position counts.",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_boqs(
    _user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    project_id: uuid.UUID = Query(..., description="Filter BOQs by project"),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    service: BOQService = Depends(_get_service),
) -> list[BOQListItem]:
    """List all BOQs for a given project with computed grand totals."""
    await _verify_project_owner_for_boq(session, project_id, _user_id, payload)
    boqs, _ = await service.list_boqs_for_project(project_id, offset=offset, limit=limit)
    # Compute grand totals + position counts via aggregate queries
    boq_ids = [b.id for b in boqs]
    # ``totals_for_boqs`` returns the full breakdown so list and detail
    # endpoints stay in lockstep (BUG-008).
    breakdown = await service.boq_repo.totals_for_boqs(boq_ids)

    # Position counts per BOQ
    from sqlalchemy import func, select

    from app.modules.boq.models import Position

    pos_counts: dict[uuid.UUID, int] = {}
    if boq_ids:
        rows = (
            await session.execute(
                select(Position.boq_id, func.count())
                .where(Position.boq_id.in_(boq_ids))
                .where(Position.unit != "")  # Exclude section headers
                .group_by(Position.boq_id)
            )
        ).all()
        for bid, cnt in rows:
            pos_counts[bid] = cnt

    results: list[BOQListItem] = []
    for b in boqs:
        money = breakdown.get(b.id, {"direct_cost": 0.0, "markups_total": 0.0, "grand_total": 0.0})
        item = BOQListItem(
            id=b.id,
            project_id=b.project_id,
            name=b.name,
            description=b.description,
            status=b.status,
            metadata=b.metadata_,
            created_at=b.created_at,
            updated_at=b.updated_at,
            direct_cost_total=money["direct_cost"],
            markups_total=money["markups_total"],
            grand_total=money["grand_total"],
            position_count=pos_counts.get(b.id, 0),
        )
        results.append(item)
    return results


# ── Templates ────────────────────────────────────────────────────────────────


@router.get(
    "/boqs/templates/",
    response_model=list[TemplateInfo],
    summary="List BOQ templates",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_templates(
    service: BOQService = Depends(_get_service),
) -> list[TemplateInfo]:
    """List all available BOQ templates.

    Returns summary information for each built-in template: name, description,
    icon, section count, and total position count.

    Templates cover common building types:
    - **residential** — Multi-family apartments, 3-5 floors
    - **office** — Commercial office, 4-8 floors
    - **warehouse** — Logistics warehouse, single-story
    - **school** — Primary/secondary school, 2-3 floors
    - **hospital** — General hospital or clinic
    - **hotel** — 3-5 star hotel with conference
    - **retail** — Shopping mall, 1-3 floors
    - **infrastructure** — Bridge / overpass
    """
    return service.list_templates()


@router.post(
    "/boqs/from-template/",
    response_model=BOQResponse,
    status_code=201,
    summary="Create BOQ from template",
    dependencies=[Depends(RequirePermission("boq.create"))],
)
async def create_boq_from_template(
    data: BOQFromTemplateRequest,
    _user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Create a complete BOQ from a built-in template.

    Provide the template ID, gross floor area (m2), and optionally a custom
    name.  The endpoint creates:
    - A new BOQ
    - Section headers for each trade group
    - All positions with quantities derived from ``area_m2 * qty_factor``

    Use ``GET /boqs/templates`` to discover available template IDs.
    """
    boq = await service.create_boq_from_template(data)
    return BOQResponse.model_validate(boq)


# ── AI-powered Classification ─────────────────────────────────────────────
# NOTE: These POST routes with literal paths (/boqs/classify, /boqs/suggest-rate)
# MUST be defined BEFORE GET /boqs/{boq_id}, otherwise FastAPI matches "classify"
# and "suggest-rate" as {boq_id} path parameters and returns 405 Method Not Allowed.


@router.post(
    "/boqs/classify/",
    response_model=ClassifyResponse,
    summary="AI: Suggest classification codes",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def classify_position(
    data: ClassifyRequest,
    service: BOQService = Depends(_get_service),
) -> ClassifyResponse:
    """Suggest classification codes for a BOQ position description.

    Uses vector similarity search against the cost database to find items
    with similar descriptions, then aggregates their classification codes
    ranked by frequency weighted by similarity score.

    Returns 3-5 suggestions ordered by confidence (highest first).
    Gracefully returns an empty list if the vector database is unavailable.

    Args:
        data: ClassifyRequest with description, unit, and project_standard.

    Returns:
        ClassifyResponse with a list of ClassificationSuggestion.
    """
    try:
        raw_suggestions = await service.classify_position(
            description=data.description,
            unit=data.unit,
            project_standard=data.project_standard,
        )
    except HTTPException:
        raise
    except Exception as exc:
        _log.exception("classify_position failed")
        # Return a structured error so the frontend can show a clear message
        # instead of an empty suggestion list that looks like "no matches".
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Classification service unavailable: {exc}",
        )

    suggestions = [
        ClassificationSuggestion(
            standard=s["standard"],
            code=s["code"],
            label=s["label"],
            confidence=s["confidence"],
        )
        for s in raw_suggestions
    ]

    return ClassifyResponse(suggestions=suggestions)


# ── Deterministic CAD Element Classification ──────────────────────────────


@router.post(
    "/boqs/classify-elements/",
    response_model=ClassifyElementsResponse,
    summary="Map CAD elements to classification codes",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def classify_elements(
    data: ClassifyElementsRequest,
) -> ClassifyElementsResponse:
    """Map CAD/BIM element categories to classification codes.

    Uses deterministic lookup tables to map Revit/IFC category names
    (e.g. ``"Walls"``, ``"Doors"``) to classification codes in the
    requested standard (DIN 276, NRM, or MasterFormat).

    This is a fast, offline operation — no AI or database access required.
    Useful for initial classification of CAD-extracted elements before
    importing them into a BOQ.

    Args:
        data: ClassifyElementsRequest with elements list and standard.

    Returns:
        ClassifyElementsResponse with classified elements and counts.
    """
    from app.modules.cad.classification_mapper import map_category_to_standard

    classified: list[ClassifiedElement] = []
    mapped_count = 0

    for elem in data.elements:
        code = map_category_to_standard(elem.category, data.standard)
        # Merge existing classification with the new mapping
        merged_classification = dict(elem.classification)
        was_mapped = False
        if code:
            merged_classification[data.standard] = code
            was_mapped = True
            mapped_count += 1

        classified.append(
            ClassifiedElement(
                id=elem.id,
                category=elem.category,
                classification=merged_classification,
                mapped=was_mapped,
            )
        )

    total = len(data.elements)
    return ClassifyElementsResponse(
        elements=classified,
        standard=data.standard,
        total=total,
        mapped_count=mapped_count,
        unmapped_count=total - mapped_count,
    )


# ── AI Cost Finder (vector search) ────────────────────────────────────────


@router.post(
    "/boqs/search-cost-items/",
    response_model=CostItemSearchResponse,
    summary="AI: Search cost items",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def search_cost_items(
    data: CostItemSearchRequest,
    service: BOQService = Depends(_get_service),
) -> CostItemSearchResponse:
    """Search cost items using AI vector similarity.

    Performs semantic search against the cost database using text embeddings.
    Returns ranked results with similarity scores, enriched with full item
    details (components, classification) from the SQL database.

    Filters: unit (exact match), region, min_score threshold.
    Gracefully returns empty results if vector DB is unavailable.
    """
    try:
        result = await service.search_cost_items(
            query=data.query,
            unit=data.unit,
            region=data.region,
            limit=data.limit,
            min_score=data.min_score,
        )
    except Exception:
        _log.exception("search_cost_items failed")
        result = {
            "results": [],
            "total_found": 0,
            "query_embedding_ms": 0,
            "search_ms": 0,
        }

    items = [
        CostItemSearchResult(
            id=r["id"],
            code=r["code"],
            description=r["description"],
            unit=r["unit"],
            rate=r["rate"],
            region=r["region"],
            score=r["score"],
            classification=r.get("classification", {}),
            components=r.get("components", []),
            currency=r.get("currency", ""),
        )
        for r in result.get("results", [])
    ]

    return CostItemSearchResponse(
        results=items,
        total_found=result.get("total_found", 0),
        query_embedding_ms=result.get("query_embedding_ms", 0),
        search_ms=result.get("search_ms", 0),
    )


# ── AI-powered Rate Suggestion ─────────────────────────────────────────────


@router.post(
    "/boqs/suggest-rate/",
    response_model=SuggestRateResponse,
    summary="AI: Suggest market rate",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def suggest_rate(
    data: SuggestRateRequest,
    service: BOQService = Depends(_get_service),
) -> SuggestRateResponse:
    """Suggest a market rate for a BOQ position based on its description.

    Uses vector similarity search to find cost items with similar
    descriptions, optionally filtered by unit and region. Computes a
    weighted average rate (weighted by similarity score) and returns
    the individual matches for transparency.

    Gracefully returns zero rate with empty matches if the vector database
    is unavailable.

    Args:
        data: SuggestRateRequest with description, unit, classification, region.

    Returns:
        SuggestRateResponse with suggested_rate, confidence, source, matches.
    """
    try:
        result = await service.suggest_rate(
            description=data.description,
            unit=data.unit,
            classification=data.classification,
            region=data.region,
        )
    except Exception:
        _log.exception("suggest_rate failed")
        result = {
            "suggested_rate": 0.0,
            "confidence": 0.0,
            "source": "vector_search",
            "matches": [],
        }

    matches = [
        RateMatch(
            code=m["code"],
            description=m["description"],
            rate=m["rate"],
            region=m["region"],
            score=m["score"],
        )
        for m in result.get("matches", [])
    ]

    return SuggestRateResponse(
        suggested_rate=result["suggested_rate"],
        confidence=result["confidence"],
        source=result["source"],
        matches=matches,
    )


# ── BOQ Anomaly Detection ─────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/check-anomalies/",
    response_model=AnomalyCheckResponse,
    summary="AI: Detect pricing anomalies",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def check_anomalies(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> AnomalyCheckResponse:
    """Check all positions in a BOQ for pricing anomalies.

    For each position with a description and unit_rate > 0, performs a
    vector search to find 5-10 similar cost items. Calculates the p25,
    median, and p75 of matched rates and flags positions whose rate
    deviates significantly from the market range:

    - **error**: rate > 3x median (likely a pricing mistake)
    - **warning**: rate > 2x median (review recommended)
    - **warning**: rate < 0.3x median (suspiciously low)

    Gracefully returns an empty anomaly list if the vector database is
    unavailable.

    Args:
        boq_id: Target BOQ identifier.

    Returns:
        AnomalyCheckResponse with anomalies list and positions_checked count.
    """
    try:
        result = await service.check_anomalies(boq_id)
    except HTTPException:
        raise
    except Exception:
        _log.exception("check_anomalies failed for BOQ %s", boq_id)
        result = {"anomalies": [], "positions_checked": 0}

    anomalies = [
        PricingAnomaly(
            position_id=a["position_id"],
            field=a["field"],
            current_value=a["current_value"],
            market_range=a["market_range"],
            severity=a["severity"],
            message=a["message"],
            suggestion=a["suggestion"],
        )
        for a in result.get("anomalies", [])
    ]

    return AnomalyCheckResponse(
        anomalies=anomalies,
        positions_checked=result.get("positions_checked", 0),
    )


# ── LLM-powered AI features ─────────────────────────────────────────────────


@router.post(
    "/boqs/enhance-description/",
    response_model=EnhanceDescriptionResponse,
    summary="AI: Enhance description",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def enhance_description(
    data: EnhanceDescriptionRequest,
    user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> EnhanceDescriptionResponse:
    """Enhance a short BOQ position description into a precise technical specification.

    Uses LLM to add material grades, standards references, and technical details.
    Requires an AI API key configured in user settings.
    """
    try:
        result = await service.enhance_description(
            user_id=user_id,
            description=data.description,
            unit=data.unit,
            classification=data.classification,
            locale=data.locale,
        )
    except HTTPException:
        raise
    except Exception:
        _log.exception("enhance_description failed")
        result = {
            "enhanced_description": data.description,
            "specifications": [],
            "standards": [],
            "confidence": 0.0,
            "model_used": "",
            "tokens_used": 0,
        }

    return EnhanceDescriptionResponse(**result)


@router.post(
    "/boqs/suggest-prerequisites/",
    response_model=SuggestPrerequisitesResponse,
    summary="AI: Suggest prerequisites",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def suggest_prerequisites(
    data: SuggestPrerequisitesRequest,
    user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> SuggestPrerequisitesResponse:
    """Suggest prerequisite and related work items for a BOQ position.

    Analyzes the target position and existing BOQ to identify commonly
    needed companion, prerequisite, and successor work items.
    """
    try:
        result = await service.suggest_prerequisites(
            user_id=user_id,
            description=data.description,
            unit=data.unit,
            classification=data.classification,
            existing_descriptions=data.existing_descriptions,
            locale=data.locale,
        )
    except HTTPException:
        raise
    except Exception:
        _log.exception("suggest_prerequisites failed")
        result = {"suggestions": [], "model_used": "", "tokens_used": 0}

    suggestions = [PrerequisiteItem(**s) for s in result.get("suggestions", [])]
    return SuggestPrerequisitesResponse(
        suggestions=suggestions,
        model_used=result.get("model_used", ""),
        tokens_used=result.get("tokens_used", 0),
    )


@router.post(
    "/boqs/{boq_id}/check-scope/",
    response_model=CheckScopeResponse,
    summary="AI: Check scope completeness",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def check_scope(
    boq_id: uuid.UUID,
    data: CheckScopeRequest,
    user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> CheckScopeResponse:
    """Analyze BOQ for scope completeness — find missing trades and work packages.

    Sends a summary of all positions to the LLM which identifies gaps:
    missing structural items, MEP, finishes, external works, preliminaries, etc.
    """
    try:
        result = await service.check_scope_completeness(
            user_id=user_id,
            boq_id=boq_id,
            project_type=data.project_type,
            region=data.region,
            currency=data.currency,
            locale=data.locale,
        )
    except HTTPException:
        raise
    except Exception:
        _log.exception("check_scope failed for BOQ %s", boq_id)
        result = {
            "completeness_score": 0.0,
            "missing_items": [],
            "warnings": ["Analysis failed"],
            "summary": "",
            "model_used": "",
            "tokens_used": 0,
        }

    missing = [ScopeMissingItem(**m) for m in result.get("missing_items", [])]
    return CheckScopeResponse(
        completeness_score=result.get("completeness_score", 0.0),
        missing_items=missing,
        warnings=result.get("warnings", []),
        summary=result.get("summary", ""),
        model_used=result.get("model_used", ""),
        tokens_used=result.get("tokens_used", 0),
    )


@router.post(
    "/boqs/escalate-rate/",
    response_model=EscalateRateResponse,
    summary="AI: Escalate rate to current prices",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def escalate_rate(
    data: EscalateRateRequest,
    user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> EscalateRateResponse:
    """Escalate a unit rate from a base year to current prices.

    Uses LLM to estimate inflation factors (material, labor, regional)
    and compute an escalated rate.
    """
    try:
        result = await service.escalate_rate(
            user_id=user_id,
            description=data.description,
            unit=data.unit,
            rate=data.rate,
            currency=data.currency,
            base_year=data.base_year,
            target_year=data.target_year,
            region=data.region,
            locale=data.locale,
        )
    except HTTPException:
        raise
    except Exception:
        _log.exception("escalate_rate failed")
        result = {
            "original_rate": data.rate,
            "escalated_rate": data.rate,
            "escalation_percent": 0.0,
            "factors": {},
            "confidence": "low",
            "reasoning": "Analysis failed",
            "model_used": "",
            "tokens_used": 0,
        }

    factors = result.get("factors", {})
    return EscalateRateResponse(
        original_rate=result["original_rate"],
        escalated_rate=result["escalated_rate"],
        escalation_percent=result["escalation_percent"],
        factors=EscalationFactors(**factors) if isinstance(factors, dict) else EscalationFactors(),
        confidence=result.get("confidence", "low"),
        reasoning=result.get("reasoning", ""),
        model_used=result.get("model_used", ""),
        tokens_used=result.get("tokens_used", 0),
    )


@router.get(
    "/boqs/{boq_id}",
    response_model=BOQWithPositions,
    summary="Get BOQ with positions",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_boq(
    boq_id: uuid.UUID,
    _user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> BOQWithPositions:
    """Get a BOQ with all its positions and grand total."""
    await _verify_boq_owner(session, boq_id, _user_id, payload)
    return await service.get_boq_with_positions(boq_id)


@router.get(
    "/boqs/{boq_id}/structured/",
    response_model=BOQWithSections,
    summary="Get BOQ with sections and markups",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_boq_structured(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> BOQWithSections:
    """Get a BOQ with hierarchical sections, subtotals, markups, and totals.

    Returns the full structured view that a professional estimator needs:
    - Sections with grouped positions and subtotals
    - Ungrouped positions (no parent section)
    - Direct cost (sum of all item totals)
    - Markup lines with computed amounts
    - Net total (direct cost + markups)
    - Grand total
    """
    return await service.get_boq_structured(boq_id)


@router.get(
    "/boqs/{boq_id}/activity/",
    response_model=ActivityLogList,
    summary="Get BOQ activity log",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_boq_activity(
    boq_id: uuid.UUID,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    service: BOQService = Depends(_get_service),
) -> ActivityLogList:
    """Get the activity log for a BOQ (paginated, newest first).

    Returns a chronological audit trail of all mutations: position additions,
    updates, deletions, markup changes, exports, etc.
    """
    return await service.get_activity_for_boq(boq_id, offset=offset, limit=limit)


@router.get(
    "/projects/{project_id}/activity/",
    response_model=ActivityLogList,
    summary="Get project activity log",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_project_activity(
    project_id: uuid.UUID,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    service: BOQService = Depends(_get_service),
) -> ActivityLogList:
    """Get the activity log for a project (paginated, newest first).

    Returns all BOQ-related activity across all BOQs in the project.
    """
    return await service.get_activity_for_project(project_id, offset=offset, limit=limit)


@router.get(
    "/projects/{project_id}/resource-by-code/",
    response_model=ResourceCodeLookupResponse,
    summary="Look up an existing resource by its code (project-wide)",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def lookup_resource_by_code(
    project_id: uuid.UUID,
    code: str,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> ResourceCodeLookupResponse:
    """Issue #133 — find the first existing resource using ``code``.

    Drives the "this code is already in use — insert the existing
    resource, or create a new one with another code?" prompt in the BOQ
    editor's manual resource form. Returns ``found=False`` when the code
    is unused anywhere in the project.
    """
    await _verify_project_owner_for_boq(session, project_id, user_id, payload)
    return await service.find_resource_by_code(project_id, code)


@router.patch(
    "/boqs/{boq_id}",
    response_model=BOQResponse,
    summary="Update BOQ",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def update_boq(
    boq_id: uuid.UUID,
    data: BOQUpdate,
    _user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Update BOQ metadata (name, description, status)."""
    await _verify_boq_owner(session, boq_id, _user_id, payload)
    boq = await service.update_boq(boq_id, data)
    return BOQResponse.model_validate(boq)


@router.delete(
    "/boqs/{boq_id}",
    status_code=204,
    summary="Delete BOQ",
    dependencies=[Depends(RequirePermission("boq.delete"))],
)
async def delete_boq(
    boq_id: uuid.UUID,
    _user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> None:
    """Delete a BOQ and all its positions."""
    await _verify_boq_owner(session, boq_id, _user_id, payload)
    await service.delete_boq(boq_id)


# ── Duplicate ────────────────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/duplicate/",
    response_model=BOQResponse,
    status_code=201,
    summary="Duplicate BOQ",
    dependencies=[Depends(RequirePermission("boq.create"))],
)
async def duplicate_boq(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Duplicate an entire BOQ with all its positions and markups.

    Creates a new BOQ named "<original> (Copy)" in the same project.
    All positions (with hierarchy) and markups are deep-copied with new IDs.
    """
    new_boq = await service.duplicate_boq(boq_id)
    return BOQResponse.model_validate(new_boq)


@router.post(
    "/positions/{position_id}/duplicate/",
    response_model=PositionResponse,
    status_code=201,
    summary="Duplicate position",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def duplicate_position(
    position_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Duplicate a single position within the same BOQ.

    Creates a copy with ordinal "<original>.1" placed after the original.
    """
    new_position = await service.duplicate_position(position_id)
    return _position_to_response(new_position)


# ── Lock & Revision ──────────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/lock/",
    response_model=BOQResponse,
    summary="Lock BOQ",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def lock_boq(
    boq_id: uuid.UUID,
    user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Lock a BOQ to prevent further edits.

    Sets is_locked=True, approved_by to the current user, approved_at to now.

    Audit B7 / CC4 — was vulnerable to a TOCTOU race. The old flow was
    READ → CHECK → UPDATE, which let two concurrent callers both pass
    the "not locked" check and both write their (different) approval
    metadata. We now use a single compare-and-swap UPDATE:

        UPDATE boqs SET is_locked=true, approved_by=:u, approved_at=:t,
                       status='final'
        WHERE id=:boq_id AND is_locked=false

    If rowcount == 0, either the BOQ doesn't exist or it was already
    locked by another caller (race winner) — both are 409 Conflict so
    the loser can distinguish the second case via the message and
    just refresh.
    """
    from datetime import datetime

    from sqlalchemy import update as sa_update

    from app.modules.boq.models import BOQ

    # Existence check up-front: gives us a clean 404 (vs the ambiguous
    # 409 you'd otherwise get from a race-loser-or-missing row).
    await service.get_boq(boq_id)  # raises 404 if missing

    now_iso = datetime.now(UTC).isoformat()
    stmt = (
        sa_update(BOQ)
        .where(BOQ.id == boq_id)
        .where(BOQ.is_locked == False)  # noqa: E712 — SQLAlchemy needs explicit ==
        .values(
            is_locked=True,
            approved_by=user_id,
            approved_at=now_iso,
            status="final",
        )
    )
    result = await service.session.execute(stmt)
    await service.session.flush()
    service.session.expire_all()

    if result.rowcount == 0:  # type: ignore[union-attr]
        # Lost the race. Surface a 409 so the client can decide whether
        # to refresh or surface an "already locked" message.
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="BOQ is already locked.",
        )

    # FSM audit row — record draft -> final transition in oe_activity_log
    # so the entity lifecycle has the same audit footprint as the other
    # FSM-managed entities. Best-effort: if audit write fails (e.g. table
    # not present in a partial migration), we still return the locked
    # BOQ — the CAS UPDATE above already committed the status change.
    try:
        from app.core.audit_log import log_activity

        await log_activity(
            service.session,
            actor_id=str(user_id),
            entity_type="boq",
            entity_id=str(boq_id),
            action="status_changed",
            from_status="draft",
            to_status="final",
            reason="Locked via /boqs/{id}/lock",
            metadata={"approved_at": now_iso, "approved_by": str(user_id)},
        )
    except Exception:
        _log.exception("FSM audit write skipped for BOQ %s lock", boq_id)

    boq = await service.get_boq(boq_id)
    return BOQResponse.model_validate(boq)


@router.post(
    "/boqs/{boq_id}/unlock/",
    response_model=BOQResponse,
    summary="Unlock BOQ",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def unlock_boq(
    boq_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Unlock a previously locked BOQ, allowing further edits.

    Only admin or manager roles can unlock. Sets is_locked=False and
    status back to "draft". Returns a warning if revisions exist.

    Audit B7 / CC4 — symmetrical CAS treatment with ``lock_boq``: the
    UPDATE only fires when ``is_locked = true``, so double-unlock
    races degrade cleanly into a 400 instead of double-recording
    "draft" reverts in the activity log.
    """
    from sqlalchemy import update as sa_update

    from app.modules.boq.models import BOQ

    role = payload.get("role", "")
    if role not in ("admin", "manager"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin or manager can unlock a BOQ.",
        )

    # Existence check first (separates 404 from race-loser).
    await service.get_boq(boq_id)

    stmt = (
        sa_update(BOQ)
        .where(BOQ.id == boq_id)
        .where(BOQ.is_locked == True)  # noqa: E712
        .values(
            is_locked=False,
            status="draft",
        )
    )
    result = await service.session.execute(stmt)
    await service.session.flush()
    service.session.expire_all()

    if result.rowcount == 0:  # type: ignore[union-attr]
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="BOQ is not locked.",
        )

    # FSM audit row — record final -> draft transition in oe_activity_log
    # for compliance traceability (regulatory dispute records).
    try:
        from app.core.audit_log import log_activity

        await log_activity(
            service.session,
            actor_id=str(user_id),
            entity_type="boq",
            entity_id=str(boq_id),
            action="status_changed",
            from_status="final",
            to_status="draft",
            reason="Unlocked via /boqs/{id}/unlock",
            metadata={"actor_role": role},
        )
    except Exception:
        _log.exception("FSM audit write skipped for BOQ %s unlock", boq_id)

    boq = await service.get_boq(boq_id)
    return BOQResponse.model_validate(boq)


@router.post(
    "/boqs/{boq_id}/create-budget/",
    status_code=201,
    summary="Create budget from BOQ",
    description="Create project budget lines from BOQ sections/positions. "
    "BOQ must be locked first. Groups positions by WBS or section.",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def create_budget_from_boq(
    boq_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> dict:
    """Create project budget lines from BOQ sections/positions.

    Groups positions by WBS (if set) or by section, creates a
    ProjectBudget entry for each group with original_budget = section total.
    Returns the list of created budget IDs.
    """
    from decimal import Decimal

    boq = await service.get_boq(boq_id)
    if not boq.is_locked:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="BOQ must be locked before creating a budget. Lock the estimate first.",
        )

    positions = boq.positions or []
    if not positions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="BOQ has no positions — nothing to budget.",
        )

    # Group positions: by wbs_id if set, otherwise by parent_id (section), else "ungrouped"
    groups: dict[str, Decimal] = {}
    for pos in positions:
        # Skip section headers (quantity=0, unit="")
        try:
            total = Decimal(str(pos.total))
        except Exception:
            total = Decimal("0")
        if total == 0 and pos.unit == "":
            continue

        group_key = pos.wbs_id or (str(pos.parent_id) if pos.parent_id else "ungrouped")
        groups[group_key] = groups.get(group_key, Decimal("0")) + total

    if not groups:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No budgetable positions found in BOQ.",
        )

    # Lazy import finance module
    created_ids: list[str] = []
    try:
        from app.modules.finance.models import ProjectBudget

        for group_key, total_amount in groups.items():
            budget = ProjectBudget(
                project_id=boq.project_id,
                wbs_id=group_key if group_key != "ungrouped" else None,
                category="other",
                original_budget=str(total_amount),
                revised_budget=str(total_amount),
                metadata_={"source": "boq", "boq_id": str(boq_id)},
            )
            session.add(budget)
            await session.flush()
            created_ids.append(str(budget.id))
    except Exception as exc:
        _log.exception("Failed to create budgets from BOQ %s: %s", boq_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create budget lines. Finance module may not be available.",
        )

    await _log_activity(
        service,
        user_id=user_id,
        action="boq.budget_created",
        target_type="boq",
        description=f"Created {len(created_ids)} budget lines from BOQ",
        boq_id=boq_id,
        project_id=boq.project_id,
    )

    _log.info(
        "Created %d budget lines from BOQ %s (project %s)",
        len(created_ids),
        boq_id,
        boq.project_id,
    )
    return {
        "created": len(created_ids),
        "budget_ids": created_ids,
        "project_id": str(boq.project_id),
    }


@router.post(
    "/boqs/{boq_id}/create-revision/",
    response_model=BOQResponse,
    status_code=201,
    summary="Create BOQ revision",
    dependencies=[Depends(RequirePermission("boq.create"))],
)
async def create_revision(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Create a new revision of a BOQ.

    Duplicates the entire BOQ (positions + markups) and links the copy
    back to the original via parent_estimate_id for revision tracking.
    """
    # Ensure Projects model is registered in SQLAlchemy metadata (FK resolution)
    from app.modules.projects import models as _proj_models  # noqa: F401

    new_boq = await service.duplicate_boq(boq_id)
    new_boq_id = new_boq.id  # Capture before session expires attributes
    # Link the new BOQ to the original as its revision parent
    await service.boq_repo.update_fields(
        new_boq_id,
        parent_estimate_id=boq_id,
        status="draft",
        is_locked=False,
    )
    refreshed = await service.get_boq(new_boq_id)
    return BOQResponse.model_validate(refreshed)


# ── Position CRUD ─────────────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/positions/",
    response_model=PositionResponse,
    status_code=201,
    summary="Add position",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def add_position(
    boq_id: uuid.UUID,
    data: PositionCreate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Add a new position to a BOQ.

    The boq_id in the URL takes precedence over the body field.
    """
    await _verify_boq_owner(session, boq_id, user_id, payload)
    # Override body boq_id with URL path parameter
    data.boq_id = boq_id
    position = await service.add_position(data)
    await _log_activity(
        service,
        user_id=user_id,
        action="position_added",
        target_type="position",
        description=f"Added position '{data.description[:60]}'",
        boq_id=boq_id,
        target_id=position.id,
    )
    # Issue #127: a reuse may have promoted an existing row to master —
    # surface linked_instance_count on the new instance's response.
    return await _position_to_response_with_links(service, position)


@router.post(
    "/boqs/{boq_id}/positions/bulk/",
    response_model=list[PositionResponse],
    status_code=201,
    summary="Bulk add positions",
    dependencies=[Depends(RequirePermission("boq.import"))],
)
async def bulk_add_positions(
    boq_id: uuid.UUID,
    payload: dict[str, Any],
    user_id: CurrentUserId,
    auth_payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> list[PositionResponse]:
    """Bulk insert multiple positions into a BOQ.

    Accepts ``{"items": [{"description": ..., "quantity": ..., "unit": ...}, ...]}``
    as sent by the Takeoff page.

    Probe-A perf fix (v2.10): batch-validate inputs into
    :class:`PositionCreate`, then call ``service.bulk_add_positions``
    which performs a single ``add_all`` + flush. Was a per-item
    ``service.add_position`` loop (51 ms/pos for 100 rows on SQLite).
    """
    await _verify_boq_owner(session, boq_id, user_id, auth_payload)
    items: list[dict[str, Any]] = payload.get("items", [])
    if not items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="'items' list is required and must not be empty",
        )

    # Determine next ordinal base from existing positions (cheap, one query)
    try:
        boq_data = await service.get_boq_with_positions(boq_id)
        existing_count = len(boq_data.positions) if boq_data.positions else 0
    except HTTPException:
        existing_count = 0

    # Validate every row up-front. A bad row aborts the whole batch with
    # 422; the previous serial loop accepted partial success which made
    # debugging harder than the perf saved.
    payloads: list[PositionCreate] = []
    errors: list[dict[str, Any]] = []
    for idx, item in enumerate(items):
        try:
            ordinal = item.get("ordinal", f"{existing_count + idx + 1:03d}")
            description = str(item.get("description", "")).strip()
            if not description:
                description = f"Position {existing_count + idx + 1}"

            try:
                quantity = float(item.get("quantity", 0))
            except (ValueError, TypeError):
                quantity = 0.0

            try:
                unit_rate = float(item.get("unit_rate", 0))
            except (ValueError, TypeError):
                unit_rate = 0.0

            # Issue #79: forward an optional ``cost_item_id`` per row so
            # bulk imports can carry CostItem linkage too. Pydantic does
            # the UUID parsing; the service validates the target.
            raw_cost_item_id = item.get("cost_item_id")
            payloads.append(
                PositionCreate(
                    boq_id=boq_id,
                    ordinal=ordinal,
                    description=description,
                    unit=item.get("unit", "pcs"),
                    quantity=quantity,
                    unit_rate=unit_rate,
                    source=item.get("source", "takeoff"),
                    classification=item.get("classification", {}),
                    metadata=item.get("metadata", {}),
                    cost_item_id=raw_cost_item_id if raw_cost_item_id else None,
                ),
            )
        except Exception as exc:
            logger.warning("Bulk import: item %d failed: %s", idx, exc)
            errors.append({"index": idx, "error": str(exc)})

    if errors and not payloads:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"All {len(errors)} items failed validation. First error: {errors[0]['error']}",
        )
    if errors:
        # Mixed batch — historically the serial path accepted partial
        # success. Preserve that contract by skipping the bad rows and
        # bulk-inserting the rest.
        logger.warning(
            "Bulk import: %d/%d rows skipped due to validation errors",
            len(errors),
            len(items),
        )

    try:
        inserted = await service.bulk_add_positions(boq_id, payloads)
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc

    return [_position_to_response(p) for p in inserted]


@router.get(
    "/positions/{position_id}",
    response_model=PositionResponse,
    summary="Get position by id",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_position(
    position_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Return a single BOQ position by id (BUG-API14).

    Previously this verb-path combination was unhandled and FastAPI fell
    through to a route that returned ``200 {}`` — a confusing 200-with-
    empty-body which masked race-condition fallout from BUG-CONCURRENCY01.
    Now it returns the full :class:`PositionResponse` or 404.
    """
    # IDOR guard: load position → derive boq_id → verify ownership chain
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    return await _position_to_response_with_links(service, existing)


@router.patch(
    "/positions/{position_id}",
    response_model=PositionResponse,
    summary="Update position",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def update_position(
    position_id: uuid.UUID,
    data: PositionUpdate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Update a BOQ position. Recalculates total if quantity or unit_rate changed."""
    # IDOR guard: load position → derive boq_id → verify ownership chain
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    # Pass actor_id through so the audit log records who made the change
    # (BUG-AUDIT01).  Without it the service falls back to anonymous and
    # the FK to ``oe_users_user`` would fail.
    try:
        position = await service.update_position(position_id, data, actor_id=user_id)
    except ValueError as exc:
        # Probe-A scenario 11 — overflow cap and similar service-layer
        # validation failures are user-facing input errors, not 500s.
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc
    # Issue #127: master propagation count / unlink flag is on the
    # position metadata; enrich linked_instance_count for masters.
    return await _position_to_response_with_links(service, position)


# ── v3.12.0 Stream A — bulk update & per-field restore ─────────────────────


@router.patch(
    "/boqs/{boq_id}/positions/bulk-update/",
    response_model=BulkUpdateResult,
    summary="Bulk update many positions",
    description=(
        "Apply one of three bulk mutations to many positions atomically: "
        "direct field assignment, multiply unit_rate by a factor, or multiply "
        "quantity by a factor. Each row goes through the normal update path "
        "so totals recompute, validation resets, and audit rows are written."
    ),
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def bulk_update_positions(
    boq_id: uuid.UUID,
    data: BulkPositionUpdate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> BulkUpdateResult:
    """Bulk-update endpoint — see :class:`BulkPositionUpdate` for payload.

    The umbrella activity-log entry uses ``action='position.bulk_<kind>'``
    and carries the full id list (plus failures) in ``changes``, so the
    activity panel renders one bulk row instead of N per-position rows.
    """
    await _verify_boq_owner(session, boq_id, user_id, payload)
    try:
        return await service.bulk_update_positions(
            boq_id, data, actor_id=user_id
        )
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc


@router.post(
    "/boqs/{boq_id}/positions/{position_id}/restore-field/",
    response_model=RestoreFieldResponse,
    summary="Restore a single field on a position from an activity-log entry",
    description=(
        "Look up the supplied log entry, verify it targets this position "
        "and recorded a change for the named field, then write the supplied "
        "value through the normal update path. A fresh log row is appended "
        "noting the source entry id so the restore chain is auditable."
    ),
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def restore_position_field(
    boq_id: uuid.UUID,
    position_id: uuid.UUID,
    data: RestoreFieldRequest,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> RestoreFieldResponse:
    """Per-cell restore from a prior :class:`BOQActivityLog` entry."""
    await _verify_boq_owner(session, boq_id, user_id, payload)
    # Cross-check the position membership before touching the service so
    # path-tampering returns 404 not 422.
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None or existing.boq_id != boq_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Position {position_id} not found in BOQ {boq_id}",
        )
    updated = await service.restore_position_field(
        position_id,
        field=data.field,
        value=data.value,
        log_id=data.log_id,
        actor_id=user_id,
    )
    # Re-read the freshest log so we can echo the restore entry id.
    new_log_id: uuid.UUID | None = None
    try:
        recent, _total = await service.activity_repo.list_for_boq(
            boq_id, offset=0, limit=1
        )
        if recent and recent[0].action == "position.field_restored":
            new_log_id = recent[0].id
    except Exception:  # noqa: BLE001 — informational only
        new_log_id = None
    return RestoreFieldResponse(
        position_id=updated.id,
        field=data.field,
        restored_value=data.value,
        source_log_id=data.log_id,
        new_log_id=new_log_id,
    )


class _ResourceVariantRepickBody(BaseModel):
    """Payload for the per-resource variant re-pick endpoint."""

    model_config = ConfigDict(str_strip_whitespace=True)

    variant_code: str = Field(
        ...,
        min_length=1,
        max_length=512,
        description=(
            "Label of the desired variant in the resource's cached "
            "``available_variants`` array. Matches ``CostVariant.label``."
        ),
    )


@router.patch(
    "/positions/{position_id}/resources/{resource_idx}/variant/",
    response_model=PositionResponse,
    summary="Re-pick variant on an existing resource entry",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def repick_resource_variant(
    position_id: uuid.UUID,
    resource_idx: int,
    data: _ResourceVariantRepickBody,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Swap the variant on an already-added BOQ resource row.

    Reads the variant set cached on
    ``position.metadata.resources[resource_idx].available_variants`` (stamped
    at apply-time by the frontend), finds the variant whose label matches
    ``variant_code``, and patches that single resource's ``unit_rate``,
    ``variant`` marker, and ``variant_snapshot``. Other resources on the
    position are untouched — their snapshots keep their original
    ``captured_at`` per the v2.6.25 immutability contract.
    """
    # IDOR guard: load position → derive boq_id → verify ownership chain.
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)

    position = await service.repick_resource_variant(
        position_id,
        resource_idx,
        data.variant_code,
        actor_id=user_id,
    )
    return _position_to_response(position)


# ── Issue #127: linked-position (code reuse) endpoints ────────────────────────


@router.get(
    "/positions/{position_id}/links/",
    response_model=PositionLinksResponse,
    summary="List linked positions sharing this code",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_position_links(
    position_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionLinksResponse:
    """List every position that reuses this position's ``reference_code``.

    Issue #127. Returns the reuse group across the WHOLE project: which
    member is the master, per-instance ordinals + quantities + totals, and
    counts. A standalone position (code used once) returns ``linked=False``.
    """
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    return await service.list_links(position_id)


@router.post(
    "/positions/{position_id}/unlink/",
    response_model=PositionResponse,
    summary="Unlink a position from its reuse group",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def unlink_position(
    position_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Detach a position from its reference-code reuse group.

    Issue #127. Values are unchanged (the position keeps its definition
    and ``reference_code``); it just stops following the master. Unlinking
    a master promotes the oldest remaining instance (or dissolves the
    group) so no instance is orphaned.
    """
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    position = await service.unlink_position(position_id, actor_id=user_id)
    return await _position_to_response_with_links(service, position)


@router.delete(
    "/positions/{position_id}",
    status_code=204,
    summary="Delete position",
    dependencies=[Depends(RequirePermission("boq.delete"))],
)
async def delete_position(
    position_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    cascade: bool = Query(
        default=False,
        description="If true, delete all descendant positions when deleting a section.",
    ),
    service: BOQService = Depends(_get_service),
) -> None:
    """Delete a single position. For sections, pass ?cascade=true to delete children."""
    # IDOR guard: load position → derive boq_id → verify ownership chain
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    await _log_activity(
        service,
        user_id=user_id,
        action="position_deleted",
        target_type="position",
        description=f"Deleted position {position_id} (cascade={cascade})",
        target_id=position_id,
    )
    await service.delete_position(position_id, cascade=cascade)


@router.post(
    "/boqs/{boq_id}/positions/reorder/",
    summary="Reorder positions",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def reorder_positions(
    boq_id: uuid.UUID,
    payload: CurrentUserPayload,
    session: SessionDep,
    data: dict = Body(...),
    user_id: CurrentUserId = None,
    service: BOQService = Depends(_get_service),
) -> dict:
    """Reorder positions within a BOQ.

    Expects ``{"position_ids": ["uuid1", "uuid2", ...]}``.
    The list order determines the new ``sort_order`` for each position.
    """
    # IDOR guard: verify BOQ ownership before any mutation
    await _verify_boq_owner(session, boq_id, user_id, payload)
    raw_ids = data.get("position_ids", [])
    position_ids = [uuid.UUID(pid) if isinstance(pid, str) else pid for pid in raw_ids]
    await service.reorder_positions(boq_id, position_ids)
    await _log_activity(
        service,
        user_id=user_id,
        action="position_updated",
        target_type="boq",
        description=f"Reordered {len(position_ids)} positions",
        boq_id=boq_id,
    )
    return {"ok": True}


# ── BOQ limits (Issue #136) ───────────────────────────────────────────────────


@router.get(
    "/limits/",
    summary="BOQ structural limits (max nesting depth, etc.)",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_boq_limits() -> dict[str, int]:
    """Return server-enforced BOQ structural limits.

    Issue #136 — the editor reads ``max_nesting_depth`` so it can disable
    "add child" / "add sub-section" once the configurable cap is reached
    and show a tooltip, keeping the UI in lock-step with the backend
    validation (single source of truth: ``service.MAX_NESTING_DEPTH``).
    """
    return {"max_nesting_depth": MAX_NESTING_DEPTH}


# ── Section CRUD ──────────────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/sections/",
    response_model=PositionResponse,
    status_code=201,
    summary="Create section",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def create_section(
    boq_id: uuid.UUID,
    data: SectionCreate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Create a section header row in a BOQ.

    Sections are positions with unit="section", quantity=0, unit_rate=0.
    They serve as grouping headers for estimating line items.
    """
    await _verify_boq_owner(session, boq_id, user_id, payload)
    section = await service.create_section(boq_id, data)
    return _position_to_response(section)


# ── Markup CRUD ───────────────────────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/markups/",
    summary="List markups",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_markups(
    boq_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> dict:
    """List all markups for a BOQ."""
    await _verify_boq_owner(session, boq_id, user_id, payload)
    markups = await service.list_markups(boq_id)
    return {"markups": [_markup_to_response(m) for m in markups]}


@router.post(
    "/boqs/{boq_id}/markups/",
    response_model=MarkupResponse,
    status_code=201,
    summary="Add markup",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def add_markup(
    boq_id: uuid.UUID,
    data: MarkupCreate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> MarkupResponse:
    """Add a markup/overhead line to a BOQ."""
    await _verify_boq_owner(session, boq_id, user_id, payload)
    markup = await service.add_markup(boq_id, data)
    return _markup_to_response(markup)


@router.patch(
    "/boqs/{boq_id}/markups/{markup_id}",
    response_model=MarkupResponse,
    summary="Update markup",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def update_markup(
    boq_id: uuid.UUID,
    markup_id: uuid.UUID,
    data: MarkupUpdate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> MarkupResponse:
    """Update a markup/overhead line on a BOQ."""
    await _verify_boq_owner(session, boq_id, user_id, payload)
    markup = await service.update_markup(markup_id, data)
    return _markup_to_response(markup)


@router.delete(
    "/boqs/{boq_id}/markups/{markup_id}",
    status_code=204,
    summary="Delete markup",
    dependencies=[Depends(RequirePermission("boq.delete"))],
)
async def delete_markup(
    boq_id: uuid.UUID,
    markup_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> None:
    """Delete a markup/overhead line from a BOQ."""
    await _verify_boq_owner(session, boq_id, user_id, payload)
    await service.delete_markup(markup_id)


@router.post(
    "/boqs/{boq_id}/markups/apply-defaults/",
    response_model=list[MarkupResponse],
    summary="Apply regional default markups",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def apply_default_markups(
    boq_id: uuid.UUID,
    region: str = Query(
        default="DEFAULT",
        description="Region code: DACH, UK, US, FR, GULF, IN, AU, JP, BR, NORDIC, RU, CN, KR, DEFAULT",
    ),
    service: BOQService = Depends(_get_service),
) -> list[MarkupResponse]:
    """Apply regional default markups to a BOQ.

    Replaces any existing markups with the standard template for the region.
    Pass region as query parameter: ``?region=DACH``.

    Supported regions: DACH, UK, US, FR, GULF, IN, AU, JP, BR, NORDIC,
    RU, CN, KR, DEFAULT.
    """
    markups = await service.apply_default_markups(boq_id, region)
    return [_markup_to_response(m) for m in markups]


# ── Feature 1: model→BOQ quantity links ───────────────────────────────────────


@router.get(
    "/positions/{position_id}/quantity-links/",
    response_model=list[QuantityLinkResponse],
    summary="List quantity links for a position",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_quantity_links(
    position_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> list[QuantityLinkResponse]:
    """List every live model→position quantity binding for a position."""
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Position not found"
        )
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    return await service.list_quantity_links(position_id)


@router.post(
    "/positions/{position_id}/quantity-links/",
    response_model=QuantityLinkResponse,
    status_code=201,
    summary="Bind a position quantity to BIM model elements",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def create_quantity_link(
    position_id: uuid.UUID,
    data: QuantityLinkCreate,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> QuantityLinkResponse:
    """Create a live binding (extraction rule) — does NOT change the quantity.

    The position's quantity is only ever changed by an explicit confirm
    (the architecture guide §7 — human-confirmed). Creating the link records the rule
    and provenance so a later model revision can be re-pulled for review.
    """
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Position not found"
        )
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    link = await service.create_quantity_link(
        position_id, data, created_by=user_id
    )
    await _log_activity(
        service,
        user_id=user_id,
        action="quantity_link_created",
        target_type="position",
        description=(
            f"Bound {data.quantity_field} → quantity from "
            f"{len(data.element_stable_ids)} model element(s)"
        ),
        boq_id=existing.boq_id,
        target_id=position_id,
    )
    return link


@router.delete(
    "/positions/{position_id}/quantity-links/{link_id}",
    status_code=204,
    summary="Delete a quantity link",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def delete_quantity_link(
    position_id: uuid.UUID,
    link_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> None:
    """Stop tracking a binding. The last-applied quantity stays put."""
    existing = await service.position_repo.get_by_id(position_id)
    if existing is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Position not found"
        )
    await _verify_boq_owner(session, existing.boq_id, user_id, payload)
    link = await service.quantity_link_repo.get_by_id(link_id)
    if link is None or link.position_id != position_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quantity link not found"
        )
    await service.delete_quantity_link(link_id)


@router.post(
    "/boqs/{boq_id}/quantity-links/refresh/",
    response_model=QuantityLinkRefreshResponse,
    summary="Re-pull bound quantities against the latest model version",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def refresh_quantity_links(
    boq_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> QuantityLinkRefreshResponse:
    """Probe every link against the latest model — flag stale, no writes.

    Returns a per-position review payload (old qty, new computed qty,
    delta, contributing elements). Applying the change is a separate,
    explicit confirm call.
    """
    await _verify_boq_owner(session, boq_id, user_id, payload)
    return await service.refresh_quantity_links(boq_id)


@router.post(
    "/boqs/{boq_id}/quantity-links/apply/",
    response_model=QuantityLinkApplyResponse,
    summary="Apply re-pulled quantities to the chosen positions",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def apply_quantity_links(
    boq_id: uuid.UUID,
    data: QuantityLinkApplyRequest,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> QuantityLinkApplyResponse:
    """Human-confirmed apply: only the listed links write to their positions.

    Each applied position records a provenance entry in
    ``metadata.model_quantity_pull`` / ``..._history`` — the figure's
    origin is auditable and never silently overwritten.
    """
    await _verify_boq_owner(session, boq_id, user_id, payload)
    result = await service.apply_quantity_links(
        boq_id, data.link_ids, applied_by=user_id
    )
    await _log_activity(
        service,
        user_id=user_id,
        action="quantity_link_applied",
        target_type="boq",
        description=(
            f"Applied {result.applied} model-driven quantity update(s)"
        ),
        boq_id=boq_id,
    )
    return result


# ── Feature 2: estimate baseline / line-level comparison ──────────────────────


@router.get(
    "/boqs/{boq_id}/compare/{other_id}",
    response_model=BOQCompareResponse,
    summary="Line-level comparison of two BOQs",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def compare_boqs(
    boq_id: uuid.UUID,
    other_id: uuid.UUID,
    user_id: CurrentUserId,
    payload: CurrentUserPayload,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> BOQCompareResponse:
    """Compare two BOQs line-by-line (added / removed / qty / rate / delta).

    Pure read. Ownership is verified on BOTH BOQs so a baseline can never
    leak positions from a project the caller does not own.
    """
    await _verify_boq_owner(session, boq_id, user_id, payload)
    await _verify_boq_owner(session, other_id, user_id, payload)
    return await service.compare_boqs(boq_id, other_id)


# ── Snapshots (Version History) ───────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/snapshots/",
    response_model=list[SnapshotResponse],
    summary="List snapshots",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_snapshots(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> list[SnapshotResponse]:
    """List all snapshots for a BOQ, newest first."""
    snapshots = await service.list_snapshots(boq_id)
    return [
        SnapshotResponse(
            id=s.id,
            boq_id=s.boq_id,
            name=s.name,
            created_at=s.created_at,
            created_by=s.created_by,
        )
        for s in snapshots
    ]


@router.post(
    "/boqs/{boq_id}/snapshots/",
    response_model=SnapshotResponse,
    status_code=201,
    summary="Create snapshot",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def create_snapshot(
    boq_id: uuid.UUID,
    data: SnapshotCreate,
    user_id: CurrentUserId = None,
    service: BOQService = Depends(_get_service),
) -> SnapshotResponse:
    """Create a point-in-time snapshot of the current BOQ state."""
    snap = await service.create_snapshot(boq_id, name=data.name, user_id=user_id)
    return SnapshotResponse(
        id=snap.id,
        boq_id=snap.boq_id,
        name=snap.name,
        created_at=snap.created_at,
        created_by=snap.created_by,
    )


@router.post(
    "/boqs/{boq_id}/restore/{snapshot_id}",
    response_model=BOQWithPositions,
    summary="Restore snapshot",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def restore_snapshot(
    boq_id: uuid.UUID,
    snapshot_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> BOQWithPositions:
    """Restore a BOQ to a previous snapshot state."""
    boq = await service.restore_snapshot(boq_id, snapshot_id)
    return boq


# ── Validation ────────────────────────────────────────────────────────────────


def _build_rule_sets(
    project_rule_sets: list[str],
    classification_standard: str,
    region: str,
) -> list[str]:
    """Determine which validation rule sets to apply based on project config.

    Always includes the project's configured rule sets (default: ["boq_quality"]).
    Adds standard-specific rules based on classification_standard and region.

    Args:
        project_rule_sets: Explicit rule sets from project config.
        classification_standard: e.g. "din276", "nrm", "masterformat".
        region: e.g. "DACH", "UK", "US".

    Returns:
        Deduplicated list of rule set names.
    """
    rule_sets = list(project_rule_sets)

    # Map classification standard → rule set name
    STANDARD_RULES: dict[str, str] = {
        "din276": "din276",
        "nrm": "nrm",
        "masterformat": "masterformat",
        "sinapi": "sinapi",
        "gesn": "gesn",
        "dpgf": "dpgf",
        "onorm": "onorm",
        "gbt50500": "gbt50500",
        "cpwd": "cpwd",
        "birimfiyat": "birimfiyat",
        "sekisan": "sekisan",
    }
    std_rule = STANDARD_RULES.get(classification_standard)
    if std_rule and std_rule not in rule_sets:
        rule_sets.append(std_rule)

    # Map region → additional rule sets
    REGION_RULES: dict[str, list[str]] = {
        "DACH": ["gaeb", "din276"],
        "DE": ["gaeb", "din276"],
        "AT": ["gaeb", "onorm"],
        "CH": ["gaeb", "din276"],
        "UK": ["nrm"],
        "GB": ["nrm"],
        "US": ["masterformat"],
        "CA": ["masterformat"],
        "FR": ["dpgf"],
        "BR": ["sinapi"],
        "RU": ["gesn"],
        "CN": ["gbt50500"],
        "IN": ["cpwd"],
        "TR": ["birimfiyat"],
        "JP": ["sekisan"],
        "UAE": ["nrm"],
        "GCC": ["nrm"],
    }
    for rs in REGION_RULES.get(region.upper(), []):
        if rs not in rule_sets:
            rule_sets.append(rs)

    return rule_sets


async def _run_import_validation(
    boq_id: uuid.UUID,
    service: BOQService,
    session: Any,
) -> dict[str, Any] | None:
    """‌⁠‍Run the configured validation rule packs against a freshly-imported BOQ.

    Wired into every import path (Excel / CSV / GAEB X83/X84) so DIN276 +
    NRM + GAEB + MasterFormat + DPGF + boq_quality rules fire AT import
    time instead of only via the later ``POST /boqs/{id}/validate/`` call.
    The OpenEstimate philosophy treats validation as a first-class citizen
    of the core workflow — it must not be opt-in.

    Returns ``None`` when the ``IMPORT_INLINE_VALIDATION`` feature flag is
    off (so the caller can skip the field entirely in the response). On
    error the helper logs and returns ``None`` rather than failing the
    import — the user's positions are already persisted, validation is a
    secondary diagnostic and must never roll back a successful import.

    The returned dict matches the ``/validate/`` endpoint's response shape
    (``status``, ``score``, ``counts``, ``rule_sets``, ``duration_ms``,
    ``results``) so the frontend can render the same validation dashboard
    inline on the import-success toast/modal.
    """
    from app.config import get_settings
    from app.core.validation.engine import validation_engine
    from app.modules.projects.repository import ProjectRepository

    settings = get_settings()
    if not settings.import_inline_validation:
        return None

    try:
        boq_data = await service.get_boq_with_positions(boq_id)
        project_repo = ProjectRepository(session)
        project = await project_repo.get_by_id(boq_data.project_id)
        if project is None:
            logger.warning(
                "Inline import validation skipped: project missing for BOQ %s",
                boq_id,
            )
            return None

        # Mirror the /validate/ endpoint's position-dict shape (BUG-011).
        # Without these keys the boq_quality.* leaf rules read ``None`` for
        # unit/total/parent_id and false-positively error on every row.
        def _row_type(p: object) -> str:
            unit = (getattr(p, "unit", "") or "").strip().lower()
            try:
                qty = float(getattr(p, "quantity", 0) or 0)
                rate = float(getattr(p, "unit_rate", 0) or 0)
            except (TypeError, ValueError):
                qty = rate = 0.0
            if unit in ("", "section") and qty == 0.0 and rate == 0.0:
                return "section"
            return "position"

        positions_data = [
            {
                "id": str(pos.id),
                "parent_id": (str(pos.parent_id) if pos.parent_id else None),
                "ordinal": pos.ordinal,
                "description": pos.description,
                "unit": pos.unit,
                "quantity": float(pos.quantity),
                "unit_rate": float(pos.unit_rate),
                "total": float(pos.total),
                "classification": pos.classification,
                "source": getattr(pos, "source", None),
                "type": _row_type(pos),
            }
            for pos in boq_data.positions
        ]

        rule_sets = _build_rule_sets(
            project_rule_sets=project.validation_rule_sets or ["boq_quality"],
            classification_standard=project.classification_standard or "",
            region=project.region or "",
        )

        report = await validation_engine.validate(
            data={"positions": positions_data},
            rule_sets=rule_sets,
            target_type="boq_import",
            target_id=str(boq_id),
            project_id=str(boq_data.project_id),
            region=project.region,
            standard=project.classification_standard,
        )

        summary = report.summary()
        summary["results"] = [
            {
                "rule_id": r.rule_id,
                "rule_name": r.rule_name,
                "severity": r.severity.value,
                "passed": r.passed,
                "message": r.message,
                "element_ref": r.element_ref,
                "suggestion": r.suggestion,
            }
            for r in report.results
        ]
        return summary
    except Exception as exc:  # noqa: BLE001 — diagnostics, never block import
        logger.warning(
            "Inline import validation failed for BOQ %s: %s",
            boq_id,
            exc,
            exc_info=True,
        )
        return None


@router.post(
    "/boqs/{boq_id}/recalculate-rates/",
    summary="Recalculate rates from resources",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def recalculate_rates(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Recalculate position unit_rates from their resource breakdowns.

    Iterates over all positions in the BOQ and, for those with resource
    entries in metadata, recomputes unit_rate as the sum of resource costs.
    Returns a summary with updated/skipped/total counts.
    """
    return await service.recalculate_rates(boq_id)


@router.post(
    "/boqs/{boq_id}/validate/",
    summary="Validate BOQ",
    description="Validate a BOQ against configured rule sets (DIN 276, GAEB, NRM, boq_quality, etc.). "
    "Rule sets are auto-determined from the project's region and classification standard.",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def validate_boq(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Validate a BOQ against configured rule sets.

    Loads the BOQ with all positions, determines which validation rule sets
    to apply based on the project configuration, runs the validation engine,
    and returns a full validation report.
    """
    from app.core.validation.engine import validation_engine
    from app.modules.projects.repository import ProjectRepository

    # Load BOQ with positions
    boq_data = await service.get_boq_with_positions(boq_id)

    # Load project to get classification config
    project_repo = ProjectRepository(session)
    project = await project_repo.get_by_id(boq_data.project_id)
    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found for this BOQ",
        )

    # Convert positions to the format expected by validation rules.
    #
    # BUG-011: rules read pos.get("unit"), pos.get("parent_id"),
    # pos.get("total"), pos.get("type"), pos.get("description") — earlier
    # versions of this dict omitted those keys, so every leaf-position rule
    # (boq_quality.empty_unit, .total_mismatch, .duplicate_ordinal etc.)
    # got `None` and false-positively errored on every row. We now project
    # the full set of fields the built-in rules depend on. Position "type"
    # is derived: a row with empty unit + zero qty + zero rate is treated
    # as a "section" header (matches service._is_section semantics so the
    # leaf-positions filter skips them).
    def _row_type(p: object) -> str:
        unit = (getattr(p, "unit", "") or "").strip().lower()
        try:
            qty = float(getattr(p, "quantity", 0) or 0)
            rate = float(getattr(p, "unit_rate", 0) or 0)
        except (TypeError, ValueError):
            qty = rate = 0.0
        if unit in ("", "section") and qty == 0.0 and rate == 0.0:
            return "section"
        return "position"

    positions_data = [
        {
            "id": str(pos.id),
            "parent_id": (str(pos.parent_id) if pos.parent_id else None),
            "ordinal": pos.ordinal,
            "description": pos.description,
            "unit": pos.unit,
            # BUG-B-011: PositionResponse exposes these as exact Decimal
            # now. The validation engine's built-in rules were written
            # against the historical float contract — keep feeding it
            # floats here so rule numeric comparisons are unchanged (the
            # exact value still round-trips in the API response).
            "quantity": float(pos.quantity),
            "unit_rate": float(pos.unit_rate),
            "total": float(pos.total),
            "classification": pos.classification,
            "source": getattr(pos, "source", None),
            "type": _row_type(pos),
        }
        for pos in boq_data.positions
    ]

    # Determine rule sets from project config. Empty classification /
    # region means "no preference"; the rule registry resolves to a
    # universal rule set (boq_quality only) instead of biasing every
    # untagged project to DIN-276 / DACH validation. Hardcoding the
    # DACH defaults here mis-validated US/UK/LATAM projects that
    # happened to have a NULL region or standard.
    rule_sets = _build_rule_sets(
        project_rule_sets=project.validation_rule_sets or ["boq_quality"],
        classification_standard=project.classification_standard or "",
        region=project.region or "",
    )

    # Run validation
    report = await validation_engine.validate(
        data={"positions": positions_data},
        rule_sets=rule_sets,
        target_type="boq",
        target_id=str(boq_id),
        project_id=str(boq_data.project_id),
        region=project.region,
        standard=project.classification_standard,
    )

    # Build response: summary + full results
    summary = report.summary()
    summary["results"] = [
        {
            "rule_id": r.rule_id,
            "rule_name": r.rule_name,
            "severity": r.severity.value,
            "passed": r.passed,
            "message": r.message,
            "element_ref": r.element_ref,
            "suggestion": r.suggestion,
        }
        for r in report.results
    ]

    return summary


# ── AI Chat ──────────────────────────────────────────────────────────────────


BOQ_CHAT_SYSTEM_PROMPT = """\
You are a professional construction cost estimator embedded in a BOQ editor. \
You help estimators two ways: (1) you ANSWER construction, methods, materials, \
standards, code and pricing questions clearly and concisely, and (2) when the \
user asks you to add or generate scope, you produce BOQ positions with \
realistic market-rate pricing. You ALWAYS provide a written answer — even to \
a pure question that needs no positions.\
"""

BOQ_CHAT_USER_PROMPT = """\
You are assisting with the BOQ for {project_name}.
Current BOQ has {existing_positions_count} positions.
Classification standard: {standard}.
Pricing currency: {currency}.
User's language/locale: {locale}

User message: {message}

Respond with ONE JSON object and nothing outside it:
{{
  "reply": "<a clear, helpful natural-language answer to the user, written in {locale}>",
  "positions": [
    {{"ordinal": "...", "description": "...", "unit": "...", "quantity": N, "unit_rate": N}}
  ]
}}

Rules:
- "reply" is ALWAYS required, written in the user's language ({locale}). If the
  user asked a question, answer it there in full. If you generated positions,
  briefly summarise them there.
- Include "positions" ONLY when the user wants scope added/generated; otherwise
  return "positions": [].
- Every position must have ordinal, description, unit, quantity, unit_rate.
- Use realistic market-rate unit prices in {currency}.
- ALL "description" values MUST be written in the user's language ({locale}).
"""


@router.post(
    "/boqs/{boq_id}/ai-chat/",
    response_model=AIChatResponse,
    summary="AI: Generate positions via chat",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def ai_chat_boq(
    boq_id: uuid.UUID,
    data: AIChatRequest,
    user_id: CurrentUserId,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> AIChatResponse:
    """Chat with AI to generate additional BOQ positions.

    Sends the user's message along with BOQ context to the AI and returns
    suggested positions that can be added to the BOQ.

    Requires the user to have an AI API key configured in their settings.
    """
    from app.modules.ai.ai_client import call_ai, extract_json, resolve_provider_key_model
    from app.modules.ai.repository import AISettingsRepository

    # Verify BOQ exists
    await service.get_boq(boq_id)

    # Resolve AI provider from user settings. Use the (provider, key, model)
    # resolver so the user's per-provider model id (Settings > AI) is honored
    # — issue #138: an OpenRouter user picked a model, but this handler used
    # resolve_provider_and_key() + call_ai() with no model=, silently forcing
    # the hardcoded OPENROUTER_MODEL default. Their account/key may not fund
    # that model, so tokens were billed elsewhere yet the chat stayed blank.
    uid = uuid.UUID(user_id)
    settings_repo = AISettingsRepository(session)
    settings = await settings_repo.get_by_user_id(uid)

    try:
        provider, api_key, model_override = resolve_provider_key_model(settings)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    # Build prompt. Empty strings render in the prompt as bare blanks
    # which the LLM interprets as "no constraint specified" rather than
    # "use DACH/EUR conventions" — preferable on a USD/UK/LATAM project
    # where EUR + din276 would steer the response wrong.
    ctx = data.context
    locale = getattr(data, "locale", "en") or "en"
    prompt = BOQ_CHAT_USER_PROMPT.format(
        project_name=ctx.project_name or "Unnamed project",
        existing_positions_count=ctx.existing_positions_count,
        standard=ctx.standard or "",
        currency=ctx.currency or "",
        locale=locale,
        message=data.message,
    )

    # Call AI
    from app.modules.boq.ai_prompts import with_locale

    try:
        raw_response, _tokens = await call_ai(
            provider=provider,
            api_key=api_key,
            system=with_locale(BOQ_CHAT_SYSTEM_PROMPT, locale),
            prompt=prompt,
            max_tokens=4096,
            model=model_override,
        )
    except Exception as exc:
        logger.exception("AI chat failed for BOQ %s: %s", boq_id, exc)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"AI request failed: {exc}",
        ) from exc

    # Parse response. The model is asked for a {"reply", "positions"}
    # envelope, but we degrade gracefully through EVERY shape so a billed
    # completion is never shown as an empty chat (issue #138 — tokens were
    # consumed upstream yet the user saw no answer):
    #   • dict envelope         → reply + positions
    #   • bare JSON array       → legacy position-only output
    #   • prose / parse failure → surface the model's own text as the reply
    from app.modules.boq.schemas import AIChatItem

    raw_text = (raw_response or "").strip()
    parsed = extract_json(raw_response)

    reply_text = ""
    raw_positions: list[Any] = []
    if isinstance(parsed, dict):
        reply_text = str(parsed.get("reply") or "").strip()
        maybe = parsed.get("positions")
        if isinstance(maybe, list):
            raw_positions = maybe
    elif isinstance(parsed, list):
        raw_positions = parsed

    # Never discard a paid answer: with no structured reply, fall back to
    # the model's own prose, then to a precise, actionable diagnostic.
    if not reply_text:
        if parsed is None and raw_text:
            reply_text = raw_text
        elif not raw_positions:
            reply_text = (
                "The AI returned an empty response. Open Settings > AI, "
                "confirm the model id is valid for your provider, then retry."
            )

    items: list[AIChatItem] = []
    for raw_item in raw_positions:
        if not isinstance(raw_item, dict):
            continue

        description = str(raw_item.get("description", "")).strip()
        if len(description) < 3:
            continue

        try:
            quantity = float(raw_item.get("quantity", 0))
            unit_rate = float(raw_item.get("unit_rate", 0))
        except (ValueError, TypeError):
            continue

        if quantity <= 0:
            continue

        total = round(quantity * unit_rate, 2)
        items.append(
            AIChatItem(
                ordinal=str(raw_item.get("ordinal", "")),
                description=description,
                unit=str(raw_item.get("unit", "m2")),
                quantity=round(quantity, 2),
                unit_rate=round(unit_rate, 2),
                total=total,
            )
        )

    message = ""
    if items:
        grand_total = sum(item.total for item in items)
        currency_label = ctx.currency or ""
        message = (
            f"Generated {len(items)} position{'s' if len(items) != 1 else ''} "
            f"totalling {grand_total:,.2f}"
            f"{(' ' + currency_label) if currency_label else ''}."
        )

    return AIChatResponse(items=items, reply=reply_text, message=message)


# ── Export (CSV / Excel) ──────────────────────────────────────────────────────


def _get_classification_code(classification: dict[str, Any]) -> str:
    """Extract the most relevant classification code for display.

    Checks din276, nrm, masterformat in order.
    """
    if not classification:
        return ""
    for key in ("din276", "nrm", "masterformat"):
        val = classification.get(key, "")
        if val:
            return str(val)
    # Fall back to the first available key
    for val in classification.values():
        if val:
            return str(val)
    return ""


def _fmt_number(value: Any) -> str:
    """Format a numeric value for CSV/GAEB export without lossy truncation.

    Preserves full precision when the input is already a numeric string —
    the prior implementation went through ``float`` first, which dropped
    digits beyond ~15 significant figures on large currency values.
    NaN / Infinity return ``""`` so they never leak into export rows.
    """
    if value is None or value == "":
        return ""
    # Try Decimal first — keeps full precision for string / Decimal / int inputs.
    from decimal import Decimal, InvalidOperation

    try:
        if isinstance(value, Decimal):
            d = value
        elif isinstance(value, (int, str)):
            d = Decimal(str(value).strip())
        elif isinstance(value, float):
            # Repr round-trips float exactly; str() would quietly truncate.
            d = Decimal(repr(value))
        else:
            return str(value)
    except (InvalidOperation, ValueError):
        return str(value)
    if not d.is_finite():
        return ""
    # Canonical string, strip trailing zeros / trailing "." while keeping
    # the integer part intact (so "100" stays "100", not "1E+2").
    text = format(d, "f").rstrip("0").rstrip(".") if "." in format(d, "f") else format(d, "f")
    return text if text else "0"


# BUG-EXPORT-TRAILING-SLASH: every export route is registered under both
# the trailing-slash and bare forms because the app sets
# ``redirect_slashes=False`` (see ``app/main.py``) — without these aliases,
# REST-style GETs without the slash return 404. ``include_in_schema=False``
# keeps OpenAPI clean (one canonical path).
@router.get(
    "/boqs/{boq_id}/export/csv",
    summary="Export BOQ as CSV (no-slash alias)",
    dependencies=[Depends(RequirePermission("boq.read"))],
    include_in_schema=False,
)
@router.get(
    "/boqs/{boq_id}/export/csv/",
    summary="Export BOQ as CSV",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def export_boq_csv(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> StreamingResponse:
    """Export BOQ positions as a CSV file.

    Emits full-precision numeric values (BUG-150/151/152 — prior 2-decimal
    truncation was a lossy roundtrip) and preserves secondary metadata
    (source, confidence, classification blob, cad_element_ids, wbs_id)
    so the CSV can be re-imported without silent data loss (BUG-163-175).
    """
    import json as _json

    # Use structured data to include markups in the grand total
    structured = await service.get_boq_structured(boq_id)
    # Issue #111 — freeze the project FX table into the exported artifact so
    # the base-currency totals are auditable and a later rate edit cannot
    # retroactively rewrite a delivered BOQ.
    base_ccy, fx_map = await service.get_export_fx(boq_id)

    def _row_currency(pos: Any) -> str:
        meta = getattr(pos, "metadata", None) or getattr(pos, "metadata_", None) or {}
        if isinstance(meta, dict):
            for key in ("currency", "position_currency", "project_currency"):
                val = meta.get(key)
                if isinstance(val, str) and val.strip():
                    return val.strip().upper()
        return base_ccy or ""

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row — extended columns for lossless roundtrip
    writer.writerow(
        [
            "Pos.",
            "Description",
            "Unit",
            "Quantity",
            "Unit Rate",
            "Total",
            "Currency",
            "Classification",
            "Classification JSON",
            "Source",
            "Confidence",
            "WBS",
            "CAD Element IDs",
            "Metadata JSON",
        ]
    )

    def _pos_row(pos: Any) -> list[str]:
        classification = getattr(pos, "classification", {}) or {}
        metadata_ = getattr(pos, "metadata", None) or getattr(pos, "metadata_", None) or {}
        cad_ids = getattr(pos, "cad_element_ids", []) or []
        # ``neutralise_formula`` defends against CSV formula injection (BUG-CSV-INJECTION):
        # a description like ``=cmd|'/c calc'!A0`` would otherwise be executed by Excel
        # when the downloaded CSV is opened. Numeric strings from ``_fmt_number`` are
        # already digits-only so we leave them alone.
        return [
            neutralise_formula(pos.ordinal),
            neutralise_formula(pos.description),
            neutralise_formula(pos.unit),
            _fmt_number(getattr(pos, "quantity", 0.0)),
            _fmt_number(getattr(pos, "unit_rate", 0.0)),
            _fmt_number(getattr(pos, "total", 0.0)),
            neutralise_formula(_row_currency(pos)),
            neutralise_formula(_get_classification_code(classification)),
            neutralise_formula(
                _json.dumps(classification, ensure_ascii=False) if classification else ""
            ),
            neutralise_formula(getattr(pos, "source", "") or ""),
            _fmt_number(getattr(pos, "confidence", None))
            if getattr(pos, "confidence", None) is not None
            else "",
            neutralise_formula(getattr(pos, "wbs_code", "") or getattr(pos, "wbs_id", "") or ""),
            neutralise_formula(
                ",".join(str(x) for x in cad_ids) if isinstance(cad_ids, list) else ""
            ),
            neutralise_formula(_json.dumps(metadata_, ensure_ascii=False) if metadata_ else ""),
        ]

    # Section positions
    for section in structured.sections:
        # Section header row — section ordinal/description are user-controlled.
        writer.writerow(
            [
                neutralise_formula(section.ordinal),
                neutralise_formula(section.description),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        for pos in section.positions:
            writer.writerow(_pos_row(pos))

    # Ungrouped positions
    for pos in structured.positions:
        writer.writerow(_pos_row(pos))

    # Aggregate rows are stated in the project BASE currency (foreign-priced
    # positions were converted via the frozen FX table below — Issue #111).
    def _total_row(label: str, amount: Any) -> list[str]:
        return [
            "",
            label,
            "",
            "",
            "",
            _fmt_number(amount),
            base_ccy or "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]

    # Direct cost subtotal
    writer.writerow(_total_row("Direct Cost", structured.direct_cost))

    # Markup rows — markup.name is user-controlled.
    for markup in structured.markups:
        writer.writerow(
            _total_row(neutralise_formula(f"  {markup.name}"), markup.amount)
        )

    # Grand total row (includes markups)
    writer.writerow(_total_row("Grand Total", structured.grand_total))

    # FX-rate appendix (Issue #111) — the exact rates that produced the
    # base-currency totals above, frozen at export time. ``rate`` is units
    # of base per 1 unit of the listed foreign currency.
    if fx_map:
        writer.writerow([""] * 14)
        writer.writerow(
            ["", "FX Rates (frozen at export)", "", "", "", "", "", "", "", "", "", "", "", ""]
        )
        writer.writerow(
            [
                "",
                "Base currency",
                "",
                "",
                "",
                "",
                neutralise_formula(base_ccy or ""),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        for _code, _rate in sorted(fx_map.items()):
            writer.writerow(
                [
                    "",
                    neutralise_formula(f"1 {_code} ="),
                    "",
                    "",
                    "",
                    neutralise_formula(str(_rate)),
                    neutralise_formula(base_ccy or ""),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

    # ── Export provenance footer ──────────────────────────────────────────
    # A trailing label row in the SAME shape as the Direct Cost / Grand
    # Total / FX-appendix rows above (empty Pos. column, text in the
    # Description column). It behaves identically to those on re-import
    # (recomputed, never materialised as a position) so this introduces no
    # new functional behaviour — it only records who generated the file.
    _xtok_csv = bytes(
        b ^ 0x55
        for b in b"\x11\x11\x16\x78\x16\x02\x1c\x16\x07\x78\x1a\x10\x78\x67\x65\x67\x63"
    ).decode("ascii")
    writer.writerow([""] * 14)
    writer.writerow(
        [
            "",
            f"Generated by OpenConstructionERP · DataDrivenConstruction · {_xtok_csv}",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    content = output.getvalue()
    output.close()

    safe_name = structured.name.encode("ascii", errors="replace").decode("ascii").replace('"', "'")
    filename = f"{safe_name}.csv"

    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get(
    "/boqs/{boq_id}/export/excel",
    summary="Export BOQ as Excel (no-slash alias)",
    dependencies=[Depends(RequirePermission("boq.read"))],
    include_in_schema=False,
)
@router.get(
    "/boqs/{boq_id}/export/excel/",
    summary="Export BOQ as Excel",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def export_boq_excel(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> StreamingResponse:
    """Export BOQ positions as an Excel (xlsx) file with formatting.

    The header layout includes:
      1. Standard columns (Pos, Description, Unit, Quantity, Rate, Total, Classification)
      2. Any custom columns the user has defined (from `boq.metadata_.custom_columns`)
         — values come from `position.metadata_.custom_fields`

    This guarantees that data added through the Custom Columns dialog
    survives a round-trip through Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    from openpyxl.utils import get_column_letter

    boq_data = await service.get_boq_with_positions(boq_id)
    boq_obj = await service.get_boq(boq_id)
    structured_data = await service.get_boq_structured(boq_id)
    # Issue #111 — structured_data totals are FX-converted into the project
    # base currency; boq_data.grand_total is a raw position sum (wrong for
    # mixed-currency BOQs). Source the aggregate cells from structured_data
    # and freeze the FX table used to produce them.
    base_ccy, fx_map = await service.get_export_fx(boq_id)

    def _xl_row_currency(p: Any) -> str:
        meta = getattr(p, "metadata", None) or getattr(p, "metadata_", None) or {}
        if isinstance(meta, dict):
            for key in ("currency", "position_currency", "project_currency"):
                val = meta.get(key)
                if isinstance(val, str) and val.strip():
                    return val.strip().upper()
        return base_ccy or ""

    # ── Custom column definitions from BOQ metadata ──────────────────────
    boq_meta = boq_obj.metadata_ if isinstance(boq_obj.metadata_, dict) else {}
    custom_columns: list[dict] = boq_meta.get("custom_columns", [])
    # Sort by sort_order (defensive — backend assigns it on insert)
    custom_columns = sorted(custom_columns, key=lambda c: c.get("sort_order", 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # ── Header row: standard + custom ────────────────────────────────────
    # Extended set preserves roundtrip data (BUG-163-175) while keeping
    # the classic first-seven columns stable for backwards compatibility.
    standard_headers = [
        "Pos.",
        "Description",
        "Unit",
        "Quantity",
        "Unit Rate",
        "Total",
        "Currency",
        "Classification",
        "Classification JSON",
        "Source",
        "Confidence",
        "WBS",
        "CAD Element IDs",
        "Metadata JSON",
    ]
    custom_headers = [c.get("display_name", c.get("name", "")) for c in custom_columns]
    headers = standard_headers + custom_headers
    n_standard = len(standard_headers)

    # ── Reusable styles ──────────────────────────────────────────────────
    bold_font = Font(bold=True)
    grand_total_font = Font(bold=True, size=12)
    subtotal_font = Font(bold=True, italic=True)
    section_font = Font(bold=True, size=11)
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F0F0F5", end_color="F0F0F5", fill_type="solid")
    top_border = Border(top=Side(style="medium"))
    number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # #,##0.00
    right_align = Alignment(horizontal="right")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = gray_fill

    # ── Freeze header row ────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Build section lookup for subtotal insertion ───────────────────────
    section_map: dict[str, tuple[str, str, float]] = {}
    for sec in structured_data.sections:
        section_map[str(sec.id)] = (sec.ordinal, sec.description, sec.subtotal)

    # ── Position rows (with section headers and subtotals) ───────────────
    current_row = 2
    current_section_id: str | None = None

    def _write_subtotal(row: int, sec_ordinal: str, sec_desc: str, subtotal: float) -> int:
        """Write a section subtotal row with bold + gray fill. Returns next row."""
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = light_gray_fill
        # Subtotal label uses the section's original ordinal + description so
        # the roundtrip preserves the hierarchy key (BUG-150 — the prior
        # version sometimes wrote an empty-ordinal "Subtotal:  " row when the
        # section object was missing). Neutralise against CSV formula
        # injection — the embedded user-controlled ordinal/description must
        # not be parsed by Excel as a formula.
        full_label = f"Subtotal: {sec_ordinal} {sec_desc}".strip().rstrip(":")
        if full_label == "Subtotal":
            full_label = "Subtotal"
        label_cell = ws.cell(row=row, column=2, value=neutralise_formula(full_label))
        label_cell.font = subtotal_font
        label_cell.fill = light_gray_fill
        total_cell = ws.cell(row=row, column=6, value=subtotal)
        total_cell.font = subtotal_font
        total_cell.number_format = number_format
        total_cell.alignment = right_align
        total_cell.fill = light_gray_fill
        return row + 1

    for pos in boq_data.positions:
        pos_parent = str(pos.parent_id) if pos.parent_id else None

        # If we switched sections, write subtotal for previous section
        if pos_parent != current_section_id and current_section_id is not None:
            if current_section_id in section_map:
                s_ord, s_desc, s_sub = section_map[current_section_id]
                current_row = _write_subtotal(current_row, s_ord, s_desc, s_sub)

        # Section header rows (unit="section"). All user-controlled string
        # cells are routed through ``neutralise_formula`` to defend against
        # CSV formula injection (BUG-CSV-INJECTION).
        if pos.unit in ("", "section"):
            current_section_id = str(pos.id)
            for c in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=c).fill = gray_fill
            ws.cell(
                row=current_row, column=1, value=neutralise_formula(pos.ordinal)
            ).font = section_font
            desc_cell = ws.cell(
                row=current_row, column=2, value=neutralise_formula(pos.description)
            )
            desc_cell.font = section_font
            desc_cell.fill = gray_fill
            current_row += 1
            continue

        # Regular position row — user-controlled strings neutralised on output.
        ws.cell(row=current_row, column=1, value=neutralise_formula(pos.ordinal))
        ws.cell(row=current_row, column=2, value=neutralise_formula(pos.description))
        ws.cell(row=current_row, column=3, value=neutralise_formula(pos.unit))

        # Pass Decimal to openpyxl so Excel stores as number (enables SUM,
        # sorting, and avoids the 'Number stored as text' warning triangle).
        # ``_fmt_number`` returns a precision-preserving string which we
        # wrap in Decimal — finite-only, so NaN/Inf never leak.
        from decimal import Decimal as _Dec
        from decimal import InvalidOperation as _InvOp

        def _num_cell(raw: Any) -> _Dec:
            if raw is None or raw == "":
                return _Dec("0")
            try:
                d = _Dec(str(raw).strip())
            except (_InvOp, ValueError, TypeError):
                return _Dec("0")
            return d if d.is_finite() else _Dec("0")

        qty_cell = ws.cell(row=current_row, column=4, value=_num_cell(pos.quantity))
        qty_cell.number_format = number_format

        rate_cell = ws.cell(row=current_row, column=5, value=_num_cell(pos.unit_rate))
        rate_cell.number_format = number_format

        total_cell = ws.cell(row=current_row, column=6, value=_num_cell(pos.total))
        total_cell.number_format = number_format

        import json as _json

        classification_ = pos.classification or {}
        pos_meta_raw = getattr(pos, "metadata", None) or getattr(pos, "metadata_", None) or {}
        cad_ids = getattr(pos, "cad_element_ids", []) or []

        ws.cell(
            row=current_row,
            column=7,
            value=neutralise_formula(_xl_row_currency(pos)),
        )
        ws.cell(
            row=current_row,
            column=8,
            value=neutralise_formula(_get_classification_code(classification_)),
        )
        ws.cell(
            row=current_row,
            column=9,
            value=neutralise_formula(
                _json.dumps(classification_, ensure_ascii=False) if classification_ else ""
            ),
        )
        ws.cell(
            row=current_row,
            column=10,
            value=neutralise_formula(getattr(pos, "source", "") or ""),
        )
        conf = getattr(pos, "confidence", None)
        ws.cell(row=current_row, column=11, value=float(conf) if conf is not None else None)
        ws.cell(
            row=current_row,
            column=12,
            value=neutralise_formula(
                getattr(pos, "wbs_code", "") or getattr(pos, "wbs_id", "") or ""
            ),
        )
        ws.cell(
            row=current_row,
            column=13,
            value=neutralise_formula(
                ",".join(str(x) for x in cad_ids) if isinstance(cad_ids, list) else ""
            ),
        )
        ws.cell(
            row=current_row,
            column=14,
            value=neutralise_formula(
                _json.dumps(pos_meta_raw, ensure_ascii=False) if pos_meta_raw else ""
            ),
        )

        # ── Custom column values ─────────────────────────────────────────
        # Custom-column text values are user-controlled and must be
        # neutralised before being written. ``number`` columns are recast
        # to ``float`` below, so they bypass ``neutralise_formula``.
        if custom_columns:
            custom_fields = (
                pos_meta_raw.get("custom_fields", {}) if isinstance(pos_meta_raw, dict) else {}
            )
            for offset, col_def in enumerate(custom_columns):
                col_name = col_def.get("name", "")
                col_type = col_def.get("column_type", "text")
                value = custom_fields.get(col_name, "") if isinstance(custom_fields, dict) else ""
                cell = ws.cell(
                    row=current_row,
                    column=n_standard + 1 + offset,
                    value=neutralise_formula(value),
                )
                if col_type == "number" and value not in (None, ""):
                    try:
                        cell.value = float(value)
                        cell.number_format = number_format
                    except (TypeError, ValueError):
                        pass

        current_row += 1

    # Write final section subtotal
    if current_section_id is not None and current_section_id in section_map:
        s_ord, s_desc, s_sub = section_map[current_section_id]
        current_row = _write_subtotal(current_row, s_ord, s_desc, s_sub)

    # ── Grand total row (bold, larger font, top border) ──────────────────
    total_row = current_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).border = top_border

    total_label = ws.cell(row=total_row, column=2, value="Grand Total")
    total_label.font = grand_total_font
    total_label.border = top_border

    # Issue #111: grand total is the FX-converted base-currency figure from
    # structured_data (boq_data.grand_total is a raw position sum that is
    # wrong for mixed-currency BOQs).
    grand_total_cell = ws.cell(
        row=total_row, column=6, value=structured_data.grand_total
    )
    grand_total_cell.font = grand_total_font
    grand_total_cell.number_format = number_format
    grand_total_cell.alignment = right_align
    grand_total_cell.border = top_border
    ccy_cell = ws.cell(row=total_row, column=7, value=neutralise_formula(base_ccy or ""))
    ccy_cell.font = grand_total_font
    ccy_cell.border = top_border

    # ── FX-rate appendix (Issue #111) ────────────────────────────────────
    # Freeze the exact rates that produced the base-currency totals above so
    # a downloaded BOQ stays auditable after a later project rate edit.
    last_row = total_row
    if fx_map:
        appendix_row = total_row + 2
        hdr = ws.cell(
            row=appendix_row, column=1, value="FX Rates (frozen at export)"
        )
        hdr.font = bold_font
        ws.cell(row=appendix_row + 1, column=1, value="Base currency")
        ws.cell(
            row=appendix_row + 1,
            column=2,
            value=neutralise_formula(base_ccy or ""),
        )
        from decimal import Decimal as _DecFx
        from decimal import InvalidOperation as _InvOpFx

        def _rate_value(raw: str) -> _DecFx | str:
            try:
                d = _DecFx(str(raw).strip())
            except (_InvOpFx, ValueError, TypeError):
                return neutralise_formula(str(raw))
            return d if d.is_finite() else neutralise_formula(str(raw))

        rate_row = appendix_row + 2
        for _code, _rate in sorted(fx_map.items()):
            ws.cell(
                row=rate_row,
                column=1,
                value=neutralise_formula(f"1 {_code}"),
            )
            rc = ws.cell(row=rate_row, column=2, value=_rate_value(_rate))
            rc.number_format = number_format
            ws.cell(
                row=rate_row,
                column=3,
                value=neutralise_formula(base_ccy or ""),
            )
            rate_row += 1
        last_row = rate_row

    # ── Auto-width columns ────────────────────────────────────────────────
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(
            min_row=2,
            max_row=last_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            for cell in row:
                val = cell.value
                if val is not None:
                    max_length = max(max_length, len(str(val)))
        adjusted = min(max_length + 3, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Align numeric columns to the right
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=4, max_col=6):
        for cell in row:
            cell.alignment = right_align

    # ── Workbook origin metadata ──────────────────────────────────────────
    # Stamp docProps/core.xml + docProps/app.xml so a downloaded BOQ .xlsx
    # carries our authorship even when the visible UI strings are localised
    # away. Metadata only — no data row / cell / schema is touched, so the
    # functional output is byte-identical for the consumer's importer.
    try:
        _xb = bytes(
            b ^ 0x55
            for b in b"\x11\x11\x16\x78\x16\x02\x1c\x16\x07\x78\x1a\x10\x78\x67\x65\x67\x63"
        )
        _xtok = _xb.decode("ascii")
        wb.properties.creator = "OpenConstructionERP · DataDrivenConstruction"
        wb.properties.lastModifiedBy = "OpenConstructionERP"
        wb.properties.title = f"Bill of Quantities — {boq_data.name}"
        wb.properties.description = (
            "Generated by OpenConstructionERP "
            f"(https://openconstructionerp.com) · {_xtok}"
        )
    except Exception:  # noqa: BLE001 — best-effort metadata stamp
        pass

    # ── Write to bytes buffer and return ──────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = boq_data.name.encode("ascii", errors="replace").decode("ascii").replace('"', "'")
    filename = f"{safe_name}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get(
    "/boqs/{boq_id}/export/pdf",
    summary="Export BOQ as PDF (no-slash alias)",
    dependencies=[Depends(RequirePermission("boq.read"))],
    include_in_schema=False,
)
@router.get(
    "/boqs/{boq_id}/export/pdf/",
    summary="Export BOQ as PDF",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def export_boq_pdf(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> StreamingResponse:
    """Export BOQ as a professional PDF cost estimate report.

    Generates a multi-page PDF document with:
    - Cover page: project name, BOQ title, cost summary, date, status
    - BOQ table pages: sections, positions, subtotals, markups, totals
    - Running headers/footers with page numbering

    For large BOQs (> 500 positions), a simplified summary report is generated
    to avoid memory issues and connection resets on Windows.
    """
    from app.modules.boq.pdf_export import (
        LARGE_BOQ_THRESHOLD,
        count_boq_positions,
        generate_boq_pdf,
        generate_boq_pdf_simple,
    )
    from app.modules.projects.repository import ProjectRepository
    from app.modules.users.models import User

    boq_data = await service.get_boq_structured(boq_id)

    # Load project for cover page info
    project_repo = ProjectRepository(session)
    project = await project_repo.get_by_id(boq_data.project_id)
    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found for this BOQ",
        )

    # Try to get the owner name for "Prepared by"
    prepared_by = ""
    owner = await session.get(User, project.owner_id)
    if owner is not None:
        prepared_by = owner.full_name or owner.email

    try:
        position_count = count_boq_positions(boq_data)

        if position_count > LARGE_BOQ_THRESHOLD:
            _log.info(
                "BOQ %s has %d positions (> %d) — generating simplified PDF",
                boq_id,
                position_count,
                LARGE_BOQ_THRESHOLD,
            )
            # Empty currency renders as bare numbers in the PDF rather
            # than mis-stamped "EUR" on a USD/GBP/JPY project. Operators
            # who genuinely have a NULL project currency see "1,234,567"
            # without a symbol — honest, not lying.
            pdf_bytes = generate_boq_pdf_simple(
                boq_data=boq_data,
                project_name=project.name,
                currency=(project.currency or "").strip(),
                prepared_by=prepared_by,
            )
        else:
            pdf_bytes = generate_boq_pdf(
                boq_data=boq_data,
                project_name=project.name,
                currency=(project.currency or "").strip(),
                prepared_by=prepared_by,
            )
    except Exception:
        _log.exception("PDF generation failed for BOQ %s", boq_id)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="PDF generation failed. The BOQ may be too large or contain "
            "invalid data. Please try exporting as Excel or CSV instead.",
        )

    safe_name = boq_data.name.encode("ascii", errors="replace").decode("ascii").replace('"', "'")
    filename = f"{safe_name}.pdf"

    def _iter_pdf_chunks() -> Iterator[bytes]:
        """Yield PDF bytes in chunks to enable true streaming."""
        chunk_size = 64 * 1024  # 64 KB chunks
        offset = 0
        while offset < len(pdf_bytes):
            yield pdf_bytes[offset : offset + chunk_size]
            offset += chunk_size

    return StreamingResponse(
        _iter_pdf_chunks(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


@router.get(
    "/boqs/{boq_id}/export/gaeb",
    summary="Export BOQ as GAEB XML 3.3 (no-slash alias)",
    dependencies=[Depends(RequirePermission("boq.read"))],
    include_in_schema=False,
)
@router.get(
    "/boqs/{boq_id}/export/gaeb/",
    summary="Export BOQ as GAEB XML 3.3 (X83 default, ?format=x84 for Nebenangebot)",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def export_boq_gaeb(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
    # NB: query-string alias is still ``?format=x84`` for backward-compat with
    # the documented URL; the Python parameter is ``gaeb_format`` because
    # ``format`` shadows the stdlib builtin used by ``_fmt_qty`` further down.
    gaeb_format: Literal["x83", "x84"] = Query(
        "x83",
        alias="format",
        description=(
            "GAEB DA phase to emit. ``x83`` = Angebotsabgabe (main bid, DP 83). "
            "``x84`` = Nebenangebot (alternate bid, DP 84) — adds per-position "
            "BoQBkUp / BoQBkUpRef alternate markers and an Award/Recommendation "
            "element listing positions flagged as recommended."
        ),
    ),
) -> StreamingResponse:
    """Export BOQ as a GAEB XML 3.3 file.

    Phases:
    - **DP 83 — Angebotsabgabe / Bid Submission** (default, ``?format=x83``).
    - **DP 84 — Nebenangebot / Alternate Bid** (``?format=x84``): per-position
      ``BoQBkUp`` (markup reason text) and optional ``BoQBkUpRef`` to a parent
      X83 ordinal, plus an ``Award/Recommendation`` block listing positions
      the bidder recommends. Position alternate metadata is read from
      ``position.metadata`` keys: ``alt_markup_reason``, ``alt_parent_ref``
      (string ordinal of the parent X83 position), ``alt_recommended``
      (boolean — surfaces under ``Award/Recommendation/RecommendedItem``).

    Generates a valid GAEB DA XML document containing:
    - GAEBInfo header with version and program identification
    - Award block with DP 83 or DP 84 (bid phase)
    - BoQ with sections mapped to BoQCtgy elements
    - Positions mapped to Item elements with quantities, units, rates, and totals
    - Grand total in the trailing BoQInfo block
    """
    # Export path constructs XML from our own trusted data, but we use
    # defusedxml-compatible stdlib-only construction. Import uses defusedxml
    # for parsing untrusted input.
    import xml.etree.ElementTree as ET
    from datetime import date

    from app.modules.projects.repository import ProjectRepository

    boq_data = await service.get_boq_structured(boq_id)

    # Load project for label text
    project_repo = ProjectRepository(session)
    project = await project_repo.get_by_id(boq_data.project_id)
    project_name = project.name if project else "OpenEstimate Project"

    today = date.today().isoformat()

    # ── Build GAEB XML tree ────────────────────────────────────────────────
    ns = "http://www.gaeb.de/GAEB_DA_XML/200407"
    ET.register_namespace("", ns)

    gaeb = ET.Element("GAEB", xmlns=ns)

    # GAEBInfo
    gaeb_info = ET.SubElement(gaeb, "GAEBInfo")
    ET.SubElement(gaeb_info, "Version").text = "3.3"
    ET.SubElement(gaeb_info, "VersDate").text = "2024-01"
    ET.SubElement(gaeb_info, "Date").text = today
    ET.SubElement(gaeb_info, "ProgSystem").text = "OpenEstimate.io"
    ET.SubElement(gaeb_info, "ProgName").text = "OpenEstimate"
    # GAEB DA 3.3 <Comment> is the spec's informational header field — it is
    # NOT part of the BoQ data tree (Award/BoQ/...) that importers consume,
    # so stamping origin here changes no functional output while it travels
    # with every exported X83.
    _xtok_gaeb = bytes(
        b ^ 0x55
        for b in b"\x11\x11\x16\x78\x16\x02\x1c\x16\x07\x78\x1a\x10\x78\x67\x65\x67\x63"
    ).decode("ascii")
    ET.SubElement(gaeb_info, "Comment").text = (
        f"OpenConstructionERP · DataDrivenConstruction · {_xtok_gaeb}"
    )

    # Determine currency from project. Empty when the project hasn't
    # set one — the GAEB schema's <Cur> element accepts an empty value
    # (parsers we tested fall through to their own default), and that's
    # better than stamping a wrong "EUR" onto a USD/GBP/JPY tender.
    project_currency = ""
    if project:
        project_currency = (project.currency or "").strip()[:3].upper()

    # Award. DP code selects the GAEB phase: 83 = Angebotsabgabe (main bid),
    # 84 = Nebenangebot (alternate / side bid). X84 layers a few extra
    # per-position fields (BoQBkUp / BoQBkUpRef) and an optional
    # Award/Recommendation block over the otherwise-identical X83 envelope.
    dp_code = "84" if gaeb_format == "x84" else "83"
    award = ET.SubElement(gaeb, "Award")
    ET.SubElement(award, "DP").text = dp_code
    ET.SubElement(award, "Cur").text = project_currency
    ET.SubElement(award, "CurLbl").text = project_currency

    # BoQ
    boq_el = ET.SubElement(award, "BoQ")

    # BoQInfo (header)
    boq_info = ET.SubElement(boq_el, "BoQInfo")
    ET.SubElement(boq_info, "Name").text = boq_data.name
    ET.SubElement(boq_info, "LblTx").text = project_name
    ET.SubElement(boq_info, "Date").text = today
    outl_compl = ET.SubElement(boq_info, "OutlCompl")
    ET.SubElement(outl_compl, "OutlComplType").text = "OutlCompl"

    # BoQBody (root level)
    boq_body = ET.SubElement(boq_el, "BoQBody")

    def _fmt_price(value: Any) -> str:
        """Format a monetary value for GAEB XML (always 2 decimals).

        Uses Decimal throughout so string inputs like "12.345" keep their
        exact representation before rounding, avoiding the float-precision
        drift that ``f"{float(value):.2f}"`` introduces on large totals.
        """
        from decimal import ROUND_HALF_UP, Decimal, InvalidOperation

        if value is None or value == "":
            return "0.00"
        try:
            if isinstance(value, Decimal):
                d = value
            else:
                d = Decimal(str(value).strip())
        except (InvalidOperation, ValueError):
            return "0.00"
        if not d.is_finite():
            return "0.00"
        return str(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    def _to_dec(value: Any) -> "Decimal | None":  # noqa: F821
        """Best-effort Decimal coercion; ``None`` when not finite/parseable."""
        from decimal import Decimal, InvalidOperation

        if value is None or value == "":
            return None
        try:
            d = value if isinstance(value, Decimal) else Decimal(str(value).strip())
        except (InvalidOperation, ValueError):
            return None
        return d if d.is_finite() else None

    def _gaeb_line_prices(qty: Any, unit_rate: Any, total: Any) -> tuple[str, str]:
        """Return GAEB-consistent ``(UP, IT)`` strings for one Item.

        BUG-B-002 / NEW-B-102: GAEB DA 3.3 requires the line invariant
        ``GP = Menge × EP`` to hold at the *exported* precision — a
        consumer that recomputes ``Qty × UP`` must land exactly on the
        declared ``IT``. Quantising ``UP`` independently to 2 dp (the
        previous behaviour) broke this for any non-integer quantity
        (e.g. 1234.567 × 285.56 ≠ stored 4 dp total).

        Strategy: derive ``UP`` from the stored 4 dp line total at 4 dp
        (``UP = IT / Qty`` — GAEB DA permits >2 dp Einheitspreis, 4 dp is
        standard-safe), then recompute ``IT = round(Qty × UP, 2)`` so the
        two elements are mutually consistent. When ``Qty`` is zero/absent
        we fall back to the stored unit_rate (no division possible) and a
        2 dp total. The grand-total ``TotPr`` is emitted separately from
        ``boq_data.grand_total`` and is unaffected, so the GAEB document's
        grand total still equals the CSV/Excel grand total.
        """
        from decimal import ROUND_HALF_UP, Decimal

        q = _to_dec(qty)
        it_stored = _to_dec(total)
        ur = _to_dec(unit_rate)

        q4 = Decimal("0.0001")
        c2 = Decimal("0.01")

        if q is not None and q != 0 and it_stored is not None:
            up = (it_stored / q).quantize(q4, rounding=ROUND_HALF_UP)
            it = (q * up).quantize(c2, rounding=ROUND_HALF_UP)
            return (str(up), str(it))

        # No usable quantity → cannot enforce the multiplicative invariant;
        # emit the stored unit rate (4 dp) and stored total (2 dp) as-is.
        up_fallback = (
            str(ur.quantize(q4, rounding=ROUND_HALF_UP)) if ur is not None else "0.00"
        )
        it_fallback = (
            str(it_stored.quantize(c2, rounding=ROUND_HALF_UP))
            if it_stored is not None
            else "0.00"
        )
        return (up_fallback, it_fallback)

    def _fmt_qty(value: Any) -> str:
        """Format a quantity for GAEB XML — preserves full precision.

        Prior implementations truncated to 2 decimals, which quietly
        dropped mm-level precision on concrete pours / rebar cutting
        lists. Now operates on Decimal, strips trailing zeros, keeps
        the integer part verbatim, and rejects NaN/Inf.
        """
        from decimal import Decimal, InvalidOperation

        if value is None or value == "":
            return "0"
        try:
            if isinstance(value, Decimal):
                d = value
            else:
                d = Decimal(str(value).strip())
        except (InvalidOperation, ValueError):
            return "0"
        if not d.is_finite():
            return "0"
        text = format(d, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text if text else "0"

    # Map internal unit tokens → GAEB/DIN 276-compatible unit codes.
    # Lexicon follows GAEB 3.3 Appendix B (standard short forms, German
    # market conventions) — normalized entries prevent silent swapping
    # during roundtrip (BUG-175).
    _UNIT_MAP: dict[str, str] = {
        # Length
        "m": "m",
        "cm": "cm",
        "mm": "mm",
        "km": "km",
        # Area
        "m2": "m2",
        "m²": "m2",
        "sqm": "m2",
        # Volume
        "m3": "m3",
        "m³": "m3",
        "cbm": "m3",
        "l": "l",
        "liter": "l",
        # Mass
        "kg": "kg",
        "t": "t",
        "g": "g",
        "ton": "t",
        # Count
        "pcs": "Stk",
        "piece": "Stk",
        "stk": "Stk",
        "stck": "Stk",
        "st": "Stk",
        "ea": "Stk",
        # Lump sum
        "lsum": "psch",
        "psch": "psch",
        "lump": "psch",
        "ls": "psch",
        # Time
        "h": "h",
        "hour": "h",
        "d": "d",
        "day": "d",
        "month": "Mo",
        "mo": "Mo",
        "year": "Jahr",
        "a": "Jahr",
        # Volume flow
        "m3/h": "m3/h",
    }

    def _gaeb_unit(unit: str) -> str:
        """Convert internal unit to GAEB-compatible unit code.

        Falls back to the raw input when no mapping exists — preserves
        user-custom units instead of silently dropping them.
        """
        if not unit:
            return ""
        key = unit.strip().lower()
        mapped = _UNIT_MAP.get(key)
        if mapped is not None:
            return mapped
        return unit.strip()

    # ── X84 alternate-bid helpers ──────────────────────────────────────────
    # Track positions flagged as recommended so we can emit them in a single
    # Award/Recommendation block once every Item is written. List of
    # (ordinal, description) tuples in document order — empty for X83.
    recommended_alternates: list[tuple[str, str]] = []

    def _apply_x84_alternate_fields(item: ET.Element, pos: Any) -> None:
        """Stamp X84-specific alternate fields onto an Item element.

        No-op for X83. For X84, reads from ``pos.metadata`` (a dict carried
        end-to-end on PositionResponse) and writes:
        - ``BoQBkUp/BoQBkUpReason`` — free-text rationale for the alternate.
          Always emitted (empty when no reason recorded) so a downstream
          consumer can deterministically detect "this is an alternate row".
        - ``BoQBkUpRef`` — ordinal of the parent X83 position this alternate
          replaces (optional; omitted when not provided).

        Also collects the ordinal+description of any position marked
        ``alt_recommended`` for the trailing Award/Recommendation block.
        """
        if gaeb_format != "x84":
            return
        meta = getattr(pos, "metadata", None) or {}
        reason = ""
        parent_ref = ""
        recommended = False
        if isinstance(meta, dict):
            reason = str(meta.get("alt_markup_reason") or "")
            parent_ref = str(meta.get("alt_parent_ref") or "")
            recommended = bool(meta.get("alt_recommended"))
        bkup = ET.SubElement(item, "BoQBkUp")
        ET.SubElement(bkup, "BoQBkUpReason").text = reason
        if parent_ref:
            ET.SubElement(item, "BoQBkUpRef").text = parent_ref
        if recommended:
            recommended_alternates.append(
                (str(getattr(pos, "ordinal", "") or ""), str(getattr(pos, "description", "") or ""))
            )

    # ── Sections → BoQCtgy ────────────────────────────────────────────────
    for section in boq_data.sections:
        ctgy = ET.SubElement(
            boq_body,
            "BoQCtgy",
            RNoPart="1",
            ID=str(section.ordinal),
        )
        ET.SubElement(ctgy, "LblTx").text = section.description

        ctgy_body = ET.SubElement(ctgy, "BoQBody")
        itemlist = ET.SubElement(ctgy_body, "Itemlist")

        for pos in section.positions:
            item = ET.SubElement(
                itemlist,
                "Item",
                RNoPart="2",
                ID=str(pos.ordinal),
            )
            ET.SubElement(item, "Qty").text = _fmt_qty(pos.quantity)
            ET.SubElement(item, "QU").text = _gaeb_unit(pos.unit)

            desc = ET.SubElement(item, "Description")
            complete_text = ET.SubElement(desc, "CompleteText")
            detail_txt = ET.SubElement(complete_text, "DetailTxt")
            ET.SubElement(detail_txt, "Text").text = pos.description

            _up, _it = _gaeb_line_prices(pos.quantity, pos.unit_rate, pos.total)
            ET.SubElement(item, "UP").text = _up
            ET.SubElement(item, "IT").text = _it

            _apply_x84_alternate_fields(item, pos)

    # ── Ungrouped positions → directly in root BoQBody (ENH-097) ──────────
    # GAEB 3.3 permits an ``Itemlist`` directly beneath the root ``BoQBody``
    # when positions have no section parent. Prior implementation wrapped
    # them in a synthetic ``BoQCtgy ID="00" LblTx="Ungrouped Positions"``
    # which polluted the outline tree and made roundtrips lossy — every
    # re-import created a phantom section. Now we write them flat.
    if boq_data.positions:
        root_itemlist = ET.SubElement(boq_body, "Itemlist")
        for pos in boq_data.positions:
            item = ET.SubElement(
                root_itemlist,
                "Item",
                RNoPart="2",
                ID=str(pos.ordinal),
            )
            ET.SubElement(item, "Qty").text = _fmt_qty(pos.quantity)
            ET.SubElement(item, "QU").text = _gaeb_unit(pos.unit)

            desc = ET.SubElement(item, "Description")
            complete_text = ET.SubElement(desc, "CompleteText")
            detail_txt = ET.SubElement(complete_text, "DetailTxt")
            ET.SubElement(detail_txt, "Text").text = pos.description

            _up, _it = _gaeb_line_prices(pos.quantity, pos.unit_rate, pos.total)
            ET.SubElement(item, "UP").text = _up
            ET.SubElement(item, "IT").text = _it

            _apply_x84_alternate_fields(item, pos)

    # ── X84: Award/Recommendation block (bidder's recommended alternates) ──
    # GAEB DA 3.3 places <Recommendation> under <Award> alongside <BoQ>. We
    # write it after the BoQ tree to keep the streaming order stable; XML
    # element ordering inside <Award> is not significant for any conformant
    # importer. Only emitted when at least one position is recommended —
    # an empty <Recommendation> tag is technically valid but adds noise.
    if gaeb_format == "x84" and recommended_alternates:
        recommendation = ET.SubElement(award, "Recommendation")
        for ord_, desc_text in recommended_alternates:
            rec_item = ET.SubElement(recommendation, "RecommendedItem")
            ET.SubElement(rec_item, "RNoPart").text = ord_
            ET.SubElement(rec_item, "LblTx").text = desc_text

    # ── Trailing BoQInfo with grand total ─────────────────────────────────
    boq_info_total = ET.SubElement(boq_el, "BoQInfo")
    ET.SubElement(boq_info_total, "TotPr").text = _fmt_price(boq_data.grand_total)

    # ── Serialize to XML string ───────────────────────────────────────────
    # XML comments are discarded by every conformant XML parser (incl. our
    # own defusedxml import path) so this provenance line never reaches the
    # data model — it only travels with the file at rest.
    xml_declaration = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml_provenance = (
        f"<!-- OpenConstructionERP · DataDrivenConstruction · {_xtok_gaeb} -->\n"
    )
    xml_body = ET.tostring(gaeb, encoding="unicode", xml_declaration=False)
    xml_content = xml_declaration + xml_provenance + xml_body

    safe_name = boq_data.name.encode("ascii", errors="replace").decode("ascii").replace('"', "'")
    ext = "X84" if gaeb_format == "x84" else "X83"
    filename = f"{safe_name}.{ext}"

    return StreamingResponse(
        iter([xml_content]),
        media_type="application/xml; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ── Import (CSV / Excel) ──────────────────────────────────────────────────────

logger = logging.getLogger(__name__)

# Column name aliases for flexible matching (all lowercased for comparison)
_COLUMN_ALIASES: dict[str, list[str]] = {
    "ordinal": ["pos", "pos.", "position", "ordinal", "nr.", "nr", "no.", "no", "#"],
    "description": [
        "description",
        "beschreibung",
        "desc",
        "text",
        "bezeichnung",
        "item",
        "item description",
    ],
    "unit": ["unit", "einheit", "me", "uom", "unit of measure"],
    "quantity": ["quantity", "qty", "menge", "amount", "qty.", "quantity (qty)"],
    "unit_rate": [
        "unit rate",
        "rate",
        "ep",
        "einheitspreis",
        "unit price",
        "unit cost",
        "price",
        "rate (ep)",
    ],
    "total": ["total", "amount", "gesamtpreis", "gp", "sum", "total price"],
    "classification": [
        "classification",
        "din 276",
        "din276",
        "kg",
        "nrm",
        "code",
        "masterformat",
        "cost code",
        "cost group",
        "class",
    ],
}


def _match_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map.

    Args:
        header: Raw column header text from the uploaded file.

    Returns:
        Canonical column key (e.g. "ordinal", "description") or None if unrecognised.
    """
    normalised = header.strip().lower()
    for canonical, aliases in _COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _detect_file_format(content_head: bytes) -> Literal["xlsx", "csv", "parquet", "unknown"]:
    """Identify an upload by its magic bytes (BUG-UPLOAD01).

    File extensions are attacker-controlled — a ``.exe`` renamed to ``.xlsx``
    would otherwise be handed to ``openpyxl`` (best case: a parse exception;
    worst case: the bytes get persisted alongside trusted attachments).
    Stdlib only — uses the existing ``app.core.file_signature.detect`` for
    container types and a UTF-8 round-trip to confirm CSV is readable text.

    Returns one of: ``"xlsx"``, ``"csv"``, ``"parquet"``, ``"unknown"``.
    """
    if not content_head:
        return "unknown"
    sig = detect_signature(content_head)
    # XLSX is an OOXML zip: starts with the standard PK\x03\x04 local-file
    # header. ``file_signature.detect`` returns ``"zip"`` for any zip
    # container; we treat that as XLSX here because it's the only zip
    # format the BOQ importer accepts.
    if sig == "zip":
        return "xlsx"
    # Apache Parquet files begin (and end) with the ``PAR1`` magic.
    if content_head[:4] == b"PAR1":
        return "parquet"
    # CSV has no magic — fall back to "is this valid UTF-8 / UTF-8-BOM /
    # latin-1 text?". Reject anything that contains a NUL byte (binary
    # garbage) or fails every common text codec. A short head is enough:
    # a 512-byte sample misclassifies astoundingly rarely on real CSVs.
    if b"\x00" in content_head:
        return "unknown"
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = content_head.decode(encoding)
        except UnicodeDecodeError:
            continue
        # A CSV header row usually contains at least one ASCII separator
        # (``,;\t|``) within the first kilobyte. A pure binary blob that
        # happens to decode as latin-1 would lack any of those.
        if any(sep in decoded for sep in (",", ";", "\t", "|", "\n")):
            return "csv"
    return "unknown"


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Parse a value to float, returning *default* on failure.

    Handles strings with comma decimal separators (e.g. "1.234,56" → 1234.56).
    """
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    # Handle European-style numbers: "1.234,56" → "1234.56"
    if "," in text and "." in text:
        # Determine which is the decimal separator (last one wins)
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            # Comma is decimal separator: "1.234,56"
            text = text.replace(".", "").replace(",", ".")
        else:
            # Dot is decimal separator: "1,234.56"
            text = text.replace(",", "")
    elif "," in text:
        # Only commas — assume comma is decimal separator: "234,56"
        text = text.replace(",", ".")
    try:
        return float(text)
    except (ValueError, TypeError):
        return default


def _parse_numeric_cell(value: Any) -> tuple[float | None, str | None]:
    """Strict numeric parse for BOQ import (BUG-IMPORT02).

    Returns ``(parsed, error)``. ``error`` is ``None`` on success.
    Empty / ``None`` cells parse to ``0.0`` (the column was simply blank);
    non-empty strings that cannot be coerced surface a row-numbered error
    so callers can return a 400 with row + offending text rather than
    silently zero-filling and crashing later in the rollup.
    """
    if value is None:
        return 0.0, None
    if isinstance(value, bool):
        # ``bool`` is an ``int`` subclass — explicitly reject so True/False
        # is not silently accepted as 1/0 in a numeric column.
        return None, f"expected a number, got boolean {value!r}"
    if isinstance(value, (int, float)):
        return float(value), None
    text = str(value).strip()
    if not text:
        return 0.0, None
    parsed = _safe_float(text, default=float("nan"))
    if parsed != parsed:  # NaN check — ``_safe_float`` returns NaN on miss
        return None, f"expected a number, got {text!r}"
    return parsed, None


def _parse_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file.

    Tries UTF-8 first, then Latin-1 as fallback (common for DACH region files).

    Returns:
        List of dicts mapping canonical column names to cell values.
    """
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file — unsupported encoding")

    # Detect delimiter by sniffing first 4KB
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_rows_from_excel(
    content_bytes: bytes,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file using openpyxl.

    Reads the first (active) worksheet. The first row is treated as headers.

    Returns:
        Tuple of (rows, import_metadata).
        rows: List of dicts mapping canonical column names to cell values.
        import_metadata: Original file structure info for round-trip export.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    sheet_names = wb.sheetnames

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    original_columns = [str(h) if h is not None else "" for h in raw_headers]
    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()

    import_metadata = {
        "original_columns": original_columns,
        "column_mapping": {str(k): v for k, v in column_map.items()},
        "sheet_names": sheet_names,
        "total_rows": len(rows),
    }

    return rows, import_metadata


@router.post(
    "/boqs/{boq_id}/import/excel/",
    summary="Import positions from Excel/CSV",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def import_boq_excel(
    boq_id: uuid.UUID,
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Import BOQ positions from an Excel or CSV file.

    Accepts a multipart file upload. The file must be .xlsx or .csv.

    Expected columns (all optional except Description):
    - **Pos / Position / Ordinal / Nr.** — position ordinal number
    - **Description / Beschreibung / Text** — description (required)
    - **Unit / Einheit / ME** — unit of measurement
    - **Quantity / Qty / Menge** — quantity
    - **Unit Rate / Rate / EP / Einheitspreis** — unit rate
    - **Total** (ignored — auto-calculated from quantity x rate)
    - **Classification / DIN 276 / KG / NRM / Code** — classification code

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    # Verify BOQ exists (raises 404 if not found)
    await service.get_boq(boq_id)

    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # BUG-UPLOAD01: validate the actual file content, not just the
    # extension. A .exe renamed to .xlsx would otherwise be handed to
    # openpyxl (parse-time crash, but the bytes still hit our buffers and
    # logs first). Stdlib magic-byte sniff against an explicit allow-list.
    detected = _detect_file_format(content[:4096])
    expected = "xlsx" if filename.endswith(".xlsx") else "csv"
    if detected != expected:
        # Special-case: a CSV uploaded with a .xlsx extension (or vice
        # versa) is a user error, not an attack. Reject with a clear hint
        # rather than the generic "format mismatch" string. Anything that
        # detects as ``unknown`` is treated as suspicious.
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"File content does not match its extension. Detected "
                f"{detected!r}, expected {expected!r}. "
                "Please upload a real .xlsx or .csv file."
            ),
        )

    # Zip-bomb guard: reject .xlsx whose uncompressed sheets exceed 50 MB.
    reject_if_xlsx_bomb(content)

    # BUG-PERF01: openpyxl's ``load_workbook`` and ``csv.reader`` are
    # synchronous and can spend seconds on a 10K-row import. Running them
    # on the event-loop thread starves every concurrent request. Defer to
    # the default thread executor so the loop stays responsive.
    import_meta: dict[str, Any] = {}
    try:
        if filename.endswith(".xlsx"):
            rows, import_meta = await asyncio.to_thread(_parse_rows_from_excel, content)
        else:
            rows = await asyncio.to_thread(_parse_rows_from_csv, content)
    except ValueError as exc:
        # ValueError covers the curated parse errors raised by our
        # ``_parse_rows_from_*`` helpers (missing header row, undecodable
        # text). Surface the message — it's already user-safe.
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not parse file: {exc}",
        )
    except (zipfile.BadZipFile, KeyError) as exc:
        # BUG-UPLOAD02: openpyxl raises ``BadZipFile`` (zip header
        # corruption) and ``KeyError`` (missing sheet xml entries) on
        # malformed xlsx. Without this catch the request returns 500 with
        # a full traceback — a footgun for log-exposed deployments.
        logger.warning("Malformed xlsx upload (%s): %s", type(exc).__name__, exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not parse file: the xlsx archive is corrupt or truncated.",
        )
    except Exception as exc:
        # BUG-UPLOAD02: log the full traceback for ops, return a sanitised
        # message to the client. ``_log.exception`` honours structlog's
        # processor chain when configured.
        logger.exception("Unexpected error parsing import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not parse file: unexpected format error.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # Import each row as a Position
    imported = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    warnings_list: list[dict[str, Any]] = []
    auto_ordinal = 1

    # Pre-compute a robust median unit-rate across the import set so we can
    # emit a "far above benchmark" warning on outliers (ENH-090). Absent a
    # reference catalog we use the imported file itself as its own baseline.
    _rate_samples: list[float] = []
    for _r in rows:
        try:
            v = _safe_float(_r.get("unit_rate"), default=0.0)
            if v > 0:
                _rate_samples.append(v)
        except Exception:
            pass
    _rate_samples.sort()
    _median_rate = _rate_samples[len(_rate_samples) // 2] if _rate_samples else 0.0

    for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        try:
            description = str(row.get("description", "")).strip()

            # Skip rows without a description (likely empty or total rows)
            if not description:
                skipped += 1
                continue

            # Skip rows that look like summary/total rows
            desc_lower = description.lower()
            if desc_lower in (
                "grand total",
                "total",
                "summe",
                "gesamt",
                "gesamtsumme",
                "subtotal",
                "zwischensumme",
            ):
                skipped += 1
                continue

            # Skip pure subtotal rows (exported with "Subtotal: <ord> <desc>")
            if desc_lower.startswith("subtotal:") or desc_lower.startswith("zwischensumme:"):
                skipped += 1
                continue

            # Build ordinal: use from file or auto-generate
            ordinal = str(row.get("ordinal", "")).strip()
            if not ordinal:
                ordinal = str(auto_ordinal)
            auto_ordinal += 1

            # Parse unit
            unit_raw = str(row.get("unit", "")).strip()
            # Parse numeric fields with strict per-row validation
            # (BUG-IMPORT02). A non-empty cell that can't be coerced to a
            # number used to silently default to 0.0 — the bad value would
            # then crash later in the cost rollup. We now surface the
            # offending row + column to the user.
            quantity_raw = row.get("quantity")
            unit_rate_raw = row.get("unit_rate")
            quantity, q_err = _parse_numeric_cell(quantity_raw)
            unit_rate, r_err = _parse_numeric_cell(unit_rate_raw)
            if q_err is not None:
                errors.append(
                    {
                        "row": row_idx,
                        "ordinal": str(row.get("ordinal", "")).strip() or str(auto_ordinal),
                        "error": f"Invalid quantity at row {row_idx}: {q_err}",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue
            if r_err is not None:
                errors.append(
                    {
                        "row": row_idx,
                        "ordinal": str(row.get("ordinal", "")).strip() or str(auto_ordinal),
                        "error": f"Invalid unit_rate at row {row_idx}: {r_err}",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue
            assert quantity is not None  # narrowed: q_err was None
            assert unit_rate is not None

            # Section detection (ENH-087): a row with a description but no
            # unit, no quantity, no unit_rate is very likely a section header
            # emitted by our own CSV/Excel exporter. Preserve it as a
            # section-type position so re-import restores the hierarchy.
            is_section_row = (
                not unit_raw
                and (quantity_raw in (None, "", 0, 0.0))
                and (unit_rate_raw in (None, "", 0, 0.0))
            )
            if is_section_row:
                section_meta: dict[str, Any] = {
                    "import_source": file.filename or "excel",
                    "import_row_index": row_idx,
                    "section_header": True,
                }
                position_data = PositionCreate(
                    boq_id=boq_id,
                    ordinal=ordinal,
                    description=description,
                    unit="section",
                    quantity=0.0,
                    unit_rate=0.0,
                    classification={},
                    source="excel_import",
                    metadata=section_meta,
                )
                await service.add_position(position_data)
                imported += 1
                continue

            unit = unit_raw or "pcs"

            # Sanity caps: reject obvious tampering / typo errors before the
            # position reaches the DB. Numbers outside these bands are either
            # fat-fingered by the client or — per QA fuzz — a deliberate
            # attempt to inflate the BOQ through an edited export file.
            _IMPORT_MAX_QUANTITY = 1e9
            _IMPORT_MAX_UNIT_RATE = 1e8  # EUR/USD per unit — a steel beam is ~10k
            if not (0 <= quantity <= _IMPORT_MAX_QUANTITY):
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Quantity out of range: {quantity}",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue
            if not (0 <= unit_rate <= _IMPORT_MAX_UNIT_RATE):
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Unit rate out of range: {unit_rate}",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue

            # Soft checks — imported, but surfaced in the UI so the user
            # can spot tampered-export attacks (ENH-090 / BUG-154) and
            # data-quality issues.
            if _median_rate > 0 and unit_rate > _median_rate * 10:
                warnings_list.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "severity": "warning",
                        "message": (
                            f"Unit rate {unit_rate:.2f} is >10× the file median "
                            f"({_median_rate:.2f}) — possible typo or tampered export."
                        ),
                    }
                )
            if quantity == 0:
                warnings_list.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "severity": "info",
                        "message": "Quantity is zero — position imported but contributes no cost.",
                    }
                )
            if unit_rate == 0:
                warnings_list.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "severity": "info",
                        "message": "Unit rate is zero — position imported without a rate.",
                    }
                )

            # Build classification from the classification column
            classification: dict[str, Any] = {}
            class_value = str(row.get("classification", "")).strip()
            if class_value:
                classification["code"] = class_value

            # Create position via service (with import metadata for round-trip)
            pos_metadata: dict[str, Any] = {}
            if import_meta:
                pos_metadata["import_source"] = file.filename or "excel"
                pos_metadata["import_row_index"] = row_idx
                pos_metadata["original_columns"] = import_meta.get("original_columns", [])

            position_data = PositionCreate(
                boq_id=boq_id,
                ordinal=ordinal,
                description=description,
                unit=unit,
                quantity=quantity,
                unit_rate=unit_rate,
                classification=classification,
                source="excel_import",
                metadata=pos_metadata,
            )
            await service.add_position(position_data)
            imported += 1

        except Exception as exc:
            errors.append(
                {
                    "row": row_idx,
                    "error": str(exc),
                    "data": {k: str(v)[:100] for k, v in row.items()},
                }
            )
            logger.warning("Import error at row %d for BOQ %s: %s", row_idx, boq_id, exc)

    # Save import metadata at BOQ level for round-trip export
    if imported > 0 and import_meta:
        try:
            boq = await service.get_boq(boq_id)
            meta = dict(boq.metadata_) if isinstance(boq.metadata_, dict) else {}
            meta["last_import"] = {
                "source_filename": file.filename,
                "source_format": "xlsx" if filename.endswith(".xlsx") else "csv",
                "original_columns": import_meta.get("original_columns", []),
                "column_mapping": import_meta.get("column_mapping", {}),
                "total_imported": imported,
                "import_date": datetime.now(UTC).isoformat(),
            }
            boq.metadata_ = meta
            await service.session.flush()
            await service.session.commit()
        except Exception:
            logger.warning("Failed to save import metadata for BOQ %s", boq_id, exc_info=True)

    logger.info(
        "BOQ import complete for %s: imported=%d, skipped=%d, errors=%d",
        boq_id,
        imported,
        skipped,
        len(errors),
    )

    # BUG-IMPORT02: when every parseable row failed validation (no
    # ``imported`` rows but errors collected), surface a 400 with the
    # first row's diagnostic so the client can show "row 2: invalid
    # quantity" rather than a confusing 200 with imported=0. Partial
    # successes still return 200 with the per-row error list intact.
    if imported == 0 and errors:
        first = errors[0]
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Import failed at row {first.get('row', '?')}: "
                f"{first.get('error', 'unknown error')}"
            ),
        )

    # Run validation inline so DIN276 / NRM / GAEB / MasterFormat / DPGF /
    # boq_quality issues surface in the import response instead of only
    # via the later /validate/ call (philosophy: validation is a first-
    # class citizen of the core workflow). Gated by IMPORT_INLINE_VALIDATION.
    validation_report = None
    if imported > 0:
        validation_report = await _run_import_validation(boq_id, service, service.session)

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "warnings": warnings_list,
        "total_rows": len(rows),
        "source_format": import_meta.get("source_format", "unknown") if import_meta else "unknown",
        "original_columns": import_meta.get("original_columns", []) if import_meta else [],
        "validation_report": validation_report,
    }


# ── GAEB XML import ──────────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/import/gaeb/",
    summary="Import positions from GAEB XML 3.3 (X83/X84)",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def import_boq_gaeb(
    boq_id: uuid.UUID,
    file: UploadFile = File(..., description="GAEB XML file (.x83, .x84, .xml)"),
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Import BOQ positions from a GAEB XML 3.3 file (BUG-153).

    Supports the GAEB DA XML formats used across DACH tendering:
      - **X83 / DP 83** — Angebotsabgabe (bid submission)
      - **X84 / DP 84** — Nebenangebote (alternative bids)
      - **X81** — Leistungsverzeichnis (BOQ skeleton)

    Namespace-agnostic parser — falls back to tag-local-name matching so
    files from different GAEB toolchains (iTWO, California.pro, Nevaris,
    etc.) all import without pre-normalization.

    Security: uses ``defusedxml`` to harden against XXE, billion-laughs,
    and other XML-parser-level attacks on user-uploaded files.
    """
    import xml.etree.ElementTree as ET

    from defusedxml.ElementTree import fromstring as _safe_fromstring

    # Verify BOQ exists (raises 404 if not found)
    await service.get_boq(boq_id)

    filename = (file.filename or "").lower()
    if not filename.endswith((".x81", ".x83", ".x84", ".xml")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Unsupported file type. Please upload a GAEB XML file (.x81, .x83, .x84, or .xml)."
            ),
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )
    # No upload size cap — per product policy.

    # Parse XML defensively via defusedxml — blocks XXE, external-entity
    # expansion, billion-laughs, and DTD-based attacks on user input.
    try:
        root = _safe_fromstring(content)
    except ET.ParseError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse GAEB XML: {exc}",
        ) from exc
    except Exception as exc:  # defusedxml raises its own subclasses for attacks
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"GAEB XML rejected by security parser: {exc}",
        ) from exc

    def _local(tag: str) -> str:
        """Strip namespace from an element tag."""
        return tag.split("}", 1)[1] if "}" in tag else tag

    def _find_child(parent: ET.Element, name: str) -> ET.Element | None:
        """Namespace-agnostic single-child lookup by local name."""
        for child in parent:
            if _local(child.tag) == name:
                return child
        return None

    def _find_all_descendants(parent: ET.Element, name: str) -> list[ET.Element]:
        """Walk the entire subtree, collect elements whose local name matches."""
        found: list[ET.Element] = []
        for el in parent.iter():
            if _local(el.tag) == name:
                found.append(el)
        return found

    def _text_of(parent: ET.Element, name: str) -> str:
        child = _find_child(parent, name)
        return (child.text or "").strip() if child is not None else ""

    def _extract_description(item: ET.Element) -> str:
        """Pull human-readable text out of GAEB's nested Description/CompleteText/DetailTxt/Text."""
        # Take the first non-empty <Text> we find anywhere in the item's
        # subtree — description structures vary wildly between exporters.
        for text_el in _find_all_descendants(item, "Text"):
            if text_el.text and text_el.text.strip():
                return text_el.text.strip()
        # Fall back to OutlineText / Outline / LblTx
        for name in ("OutlineText", "OutlTxt", "LblTx"):
            val = _text_of(item, name)
            if val:
                return val
        return ""

    # Build reverse map from the export lexicon so GAEB unit codes round-trip
    # back to our internal tokens (BUG-175 — "Stk" → "pcs", "psch" → "lsum").
    _GAEB_TO_INTERNAL: dict[str, str] = {
        "stk": "pcs",
        "st": "pcs",
        "psch": "lsum",
        "jahr": "year",
        "mo": "month",
    }

    def _normalize_unit(unit: str) -> str:
        key = (unit or "").strip().lower()
        return _GAEB_TO_INTERNAL.get(key, unit.strip()) if key else ""

    # Locate the *top-level* BoQBody — the one directly inside <BoQ>.
    # A GAEB tree nests BoQBody recursively under each BoQCtgy, so
    # traversing ``_find_all_descendants`` would double-visit every Item.
    top_body: ET.Element | None = None
    for el in root.iter():
        if _local(el.tag) == "BoQ":
            top_body = _find_child(el, "BoQBody")
            if top_body is not None:
                break
    if top_body is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No <BoQBody> element found. Is this a valid GAEB DA XML?",
        )
    boq_bodies = [top_body]

    imported = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    sections_seen: list[dict[str, str]] = []

    # Capture currency for round-trip metadata. Empty when the source
    # GAEB doesn't carry <Cur> — preferable to a EUR fallback that
    # mis-stamps non-Eurozone tenders. Downstream code that needs a
    # currency falls back to project.currency at the consumer side.
    award = None
    for el in root.iter():
        if _local(el.tag) == "Award":
            award = el
            break
    currency = (_text_of(award, "Cur") if award is not None else "") or ""

    def _process_category(ctgy: ET.Element, parent_ordinal: str = "") -> None:
        nonlocal imported, skipped
        ord_ = (ctgy.get("ID") or "").strip() or parent_ordinal
        label = _text_of(ctgy, "LblTx") or "Section"
        sections_seen.append({"ordinal": ord_, "label": label})

        # Each BoQCtgy has its own BoQBody containing Itemlist/Item.
        inner_body = _find_child(ctgy, "BoQBody")
        if inner_body is not None:
            # Nested categories (recursion for multi-level hierarchies).
            for child in inner_body:
                local = _local(child.tag)
                if local == "BoQCtgy":
                    _process_category(child, parent_ordinal=ord_)
                elif local == "Itemlist":
                    for item in child:
                        if _local(item.tag) == "Item":
                            _import_item(item, section_ordinal=ord_)

    def _import_item(item: ET.Element, *, section_ordinal: str = "") -> None:
        nonlocal imported, skipped
        try:
            pos_ordinal = (item.get("ID") or "").strip() or str(imported + 1)
            description = _extract_description(item)
            if not description:
                skipped += 1
                return None

            unit_raw = _text_of(item, "QU")
            unit = _normalize_unit(unit_raw) or "pcs"
            quantity = _safe_float(_text_of(item, "Qty"), default=0.0)
            unit_rate = _safe_float(_text_of(item, "UP"), default=0.0)

            if not (0 <= quantity <= 1e9):
                errors.append(
                    {
                        "ordinal": pos_ordinal,
                        "error": f"Quantity out of range: {quantity}",
                    }
                )
                return None
            if not (0 <= unit_rate <= 1e8):
                errors.append(
                    {
                        "ordinal": pos_ordinal,
                        "error": f"Unit rate out of range: {unit_rate}",
                    }
                )
                return None

            position_data = PositionCreate(
                boq_id=boq_id,
                ordinal=pos_ordinal,
                description=description,
                unit=unit,
                quantity=quantity,
                unit_rate=unit_rate,
                classification={"gaeb_section": section_ordinal} if section_ordinal else {},
                source="gaeb_import",
                metadata={
                    "import_source": file.filename or "gaeb",
                    "gaeb_ordinal": pos_ordinal,
                    "gaeb_section": section_ordinal,
                    "gaeb_unit_original": unit_raw,
                    "gaeb_currency": currency,
                },
            )
            # add_position is async — run it via await below.
            return position_data
        except Exception as exc:  # noqa: BLE001 — narrow at caller
            errors.append({"error": str(exc), "ordinal": ""})
            return None

    # Walk the top-level BoQBody: may contain direct Item elements OR BoQCtgy.
    for body in boq_bodies:
        for child in body:
            local = _local(child.tag)
            if local == "BoQCtgy":
                # The original _process_category helper builds positions but
                # can't await — so refactor: collect items, then insert.
                pass

    # Second, simpler pass: collect every Item anywhere in the tree, attribute
    # it to the nearest ancestor BoQCtgy's ID for section ordinal.
    def _ancestor_ctgy_id(el: ET.Element, ancestors: list[ET.Element]) -> str:
        for anc in reversed(ancestors):
            if _local(anc.tag) == "BoQCtgy":
                return (anc.get("ID") or "").strip()
        return ""

    def _walk_and_collect(
        el: ET.Element, ancestors: list[ET.Element]
    ) -> list[tuple[ET.Element, str]]:
        found: list[tuple[ET.Element, str]] = []
        for child in el:
            if _local(child.tag) == "Item":
                found.append((child, _ancestor_ctgy_id(child, ancestors + [el])))
            else:
                found.extend(_walk_and_collect(child, ancestors + [el]))
        return found

    collected: list[tuple[ET.Element, str]] = []
    for body in boq_bodies:
        collected.extend(_walk_and_collect(body, []))

    auto_counter = 0
    for item, section_ordinal in collected:
        auto_counter += 1
        try:
            pos_ordinal = (item.get("ID") or "").strip() or str(auto_counter)
            description = _extract_description(item)
            if not description:
                skipped += 1
                continue

            unit_raw = _text_of(item, "QU")
            unit = _normalize_unit(unit_raw) or "pcs"
            quantity = _safe_float(_text_of(item, "Qty"), default=0.0)
            unit_rate = _safe_float(_text_of(item, "UP"), default=0.0)

            if not (0 <= quantity <= 1e9):
                errors.append(
                    {"ordinal": pos_ordinal, "error": f"Quantity out of range: {quantity}"}
                )
                continue
            if not (0 <= unit_rate <= 1e8):
                errors.append(
                    {"ordinal": pos_ordinal, "error": f"Unit rate out of range: {unit_rate}"}
                )
                continue

            classification: dict[str, Any] = {}
            if section_ordinal:
                classification["gaeb_section"] = section_ordinal

            position_data = PositionCreate(
                boq_id=boq_id,
                ordinal=pos_ordinal,
                description=description,
                unit=unit,
                quantity=quantity,
                unit_rate=unit_rate,
                classification=classification,
                source="gaeb_import",
                metadata={
                    "import_source": file.filename or "gaeb",
                    "gaeb_ordinal": pos_ordinal,
                    "gaeb_section": section_ordinal,
                    "gaeb_unit_original": unit_raw,
                    "gaeb_currency": currency,
                },
            )
            await service.add_position(position_data)
            imported += 1
        except Exception as exc:
            errors.append({"ordinal": item.get("ID") or "", "error": str(exc)})
            logger.warning("GAEB import error for BOQ %s: %s", boq_id, exc)

    # Persist lightweight import metadata at the BOQ level.
    if imported > 0:
        try:
            boq_obj = await service.get_boq(boq_id)
            meta = dict(boq_obj.metadata_) if isinstance(boq_obj.metadata_, dict) else {}
            meta["last_import"] = {
                "source_filename": file.filename,
                "source_format": "gaeb",
                "gaeb_currency": currency,
                "total_imported": imported,
                "total_sections": len(sections_seen),
                "import_date": datetime.now(UTC).isoformat(),
            }
            boq_obj.metadata_ = meta
            await service.session.flush()
            await service.session.commit()
        except Exception:
            logger.warning(
                "Failed to persist GAEB import metadata for BOQ %s", boq_id, exc_info=True
            )

    logger.info(
        "GAEB import complete for %s: imported=%d, skipped=%d, errors=%d, sections=%d",
        boq_id,
        imported,
        skipped,
        len(errors),
        len(sections_seen),
    )

    # Run validation inline against the freshly-imported GAEB BOQ so
    # DIN276 / GAEB / boq_quality rule packs fire AT import time, not
    # later via the standalone /validate/ endpoint. For DACH GAEB files
    # the project region is almost always DE/AT/CH so _build_rule_sets
    # selects the gaeb + din276 rule packs automatically.
    validation_report = None
    if imported > 0:
        validation_report = await _run_import_validation(boq_id, service, service.session)

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "sections": sections_seen,
        "source_format": "gaeb",
        "currency": currency,
        "validation_report": validation_report,
    }


# ── Smart import helpers ─────────────────────────────────────────────────────


def _extract_from_pdf(content: bytes) -> dict[str, Any]:
    """Extract text and tables from a PDF file.

    Uses pdfplumber to extract tabular data first, falling back to plain text.

    Args:
        content: Raw PDF file bytes.

    Returns:
        Dict with ``text`` (extracted content) and ``structured`` flag.
    """
    import pdfplumber

    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        text_parts.append("\t".join(str(cell or "") for cell in row))
            else:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)

    return {"text": "\n".join(text_parts), "structured": False}


def _extract_from_image(content: bytes, ext: str) -> dict[str, Any]:
    """Prepare an image for AI vision analysis.

    Encodes the raw image bytes as base64 and determines the MIME type.

    Args:
        content: Raw image file bytes.
        ext: File extension (e.g. ``jpg``, ``png``).

    Returns:
        Dict with ``image_base64``, ``mime``, empty ``text``, and ``structured`` flag.
    """
    import base64

    mime = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"
    b64 = base64.b64encode(content).decode()
    return {"text": "", "image_base64": b64, "mime": mime, "structured": False}


def _extract_from_excel_for_smart(content: bytes) -> dict[str, Any]:
    """Extract data from Excel for smart import.

    Tries to parse as structured rows first. If columns can be detected, returns
    structured data. Otherwise returns the raw cell text for AI processing.

    Args:
        content: Raw Excel file bytes.

    Returns:
        Dict with ``text``, ``structured`` flag, and optionally ``rows``.
    """
    try:
        rows, _meta = _parse_rows_from_excel(content)
        if rows:
            # Check if we have enough structure for a direct import
            has_description = any(r.get("description") for r in rows)
            if has_description:
                return {"text": "", "structured": True, "rows": rows}
    except Exception:
        logger.debug("Smart import: structured Excel parsing failed, using raw text", exc_info=True)

    # Fall back to extracting raw text from all cells
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    text_parts: list[str] = []
    if ws is not None:
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell or "") for cell in row]
            line = "\t".join(cells).strip()
            if line:
                text_parts.append(line)
    wb.close()
    return {"text": "\n".join(text_parts), "structured": False}


def _extract_from_csv_for_smart(content: bytes) -> dict[str, Any]:
    """Extract data from CSV for smart import.

    Tries structured parsing first. If column detection fails, falls back to
    raw text extraction for AI processing.

    Args:
        content: Raw CSV file bytes.

    Returns:
        Dict with ``text``, ``structured`` flag, and optionally ``rows``.
    """
    try:
        rows = _parse_rows_from_csv(content)
        if rows:
            has_description = any(r.get("description") for r in rows)
            if has_description:
                return {"text": "", "structured": True, "rows": rows}
    except Exception:
        logger.debug("Smart import: structured CSV parsing failed, using raw text", exc_info=True)

    # Fall back to raw text
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode("latin-1", errors="replace")

    return {"text": text, "structured": False}


async def _extract_from_cad(content: bytes, ext: str, filename: str) -> dict[str, Any]:
    """Extract data from a CAD/BIM file using DDC Community converters.

    Saves the uploaded file to a temp directory, runs the appropriate DDC
    converter (.exe) to produce an Excel file, then parses the Excel output
    into an element summary for AI processing.

    Args:
        content: Raw CAD file bytes.
        ext: Lowercase file extension without dot (e.g. ``"rvt"``).
        filename: Original file name for temp file creation.

    Returns:
        Dict with ``text`` (element summary), ``structured`` flag, and metadata.
        If no converter is installed, returns a helpful message with download link.
    """
    from app.modules.boq.cad_import import (
        convert_cad_to_excel,
        find_converter,
        parse_cad_excel,
        summarize_cad_elements,
    )

    converter = find_converter(ext)
    if not converter:
        return {
            "text": (
                f"CAD file detected (.{ext}) but no DDC converter found.\n"
                f"Download DDC converters from:\n"
                f"https://github.com/datadrivenconstruction/ddc-community-toolkit/releases\n"
                f"Place .exe files in one of these locations:\n"
                f"  - converters/bin/ (project root)\n"
                f"  - ~/.openestimator/converters/\n"
                f"  - Set OPENESTIMATOR_CONVERTERS_DIR environment variable"
            ),
            "structured": False,
            "cad_no_converter": True,
        }

    # Save to temp dir, convert, parse
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / filename
        input_path.write_bytes(content)

        output_dir = Path(tmpdir) / "output"
        output_dir.mkdir()

        excel_path = await convert_cad_to_excel(input_path, output_dir, ext)
        if not excel_path:
            return {
                "text": (
                    f"CAD conversion failed for .{ext} file. "
                    "Check that the converter is properly installed and the file is valid."
                ),
                "structured": False,
            }

        elements = parse_cad_excel(excel_path)
        summary = summarize_cad_elements(elements)

        return {
            "text": summary,
            "structured": False,
            "cad_elements": len(elements),
            "cad_format": ext,
        }


# ── Smart import endpoint ────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/import/smart/",
    summary="Smart import: any file via AI",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def smart_import(
    boq_id: uuid.UUID,
    user_id: CurrentUserId,
    file: UploadFile = File(
        ...,
        description="Any document file (Excel, CSV, PDF, image, or CAD/BIM: .rvt, .ifc, .dwg, .dgn)",
    ),
    service: BOQService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Smart import: parse ANY file into BOQ positions using AI.

    Accepts Excel (.xlsx), CSV (.csv), PDF (.pdf), image files
    (.jpg, .jpeg, .png, .tiff, .bmp), and CAD/BIM files
    (.rvt, .ifc, .dwg, .dgn). For structured Excel/CSV with
    recognisable column headers, performs a direct import. For CAD/BIM
    files, runs a DDC converter to extract element data. Otherwise,
    sends the extracted text (or image) to the user's configured AI
    provider for intelligent parsing into BOQ positions.

    Returns:
        Summary with imported/error counts, method used, and AI model if applicable.
    """
    # Verify BOQ exists, capture project currency for downstream LLM prompts.
    boq_obj = await service.get_boq(boq_id)
    _project_currency: str = ""
    try:
        from app.modules.projects.repository import ProjectRepository

        _proj = await ProjectRepository(session).get_by_id(boq_obj.project_id)
        _project_currency = (getattr(_proj, "currency", "") or "").strip()
    except Exception:  # noqa: BLE001 — currency is best-effort, prompt tolerates blank
        _project_currency = ""

    filename = (file.filename or "unknown").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # No upload size cap — per product policy.

    # ── 1. Extract text/data based on file type ────────────────────────
    if ext in ("xlsx", "xls"):
        # BUG-UPLOAD01b: smart-import path used to skip the xlsx-bomb
        # guard that import_boq_excel calls — same DoS surface via this
        # endpoint. Apply the same defence here before parsing.
        from app.core.upload_guards import reject_if_xlsx_bomb

        reject_if_xlsx_bomb(content)
        extracted = _extract_from_excel_for_smart(content)
    elif ext == "csv":
        extracted = _extract_from_csv_for_smart(content)
    elif ext == "pdf":
        extracted = _extract_from_pdf(content)
    elif ext in ("jpg", "jpeg", "png", "tiff", "bmp"):
        extracted = _extract_from_image(content, ext)
    elif ext in ("rvt", "ifc", "dwg", "dgn"):
        extracted = await _extract_from_cad(content, ext, file.filename or f"model.{ext}")
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Unsupported file type: .{ext}. Supported: xlsx, csv, pdf, jpg, png, tiff, rvt, ifc, dwg, dgn."
            ),
        )

    # ── 1b. Handle missing CAD converter (return early) ────────────────
    if extracted.get("cad_no_converter"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=extracted["text"],
        )

    # ── 2. Direct import for structured Excel/CSV ──────────────────────
    if extracted.get("structured") and extracted.get("rows"):
        rows = extracted["rows"]
        imported = 0
        skipped = 0
        errors: list[dict[str, Any]] = []
        auto_ordinal = 1

        for row_idx, row in enumerate(rows, start=2):
            try:
                description = str(row.get("description", "")).strip()
                if not description:
                    skipped += 1
                    continue

                desc_lower = description.lower()
                if desc_lower in (
                    "grand total",
                    "total",
                    "summe",
                    "gesamt",
                    "gesamtsumme",
                    "subtotal",
                    "zwischensumme",
                ):
                    skipped += 1
                    continue

                ordinal = str(row.get("ordinal", "")).strip()
                if not ordinal:
                    ordinal = str(auto_ordinal)
                auto_ordinal += 1

                unit = str(row.get("unit", "pcs")).strip() or "pcs"
                quantity = _safe_float(row.get("quantity"), default=0.0)
                unit_rate = _safe_float(row.get("unit_rate"), default=0.0)

                classification: dict[str, Any] = {}
                class_value = str(row.get("classification", "")).strip()
                if class_value:
                    classification["code"] = class_value

                position_data = PositionCreate(
                    boq_id=boq_id,
                    ordinal=ordinal,
                    description=description,
                    unit=unit,
                    quantity=quantity,
                    unit_rate=unit_rate,
                    classification=classification,
                    source="smart_import",
                )
                await service.add_position(position_data)
                imported += 1

            except Exception as exc:
                errors.append(
                    {
                        "row": row_idx,
                        "error": str(exc),
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )

        logger.info(
            "Smart import (direct) for BOQ %s: imported=%d, skipped=%d, errors=%d",
            boq_id,
            imported,
            skipped,
            len(errors),
        )

        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "total_items": len(rows),
            "method": "direct",
            "model_used": None,
        }

    # ── 3. AI-powered import ───────────────────────────────────────────
    from app.modules.ai.ai_client import call_ai, extract_json, resolve_provider_key_model
    from app.modules.ai.prompts import (
        CAD_IMPORT_PROMPT,
        SMART_IMPORT_PROMPT,
        SMART_IMPORT_VISION_PROMPT,
        SYSTEM_PROMPT,
    )
    from app.modules.ai.repository import AISettingsRepository

    settings_repo = AISettingsRepository(session)
    user_uuid = uuid.UUID(user_id)
    ai_settings = await settings_repo.get_by_user_id(user_uuid)

    try:
        provider, api_key, model_override = resolve_provider_key_model(ai_settings)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        ) from exc

    # Build prompt
    image_b64: str | None = None
    image_mime: str = "image/jpeg"

    if extracted.get("image_base64"):
        # Image-based: use vision prompt
        prompt = SMART_IMPORT_VISION_PROMPT.format(filename=file.filename or "image")
        image_b64 = extracted["image_base64"]
        image_mime = extracted.get("mime", "image/jpeg")
    elif extracted.get("cad_format"):
        # CAD/BIM data: use specialized CAD prompt with larger context
        text_content = extracted.get("text", "")[:12000]
        if not text_content.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CAD conversion produced no element data. The file may be empty or corrupt.",
            )
        prompt = CAD_IMPORT_PROMPT.format(text=text_content, currency=_project_currency)
    else:
        # Text-based: use text prompt (truncate to 8000 chars for context window)
        text_content = extracted.get("text", "")[:8000]
        if not text_content.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not extract any text from the file. Try a different format.",
            )
        prompt = SMART_IMPORT_PROMPT.format(
            text=text_content,
            filename=file.filename or "document",
        )

    # Call AI
    try:
        raw_response, _tokens = await call_ai(
            provider=provider,
            api_key=api_key,
            system=SYSTEM_PROMPT,
            prompt=prompt,
            image_base64=image_b64,
            image_media_type=image_mime,
            max_tokens=4096,
            model=model_override,
        )
    except Exception as exc:
        logger.exception("Smart import AI call failed for BOQ %s: %s", boq_id, exc)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"AI service error: {exc}",
        ) from exc

    items = extract_json(raw_response)
    if not isinstance(items, list):
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI did not return a valid list of items. Please try again.",
        )

    # ── 4. Create positions from AI response ───────────────────────────
    is_cad_import = bool(extracted.get("cad_format"))
    import_source = "cad_import_ai" if is_cad_import else "smart_import_ai"

    imported = 0
    errors = []
    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        try:
            description = str(item.get("description", "")).strip()
            if len(description) < 2:
                continue

            ordinal = str(item.get("ordinal", "")).strip()
            if not ordinal:
                prefix = "CAD" if is_cad_import else "AI"
                ordinal = f"{prefix}.{imported + 1:04d}"

            unit = str(item.get("unit", "lsum")).strip() or "lsum"

            try:
                quantity = float(item.get("quantity", 0))
            except (ValueError, TypeError):
                quantity = 0.0

            try:
                unit_rate = float(item.get("unit_rate", 0))
            except (ValueError, TypeError):
                unit_rate = 0.0

            classification = item.get("classification", {})
            if not isinstance(classification, dict):
                classification = {}

            position_data = PositionCreate(
                boq_id=boq_id,
                ordinal=ordinal,
                description=description,
                unit=unit,
                quantity=max(quantity, 0.0),
                unit_rate=max(unit_rate, 0.0),
                classification=classification,
                source=import_source,
                confidence=0.7,
            )
            await service.add_position(position_data)
            imported += 1
        except Exception as exc:
            errors.append(
                {
                    "item": str(item.get("description", "?"))[:100],
                    "error": str(exc),
                }
            )
            logger.warning(
                "Smart import AI item error at index %d for BOQ %s: %s",
                idx,
                boq_id,
                exc,
            )

    method = "cad_ai" if is_cad_import else "ai"
    logger.info(
        "Smart import (%s) for BOQ %s: imported=%d, errors=%d, provider=%s",
        method,
        boq_id,
        imported,
        len(errors),
        provider,
    )

    result: dict[str, Any] = {
        "imported": imported,
        "errors": errors,
        "total_items": len(items),
        "method": method,
        "model_used": provider,
    }
    if is_cad_import:
        result["cad_format"] = extracted.get("cad_format")
        result["cad_elements"] = extracted.get("cad_elements", 0)
    return result


# ── Sustainability / CO2 Calculator ──────────────────────────────────────────

from app.modules.boq.epd_materials import (
    EPD_CATEGORIES,
    EPD_INDEX,
    EU_CPR_BENCHMARKS,
    detect_epd_material,
    search_epd_materials,
)


def _co2_rating(benchmark_per_m2: float) -> tuple[str, str]:
    """Determine CO2 rating based on benchmark per m2."""
    if benchmark_per_m2 < 80:
        return "A", "Excellent"
    if benchmark_per_m2 < 150:
        return "B", "Good"
    if benchmark_per_m2 < 250:
        return "C", "Average"
    return "D", "Poor"


def _eu_cpr_compliance(gwp_per_m2_year: float) -> str:
    """Determine EU CPR 2024/3110 compliance level."""
    if gwp_per_m2_year <= EU_CPR_BENCHMARKS["excellent"]:
        return "excellent"
    if gwp_per_m2_year <= EU_CPR_BENCHMARKS["good"]:
        return "good"
    if gwp_per_m2_year <= EU_CPR_BENCHMARKS["acceptable"]:
        return "acceptable"
    return "non-compliant"


def _get_position_co2(pos: Any) -> dict[str, Any] | None:
    """Extract stored CO2 data from position metadata, or auto-detect."""
    meta = pos.metadata_ if hasattr(pos, "metadata_") else (pos.metadata or {})
    if isinstance(meta, dict) and "co2" in meta:
        return meta["co2"]
    return None


# ── Resource Summary ──────────────────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/resource-summary/",
    response_model=ResourceSummaryResponse,
    summary="Get resource summary",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_resource_summary(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> ResourceSummaryResponse:
    """Aggregate all resources across a BOQ's positions.

    Loads every position's ``metadata_.resources`` list, combines resources
    that share the same (name, type) key, sums quantities and costs, and
    groups totals by resource type (material, labor, equipment, etc.).

    When a position has no explicit resources in metadata, falls back to
    looking up the matching cost item from the database and using its
    ``components`` list as the resource data.

    Returns:
        ResourceSummaryResponse with per-type counts/totals and a flat
        resource list sorted by total_cost descending.
    """
    boq_data = await service.get_boq_with_positions(boq_id)

    # Aggregation key: (name_lower, type_lower) → accumulator
    agg: dict[tuple[str, str], dict[str, Any]] = {}

    def _add_resource(
        raw: dict[str, Any],
        pos_id: str,
        pos_qty: float = 1.0,
        resource_idx: int | None = None,
    ) -> None:
        """Add a single resource dict to the aggregation map.

        ``resource_idx`` is captured so the frontend can fan a re-pick out
        to every (position, slot) where this aggregated resource lives.
        It is ``None`` only for synthetic resources derived from
        ``position.description`` (no real slot to patch).
        """
        name = str(raw.get("name", raw.get("code", ""))).strip()
        rtype = str(raw.get("type", "other")).strip().lower()
        if not name:
            return

        unit = str(raw.get("unit", "")).strip()
        try:
            qty = float(raw.get("quantity", 1.0))
            rate = float(raw.get("unit_rate", 0))
        except (ValueError, TypeError):
            return

        cost = qty * rate * max(pos_qty, 1.0)
        key = (name.lower(), rtype)

        if key not in agg:
            agg[key] = {
                "name": name,
                "type": rtype,
                "unit": unit,
                "total_quantity": 0.0,
                "total_cost": 0.0,
                "rates": [],
                "positions": set(),
                # Variant surface — first-seen wins for the catalog/stats,
                # since variants are intrinsic to the abstract resource.
                "available_variants": None,
                "variant_stats": None,
                "currency": None,
                "resource_code": None,
                # Distinct (label, default) tuples observed across positions:
                # if all entries agree we surface that pick; mixed → "__mixed__".
                "variant_labels": set(),
                "variant_defaults": set(),
                "position_refs": [],
            }

        entry = agg[key]
        entry["total_quantity"] += qty * max(pos_qty, 1.0)
        entry["total_cost"] += cost
        entry["rates"].append(rate)
        entry["positions"].add(pos_id)

        # Capture variant catalog / stats / currency / code on first sighting.
        avail = raw.get("available_variants")
        if entry["available_variants"] is None and isinstance(avail, list) and len(avail) >= 2:
            entry["available_variants"] = avail
        vstats = raw.get("variant_stats") or raw.get("available_variant_stats")
        if entry["variant_stats"] is None and isinstance(vstats, dict):
            entry["variant_stats"] = vstats
        cur = raw.get("currency")
        if entry["currency"] is None and isinstance(cur, str) and cur:
            entry["currency"] = cur
        rc = raw.get("code")
        if entry["resource_code"] is None and isinstance(rc, str) and rc.strip():
            entry["resource_code"] = rc.strip()

        # Track current pick / default per position so the UI can show a
        # consistent pill (or flag "mixed" when positions disagree).
        variant = raw.get("variant")
        if isinstance(variant, dict):
            label = variant.get("label")
            if isinstance(label, str) and label:
                entry["variant_labels"].add(label)
        else:
            entry["variant_labels"].add("__unset__")
        vdef = raw.get("variant_default")
        if isinstance(vdef, str) and vdef in ("mean", "median"):
            entry["variant_defaults"].add(vdef)

        # Record a pointer back to this exact resource slot for fan-out
        # re-pick. Skip for synthetic rows (resource_idx is None) — they
        # have no slot to patch.
        if resource_idx is not None:
            entry["position_refs"].append(
                ResourcePositionRef(position_id=pos_id, resource_idx=resource_idx)
            )

    for pos in boq_data.positions:
        meta = pos.metadata or {}
        resources = meta.get("resources")
        pos_qty = float(pos.quantity or 0) if hasattr(pos, "quantity") else 1.0

        if isinstance(resources, list) and len(resources) > 0:
            for idx, raw in enumerate(resources):
                if not isinstance(raw, dict):
                    continue
                _add_resource(raw, str(pos.id), resource_idx=idx)
        else:
            # Fast heuristic: classify by description and create a synthetic resource
            desc = pos.description or ""
            if not desc.strip():
                continue
            rate = float(pos.unit_rate or 0) if hasattr(pos, "unit_rate") else 0.0
            total = rate * max(pos_qty, 1.0)
            if total <= 0:
                continue
            cat = BOQService._classify_position_category(desc)
            _add_resource(
                {
                    "name": desc[:80],
                    "type": cat,
                    "unit": str(getattr(pos, "unit", "") or ""),
                    "quantity": pos_qty,
                    "unit_rate": rate,
                },
                str(pos.id),
            )

    # Build flat resource list sorted by total_cost descending
    resource_items: list[ResourceSummaryItem] = []
    for entry in agg.values():
        rates_list: list[float] = entry["rates"]
        avg_rate = sum(rates_list) / len(rates_list) if rates_list else 0.0

        # Resolve consensus pick across positions:
        #   * single non-"__unset__" label  → that pick
        #   * many                          → "__mixed__"
        #   * only "__unset__"              → None (no explicit pick)
        labels: set[str] = entry["variant_labels"]
        explicit_labels = labels - {"__unset__"}
        if len(explicit_labels) == 1 and "__unset__" not in labels:
            current_label = next(iter(explicit_labels))
        elif len(explicit_labels) >= 2:
            current_label = "__mixed__"
        else:
            current_label = None

        defaults: set[str] = entry["variant_defaults"]
        variant_default = next(iter(defaults)) if len(defaults) == 1 else None

        resource_items.append(
            ResourceSummaryItem(
                name=entry["name"],
                type=entry["type"],
                unit=entry["unit"],
                total_quantity=round(entry["total_quantity"], 3),
                avg_unit_rate=round(avg_rate, 2),
                total_cost=round(entry["total_cost"], 2),
                positions_used=len(entry["positions"]),
                available_variants=entry["available_variants"],
                variant_stats=entry["variant_stats"],
                current_variant_label=current_label,
                variant_default=variant_default,
                currency=entry["currency"],
                resource_code=entry["resource_code"],
                position_refs=entry["position_refs"],
            )
        )

    resource_items.sort(key=lambda r: r.total_cost, reverse=True)

    # Dedupe variant pickers across summary rows. Two collapse scenarios:
    #   1. Two rows share the same ``resource_code`` (CWICR
    #      KADX_KATO_KAKASA_KATO: two component rows under KALI-RI-KATO-KANE
    #      with identical 3-variant catalogs).
    #   2. Two rows carry the same variant-label set even with different
    #      codes — happens when a position persisted the synthetic top-level
    #      resource alongside a component that mirrors it (BG_SOFIA shape:
    #      "Стоманени конструкции" appears as both the cost item's top
    #      variants and as component[0]).
    # In both cases strip ``available_variants`` / ``variant_stats`` from
    # secondary rows so the UI only renders one ▾N picker per unique
    # catalog. Picker fan-out via ``position_refs`` covers all linked
    # positions already.
    seen_codes: set[str] = set()
    seen_hashes: set[str] = set()
    for it in resource_items:
        if not it.available_variants:
            continue
        label_hash = "|".join(
            (v.get("label") or "").strip() for v in it.available_variants if isinstance(v, dict)
        )
        already = (it.resource_code and it.resource_code in seen_codes) or (
            label_hash and label_hash in seen_hashes
        )
        if already:
            it.available_variants = None
            it.variant_stats = None
            it.current_variant_label = None
            it.variant_default = None
        else:
            if it.resource_code:
                seen_codes.add(it.resource_code)
            if label_hash:
                seen_hashes.add(label_hash)

    # Build by_type summary
    by_type: dict[str, ResourceTypeSummary] = {}
    for item in resource_items:
        if item.type not in by_type:
            by_type[item.type] = ResourceTypeSummary(count=0, total_cost=0.0)
        by_type[item.type].count += 1
        by_type[item.type].total_cost = round(by_type[item.type].total_cost + item.total_cost, 2)

    # Issue #106 — Pareto / ABC analysis. Items are already sorted by total_cost
    # descending above, so we walk the cumulative percentage and assign the
    # standard 80/15/5 buckets. The thresholds are conventional, not
    # ISO-prescribed; they match the user's "what hurts the budget most" intent
    # (A = ~top 20 % of items that drive ~80 % of cost). When grand_total is 0
    # (e.g. fresh BOQ with no rates yet) we skip ABC entirely so we don't
    # divide by zero.
    grand_total = round(sum(it.total_cost for it in resource_items), 2)
    if grand_total > 0:
        cumulative = 0.0
        for item in resource_items:
            pct = (item.total_cost / grand_total) * 100.0
            item.abc_percentage = round(pct, 2)
            cumulative += pct
            # Use the cumulative threshold *before* this item rather than
            # after — otherwise the single biggest item would always be
            # classified A even on a flat distribution. Standard practice.
            if cumulative <= 80.0:
                item.abc_class = "A"
            elif cumulative <= 95.0:
                item.abc_class = "B"
            else:
                item.abc_class = "C"

    return ResourceSummaryResponse(
        total_resources=len(resource_items),
        by_type=by_type,
        resources=resource_items,
        grand_total=grand_total,
    )


@router.post(
    "/boqs/{boq_id}/enrich-resources/",
    summary="Enrich position resources",
    dependencies=[Depends(RequirePermission("boq.write"))],
)
async def enrich_resources(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Auto-populate metadata.resources for positions that don't have them.

    For each position without explicit resources:
    1. If metadata has cost_item_code → look up cost item → copy components
    2. Else → fuzzy match by description via _lookup_cost_item_components

    Returns count of enriched positions.
    """
    boq_data = await service.get_boq_with_positions(boq_id)
    cost_repo = CostItemRepository(session)
    enriched_count = 0
    total_positions = 0

    for pos in boq_data.positions:
        # Skip sections (positions with children / no unit)
        if not pos.unit:
            continue
        total_positions += 1

        # Pydantic strips the trailing underscore from the SQLAlchemy
        # column name — PositionResponse exposes it as `metadata`.
        raw_meta = getattr(pos, "metadata", None) or getattr(pos, "metadata_", None)
        meta = dict(raw_meta) if raw_meta else {}
        existing_resources = meta.get("resources")
        if isinstance(existing_resources, list) and len(existing_resources) > 0:
            continue  # Already has resources

        # Try lookup by cost_item_code first
        components: list[dict[str, Any]] = []
        cost_item_code = meta.get("cost_item_code")
        if cost_item_code:
            try:
                items, _, _ = await cost_repo.search(q=str(cost_item_code), limit=1)
                if items and items[0].components:
                    raw = items[0].components
                    if isinstance(raw, str):
                        import json as _json

                        raw = _json.loads(raw)
                    if isinstance(raw, list):
                        components = raw
            except Exception:
                logger.debug("Assembly expand: component lookup by code failed", exc_info=True)

        # Fallback: lookup by description
        if not components:
            components = await BOQService._lookup_cost_item_components(
                cost_repo, pos.description or ""
            )

        if components:
            resources = []
            for c in components:
                res = {
                    "name": c.get("name", ""),
                    "code": c.get("code", ""),
                    "type": c.get("type", "other"),
                    "unit": c.get("unit", ""),
                    "quantity": float(c.get("quantity", 0)),
                    "unit_rate": float(c.get("unit_rate", 0)),
                    "total": float(c.get("cost", 0))
                    or float(c.get("quantity", 0)) * float(c.get("unit_rate", 0)),
                }
                resources.append(res)

            meta["resources"] = resources
            await service.position_repo.update_fields(pos.id, metadata_=meta)
            enriched_count += 1

    await session.commit()

    return {
        "enriched_count": enriched_count,
        "total_positions": total_positions,
    }


@router.get(
    "/epd-materials/",
    summary="List EPD materials",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_epd_materials(
    category: str | None = Query(default=None, description="Filter by category"),
    search: str | None = Query(default=None, description="Search by name or ID"),
) -> dict[str, Any]:
    """List available EPD materials with optional filtering."""
    materials = search_epd_materials(category=category, query=search)
    return {
        "materials": materials,
        "categories": EPD_CATEGORIES,
        "total": len(materials),
    }


@router.post(
    "/boqs/{boq_id}/enrich-co2/",
    response_model=CO2EnrichResponse,
    summary="Enrich BOQ with CO2 data",
    dependencies=[Depends(RequirePermission("boq.write"))],
)
async def enrich_co2(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> CO2EnrichResponse:
    """Auto-detect EPD materials for all BOQ positions and store CO2 data.

    Loops through positions, matches descriptions to the 77 EPD materials,
    calculates GWP totals, and stores results in position metadata.
    Skips positions that already have manually assigned CO2 data.
    """
    boq_data = await service.get_boq_with_positions(boq_id)
    enriched = 0
    skipped = 0
    total = 0

    for pos in boq_data.positions:
        if not pos.unit:
            continue  # Skip section headers
        total += 1

        # Same Pydantic-vs-ORM duality as enrich-resources above.
        raw_meta = getattr(pos, "metadata", None) or getattr(pos, "metadata_", None)
        meta = dict(raw_meta) if raw_meta else {}
        existing_co2 = meta.get("co2", {})

        # Skip manually assigned (source=manual)
        if existing_co2.get("source") == "manual":
            skipped += 1
            continue

        epd = detect_epd_material(pos.description)
        if epd is None:
            skipped += 1
            continue

        qty = float(pos.quantity) if pos.quantity else 0.0
        gwp_total = qty * epd["gwp"]

        meta["co2"] = {
            "epd_id": epd["id"],
            "epd_name": epd["name"],
            "category": epd["category"],
            "gwp_per_unit": epd["gwp"],
            "gwp_total": round(gwp_total, 2),
            "unit": epd["unit"],
            "source": "auto",
            "stages": epd["stages"],
            "data_source": epd["source"],
        }
        await service.position_repo.update_fields(pos.id, metadata_=meta)
        enriched += 1

    await session.commit()
    return CO2EnrichResponse(enriched=enriched, skipped=skipped, total=total)


@router.put(
    "/positions/{position_id}/co2/",
    summary="Assign CO2 data to position",
    dependencies=[Depends(RequirePermission("boq.write"))],
)
async def assign_position_co2(
    position_id: uuid.UUID,
    payload: CO2AssignRequest,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Manually assign an EPD material to a BOQ position.

    Updates the position's metadata with CO2 data from the specified EPD material.
    """
    from fastapi import HTTPException

    epd = EPD_INDEX.get(payload.epd_id)
    if not epd:
        raise HTTPException(
            status_code=404,
            detail=f"EPD material '{payload.epd_id}' not found. Use GET /epd-materials to list available materials.",
        )

    pos = await service.position_repo.get_by_id(position_id)
    if not pos:
        raise HTTPException(status_code=404, detail="Position not found")

    meta = dict(pos.metadata_) if pos.metadata_ else {}
    qty = float(pos.quantity) if pos.quantity else 0.0
    gwp_total = qty * epd["gwp"]

    meta["co2"] = {
        "epd_id": epd["id"],
        "epd_name": epd["name"],
        "category": epd["category"],
        "gwp_per_unit": epd["gwp"],
        "gwp_total": round(gwp_total, 2),
        "unit": epd["unit"],
        "source": "manual",
        "stages": epd["stages"],
        "data_source": epd["source"],
    }
    await service.position_repo.update_fields(position_id, metadata_=meta)
    await session.commit()

    return {"status": "ok", "co2": meta["co2"]}


@router.get(
    "/boqs/{boq_id}/sustainability/",
    response_model=SustainabilityResponse,
    summary="Get sustainability analysis",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_sustainability(
    boq_id: uuid.UUID,
    area_m2: float = Query(default=0.0, ge=0.0, description="Project gross floor area in m2"),
    service: BOQService = Depends(_get_service),
) -> SustainabilityResponse:
    """Calculate CO2 emissions for a BOQ using EPD data.

    Two-pass approach:
    1. Use stored metadata.co2 data (from enrich-co2 or manual assignment)
    2. Fall back to auto-detection for positions without stored CO2

    Returns per-position detail, category breakdown, benchmarks, and EU CPR compliance.
    """
    boq_data = await service.get_boq_with_positions(boq_id)

    positions_detail: list[PositionCO2Detail] = []
    category_totals: dict[str, dict[str, Any]] = {}  # category -> {co2, qty, count}
    positions_matched = 0
    enriched_count = 0
    total_positions = 0

    for pos in boq_data.positions:
        if not pos.unit:
            continue  # Skip section headers
        total_positions += 1
        qty = float(pos.quantity) if pos.quantity else 0.0

        # Try stored CO2 data first
        stored = _get_position_co2(pos)
        if stored and stored.get("epd_id"):
            epd_id = stored["epd_id"]
            epd_name = stored.get("epd_name", "")
            gwp_per_unit = float(stored.get("gwp_per_unit", 0))
            gwp_total = qty * gwp_per_unit  # Recalculate with current quantity
            category = stored.get("category", "")
            source = "enriched" if stored.get("source") != "manual" else "manual"
            enriched_count += 1
            positions_matched += 1
        else:
            # Auto-detect fallback
            epd = detect_epd_material(pos.description)
            if epd:
                epd_id = epd["id"]
                epd_name = epd["name"]
                gwp_per_unit = epd["gwp"]
                gwp_total = qty * gwp_per_unit
                category = epd["category"]
                source = "auto-detected"
                positions_matched += 1
            else:
                epd_id = None
                epd_name = None
                gwp_per_unit = 0.0
                gwp_total = 0.0
                category = ""
                source = "none"

        positions_detail.append(
            PositionCO2Detail(
                position_id=str(pos.id),
                ordinal=pos.ordinal,
                description=pos.description[:120],
                quantity=qty,
                unit=pos.unit,
                epd_id=epd_id,
                epd_name=epd_name,
                gwp_per_unit=round(gwp_per_unit, 4),
                gwp_total=round(gwp_total, 2),
                category=category,
                source=source,
            )
        )

        # Accumulate by category
        if category and gwp_total != 0:
            if category not in category_totals:
                cat_info = next((c for c in EPD_CATEGORIES if c["id"] == category), {})
                category_totals[category] = {
                    "label": cat_info.get("label", category),
                    "co2": 0.0,
                    "qty": 0.0,
                    "count": 0,
                    "unit": pos.unit,
                }
            category_totals[category]["co2"] += gwp_total
            category_totals[category]["qty"] += qty
            category_totals[category]["count"] += 1

    total_co2_kg = sum(ct["co2"] for ct in category_totals.values())
    total_co2_tons = total_co2_kg / 1000.0
    abs_total = sum(abs(ct["co2"]) for ct in category_totals.values()) or 1.0

    # Build breakdown sorted by absolute CO2 descending
    breakdown: list[CO2MaterialBreakdown] = []
    for cat, info in sorted(category_totals.items(), key=lambda x: abs(x[1]["co2"]), reverse=True):
        breakdown.append(
            CO2MaterialBreakdown(
                material=info["label"],
                category=cat,
                quantity=round(info["qty"], 2),
                unit=info["unit"],
                co2_kg=round(info["co2"], 1),
                percentage=round(abs(info["co2"]) / abs_total * 100, 1),
                positions_count=info["count"],
            )
        )

    # Benchmark per m2 and rating
    benchmark: float | None = None
    rating = ""
    rating_label = ""
    eu_cpr_compliance = ""
    eu_cpr_gwp_per_m2_year: float | None = None

    if area_m2 > 0:
        benchmark = round(total_co2_kg / area_m2, 1)
        rating, rating_label = _co2_rating(benchmark)
        # EU CPR uses annualized value (50-year reference service period)
        eu_cpr_gwp_per_m2_year = round(total_co2_kg / area_m2 / 50, 2)
        eu_cpr_compliance = _eu_cpr_compliance(eu_cpr_gwp_per_m2_year)

    # Data quality
    if enriched_count == positions_matched and enriched_count > 0:
        data_quality = "enriched"
    elif enriched_count > 0:
        data_quality = "mixed"
    else:
        data_quality = "estimated"

    return SustainabilityResponse(
        total_co2_kg=round(total_co2_kg, 1),
        total_co2_tons=round(total_co2_tons, 2),
        breakdown=breakdown,
        benchmark_per_m2=benchmark,
        rating=rating,
        rating_label=rating_label,
        project_area_m2=area_m2 if area_m2 > 0 else None,
        positions_analyzed=total_positions,
        positions_matched=positions_matched,
        lifecycle_stages="A1-A3",
        data_quality=data_quality,
        positions_detail=positions_detail,
        eu_cpr_compliance=eu_cpr_compliance,
        eu_cpr_gwp_per_m2_year=eu_cpr_gwp_per_m2_year,
    )


# ── Cost Breakdown ──────────────────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/cost-breakdown/",
    response_model=CostBreakdownResponse,
    summary="Get cost breakdown",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_cost_breakdown(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> CostBreakdownResponse:
    """Get a cost breakdown for a BOQ split by resource category.

    Analyzes all positions in the BOQ and aggregates costs into categories:
    material, labor, equipment, subcontractor, and other. Each position's
    ``metadata.resources`` list is used when available; otherwise, the position
    description is classified via keyword heuristics.

    Returns:
        CostBreakdownResponse with direct cost categories, markup lines,
        grand total, and top 10 most expensive resources.
    """
    return await service.get_cost_breakdown(boq_id)


# ── Statistics ──────────────────────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/statistics/",
    summary="Get BOQ statistics",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_boq_statistics(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> dict:
    """Get aggregated statistics for a BOQ.

    Returns position count, section count, direct cost, grand total, average
    unit rate, completion percentage, unit/source breakdowns, and
    classification coverage.
    """

    result = await service.get_statistics(boq_id)
    return result.model_dump()


# ── Sensitivity Analysis (Tornado Chart) ─────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/sensitivity/",
    response_model=SensitivityResponse,
    summary="Get sensitivity analysis",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_sensitivity(
    boq_id: uuid.UUID,
    variation_pct: float = Query(
        default=10.0, gt=0.0, le=100.0, description="Cost variation percentage"
    ),
    top_n: int = Query(default=15, ge=1, le=50, description="Number of top positions to return"),
    service: BOQService = Depends(_get_service),
) -> SensitivityResponse:
    """Compute sensitivity analysis (tornado chart data) for a BOQ.

    For each non-section position, calculates the share of the total cost and
    the impact of a ``variation_pct`` increase or decrease.  Returns the top N
    positions sorted by descending impact.

    Args:
        boq_id: Target BOQ identifier.
        variation_pct: Percentage to vary each position's cost (default 10%).
        top_n: Maximum number of positions to return (default 15).

    Returns:
        SensitivityResponse with base_total, variation_pct, and ranked items.
    """
    boq_data = await service.get_boq_with_positions(boq_id)

    # Filter to non-section positions (positions that have a unit)
    items = [p for p in boq_data.positions if p.unit and p.unit.strip() != ""]

    # BUG-B-011: PositionResponse.total is now an exact Decimal. This
    # sensitivity model multiplies by a float factor, so work in float
    # locally (the exact value is preserved in storage / JSON response —
    # a ±10% sensitivity band does not need sub-cent exactness).
    base_total = float(sum(p.total for p in items))

    if base_total == 0 or len(items) == 0:
        return SensitivityResponse(
            base_total=0.0,
            variation_pct=variation_pct,
            items=[],
        )

    factor = variation_pct / 100.0

    sensitivity_items: list[SensitivityItem] = []
    for pos in items:
        pos_total = float(pos.total)
        share_pct = round(pos_total / base_total * 100, 2)
        impact = round(pos_total * factor, 2)
        sensitivity_items.append(
            SensitivityItem(
                ordinal=pos.ordinal,
                description=pos.description,
                total=round(pos_total, 2),
                share_pct=share_pct,
                impact_low=round(-impact, 2),
                impact_high=round(impact, 2),
            )
        )

    # Sort by absolute impact descending, take top N
    sensitivity_items.sort(key=lambda x: abs(x.impact_high), reverse=True)
    sensitivity_items = sensitivity_items[:top_n]

    return SensitivityResponse(
        base_total=round(base_total, 2),
        variation_pct=variation_pct,
        items=sensitivity_items,
    )


# ── AACE Estimate Classification ─────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/classification/",
    response_model=EstimateClassificationResponse,
    summary="Get AACE estimate classification",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_estimate_classification(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> EstimateClassificationResponse:
    """Get the AACE 18R-97 estimate classification for a BOQ.

    Auto-detects the estimate class (1-5) based on the number of positions,
    rate completeness, resource completeness, and classification coverage.

    Returns:
        EstimateClassificationResponse with class, accuracy range, definition
        level, methodology description, and underlying metrics.
    """
    return await service.get_estimate_classification(boq_id)


# ── Monte Carlo Cost Risk Analysis ───────────────────────────────────────────


def _pert_sample(low: float, mode: float, high: float) -> float:
    """Sample from a Beta-PERT distribution.

    Uses the standard PERT parameterization with lambda=4.

    Args:
        low: Minimum value (optimistic).
        mode: Most likely value.
        high: Maximum value (pessimistic).

    Returns:
        A random sample from the PERT distribution in [low, high].
    """
    if high <= low:
        return mode
    lam = 4.0
    alpha = 1.0 + lam * (mode - low) / (high - low)
    beta_param = 1.0 + lam * (high - mode) / (high - low)
    sample = random.betavariate(alpha, beta_param)
    return low + (high - low) * sample


@router.get(
    "/boqs/{boq_id}/cost-risk/",
    response_model=CostRiskResponse,
    summary="Monte Carlo cost risk simulation",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_cost_risk(
    boq_id: uuid.UUID,
    iterations: int = Query(
        default=1000, ge=100, le=10000, description="Number of Monte Carlo iterations"
    ),
    optimistic_pct: float = Query(
        default=15.0, ge=0.0, le=50.0, description="Optimistic cost reduction %"
    ),
    pessimistic_pct: float = Query(
        default=25.0, ge=0.0, le=100.0, description="Pessimistic cost increase %"
    ),
    service: BOQService = Depends(_get_service),
) -> CostRiskResponse:
    """Run a Monte Carlo cost risk simulation for a BOQ.

    For each iteration, every non-section position's total cost is sampled
    using a Beta-PERT distribution with:
        - optimistic = total * (1 - optimistic_pct/100)
        - most_likely = total
        - pessimistic = total * (1 + pessimistic_pct/100)

    After all iterations, percentiles (P10..P90) are computed, a histogram
    with ~20 bins is built, and the top risk drivers (positions contributing
    most to total variance) are identified.

    Contingency is defined as P80 - P50.  Recommended budget is P80.

    Args:
        boq_id: Target BOQ identifier.
        iterations: Number of simulation iterations (default 1000).
        optimistic_pct: Optimistic cost reduction percentage (default 15).
        pessimistic_pct: Pessimistic cost increase percentage (default 25).

    Returns:
        CostRiskResponse with percentiles, histogram, contingency, and risk drivers.
    """
    boq_data = await service.get_boq_with_positions(boq_id)

    # Filter to non-section positions (positions that have a unit)
    items = [p for p in boq_data.positions if p.unit and p.unit.strip() != ""]
    # BUG-B-011: PositionResponse.total is now an exact Decimal; this
    # Monte-Carlo model runs in float (the exact value is preserved in
    # storage / JSON response — a stochastic risk band does not need
    # sub-cent exactness).
    base_total = float(sum(p.total for p in items))

    if base_total == 0 or len(items) == 0:
        return CostRiskResponse(
            iterations=iterations,
            base_total=0.0,
            percentiles=CostRiskPercentiles(p10=0.0, p25=0.0, p50=0.0, p75=0.0, p80=0.0, p90=0.0),
            contingency_p80=0.0,
            contingency_pct=0.0,
            recommended_budget=0.0,
            histogram=[],
            risk_drivers=[],
        )

    opt_factor = 1.0 - optimistic_pct / 100.0
    pess_factor = 1.0 + pessimistic_pct / 100.0

    # Pre-compute per-position bounds
    position_bounds: list[tuple[float, float, float, str, str]] = []
    for pos in items:
        t = float(pos.total)
        position_bounds.append((t * opt_factor, t, t * pess_factor, pos.ordinal, pos.description))

    # Run Monte Carlo simulation
    iteration_totals: list[float] = []
    # Track per-position sampled values for variance analysis
    n_positions = len(position_bounds)
    position_sums: list[float] = [0.0] * n_positions
    position_sq_sums: list[float] = [0.0] * n_positions

    for _ in range(iterations):
        iter_total = 0.0
        for idx, (low, mode, high, _ordinal, _desc) in enumerate(position_bounds):
            sampled = _pert_sample(low, mode, high)
            iter_total += sampled
            position_sums[idx] += sampled
            position_sq_sums[idx] += sampled * sampled
        iteration_totals.append(iter_total)

    # Sort for percentile extraction
    iteration_totals.sort()

    def _percentile(sorted_data: list[float], pct: float) -> float:
        """Extract a percentile from sorted data using linear interpolation."""
        n = len(sorted_data)
        idx = pct / 100.0 * (n - 1)
        lower = int(idx)
        upper = min(lower + 1, n - 1)
        frac = idx - lower
        return sorted_data[lower] + frac * (sorted_data[upper] - sorted_data[lower])

    p10 = round(_percentile(iteration_totals, 10), 2)
    p25 = round(_percentile(iteration_totals, 25), 2)
    p50 = round(_percentile(iteration_totals, 50), 2)
    p75 = round(_percentile(iteration_totals, 75), 2)
    p80 = round(_percentile(iteration_totals, 80), 2)
    p90 = round(_percentile(iteration_totals, 90), 2)

    contingency_p80 = round(p80 - p50, 2)
    contingency_pct = round((contingency_p80 / p50 * 100) if p50 > 0 else 0.0, 1)

    # Build histogram with ~20 bins
    min_val = iteration_totals[0]
    max_val = iteration_totals[-1]
    num_bins = 20
    bin_width = (max_val - min_val) / num_bins if max_val > min_val else 1.0

    histogram: list[CostRiskHistogramBin] = []
    for i in range(num_bins):
        bin_start = min_val + i * bin_width
        bin_end = min_val + (i + 1) * bin_width
        count = 0
        for val in iteration_totals:
            if i == num_bins - 1:
                # Last bin includes the upper bound
                if bin_start <= val <= bin_end:
                    count += 1
            else:
                if bin_start <= val < bin_end:
                    count += 1
        histogram.append(
            CostRiskHistogramBin(
                bin_start=round(bin_start, 2),
                bin_end=round(bin_end, 2),
                count=count,
            )
        )

    # Calculate risk drivers — positions sorted by their share of total variance
    position_variances: list[tuple[float, str, str]] = []
    for idx in range(n_positions):
        mean = position_sums[idx] / iterations
        variance = (position_sq_sums[idx] / iterations) - (mean * mean)
        ordinal = position_bounds[idx][3]
        description = position_bounds[idx][4]
        position_variances.append((variance, ordinal, description))

    total_variance = sum(v[0] for v in position_variances)

    risk_drivers: list[CostRiskDriver] = []
    if total_variance > 0:
        position_variances.sort(key=lambda x: x[0], reverse=True)
        for variance, ordinal, description in position_variances[:10]:
            contribution_pct = round(variance / total_variance * 100, 1)
            risk_drivers.append(
                CostRiskDriver(
                    ordinal=ordinal,
                    description=description,
                    contribution_pct=contribution_pct,
                )
            )

    return CostRiskResponse(
        iterations=iterations,
        base_total=round(base_total, 2),
        percentiles=CostRiskPercentiles(p10=p10, p25=p25, p50=p50, p75=p75, p80=p80, p90=p90),
        contingency_p80=contingency_p80,
        contingency_pct=contingency_pct,
        recommended_budget=p80,
        histogram=histogram,
        risk_drivers=risk_drivers,
    )


# ── Custom Column Definitions ────────────────────────────────────────────────


@router.get(
    "/boqs/{boq_id}/columns/",
    summary="List custom columns",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_custom_columns(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> list[dict]:
    """List custom column definitions for a BOQ."""
    boq = await service.get_boq(boq_id)
    meta = boq.metadata_ if isinstance(boq.metadata_, dict) else {}
    return meta.get("custom_columns", [])


class CustomColumnCreate(BaseModel):
    """Request body for ``POST /boqs/{boq_id}/columns/``.

    Typed so a typo'd field (e.g. ``column_typ`` instead of ``column_type``)
    is rejected with a clear 422 instead of being silently dropped — the
    previous ``data: dict = Body(...)`` shape happily accepted unknown
    keys, which was a UX trap when the frontend evolved its schema. Using
    ``model_config = ConfigDict(extra='forbid')`` makes any unexpected
    field a validation error the user can act on.
    """

    model_config = ConfigDict(extra="forbid")

    name: str
    display_name: str | None = None
    column_type: Literal["text", "number", "date", "select", "calculated"] = "text"
    options: list[str] = Field(default_factory=list)
    sort_order: int | None = None  # Server assigns; accepted but overwritten.
    formula: str | None = None
    decimals: int | None = None
    # v2.9.x — semantic hints for region-specific number columns. Backend
    # stores them verbatim; the frontend's value getter does the maths
    # against ``position.metadata.resources[]`` at render time. Storing
    # the hint (rather than a precomputed value) keeps section subtotals
    # and live-editing of resources working without an invalidation step.
    derived: Literal["resource_sum", "percentage_of_unit_rate"] | None = None
    # `resource_role` filters which `metadata.resources[]` entries a derived
    # column aggregates. Accept either a single role or a list — the GAEB
    # "Sonstiges-EP" preset sweeps everything that isn't labor / material /
    # equipment, so it stores ``["other", "operator", "subcontractor"]`` to
    # keep Lohn + Material + Geräte + Sonstiges = unit_rate.
    resource_role: (
        Literal["material", "labor", "equipment", "operator", "subcontractor", "other"]
        | list[
            Literal["material", "labor", "equipment", "operator", "subcontractor", "other"]
        ]
        | None
    ) = None


@router.post(
    "/boqs/{boq_id}/columns/",
    summary="Add custom column",
    status_code=201,
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def add_custom_column(
    boq_id: uuid.UUID,
    payload: CustomColumnCreate,
    service: BOQService = Depends(_get_service),
) -> dict:
    """Add a custom column definition to a BOQ.

    Body: {"name": "supplier", "display_name": "Supplier", "column_type": "text", "options": []}
    """
    name = payload.name.strip().lower().replace(" ", "_")
    if not name or not name.isidentifier():
        raise HTTPException(400, "Invalid column name — use alphanumeric + underscore")

    reserved = {
        "ordinal",
        "description",
        "unit",
        "quantity",
        "unit_rate",
        "total",
        "id",
        "parent_id",
    }
    if name in reserved:
        raise HTTPException(400, f"Column name '{name}' is reserved")

    display_name = payload.display_name or name.replace("_", " ").title()
    column_type = payload.column_type

    options = list(payload.options or [])
    # v2.7.0/E — calculated columns carry a user-authored formula evaluated
    # client-side by the BOQ formula engine. Backend is purely a passthrough:
    # we store the formula string + display decimals and trust the frontend
    # to evaluate (the engine is CSP-safe and lives in the browser anyway).
    formula = (payload.formula or "") if column_type == "calculated" else ""
    decimals: int | None
    if column_type == "calculated":
        try:
            decimals = int(payload.decimals) if payload.decimals is not None else 2
        except (TypeError, ValueError):
            decimals = 2
        decimals = max(0, min(6, decimals))
    else:
        decimals = None

    from sqlalchemy.orm.attributes import flag_modified

    boq = await service.get_boq(boq_id)
    # Build a fresh metadata dict with a fresh list — both via deep copy so
    # SQLAlchemy doesn't see "same identity = no change". We then explicitly
    # flag_modified to defeat the JSON column's value-based dirty detection
    # which otherwise misses nested mutations.
    existing_meta = boq.metadata_ if isinstance(boq.metadata_, dict) else {}
    existing_columns = list(existing_meta.get("custom_columns", []))

    # Check uniqueness
    if any(c.get("name") == name for c in existing_columns):
        raise HTTPException(400, f"Column '{name}' already exists")

    col_def: dict = {
        "name": name,
        "display_name": display_name,
        "column_type": column_type,
        "options": list(options),
        "sort_order": len(existing_columns),
    }
    if column_type == "calculated":
        col_def["formula"] = str(formula)
        col_def["decimals"] = decimals
    # Forward semantic-derivation hints when supplied (GAEB Lohn/Material/
    # Geräte EP columns, ÖNORM Lohn-Anteil %). Frontend reads them on
    # render to compute the value from the position's resources.
    if payload.derived is not None:
        col_def["derived"] = payload.derived
    if payload.resource_role is not None:
        col_def["resource_role"] = payload.resource_role
    new_columns = [*existing_columns, col_def]
    new_meta = {**existing_meta, "custom_columns": new_columns}

    boq.metadata_ = new_meta
    flag_modified(boq, "metadata_")
    await service.session.flush()
    await service.session.commit()

    return col_def


@router.delete(
    "/boqs/{boq_id}/columns/{column_name}",
    summary="Delete custom column",
    status_code=204,
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def delete_custom_column(
    boq_id: uuid.UUID,
    column_name: str,
    service: BOQService = Depends(_get_service),
) -> None:
    """Remove a custom column definition (data in positions preserved)."""
    from sqlalchemy.orm.attributes import flag_modified

    boq = await service.get_boq(boq_id)
    existing_meta = boq.metadata_ if isinstance(boq.metadata_, dict) else {}
    existing_columns = list(existing_meta.get("custom_columns", []))
    new_columns = [c for c in existing_columns if c.get("name") != column_name]
    new_meta = {**existing_meta, "custom_columns": new_columns}

    boq.metadata_ = new_meta
    flag_modified(boq, "metadata_")
    await service.session.flush()
    await service.session.commit()


# ── Per-BOQ named variables ($GFA, $LABOR_RATE, …) ───────────────────────────
#
# Variables are scoped to a single BOQ document and live on
# ``boq.metadata_["variables"]`` as a JSON array. Two BOQs in the same
# project can have independent ``$GFA`` values. Used by the formula
# engine — ``=$GFA * 0.15`` resolves to a literal number at evaluation
# time. See plan ``inherited-knitting-dahl.md`` Phase B for design notes.

# Match the spec from the plan: uppercase, alnum + underscore, max 32 chars.
_VARIABLE_NAME_RE = re.compile(r"^[A-Z][A-Z0-9_]{0,31}$")
_MAX_VARIABLES_PER_BOQ = 50
_VARIABLE_TYPES = ("number", "text", "date")


class BOQVariable(BaseModel):
    """One named variable scoped to a BOQ document."""

    name: str = Field(
        ...,
        description="Uppercase identifier without the leading '$', e.g. 'GFA'.",
    )
    type: Literal["number", "text", "date"]
    value: str | float | int | None = None
    description: str | None = None

    @field_validator("name")
    @classmethod
    def _validate_name(cls, raw: str) -> str:
        # Strip a stray leading "$" — easier than rejecting it.
        cleaned = raw[1:] if raw.startswith("$") else raw
        cleaned = cleaned.strip()
        if not _VARIABLE_NAME_RE.match(cleaned):
            raise ValueError(
                "Variable name must be UPPER_SNAKE_CASE, 1–32 chars, "
                "starting with a letter. Got: " + raw,
            )
        return cleaned


def _coerce_variable_value(var: BOQVariable) -> str | float | int | None:
    """Sanitise a value to match the declared type. Stored values are
    used directly by the formula engine, so it's important that
    ``type=number`` actually means a number — not the string ``"42"``."""
    if var.value is None or var.value == "":
        return None
    if var.type == "number":
        try:
            return float(var.value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(
                400,
                f"Variable '${var.name}' is type=number but value is not numeric",
            ) from exc
    if var.type == "date":
        # Don't parse — accept any non-empty string. The formula engine
        # treats date variables opaquely (mostly for display in tooltips).
        return str(var.value)
    return str(var.value)


@router.get(
    "/boqs/{boq_id}/variables/",
    summary="List per-BOQ named variables",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_boq_variables(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> list[dict]:
    """Return the variables registered on this BOQ."""
    boq = await service.get_boq(boq_id)
    meta = boq.metadata_ if isinstance(boq.metadata_, dict) else {}
    raw = meta.get("variables", [])
    if not isinstance(raw, list):
        return []
    return raw


@router.put(
    "/boqs/{boq_id}/variables/",
    summary="Replace per-BOQ named variables",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def replace_boq_variables(
    boq_id: uuid.UUID,
    variables: list[BOQVariable] = Body(...),
    service: BOQService = Depends(_get_service),
) -> list[dict]:
    """Replace the entire variable list for a BOQ.

    The plan calls for whole-list replacement (vs per-row CRUD) — the
    list is small (≤50) and the editor UI sends the whole table back
    on save, so a single round-trip keeps state simple.
    """
    if len(variables) > _MAX_VARIABLES_PER_BOQ:
        raise HTTPException(
            400,
            f"Too many variables ({len(variables)} > {_MAX_VARIABLES_PER_BOQ})",
        )

    seen: set[str] = set()
    sanitised: list[dict] = []
    for var in variables:
        if var.name in seen:
            raise HTTPException(400, f"Duplicate variable name: ${var.name}")
        seen.add(var.name)
        sanitised.append(
            {
                "name": var.name,
                "type": var.type,
                "value": _coerce_variable_value(var),
                "description": (var.description or None),
            },
        )

    from sqlalchemy.orm.attributes import flag_modified

    boq = await service.get_boq(boq_id)
    existing_meta = dict(boq.metadata_) if isinstance(boq.metadata_, dict) else {}
    existing_meta["variables"] = sanitised
    boq.metadata_ = existing_meta
    flag_modified(boq, "metadata_")
    await service.session.flush()
    await service.session.commit()

    return sanitised


# ── Renumber positions (gap-of-10 ordinal scheme) ───────────────────────────


class RenumberRequest(BaseModel):
    """Options for the renumber endpoint.

    All fields are optional — omitting the body keeps the legacy behaviour
    (gap-of-10 scheme, padded ordinals) so existing clients keep working.
    """

    scheme: Literal["gap10", "gap100", "sequential", "dotted"] = "gap10"
    pad: bool = True


@router.post(
    "/boqs/{boq_id}/renumber/",
    summary="Renumber positions",
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def renumber_positions(
    boq_id: uuid.UUID,
    options: RenumberRequest | None = None,
    service: BOQService = Depends(_get_service),
) -> dict:
    """Renumber every position in a BoQ using one of several professional schemes.

    Supported schemes:

    * ``gap10`` (default) — ``01, 01.10, 01.20, 01.30, 02, 02.10`` — leaves
      room to insert ``01.15`` between two positions later without
      renumbering everything else. Standard German tender output convention.
    * ``gap100`` — ``01, 01.100, 01.200`` — same idea, even more headroom
      for very large BOQs that may grow significantly post-tender.
    * ``sequential`` — ``01, 01.01, 01.02, 01.03`` — compact and traditional;
      good for fixed-scope BOQs that won't get extra positions later.
    * ``dotted`` — ``1, 1.1, 1.2, 1.3`` — short-form decimal numbering
      common in NRM-style measurement.

    The ``pad`` option controls whether top-level section numbers are
    zero-padded to two digits (``01`` vs ``1``).

    Positions are processed in their current ``sort_order`` so the user's
    drag-and-drop order is preserved. Only the ``ordinal`` field is rewritten.
    """
    opts = options or RenumberRequest()

    # Step (gap) per scheme. Sequential and dotted have step=1; gap10/gap100
    # leave room to insert.
    step_per_scheme: dict[str, int] = {
        "gap10": 10,
        "gap100": 100,
        "sequential": 1,
        "dotted": 1,
    }
    step = step_per_scheme[opts.scheme]
    use_dotted = opts.scheme == "dotted"

    def _fmt_section(idx: int) -> str:
        if not opts.pad:
            return str(idx)
        return f"{idx:02d}"

    def _fmt_leaf_value(parent_ord: str, value: int) -> str:
        if use_dotted:
            return f"{parent_ord}.{value}"
        # Width: 2 digits for gap10/sequential, 3 digits for gap100
        width = 3 if opts.scheme == "gap100" else 2
        return f"{parent_ord}.{value:0{width}d}"

    def _fmt_top_leaf(value: int) -> str:
        # Top-level leaves without a parent section.
        if use_dotted:
            return str(value)
        width = 4 if opts.scheme in ("gap10", "gap100") else 2
        return f"{value:0{width}d}"

    boq_data = await service.get_boq_with_positions(boq_id)
    positions = list(boq_data.positions)

    def _is_section(pos: object) -> bool:
        """Mirror the canonical frontend isSection check (api.ts:136).

        Sections are stored with EITHER unit="" (demo seed convention) OR
        unit="section" (create_section endpoint convention) — handle both.
        """
        u = (getattr(pos, "unit", "") or "").strip().lower()
        return u == "" or u == "section"

    # Build hierarchy by parent_id. Top-level positions have parent_id=None.
    by_parent: dict[str | None, list] = {}
    for p in positions:
        key = str(p.parent_id) if p.parent_id else None
        by_parent.setdefault(key, []).append(p)

    # Sort each group by sort_order so renumber matches the on-screen order.
    for key in by_parent:
        by_parent[key].sort(key=lambda x: (x.sort_order or 0, x.ordinal or ""))

    updates: list[tuple[uuid.UUID, str]] = []  # (id, new_ordinal)
    section_idx = 0

    def _walk(parent_key: str | None, parent_ordinal: str | None) -> None:
        """Walk one branch of the hierarchy and assign ordinals."""
        nonlocal section_idx
        children = by_parent.get(parent_key, [])
        leaf_idx = 0
        for child in children:
            is_section = _is_section(child)
            if is_section:
                if parent_key is None:
                    section_idx += 1
                    new_ord = _fmt_section(section_idx)
                else:
                    leaf_idx += 1
                    new_ord = _fmt_leaf_value(parent_ordinal or "", leaf_idx * step)
            else:
                leaf_idx += 1
                if parent_ordinal:
                    new_ord = _fmt_leaf_value(parent_ordinal, leaf_idx * step)
                else:
                    new_ord = _fmt_top_leaf(leaf_idx * step)
            updates.append((child.id, new_ord))
            _walk(str(child.id), new_ord)

    _walk(None, None)

    # Apply via repository
    n_updated = 0
    for pos_id, new_ordinal in updates:
        await service.position_repo.update_fields(pos_id, ordinal=new_ordinal)
        n_updated += 1

    await service.session.commit()

    return {
        "renumbered": n_updated,
        "scheme": opts.scheme,
    }


# ── Vector / semantic memory endpoints ───────────────────────────────────
#
# These three routes plug the BOQ module into the cross-module semantic
# memory layer (see ``app/core/vector_index.py``).  They are intentionally
# uniform across every module that participates — only the adapter and
# the row loader differ.


@router.get(
    "/vector/status/",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def boq_vector_status() -> dict[str, Any]:
    """Return health + row count for the ``oe_boq_positions`` collection.

    Used by the admin panel and the global search status widget so the
    user can tell at a glance whether semantic search over BOQ positions
    is ready, partially indexed or empty.
    """
    from app.core.vector_index import COLLECTION_BOQ, collection_status

    return collection_status(COLLECTION_BOQ)


@router.post(
    "/vector/reindex/",
    dependencies=[Depends(RequirePermission("boq.write"))],
)
async def boq_vector_reindex(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    boq_id: uuid.UUID | None = Query(default=None),
    purge_first: bool = Query(default=False),
) -> dict[str, Any]:
    """Backfill the BOQ vector collection.

    Optional filters narrow the scope so users can reindex one project or
    even one BOQ at a time without re-embedding the entire tenant.  Set
    ``purge_first=true`` to wipe the matching subset before re-encoding —
    useful when the embedding model has changed.
    """
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from app.core.vector_index import reindex_collection
    from app.modules.boq.models import BOQ as BOQModel  # noqa: N811  -- domain class, not constant
    from app.modules.boq.models import Position
    from app.modules.boq.vector_adapter import boq_position_adapter

    stmt = select(Position).options(selectinload(Position.boq))
    if boq_id is not None:
        stmt = stmt.where(Position.boq_id == boq_id)
    elif project_id is not None:
        stmt = stmt.join(BOQModel, Position.boq_id == BOQModel.id).where(
            BOQModel.project_id == project_id
        )

    rows = list((await session.execute(stmt)).scalars().all())
    return await reindex_collection(
        boq_position_adapter,
        rows,
        purge_first=purge_first,
    )


@router.get(
    "/positions/{position_id}/similar/",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def boq_position_similar(
    position_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
    limit: int = Query(default=5, ge=1, le=20),
    cross_project: bool = Query(default=True),
) -> dict[str, Any]:
    """Return BOQ positions semantically similar to the given one.

    By default the search is **cross-project** — that's the highest-value
    use case: estimators want to find how a similar position was priced
    in past projects so they can reuse the unit rate.  Pass
    ``cross_project=false`` to limit the search to the same project.

    Returns a list of :class:`VectorHit` dicts plus the original row id
    so the frontend can highlight the source.
    """
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from app.core.vector_index import find_similar
    from app.modules.boq.models import Position
    from app.modules.boq.vector_adapter import boq_position_adapter

    stmt = select(Position).options(selectinload(Position.boq)).where(Position.id == position_id)
    row = (await session.execute(stmt)).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="Position not found")

    project_id = (
        str(row.boq.project_id) if row.boq is not None and row.boq.project_id is not None else None
    )
    hits = await find_similar(
        boq_position_adapter,
        row,
        project_id=project_id,
        cross_project=cross_project,
        limit=limit,
    )
    return {
        "source_id": str(position_id),
        "limit": limit,
        "cross_project": cross_project,
        "hits": [h.to_dict() for h in hits],
    }


# ── Project Intelligence (RFC 25) ───────────────────────────────────────────


@router.get(
    "/line-items/",
    response_model=list[LineItemResponse],
    summary="Top cost drivers by BOQ line (RFC 25)",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_line_items(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project scope"),
    group: str = Query(
        default="cost", description="Grouping strategy — reserved; defaults to 'cost'"
    ),
    top_n: int = Query(default=20, ge=1, le=200),
    service: BOQService = Depends(_get_service),
) -> list[LineItemResponse]:
    """Return the top-N cost drivers across every BOQ in the project."""
    await verify_project_access(project_id, user_id, session)
    rows = await service.get_line_items(project_id, group=group, top_n=top_n)
    return [LineItemResponse(**r) for r in rows]


@router.get(
    "/cost-rollup/",
    response_model=list[CostRollupItem],
    summary="Cost rollup by classification (RFC 25)",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_cost_rollup(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project scope"),
    group_by: str = Query(
        default="din276",
        description="Classification key: din276 | nrm | masterformat | cost_code",
    ),
    service: BOQService = Depends(_get_service),
) -> list[CostRollupItem]:
    """Roll up BOQ totals by classification code (DIN 276 by default)."""
    await verify_project_access(project_id, user_id, session)
    rows = await service.get_cost_rollup(project_id, group_by=group_by)
    return [CostRollupItem(**r) for r in rows]


@router.get(
    "/anomalies/",
    response_model=list[AnomalyResponse],
    summary="BOQ anomaly flags (RFC 25, statistical)",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_anomalies(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project scope"),
    service: BOQService = Depends(_get_service),
) -> list[AnomalyResponse]:
    """Statistical anomaly detection (z-score + IQR + format checks)."""
    await verify_project_access(project_id, user_id, session)
    rows = await service.get_anomalies(project_id)
    return [AnomalyResponse(**r) for r in rows]
