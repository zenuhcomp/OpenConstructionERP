"""‌⁠‍Tasks API routes.

Endpoints:
    GET    /                    - List tasks for a project
    POST   /                    - Create task
    GET    /my-tasks             - List tasks for the current user
    GET    /export               - Export tasks as Excel file
    GET    /template             - Download import template Excel file
    POST   /import/file          - Import tasks from Excel/CSV file
    GET    /{task_id}            - Get single task
    PATCH  /{task_id}            - Update task
    DELETE /{task_id}            - Delete task
    POST   /{task_id}/complete   - Mark task as completed
    PATCH  /{task_id}/bim-links  - Replace linked BIM element ids
"""

from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import UTC, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse

from app.core.bulk_ops import BulkAssignRequest, BulkDeleteRequest, BulkStatusRequest
from app.dependencies import CurrentUserId, RequirePermission, SessionDep, verify_project_access
from app.modules.tasks.schemas import (
    TaskBimLinkRequest,
    TaskCompleteRequest,
    TaskCreate,
    TaskResponse,
    TaskStatsResponse,
    TaskUpdate,
)
from app.modules.tasks.service import TaskService

router = APIRouter(tags=["tasks"])
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> TaskService:
    return TaskService(session)


def _compute_checklist_progress(checklist: list | None) -> float:
    """‌⁠‍Return completion percentage (0.0 - 100.0) for a checklist."""
    if not checklist:
        return 0.0
    total = len(checklist)
    done = sum(1 for c in checklist if isinstance(c, dict) and c.get("completed"))
    return round(done / total * 100, 1) if total > 0 else 0.0


def _compute_is_overdue(item: object) -> bool:
    """‌⁠‍Determine if a task is overdue based on due_date and status."""
    status = getattr(item, "status", "")
    if status == "completed":
        return False
    due_date = getattr(item, "due_date", None)
    if not due_date:
        return False
    try:
        today_str = datetime.now(UTC).strftime("%Y-%m-%d")
        return str(due_date) < today_str
    except (ValueError, TypeError):
        return False


def _to_response(
    item: object,
    name_map: dict[str, str] | None = None,
) -> TaskResponse:
    checklist = item.checklist or []  # type: ignore[attr-defined]
    responsible_id = str(item.responsible_id) if item.responsible_id else None  # type: ignore[attr-defined]
    assigned_to_name = name_map.get(responsible_id) if name_map is not None and responsible_id is not None else None
    # Derive completed_at from updated_at when status is "completed" (the model
    # doesn't have a dedicated column, so we approximate).
    completed_at: str | None = None
    if getattr(item, "status", "") == "completed" and getattr(item, "updated_at", None):
        completed_at = item.updated_at.isoformat()  # type: ignore[attr-defined]
    return TaskResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        task_type=item.task_type,  # type: ignore[attr-defined]
        title=item.title,  # type: ignore[attr-defined]
        description=item.description,  # type: ignore[attr-defined]
        checklist=checklist,
        checklist_progress=_compute_checklist_progress(checklist),
        responsible_id=responsible_id,
        persons_involved=item.persons_involved or [],  # type: ignore[attr-defined]
        due_date=item.due_date,  # type: ignore[attr-defined]
        milestone_id=item.milestone_id,  # type: ignore[attr-defined]
        meeting_id=item.meeting_id,  # type: ignore[attr-defined]
        status=item.status,  # type: ignore[attr-defined]
        priority=item.priority,  # type: ignore[attr-defined]
        result=item.result,  # type: ignore[attr-defined]
        is_private=item.is_private,  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        assigned_to=responsible_id,
        assigned_to_name=assigned_to_name,
        bim_element_ids=[str(x) for x in (getattr(item, "bim_element_ids", None) or [])],
        metadata=getattr(item, "metadata_", {}),
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
        completed_at=completed_at,
        is_overdue=_compute_is_overdue(item),
    )


async def _to_response_resolved(
    item: object,
    service: TaskService,
) -> TaskResponse:
    """``_to_response`` plus a single-row assignee-name lookup.

    Used by the single-item endpoints so a task assigned by UUID still
    renders the assignee's display name (not just the raw id).
    """
    names = await service.resolve_assignee_names([item])  # type: ignore[list-item]
    return _to_response(item, names)


@router.get("/", response_model=list[TaskResponse])
async def list_tasks(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    type_filter: str | None = Query(default=None, alias="type"),
    status_filter: str | None = Query(default=None, alias="status"),
    priority: str | None = Query(default=None),
    responsible_id: str | None = Query(default=None),
    meeting_id: str | None = Query(default=None),
    bim_element_id: str | None = Query(
        default=None,
        description=(
            "Filter tasks to those whose bim_element_ids JSON array contains "
            "this element id (used by the BIM viewer to show linked defects)."
        ),
    ),
    search: str | None = Query(
        default=None,
        max_length=200,
        description="Free-text search across title, description, and result fields.",
    ),
    service: TaskService = Depends(_get_service),
) -> list[TaskResponse]:
    """List tasks for a project with optional filters.

    Private tasks are only visible to their creator. When ``bim_element_id``
    is supplied, the result is filtered to tasks whose ``bim_element_ids``
    list includes that element id. All other filters still apply on top.
    """
    await verify_project_access(project_id, user_id, session)
    if bim_element_id:
        # JSON-contains filter — delegated to the service for dialect handling.
        tasks_with_bim = await service.get_tasks_for_bim_element(
            bim_element_id,
            project_id=project_id,
            current_user_id=user_id,
        )
        # Apply the remaining lightweight filters in memory. The element
        # filter is typically very selective (O(handful)) so this is fine.
        if type_filter is not None:
            tasks_with_bim = [t for t in tasks_with_bim if t.task_type == type_filter]
        if status_filter is not None:
            tasks_with_bim = [t for t in tasks_with_bim if t.status == status_filter]
        if priority is not None:
            tasks_with_bim = [t for t in tasks_with_bim if t.priority == priority]
        if responsible_id is not None:
            tasks_with_bim = [t for t in tasks_with_bim if str(t.responsible_id or "") == responsible_id]
        if meeting_id is not None:
            tasks_with_bim = [t for t in tasks_with_bim if t.meeting_id == meeting_id]
        if search and search.strip():
            needle = search.strip().lower()
            tasks_with_bim = [
                t
                for t in tasks_with_bim
                if needle in (t.title or "").lower()
                or needle in (t.description or "").lower()
                or needle in (t.result or "").lower()
            ]
        page = tasks_with_bim[offset : offset + limit]
        names = await service.resolve_assignee_names(page)
        return [_to_response(t, names) for t in page]

    tasks, _ = await service.list_tasks(
        project_id,
        current_user_id=user_id,
        offset=offset,
        limit=limit,
        task_type=type_filter,
        status_filter=status_filter,
        priority=priority,
        responsible_id=responsible_id,
        meeting_id=meeting_id,
        search=search,
    )
    names = await service.resolve_assignee_names(tasks)
    return [_to_response(t, names) for t in tasks]


@router.get("/my-tasks/", response_model=list[TaskResponse])
async def my_tasks(
    user_id: CurrentUserId,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    status_filter: str | None = Query(default=None, alias="status"),
    service: TaskService = Depends(_get_service),
) -> list[TaskResponse]:
    """List tasks assigned to the current user across all projects."""
    tasks, _ = await service.list_my_tasks(
        user_id,
        offset=offset,
        limit=limit,
        status_filter=status_filter,
    )
    names = await service.resolve_assignee_names(tasks)
    return [_to_response(t, names) for t in tasks]


@router.post("/", response_model=TaskResponse, status_code=201)
async def create_task(
    data: TaskCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("tasks.create")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Create a new task."""
    await verify_project_access(data.project_id, user_id, session)
    task = await service.create_task(data, user_id=user_id)
    return await _to_response_resolved(task, service)


@router.get("/stats/", response_model=TaskStatsResponse)
async def task_stats(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TaskService = Depends(_get_service),
) -> TaskStatsResponse:
    """Return summary statistics for tasks in a project.

    Includes total, breakdown by status/type/priority, overdue count,
    and average checklist progress.
    """
    await verify_project_access(project_id, user_id, session)
    return await service.get_stats(project_id, current_user_id=user_id)


# ── Export tasks as Excel ────────────────────────────────────────────────────


@router.get("/export/")
async def export_tasks(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TaskService = Depends(_get_service),
) -> StreamingResponse:
    """Export tasks for a project as Excel file."""
    await verify_project_access(project_id, user_id, session)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    tasks, _ = await service.list_tasks(project_id, current_user_id=user_id, offset=0, limit=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = [
        "Title",
        "Type",
        "Status",
        "Priority",
        "Assignee",
        "Due Date",
        "Created",
        "Checklist Progress",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, task in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=task.title)  # type: ignore[attr-defined]
        ws.cell(row=row_idx, column=2, value=task.task_type)  # type: ignore[attr-defined]
        ws.cell(row=row_idx, column=3, value=task.status)  # type: ignore[attr-defined]
        ws.cell(row=row_idx, column=4, value=task.priority)  # type: ignore[attr-defined]
        ws.cell(
            row=row_idx,
            column=5,
            value=str(task.responsible_id) if task.responsible_id else "",  # type: ignore[attr-defined]
        )
        ws.cell(row=row_idx, column=6, value=task.due_date)  # type: ignore[attr-defined]
        ws.cell(
            row=row_idx,
            column=7,
            value=str(task.created_at) if task.created_at else "",  # type: ignore[attr-defined]
        )
        # Checklist progress
        checklist = task.checklist or []  # type: ignore[attr-defined]
        total = len(checklist)
        done = sum(1 for c in checklist if isinstance(c, dict) and c.get("completed"))
        ws.cell(
            row=row_idx,
            column=8,
            value=f"{done}/{total}" if total > 0 else "",
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tasks_export.xlsx"'},
    )


# ── Import template ─────────────────────────────────────────────────────────


@router.get("/template/")
async def download_task_template() -> StreamingResponse:
    """Download an Excel template for task import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ["Title", "Type", "Status", "Priority", "Due Date", "Description"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Example row
    ws.cell(row=2, column=1, value="Review structural drawings for Level 5")
    ws.cell(row=2, column=2, value="task")
    ws.cell(row=2, column=3, value="open")
    ws.cell(row=2, column=4, value="high")
    ws.cell(row=2, column=5, value="2026-06-15")
    ws.cell(row=2, column=6, value="Check all beam dimensions against the spec")

    # Add a note sheet with valid values
    notes = wb.create_sheet("Valid Values")
    notes.cell(row=1, column=1, value="Type values:").font = Font(bold=True)
    for i, v in enumerate(["task", "topic", "information", "decision", "personal"], 2):
        notes.cell(row=i, column=1, value=v)
    notes.cell(row=1, column=3, value="Status values:").font = Font(bold=True)
    for i, v in enumerate(["draft", "open", "in_progress", "completed"], 2):
        notes.cell(row=i, column=3, value=v)
    notes.cell(row=1, column=5, value="Priority values:").font = Font(bold=True)
    for i, v in enumerate(["low", "normal", "high", "urgent"], 2):
        notes.cell(row=i, column=5, value=v)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tasks_import_template.xlsx"'},
    )


# ── Import tasks from file ──────────────────────────────────────────────────

_TASK_COLUMN_MAP: dict[str, str] = {
    "title": "title",
    "name": "title",
    "task name": "title",
    "task": "title",
    "titel": "title",
    "aufgabe": "title",
    "type": "task_type",
    "task type": "task_type",
    "typ": "task_type",
    "status": "status",
    "priority": "priority",
    "priorität": "priority",
    "due date": "due_date",
    "due": "due_date",
    "fällig": "due_date",
    "deadline": "due_date",
    "description": "description",
    "beschreibung": "description",
    "details": "description",
    "название": "title",
    "заголовок": "title",
    "описание": "description",
    "статус": "status",
    "приоритет": "priority",
    "исполнитель": "assigned_to",
    "назначено": "assigned_to",
    "срок": "due_date",
    "дата окончания": "due_date",
    "дедлайн": "due_date",
    "тип": "task_type",
    "часы": "estimated_hours",
    "оценка часов": "estimated_hours",
    "titre": "title",
    "nom": "title",
    "statut": "status",
    "état": "status",
    "priorité": "priority",
    "assigné à": "assigned_to",
    "responsable": "assigned_to",
    "échéance": "due_date",
    "date limite": "due_date",
    "heures estimées": "estimated_hours",
    "título": "title",
    "descripción": "description",
    "estado": "status",
    "prioridad": "priority",
    "asignado a": "assigned_to",
    "fecha límite": "due_date",
    "vencimiento": "due_date",
    "tipo": "task_type",
    "horas estimadas": "estimated_hours",
    "titolo": "title",
    "descrizione": "description",
    "stato": "status",
    "priorità": "priority",
    "assegnato a": "assigned_to",
    "scadenza": "due_date",
    "ore stimate": "estimated_hours",
    "タイトル": "title",
    "説明": "description",
    "ステータス": "status",
    "状態": "status",
    "優先度": "priority",
    "担当者": "assigned_to",
    "期限": "due_date",
    "締切": "due_date",
    "タイプ": "task_type",
    "種類": "task_type",
    "見積時間": "estimated_hours",
    "标题": "title",
    "名称": "title",
    "描述": "description",
    "说明": "description",
    "状态": "status",
    "优先级": "priority",
    "负责人": "assigned_to",
    "分配给": "assigned_to",
    "截止日期": "due_date",
    "到期日": "due_date",
    "类型": "task_type",
    "预估工时": "estimated_hours",
    "descrição": "description",
    "atribuído a": "assigned_to",
    "responsável": "assigned_to",
    "prazo": "due_date",
    "data limite": "due_date",
    "prioridade": "priority",
    "beschrijving": "description",
    "prioriteit": "priority",
    "toegewezen aan": "assigned_to",
    "vervaldatum": "due_date",
    "geschatte uren": "estimated_hours",
    "제목": "title",
    "설명": "description",
    "상태": "status",
    "우선순위": "priority",
    "담당자": "assigned_to",
    "마감일": "due_date",
    "유형": "task_type",
    "예상 시간": "estimated_hours",
}

_VALID_TASK_TYPES = {"task", "topic", "information", "decision", "personal"}
_VALID_STATUSES = {"draft", "open", "in_progress", "completed"}
_VALID_PRIORITIES = {"low", "normal", "high", "urgent"}
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _match_task_column(header: str) -> str | None:
    """Match a header string to a canonical task column name."""
    return _TASK_COLUMN_MAP.get(header.strip().lower().replace("_", " "))


def _parse_task_rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    """Parse CSV content into a list of row dicts."""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("No headers found")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_task_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None and str(val).strip():
                row[canonical] = str(val).strip()
        if row:
            rows.append(row)
    return rows


def _parse_task_rows_from_excel(content: bytes) -> list[dict[str, Any]]:
    """Parse Excel (.xlsx) content into a list of row dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("No active sheet found")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        wb.close()
        raise ValueError("No headers found")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_task_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None and str(val).strip():
                row[canonical] = str(val).strip()
        if row:
            rows.append(row)

    wb.close()
    return rows


@router.post("/import/file/")
async def import_tasks_file(
    user_id: CurrentUserId,
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    _perm: None = Depends(RequirePermission("tasks.create")),
    service: TaskService = Depends(_get_service),
) -> dict[str, Any]:
    """Import tasks from an Excel or CSV file upload.

    Expected columns (flexible auto-detection):
    - **Title / Task Name** -- task title (required)
    - **Type / Task Type** -- task type (task, topic, information, decision, personal)
    - **Status** -- status (draft, open, in_progress, completed)
    - **Priority** -- priority (low, normal, high, urgent)
    - **Due Date / Deadline** -- due date in YYYY-MM-DD format
    - **Description / Details** -- task description

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    await verify_project_access(project_id, user_id, session)
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # Magic-byte sniff — the extension is hostile-supplied, so reject any
    # payload whose first bytes don't match the declared format before we
    # hand the buffer to openpyxl / csv.reader. Mirrors the contacts
    # importer (R7 audit pattern). CSV has no canonical signature; reject
    # an obvious-binary leading byte instead.
    head = content[:8]
    if filename.endswith(".xlsx"):
        if not head.startswith(b"PK\x03\x04"):
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="File does not look like a valid .xlsx (missing ZIP signature).",
            )
    elif filename.endswith(".xls"):
        if not head.startswith(b"\xd0\xcf\x11\xe0"):
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="File does not look like a valid .xls (missing OLE signature).",
            )
    else:  # .csv
        for sig in (b"MZ", b"\x7fELF", b"\xca\xfe\xba\xbe", b"PK\x03\x04", b"\xd0\xcf\x11\xe0"):
            if head.startswith(sig):
                raise HTTPException(
                    status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                    detail="File does not look like CSV (binary signature detected).",
                )

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_task_rows_from_excel(content)
        else:
            rows = _parse_task_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing task import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    imported_count = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            title = str(row.get("title", "")).strip()
            if not title:
                skipped += 1
                continue

            task_type = str(row.get("task_type", "task")).strip().lower()
            if task_type not in _VALID_TASK_TYPES:
                task_type = "task"

            task_status = str(row.get("status", "open")).strip().lower()
            if task_status not in _VALID_STATUSES:
                task_status = "open"

            priority = str(row.get("priority", "normal")).strip().lower()
            if priority not in _VALID_PRIORITIES:
                priority = "normal"

            due_date = str(row.get("due_date", "")).strip() or None
            if due_date and not _DATE_RE.match(due_date):
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Invalid date format: {due_date} (expected YYYY-MM-DD)",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue

            description = str(row.get("description", "")).strip() or None

            create_data = TaskCreate(
                project_id=project_id,
                task_type=task_type,
                title=title,
                description=description,
                status=task_status,
                priority=priority,
                due_date=due_date,
            )
            await service.create_task(create_data, user_id=user_id)
            imported_count += 1
        except Exception as exc:
            errors.append(
                {
                    "row": row_idx,
                    "error": str(exc)[:200],
                    "data": {k: str(v)[:100] for k, v in row.items()},
                }
            )

    return {
        "imported": imported_count,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Bulk operations (must come BEFORE parametric /{task_id}) ───────────


async def _filter_owned_task_ids(
    session: SessionDep,
    user_id: str,
    task_ids: list[uuid.UUID],
) -> list[uuid.UUID]:
    """Return the subset of task_ids that belong to projects owned by user_id."""
    from sqlalchemy import select as _select

    from app.modules.projects.repository import ProjectRepository
    from app.modules.tasks.models import Task

    proj_repo = ProjectRepository(session)
    owned_projects, _ = await proj_repo.list_for_user(owner_id=user_id, offset=0, limit=10000, exclude_archived=False)
    owned_project_ids = {str(p.id) for p in owned_projects}

    rows = (await session.execute(_select(Task.id, Task.project_id).where(Task.id.in_(task_ids)))).all()
    return [r[0] for r in rows if str(r[1]) in owned_project_ids]


@router.post(
    "/batch/delete/",
    status_code=200,
    dependencies=[Depends(RequirePermission("tasks.delete"))],
)
async def batch_delete_tasks(
    body: BulkDeleteRequest,
    user_id: CurrentUserId,
    session: SessionDep,
) -> dict:
    """Delete multiple tasks in one request. Only project-owned tasks are deleted."""
    from app.core.bulk_ops import bulk_delete
    from app.modules.tasks.models import Task

    allowed = await _filter_owned_task_ids(session, user_id, body.ids)
    deleted = await bulk_delete(session, Task, allowed)
    logger.info(
        "Bulk delete tasks: requested=%d deleted=%d user=%s",
        len(body.ids),
        deleted,
        user_id,
    )
    return {"requested": len(body.ids), "deleted": deleted}


@router.patch(
    "/batch/status/",
    status_code=200,
    dependencies=[Depends(RequirePermission("tasks.update"))],
)
async def batch_update_task_status(
    body: BulkStatusRequest,
    user_id: CurrentUserId,
    session: SessionDep,
) -> dict:
    """Bulk-update status on multiple tasks."""
    from app.core.bulk_ops import bulk_update_status
    from app.modules.tasks.models import Task

    allowed_statuses = {"draft", "open", "in_progress", "completed"}
    if body.status not in allowed_statuses:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid status. Allowed: {sorted(allowed_statuses)}",
        )

    allowed_ids = await _filter_owned_task_ids(session, user_id, body.ids)
    updated = await bulk_update_status(session, Task, allowed_ids, body.status, allowed_statuses=allowed_statuses)
    logger.info(
        "Bulk update task status: requested=%d updated=%d new_status=%s user=%s",
        len(body.ids),
        updated,
        body.status,
        user_id,
    )
    return {"requested": len(body.ids), "updated": updated, "status": body.status}


@router.post(
    "/batch/assign/",
    status_code=200,
    dependencies=[Depends(RequirePermission("tasks.update"))],
)
async def batch_assign_tasks(
    body: BulkAssignRequest,
    user_id: CurrentUserId,
    session: SessionDep,
) -> dict:
    """Bulk-assign multiple tasks to a single user."""
    from app.core.bulk_ops import bulk_update_fields
    from app.modules.tasks.models import Task

    allowed_ids = await _filter_owned_task_ids(session, user_id, body.ids)
    updated = await bulk_update_fields(session, Task, allowed_ids, {"responsible_id": body.assignee_id})
    logger.info(
        "Bulk assign tasks: requested=%d updated=%d assignee=%s user=%s",
        len(body.ids),
        updated,
        body.assignee_id,
        user_id,
    )
    return {"requested": len(body.ids), "updated": updated, "assignee_id": body.assignee_id}


@router.get("/{task_id}", response_model=TaskResponse)
async def get_task(
    task_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.read")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Get a single task. Private tasks are only visible to their creator."""
    task = await service.get_task(task_id, current_user_id=user_id)
    await verify_project_access(task.project_id, user_id, session)
    return await _to_response_resolved(task, service)


@router.patch("/{task_id}", response_model=TaskResponse)
async def update_task(
    task_id: uuid.UUID,
    data: TaskUpdate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.update")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Update a task."""
    existing = await service.get_task(task_id, current_user_id=user_id)
    await verify_project_access(existing.project_id, user_id, session)
    task = await service.update_task(task_id, data, current_user_id=user_id)
    return await _to_response_resolved(task, service)


@router.delete("/{task_id}", status_code=204)
async def delete_task(
    task_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.delete")),
    service: TaskService = Depends(_get_service),
) -> None:
    """Delete a task."""
    existing = await service.get_task(task_id, current_user_id=user_id)
    await verify_project_access(existing.project_id, user_id, session)
    await service.delete_task(task_id, current_user_id=user_id)


@router.post("/{task_id}/complete/", response_model=TaskResponse)
async def complete_task(
    task_id: uuid.UUID,
    session: SessionDep,
    body: TaskCompleteRequest | None = None,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.update")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Mark a task as completed with optional result text."""
    existing = await service.get_task(task_id, current_user_id=user_id)
    await verify_project_access(existing.project_id, user_id, session)
    result = body.result if body else None
    task = await service.complete_task(task_id, result=result, current_user_id=user_id)
    return await _to_response_resolved(task, service)


@router.patch("/{task_id}/bim-links/", response_model=TaskResponse)
async def update_task_bim_links(
    task_id: uuid.UUID,
    body: TaskBimLinkRequest,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.update")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Replace the full set of BIM element ids linked to this task.

    Idempotent set semantics — the incoming ``bim_element_ids`` list
    fully overwrites the previously stored list. Sending an empty list
    clears all links.
    """
    existing = await service.get_task(task_id, current_user_id=user_id)
    await verify_project_access(existing.project_id, user_id, session)
    task = await service.update_bim_links(
        task_id,
        body.bim_element_ids,
        current_user_id=user_id,
    )
    return await _to_response_resolved(task, service)


# ── Vector / semantic memory endpoints ───────────────────────────────────
#
# ``/vector/status/`` + ``/vector/reindex/`` are wired via the shared
# factory in ``app.core.vector_routes`` (see the ``include_router`` call
# at the bottom of this file).  The ``/{id}/similar/`` endpoint below
# stays module-specific.


@router.get(
    "/{task_id}/similar/",
    dependencies=[Depends(RequirePermission("tasks.read"))],
)
async def tasks_similar(
    task_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
    limit: int = Query(default=5, ge=1, le=20),
    cross_project: bool = Query(default=True),
) -> dict[str, Any]:
    """Return tasks semantically similar to the given one.

    By default the search is **cross-project** — that's the highest-value
    use case: users want to find how a similar task / defect / inspection
    was handled in past projects so they can reuse the resolution.  Pass
    ``cross_project=false`` to limit the search to the same project.

    Returns a list of :class:`VectorHit` dicts plus the original row id
    so the frontend can highlight the source.
    """
    from sqlalchemy import select

    from app.core.vector_index import find_similar
    from app.modules.tasks.models import Task
    from app.modules.tasks.vector_adapter import task_vector_adapter

    stmt = select(Task).where(Task.id == task_id)
    row = (await session.execute(stmt)).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="Task not found")

    project_id = str(row.project_id) if row.project_id is not None else None
    hits = await find_similar(
        task_vector_adapter,
        row,
        project_id=project_id,
        cross_project=cross_project,
        limit=limit,
    )
    return {
        "source_id": str(task_id),
        "limit": limit,
        "cross_project": cross_project,
        "hits": [h.to_dict() for h in hits],
    }


# ── Mount vector status + reindex via the shared factory ────────────────
from app.core.vector_index import COLLECTION_TASKS  # noqa: E402
from app.core.vector_routes import create_vector_routes  # noqa: E402
from app.modules.tasks.models import Task as _TaskModel  # noqa: E402
from app.modules.tasks.vector_adapter import (  # noqa: E402
    task_vector_adapter as _task_vector_adapter,
)

router.include_router(
    create_vector_routes(
        collection=COLLECTION_TASKS,
        adapter=_task_vector_adapter,
        model=_TaskModel,
        read_permission="tasks.read",
        write_permission="tasks.update",
        project_id_attr="project_id",
    )
)
