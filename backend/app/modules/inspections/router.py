"""‌⁠‍Inspections API routes.

Endpoints:
    GET    /                         - List inspections for a project
    POST   /                         - Create inspection
    GET    /export                   - Export inspections as Excel
    GET    /{inspection_id}          - Get single inspection
    PATCH  /{inspection_id}          - Update inspection
    DELETE /{inspection_id}          - Delete inspection
    POST   /{inspection_id}/complete - Mark inspection as completed
"""

import io
import logging
import uuid

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from app.dependencies import CurrentUserId, RequirePermission, SessionDep, verify_project_access
from app.modules.inspections.schemas import (
    InspectionCreate,
    InspectionResponse,
    InspectionUpdate,
)
from app.modules.inspections.service import InspectionService

router = APIRouter(tags=["inspections"])
logger = logging.getLogger(__name__)


class CompleteInspectionRequest(BaseModel):
    """‌⁠‍Request body for completing an inspection."""

    result: str = Field(default="pass", pattern=r"^(pass|fail|partial)$")


def _get_service(session: SessionDep) -> InspectionService:
    return InspectionService(session)


def _to_response(item: object) -> InspectionResponse:
    """‌⁠‍Build an InspectionResponse from a QualityInspection ORM object."""
    return InspectionResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        inspection_number=item.inspection_number,  # type: ignore[attr-defined]
        inspection_type=item.inspection_type,  # type: ignore[attr-defined]
        title=item.title,  # type: ignore[attr-defined]
        description=item.description,  # type: ignore[attr-defined]
        location=item.location,  # type: ignore[attr-defined]
        wbs_id=item.wbs_id,  # type: ignore[attr-defined]
        inspector_id=str(item.inspector_id) if item.inspector_id else None,  # type: ignore[attr-defined]
        inspection_date=item.inspection_date,  # type: ignore[attr-defined]
        status=item.status,  # type: ignore[attr-defined]
        result=item.result,  # type: ignore[attr-defined]
        checklist_data=item.checklist_data or [],  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        metadata=getattr(item, "metadata_", {}),
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
    )


@router.get("/", response_model=list[InspectionResponse])
async def list_inspections(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    type_filter: str | None = Query(default=None, alias="type"),
    status_filter: str | None = Query(default=None, alias="status"),
    service: InspectionService = Depends(_get_service),
) -> list[InspectionResponse]:
    """List inspections for a project with optional filters."""
    await verify_project_access(project_id, user_id, session)
    inspections, _ = await service.list_inspections(
        project_id,
        offset=offset,
        limit=limit,
        inspection_type=type_filter,
        status_filter=status_filter,
    )
    return [_to_response(i) for i in inspections]


@router.post("/", response_model=InspectionResponse, status_code=201)
async def create_inspection(
    data: InspectionCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("inspections.create")),
    service: InspectionService = Depends(_get_service),
) -> InspectionResponse:
    """Create a new quality inspection."""
    await verify_project_access(data.project_id, user_id, session)
    inspection = await service.create_inspection(data, user_id=user_id)
    return _to_response(inspection)


@router.get("/export/")
async def export_inspections(
    project_id: uuid.UUID = Query(...),
    session: SessionDep = None,  # type: ignore[assignment]
    _user: CurrentUserId = None,  # type: ignore[assignment]
) -> StreamingResponse:
    """Export inspections for a project as Excel."""
    await verify_project_access(project_id, _user, session)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from sqlalchemy import select

    from app.modules.inspections.models import QualityInspection

    result = await session.execute(
        select(QualityInspection)
        .where(QualityInspection.project_id == project_id)
        .order_by(QualityInspection.inspection_number)
        .limit(50000)
    )
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspections"

    headers = [
        "Inspection #",
        "Title",
        "Type",
        "Inspector",
        "Date",
        "Location",
        "Status",
        "Result",
        "Checklist Items (pass/fail)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.inspection_number)
        ws.cell(row=row_idx, column=2, value=item.title)
        ws.cell(row=row_idx, column=3, value=item.inspection_type)
        ws.cell(row=row_idx, column=4, value=str(item.inspector_id) if item.inspector_id else "")
        ws.cell(row=row_idx, column=5, value=item.inspection_date or "")
        ws.cell(row=row_idx, column=6, value=item.location or "")
        ws.cell(row=row_idx, column=7, value=item.status)
        ws.cell(row=row_idx, column=8, value=item.result or "")
        # Checklist pass/fail count
        checklist = item.checklist_data or []
        passed = sum(1 for ci in checklist if isinstance(ci, dict) and ci.get("passed"))
        failed = len(checklist) - passed
        ws.cell(row=row_idx, column=9, value=f"{passed}/{failed}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="inspections.xlsx"'},
    )


@router.get("/{inspection_id}", response_model=InspectionResponse)
async def get_inspection(
    inspection_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: InspectionService = Depends(_get_service),
) -> InspectionResponse:
    """Get a single inspection."""
    inspection = await service.get_inspection(inspection_id)
    await verify_project_access(inspection.project_id, str(user_id), session)
    return _to_response(inspection)


@router.patch("/{inspection_id}", response_model=InspectionResponse)
async def update_inspection(
    inspection_id: uuid.UUID,
    data: InspectionUpdate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("inspections.update")),
    service: InspectionService = Depends(_get_service),
) -> InspectionResponse:
    """Update an inspection."""
    existing = await service.get_inspection(inspection_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    inspection = await service.update_inspection(inspection_id, data)
    return _to_response(inspection)


@router.delete("/{inspection_id}", status_code=204)
async def delete_inspection(
    inspection_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("inspections.delete")),
    service: InspectionService = Depends(_get_service),
) -> None:
    """Delete an inspection."""
    existing = await service.get_inspection(inspection_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    await service.delete_inspection(inspection_id)


@router.post("/{inspection_id}/create-defect/", status_code=201)
async def create_defect_from_inspection(
    inspection_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("inspections.update")),
    service: InspectionService = Depends(_get_service),
) -> dict:
    """Create a punchlist item pre-filled from a failed inspection.

    The inspection must have result='fail' or 'partial'. Pre-fills the
    punchlist item with the inspection's title, location, and checklist
    details for failed items.
    """
    inspection = await service.get_inspection(inspection_id)
    # Cross-tenant guard: resolve the inspection's project, then run the
    # standard ownership check. ``verify_project_access`` raises 404 on
    # both "missing project" and "not yours" so this endpoint cannot be
    # used as an inspection-UUID enumeration oracle.
    await verify_project_access(inspection.project_id, str(user_id), session)

    if inspection.result not in ("fail", "partial"):
        from fastapi import HTTPException

        raise HTTPException(
            status_code=400,
            detail="Only failed or partial inspections can create defects.",
        )

    # Accept both ``passed`` (legacy) and ``response`` (schema) fields
    # when deciding which checklist items failed. ``response`` of
    # no/fail/false/0/failed counts as a failure.
    def _is_failed(ci: dict) -> bool:
        if "passed" in ci:
            return ci.get("passed") is False
        resp = str(ci.get("response", "")).strip().lower()
        if not resp:
            return False
        return resp in {"no", "fail", "false", "0", "failed"}

    # Build description from failed checklist items
    checklist = inspection.checklist_data or []
    failed_items = [
        ci.get("question", ci.get("description", "Unknown item"))
        for ci in checklist
        if isinstance(ci, dict) and _is_failed(ci)
    ]
    description_parts = [
        f"Defect from inspection {inspection.inspection_number}: {inspection.title}",
    ]
    if failed_items:
        description_parts.append("Failed items:")
        for item in failed_items:
            description_parts.append(f"  - {item}")

    description = "\n".join(description_parts)

    # Lazy import punchlist module
    try:
        from app.modules.punchlist.models import PunchItem

        punch_item = PunchItem(
            project_id=inspection.project_id,
            title=f"Defect: {inspection.title}",
            description=description,
            priority="high" if inspection.result == "fail" else "medium",
            status="open",
            category=inspection.inspection_type
            if inspection.inspection_type in ("structural", "electrical", "plumbing", "fire_safety", "general")
            else "general",
            trade=inspection.inspection_type,
            created_by=str(user_id),
            metadata_={
                "source": "inspection",
                "inspection_id": str(inspection_id),
                "inspection_number": inspection.inspection_number,
            },
        )
        # Copy location if available
        if inspection.location:
            punch_item.title = f"Defect: {inspection.title} @ {inspection.location}"
        session.add(punch_item)
        await session.flush()

        logger.info(
            "Created punchlist item %s from inspection %s",
            punch_item.id,
            inspection_id,
        )
        return {
            "punch_item_id": str(punch_item.id),
            "inspection_id": str(inspection_id),
            "title": punch_item.title,
        }
    except ImportError:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=501,
            detail="Punchlist module is not available.",
        )
    except Exception as exc:
        logger.exception("Failed to create defect from inspection %s: %s", inspection_id, exc)
        from fastapi import HTTPException

        raise HTTPException(
            status_code=500,
            detail="Failed to create punchlist item from inspection.",
        )


@router.post("/{inspection_id}/create-ncr/", status_code=201)
async def create_ncr_from_inspection(
    inspection_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("inspections.update")),
    service: InspectionService = Depends(_get_service),
) -> dict:
    """Pre-fill a Non-Conformance Report from a failed inspection.

    Punchlist (``create-defect``) covers minor defects; NCR is the
    formal channel for non-conformances that need root-cause analysis,
    corrective + preventive action, and engineer signoff. ``critical``
    failed checklist items lift severity from ``major`` to ``critical``.

    Idempotent: an NCR linked to this inspection (``linked_inspection_id``)
    is returned as-is rather than duplicated. Re-firing on a re-failed
    re-inspection therefore needs the prior NCR closed first.
    """
    inspection = await service.get_inspection(inspection_id)
    # Cross-tenant guard — must precede any business-logic branch that
    # could mutate (or leak) the inspection's parent project.
    await verify_project_access(inspection.project_id, str(user_id), session)

    if inspection.result not in ("fail", "partial"):
        from fastapi import HTTPException

        raise HTTPException(
            status_code=400,
            detail="Only failed or partial inspections can create an NCR.",
        )

    # Idempotency check: existing NCR linked to this inspection?
    from sqlalchemy import select

    from app.modules.ncr.models import NCR
    from app.modules.ncr.schemas import NCRCreate
    from app.modules.ncr.service import NCRService

    existing = (
        (await session.execute(select(NCR).where(NCR.linked_inspection_id == str(inspection_id)))).scalars().first()
    )
    if existing is not None:
        return {
            "ncr_id": str(existing.id),
            "ncr_number": existing.ncr_number,
            "inspection_id": str(inspection_id),
            "status": existing.status,
            "created": False,
        }

    # Accept both conventions: ChecklistEntry schema uses ``response``
    # (yes/no/pass/fail/...) but some clients embed a legacy ``passed``
    # boolean. Treat absent / yes / pass / true as passed.
    def _is_failed(ci: dict) -> bool:
        if "passed" in ci:
            return ci.get("passed") is False
        resp = str(ci.get("response", "")).strip().lower()
        if not resp:
            return False
        return resp in {"no", "fail", "false", "0", "failed"}

    checklist = inspection.checklist_data or []
    failed_items = [ci for ci in checklist if isinstance(ci, dict) and _is_failed(ci)]
    has_critical_failure = any(bool(ci.get("critical")) for ci in failed_items)

    description_parts = [
        f"Auto-generated from inspection {inspection.inspection_number}: {inspection.title}",
    ]
    if inspection.description:
        description_parts.append(f"\nInspection notes: {inspection.description}")
    if failed_items:
        description_parts.append("\nFailed checklist items:")
        for item in failed_items:
            label = item.get("question") or item.get("description") or "Unknown item"
            crit = " (critical)" if item.get("critical") else ""
            note = item.get("notes") or ""
            note_suffix = f" — {note}" if note else ""
            description_parts.append(f"  - {label}{crit}{note_suffix}")
    description = "\n".join(description_parts)

    # Map inspection_type → ncr_type. Inspection types are open-text-ish
    # (structural / electrical / plumbing / fire_safety / general / quality)
    # whereas NCR's ncr_type is constrained to a regulated set.
    ncr_type = "workmanship"
    itype = (inspection.inspection_type or "").lower()
    if itype in ("design", "documentation", "safety", "material"):
        ncr_type = itype
    elif itype in ("structural", "electrical", "plumbing", "fire_safety", "general", "quality"):
        ncr_type = "workmanship"

    severity = "critical" if has_critical_failure or inspection.result == "fail" else "major"

    payload = NCRCreate(
        project_id=inspection.project_id,
        title=f"NCR from inspection {inspection.inspection_number}: {inspection.title}"[:500],
        description=description[:10000],
        ncr_type=ncr_type,
        severity=severity,
        location_description=inspection.location,
        linked_inspection_id=str(inspection_id),
        status="identified",
        metadata={
            "source": "inspection",
            "inspection_id": str(inspection_id),
            "inspection_number": inspection.inspection_number,
            "inspection_type": inspection.inspection_type,
            "inspection_result": inspection.result,
            "failed_item_count": len(failed_items),
            "critical_failure": has_critical_failure,
        },
    )

    ncr_service = NCRService(session)
    ncr = await ncr_service.create_ncr(payload, user_id=str(user_id))
    logger.info(
        "Created NCR %s from inspection %s (severity=%s)",
        ncr.ncr_number,
        inspection_id,
        severity,
    )
    return {
        "ncr_id": str(ncr.id),
        "ncr_number": ncr.ncr_number,
        "inspection_id": str(inspection_id),
        "severity": severity,
        "ncr_type": ncr_type,
        "created": True,
    }


@router.post("/{inspection_id}/complete/", response_model=InspectionResponse)
async def complete_inspection(
    inspection_id: uuid.UUID,
    session: SessionDep,
    body: CompleteInspectionRequest | None = None,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("inspections.update")),
    service: InspectionService = Depends(_get_service),
) -> InspectionResponse:
    """Mark an inspection as completed with a pass/fail/partial result.

    Cross-tenant guard: resolves the inspection's parent project before
    flipping any state so one tenant cannot mutate another tenant's
    inspection. Returns 404 (via ``verify_project_access``) on a UUID
    that exists but belongs to a project the caller does not own.
    """
    # Look the inspection up first to get the project_id, then enforce
    # the project-ownership gate BEFORE service.complete mutates state.
    existing = await service.get_inspection(inspection_id)
    await verify_project_access(existing.project_id, str(user_id), session)

    result = body.result if body else "pass"
    inspection = await service.complete_inspection(inspection_id, result=result)
    return _to_response(inspection)
