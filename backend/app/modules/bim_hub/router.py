"""‌⁠‍BIM Hub API routes.

Endpoint convention
-------------------
**Canonical** path for any per-model endpoint is ``/models/{model_id}/...``
to match the spatial intent (``models`` is a collection of resources). The
older flat ``/{model_id}/...`` paths are kept as back-compat aliases so
existing SDK callers don't break — both paths resolve to the same handler.
New endpoints SHOULD use the ``/models/{model_id}/...`` form.

Endpoints:
    Models:
        GET    /                                — List models for a project
        POST   /                                — Create model
        POST   /upload                          — Upload BIM data (DataFrame + optional DAE)
        GET    /models/{model_id}                — Get single model (canonical)
        GET    /{model_id}                       — Get single model (alias)
        PATCH  /models/{model_id}                — Update model (canonical)
        PATCH  /{model_id}                       — Update model (alias)
        DELETE /models/{model_id}                — Delete model (canonical)
        DELETE /{model_id}                       — Delete model (alias)
        GET    /models/{model_id}/geometry       — Serve DAE geometry file

    Elements:
        GET    /models/{model_id}/elements      — List elements (paginated, filterable)
        POST   /models/{model_id}/elements      — Bulk import elements
        GET    /{model_id}/elements             — List elements (alias)
        GET    /elements/{element_id}            — Get single element

    BOQ Links:
        GET    /links                            — List links for a BOQ position
        POST   /links                            — Create link
        DELETE /links/{link_id}                  — Delete link

    Quantity Maps:
        GET    /quantity-maps                    — List quantity map rules
        POST   /quantity-maps                    — Create quantity map rule
        PATCH  /quantity-maps/{map_id}           — Update quantity map rule
        POST   /quantity-maps/apply              — Apply rules on model

    Diffs:
        POST   /models/{model_id}/diff/{old_id}  — Compute diff
        GET    /diffs/{diff_id}                   — Get diff

    Element Groups (saved selections):
        GET    /element-groups/                   — List groups for a project
        POST   /element-groups/                   — Create a group
        PATCH  /element-groups/{group_id}         — Update a group
        DELETE /element-groups/{group_id}         — Delete a group

    Dataframe (Parquet + DuckDB analytical queries):
        GET    /models/{model_id}/dataframe/schema/              — Column names + types
        POST   /models/{model_id}/dataframe/query/               — Query via DuckDB SQL
        GET    /models/{model_id}/dataframe/columns/{col}/values — Value counts for a column
"""

import csv
import gzip as _gzip
import io
import json
import logging
import pathlib
import uuid
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Header, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.rate_limiter import upload_limiter
from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.bim_hub import file_storage as bim_file_storage
from app.modules.bim_hub.schemas import (
    AssetInfoUpdateRequest,
    AssetListResponse,
    AssetSummary,
    BIMElementBulkImport,
    BIMElementGroupCreate,
    BIMElementGroupResponse,
    BIMElementGroupUpdate,
    BIMElementListResponse,
    BIMElementResponse,
    BIMModelBOQLinkAggregate,
    BIMModelBOQLinksResponse,
    BIMModelCreate,
    BIMModelDiffResponse,
    BIMModelListResponse,
    BIMModelResponse,
    BIMModelUpdate,
    BIMQuantityMapCreate,
    BIMQuantityMapListResponse,
    BIMQuantityMapResponse,
    BIMQuantityMapUpdate,
    BOQElementLinkBrief,
    BOQElementLinkCreate,
    BOQElementLinkListResponse,
    BOQElementLinkResponse,
    FederationCreate,
    FederationFullResponse,
    FederationListResponse,
    FederationModelAdd,
    FederationModelResponse,
    FederationResponse,
    FederationTypeTreeResponse,
    FederationUpdate,
    QuantityMapApplyRequest,
    QuantityMapApplyResult,
)
from app.modules.bim_hub.service import BIMHubService

logger = logging.getLogger(__name__)

router = APIRouter()


def _quick_validate_geometry_bytes(blob: bytes, ext: str) -> tuple[bool, str]:
    """Fast, magic-byte / structural pre-check for served geometry.

    Mirrors the heavier ``_validate_geometry_file`` (which runs on ingest
    against a Path) but works on an already-in-memory byte buffer. We
    keep it cheap: only the first ~4 KB are inspected. Returns
    ``(ok, reason)``; the caller raises 422 with the reason when ok is
    False so the BIM viewer surfaces an actionable error instead of
    feeding garbage to Three.js loaders.

    Bug context: external user (Downtown Medical Center / Projet1, RVT)
    reported "Impossible de charger la géométrie 3D" with magic bytes
    ``3c 3f 78 6d 6c`` (``<?xml``) — the stored ``geometry.dae`` was
    XML but not COLLADA. Ingest-time validation existed but only for
    *new* uploads; old corrupt blobs kept streaming. This guard closes
    that gap for every read.
    """
    if not blob:
        return False, "empty buffer"
    if len(blob) < 200:
        return False, f"file suspiciously small ({len(blob)} bytes)"

    ext_norm = ext.lower()
    if ext_norm == ".glb":
        if blob[:4] != b"glTF":
            return False, (
                f"GLB magic mismatch — first 4 bytes are {blob[:4]!r}, expected b'glTF'"
            )
        # Version is the 4-byte LE integer at offset 4.
        if len(blob) >= 12:
            version = int.from_bytes(blob[4:8], "little", signed=False)
            if version != 2:
                return False, f"unsupported GLB version {version} (expected 2)"
        return True, "ok"

    if ext_norm == ".dae":
        # Peek at the first 4 KB and verify a COLLADA root tag exists.
        # We deliberately do NOT do a full XML parse here — we trust
        # the in-memory tax of a 4 KB head scan and let the browser do
        # the heavy lifting once the file is known-good shape.
        # Accept namespace-prefixed roots like `<ns0:COLLADA>` (Revit /
        # DDC pipeline) as well as the bare `<COLLADA>` — both are valid
        # COLLADA per the XML namespace spec. Closes issue #153.
        import re as _re

        head = blob[:4096]
        try:
            head_text = head.decode("utf-8", errors="replace")
        except Exception as exc:  # pragma: no cover — utf-8 with errors='replace' can't raise
            return False, f"DAE head undecodable: {exc}"
        if not _re.search(
            r"<(?:[a-zA-Z_][\w.-]*:)?COLLADA\b", head_text, _re.IGNORECASE
        ):
            # Surface what we DID find so the user/admin can recognise it
            # (e.g. "<ifcxml", "<gbxml", "<!doctype html").
            first_tag_match = _re.search(r"<([a-zA-Z_:][\w:.-]{0,40})", head_text)
            first_tag = (
                f"<{first_tag_match.group(1)}>" if first_tag_match else "no root tag"
            )
            return False, (
                f"DAE has no <COLLADA> root in first 4 KB (first tag found: {first_tag})"
            )
        return True, "ok"

    if ext_norm == ".gltf":
        # gltf JSON — must parse as JSON object with an "asset" key.
        try:
            head = blob[: min(len(blob), 16384)]
            obj = json.loads(head.decode("utf-8", errors="replace"))
        except Exception as exc:
            return False, f"glTF JSON parse failed: {exc}"
        if not isinstance(obj, dict) or "asset" not in obj:
            return False, "glTF JSON missing required 'asset' field"
        return True, "ok"

    # Unknown extension — let it through (preserves prior behaviour for
    # any future extension we add without remembering to update this).
    return True, f"unknown extension {ext_norm}; skipped checks"


def _to_qty_float(val: object) -> float:
    """Best-effort numeric coercion for quantity-presence checks.

    Used by the upload-cad honesty gate (BUG-V320-DDC-01) to decide
    whether *any* imported element carries a real (non-zero, finite)
    quantity.  A string ``"0"`` or ``""`` or a NaN must read as 0.0 so a
    quantity-less import is not mistaken for a successful one.
    """
    if val is None or isinstance(val, bool):
        return 0.0
    try:
        f = float(val)
    except (ValueError, TypeError):
        return 0.0
    if f != f or f in (float("inf"), float("-inf")):
        return 0.0
    return f

# Legacy on-disk path kept only for backward compatibility with any
# external code that may still import ``_BIM_DATA_DIR``.  New code MUST
# go through :mod:`app.modules.bim_hub.file_storage` which wraps the
# pluggable :class:`~app.core.storage.StorageBackend`.
_BIM_DATA_DIR = pathlib.Path(__file__).resolve().parents[4] / "data" / "bim"


def _get_service(session: SessionDep) -> BIMHubService:
    return BIMHubService(session)


# ═══════════════════════════════════════════════════════════════════════════════
# Project-ownership authorization helper
#
# Every BIM endpoint that touches a project (directly via ?project_id= or
# indirectly via a model/element/diff that belongs to a project) MUST call
# ``_verify_project_access`` before returning data or mutating state.
#
# This closes the IDOR from the v1.3.13 audit: previously any authenticated
# user could read/modify/delete models belonging to projects they do not own
# simply by guessing UUIDs. We now resolve the underlying project, verify
# ownership (or admin bypass) and return a 404 — not a 403 — so we also don't
# leak the existence of UUIDs the caller is not allowed to see.
# ═══════════════════════════════════════════════════════════════════════════════


async def _verify_project_access(
    session: AsyncSession,
    project_id: uuid.UUID,
    user_id: str,
) -> None:
    """‌⁠‍Raise 404 if the user is not the owner or an admin of the project.

    Mirrors the central helper in ``erp_chat.tools._require_project_access``
    but returns an HTTPException suitable for router use. Emits 404 (not 403)
    on both "project missing" and "access denied" to avoid UUID enumeration.
    """
    from app.modules.projects.repository import ProjectRepository
    from app.modules.users.repository import UserRepository

    proj_repo = ProjectRepository(session)
    project = await proj_repo.get_by_id(project_id)
    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    # Admin bypass — admins can touch any project regardless of ownership.
    try:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_id(uuid.UUID(str(user_id)))
        if user is not None and getattr(user, "role", "") == "admin":
            return
    except Exception:
        # If the role lookup explodes, fall through to the ownership check —
        # never silently bypass authorization.
        logger.exception("Admin-role lookup failed during BIM access check")

    if str(project.owner_id) != str(user_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )


async def _verify_model_access(
    service: "BIMHubService",
    model_id: uuid.UUID,
    user_id: str,
) -> Any:
    """‌⁠‍Load a BIM model and verify the caller owns its project.

    Returns the model object so callers can reuse it without a second query.
    Raises 404 if the model is missing or the user has no access.
    """
    model = await service.get_model(model_id)
    if model is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Model not found",
        )
    await _verify_project_access(service.session, model.project_id, user_id)
    return model


# ═══════════════════════════════════════════════════════════════════════════════
# DataFrame column alias detection (flexible header matching)
# ═══════════════════════════════════════════════════════════════════════════════

_BIM_COLUMN_ALIASES: dict[str, list[str]] = {
    "element_id": [
        "element_id",
        "elementid",
        "id",
        "guid",
        "ifc_guid",
        "ifcguid",
        "global_id",
        "globalid",
        "stable_id",
        "stableid",
        "unique_id",
        "uniqueid",
        "revit_id",
        "elem_id",
    ],
    "element_type": [
        "element_type",
        "elementtype",
        "ifc_type",
        "ifctype",
        "object_type",
        "objecttype",
    ],
    # _-prefixed alias groups are NOT top-level BIMElement columns.
    # _rows_to_elements promotes them into the properties JSONB under
    # clean keys (category, family, type_name) so the frontend can
    # build Revit Browser-style hierarchy without data collisions.
    "_category": [
        "category",
        "elementcategory",
        "revit_category",
        "revitcategory",
        "ifc_class",
        "ifcclass",
        "class",
    ],
    "_family": [
        "family",
        "family_name",
        "familyname",
        "revit_family",
        "revitfamily",
        "family_and_type",
        "familyandtype",
    ],
    "_type_name": [
        "type_name",
        "typename",
        "type",
        "revit_type",
        "revittype",
    ],
    "name": [
        "name",
        "element_name",
        "elementname",
        "description",
        "bezeichnung",
        "label",
        "title",
    ],
    "storey": [
        "storey",
        "story",
        "level",
        "level_name",
        "levelname",
        "host_level_name",
        "hostlevelname",
        "floor",
        "floor_name",
        "etage",
        "geschoss",
        "building_storey",
        "buildingstorey",
        "ifc_storey",
        "ifcstorey",
        "base_constraint",
        "baseconstraint",
        "base_level",
        "baselevel",
        "reference_level",
        "referencelevel",
        "associated_level",
        "associatedlevel",
        "schedule_level",
        "schedulelevel",
    ],
    "mesh_ref": [
        "mesh_ref",
        "meshref",
        "mesh_id",
        "meshid",
        "node_id",
        "nodeid",
        "dae_node",
        "daenode",
        "collada_node",
        "colladanode",
        "geometry_ref",
        "geometryref",
    ],
    "bbox_min_x": [
        "bbox_min_x", "bboxminx", "min_x", "minx",
        "bounding_box_min_x", "boundingboxminx",
        "bb_min_x", "bbminx", "xmin",
    ],
    "bbox_min_y": [
        "bbox_min_y", "bboxminy", "min_y", "miny",
        "bounding_box_min_y", "boundingboxminy",
        "bb_min_y", "bbminy", "ymin",
    ],
    "bbox_min_z": [
        "bbox_min_z", "bboxminz", "min_z", "minz",
        "bounding_box_min_z", "boundingboxminz",
        "bb_min_z", "bbminz", "zmin",
    ],
    "bbox_max_x": [
        "bbox_max_x", "bboxmaxx", "max_x", "maxx",
        "bounding_box_max_x", "boundingboxmaxx",
        "bb_max_x", "bbmaxx", "xmax",
    ],
    "bbox_max_y": [
        "bbox_max_y", "bboxmaxy", "max_y", "maxy",
        "bounding_box_max_y", "boundingboxmaxy",
        "bb_max_y", "bbmaxy", "ymax",
    ],
    "bbox_max_z": [
        "bbox_max_z", "bboxmaxz", "max_z", "maxz",
        "bounding_box_max_z", "boundingboxmaxz",
        "bb_max_z", "bbmaxz", "zmax",
    ],
    "bounding_box": [
        "bounding_box",
        "boundingbox",
        "bbox",
        "aabb",
    ],
    "discipline": [
        "discipline",
        "disziplin",
        "trade",
        "gewerk",
        "domain",
        "system",
    ],
    "area_m2": [
        "area_m2",
        "area",
        "flaeche",
        "fläche",
        "surface_area",
        "surfacearea",
        "gross_area",
        "grossarea",
        "net_area",
        "netarea",
    ],
    "volume_m3": [
        "volume_m3",
        "volume",
        "volumen",
        "gross_volume",
        "grossvolume",
        "net_volume",
        "netvolume",
    ],
    "length_m": [
        "length_m",
        "length",
        "laenge",
        "länge",
        "span",
    ],
    "weight_kg": [
        "weight_kg",
        "weight",
        "gewicht",
        "mass",
        "masse",
    ],
    "properties": [
        "properties",
        "props",
        "attributes",
        "parameters",
        "pset",
    ],
}


def _match_bim_column(header: str) -> str | None:
    """Match a header string to a canonical BIM column name."""
    normalised = header.strip().lower().replace(" ", "_").replace("-", "_")
    for canonical, aliases in _BIM_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return normalised if normalised else None


# Properties-blob keys we will scan, in priority order, when the upload
# row has no top-level storey/level column. Most Revit and IFC exporters
# put the building level under "Level"; some put the host constraint
# under "Base Constraint" / "Reference Level" instead. Matched
# case-insensitively against the props dict.
_STOREY_PROPERTY_FALLBACK_KEYS: tuple[str, ...] = (
    "level",
    "base level",
    "baselevel",
    "base constraint",
    "baseconstraint",
    "reference level",
    "referencelevel",
    "host level",
    "hostlevel",
    "schedule level",
    "schedulelevel",
    "associated level",
    "associatedlevel",
    "building storey",
    "buildingstorey",
    "ifcbuildingstorey",
    "storey",
    "story",
    "floor",
    "etage",
    "geschoss",
)

# Literal-string sentinels that mean "no storey assigned" — Revit
# exports often write "None" / "<None>" instead of leaving the cell
# blank.  Matched case-insensitively after stripping.
_STOREY_NULL_LITERALS: frozenset[str] = frozenset({
    "", "none", "null", "<none>", "n/a", "na", "-", "—",
})


def _normalise_storey(raw: Any) -> str | None:
    """Coerce a raw storey value to a clean string or None.

    Trims whitespace and treats common "no value" literals
    (``"None"``, ``"<None>"``, ``"N/A"``, ``"-"``, …) as None so
    they don't pollute the BIMFilterPanel storey list with a
    bogus bucket.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if text.lower() in _STOREY_NULL_LITERALS:
        return None
    return text


def _extract_storey(row: dict[str, Any], props: dict[str, Any]) -> str | None:
    """Resolve the building level for an element row.

    Priority:
        1. Top-level ``storey`` column (already aliased from ``level`` /
           ``base_level`` / etc. via :data:`_BIM_COLUMN_ALIASES`).
        2. Case-insensitive match against
           :data:`_STOREY_PROPERTY_FALLBACK_KEYS` inside the
           ``properties`` JSON blob — Revit/IFC exports frequently
           bury "Level" inside the property bag instead of promoting
           it to a column.

    Returns None if nothing usable is found, so downstream consumers
    can render the element as "no level" rather than crashing.
    """
    primary = _normalise_storey(row.get("storey"))
    if primary:
        return primary

    if not props:
        return None

    # Build a lower-cased lookup once so we can match keys regardless
    # of the export's casing convention ("Level" vs "level" vs "LEVEL").
    lc_props = {str(k).strip().lower(): v for k, v in props.items()}
    for key in _STOREY_PROPERTY_FALLBACK_KEYS:
        if key in lc_props:
            resolved = _normalise_storey(lc_props[key])
            if resolved:
                return resolved
    return None


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Parse a value to float, returning *default* on failure."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except (ValueError, TypeError):
        return default


def _parse_bim_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for BIM element import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file — unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_bim_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_bim_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for BIM element import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_bim_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


def _rows_to_elements(
    rows: list[dict[str, Any]],
    has_geometry: bool,
) -> list[dict[str, Any]]:
    """Convert parsed rows into BIMElement-compatible dicts.

    Builds quantities JSON from area_m2, volume_m3, length_m, weight_kg columns.
    Parses properties column as JSON if present.
    """
    elements: list[dict[str, Any]] = []
    quantity_keys = {"area_m2", "volume_m3", "length_m", "weight_kg"}

    for row in rows:
        eid = str(row.get("element_id", "")).strip()
        if not eid:
            continue

        # Parse quantities
        quantities: dict[str, float] = {}
        for qk in quantity_keys:
            val = row.get(qk)
            if val is not None:
                fval = _safe_float(val)
                if fval != 0.0:
                    quantities[qk] = fval

        # Parse properties (could be JSON string)
        raw_props = row.get("properties")
        props: dict[str, Any] = {}
        if isinstance(raw_props, str) and raw_props.strip():
            try:
                props = json.loads(raw_props)
            except (json.JSONDecodeError, ValueError):
                props = {"raw": raw_props}
        elif isinstance(raw_props, dict):
            props = raw_props

        # Explicit mesh_ref column wins, otherwise fall back to element_id
        # (which matches DDC RvtExporter's DAE ``<node id="...">`` pattern).
        raw_mesh_ref = row.get("mesh_ref")
        if raw_mesh_ref is not None and str(raw_mesh_ref).strip():
            mesh_ref: str | None = str(raw_mesh_ref).strip()
        elif has_geometry:
            mesh_ref = eid
        else:
            mesh_ref = None

        # Explicit bbox — either a pre-built JSON blob in ``bounding_box`` or
        # six individual min/max columns.
        bbox: dict[str, float] | None = None
        raw_bbox = row.get("bounding_box")
        if isinstance(raw_bbox, dict):
            bbox = {k: float(v) for k, v in raw_bbox.items() if v is not None}
        elif isinstance(raw_bbox, str) and raw_bbox.strip():
            try:
                parsed = json.loads(raw_bbox)
                if isinstance(parsed, dict):
                    bbox = {k: float(v) for k, v in parsed.items() if v is not None}
            except (json.JSONDecodeError, ValueError, TypeError):
                pass
        if bbox is None:
            bbox_keys = ("bbox_min_x", "bbox_min_y", "bbox_min_z",
                         "bbox_max_x", "bbox_max_y", "bbox_max_z")
            if any(row.get(k) is not None for k in bbox_keys):
                bbox = {
                    "min_x": _safe_float(row.get("bbox_min_x")),
                    "min_y": _safe_float(row.get("bbox_min_y")),
                    "min_z": _safe_float(row.get("bbox_min_z")),
                    "max_x": _safe_float(row.get("bbox_max_x")),
                    "max_y": _safe_float(row.get("bbox_max_y")),
                    "max_z": _safe_float(row.get("bbox_max_z")),
                }
                # Heuristic: if the numbers look like millimetres (range >10000
                # in any axis) convert to metres.
                ranges = [
                    abs(bbox["max_x"] - bbox["min_x"]),
                    abs(bbox["max_y"] - bbox["min_y"]),
                    abs(bbox["max_z"] - bbox["min_z"]),
                ]
                if any(r > 10_000 for r in ranges):
                    bbox = {k: v / 1000.0 for k, v in bbox.items()}

        # Promote _-prefixed canonical keys into properties under clean names.
        # These come from the split alias groups (_category, _family, _type_name)
        # that used to collide on element_type.
        _PROMOTE_TO_PROPS = {
            "_category": "category",
            "_family": "family",
            "_type_name": "type_name",
        }
        for raw_key, clean_key in _PROMOTE_TO_PROPS.items():
            val = row.get(raw_key)
            if val is not None:
                cleaned = str(val).strip()
                if cleaned and cleaned.lower() not in ("none", "null", "n/a", "-"):
                    props[clean_key] = cleaned

        # Collect any extra columns not in known canonical keys as properties
        bbox_col_keys = {
            "bounding_box", "bbox_min_x", "bbox_min_y", "bbox_min_z",
            "bbox_max_x", "bbox_max_y", "bbox_max_z",
        }
        known_keys = {
            "element_id", "element_type", "name", "storey",
            "discipline", "properties", "mesh_ref",
        } | quantity_keys | bbox_col_keys | set(_PROMOTE_TO_PROPS.keys())
        for k, v in row.items():
            if k not in known_keys and v is not None and str(v).strip():
                props[k] = v

        # If element_type is empty after the alias split (because the Excel
        # only had "Category" and "Type" columns, both now redirected away
        # from element_type), fall back to category (the broadest
        # classification) so we don't store rows with no type at all.
        raw_element_type = str(row.get("element_type", "")).strip() or None
        if not raw_element_type and props.get("category"):
            raw_element_type = props["category"]

        element: dict[str, Any] = {
            "stable_id": eid,
            "element_type": raw_element_type,
            "name": str(row.get("name", "")).strip() or None,
            # Resolve storey from the top-level column first; if absent
            # (or a "None"/"-" sentinel), fall back to scanning the
            # properties blob for a "Level" / "Base Constraint" / etc.
            # key. See _extract_storey for the full priority chain.
            "storey": _extract_storey(row, props),
            "discipline": str(row.get("discipline", "")).strip() or None,
            "quantities": quantities,
            "properties": props,
            "mesh_ref": mesh_ref,
            "bounding_box": bbox,
        }
        elements.append(element)

    return elements


# ═══════════════════════════════════════════════════════════════════════════════
# Upload (DataFrame + optional DAE geometry)
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/upload/", status_code=201)
async def upload_bim_data(
    project_id: str = Query(..., description="Project UUID"),
    name: str = Query(default="Imported Model", max_length=255),
    discipline: str = Query(default="architecture", max_length=50),
    data_file: UploadFile = File(..., description="CSV or Excel file with element data"),
    geometry_file: UploadFile | None = File(
        default=None, description="DAE/COLLADA geometry file"
    ),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Upload BIM data from Cad2Data converter output.

    Accepts a DataFrame file (CSV/Excel) with one row per building element
    and an optional COLLADA (.dae) geometry file where each mesh node has
    an ID matching ``element_id`` from the DataFrame.

    Expected DataFrame columns (flexible auto-detection via aliases):
    - **element_id / id / guid** -- unique element identifier (required)
    - **element_type / type / category** -- element classification
    - **name / description** -- human-readable name
    - **storey / level / floor** -- building storey assignment
    - **discipline / trade** -- discipline (architecture, structural, ...)
    - **area_m2 / area** -- area in m2
    - **volume_m3 / volume** -- volume in m3
    - **length_m / length** -- length in m
    - **weight_kg / weight** -- weight in kg
    - **properties** -- JSON string of additional properties

    Returns:
        Summary with model_id, element_count, storeys, and disciplines.
    """
    # --- Verify project access (IDOR guard) ---
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid project_id: {exc}",
        ) from exc
    await _verify_project_access(service.session, project_uuid, user_id or "")

    allowed, _ = upload_limiter.is_allowed(str(user_id or "anon"))
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many uploads. Please wait a moment and try again.",
            headers={"Retry-After": "60"},
        )

    # --- Validate data file ---
    data_filename = (data_file.filename or "").lower()
    if not data_filename.endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported data file type. Please upload CSV (.csv) or Excel (.xlsx) file.",
        )

    data_content = await data_file.read()
    if not data_content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded data file is empty.",
        )

    # No upload size cap — per product policy.

    # --- Validate geometry file (if provided) ---
    has_geometry = False
    geometry_content: bytes | None = None
    if geometry_file is not None:
        geo_filename = (geometry_file.filename or "").lower()
        if not geo_filename.endswith((".dae", ".glb", ".gltf")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported geometry file type. Please upload DAE (.dae), GLB (.glb), or glTF (.gltf) file.",
            )
        geometry_content = await geometry_file.read()
        if geometry_content:
            has_geometry = True

    # --- Parse data file ---
    try:
        if data_filename.endswith((".xlsx", ".xls")):
            rows = _parse_bim_rows_from_excel(data_content)
        else:
            rows = _parse_bim_rows_from_csv(data_content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse data file: {exc}",
        ) from exc
    except Exception as exc:
        logger.exception("Unexpected error parsing BIM data file")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse data file: {exc}",
        ) from exc

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in the uploaded file.",
        )

    # --- Convert rows to element dicts ---
    element_dicts = _rows_to_elements(rows, has_geometry=has_geometry)
    if not element_dicts:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid elements found. Ensure the file has an 'element_id' column.",
        )

    # --- Determine format from geometry file extension ---
    geo_ext = ""
    if geometry_file and geometry_file.filename:
        geo_ext = pathlib.Path(geometry_file.filename).suffix.lstrip(".").lower()

    model_format = geo_ext if geo_ext else "csv"

    # --- Create BIM model ---
    from app.modules.bim_hub.schemas import BIMModelCreate

    model_data = BIMModelCreate(
        project_id=uuid.UUID(project_id),
        name=name,
        discipline=discipline,
        model_format=model_format,
        status="processing",
    )
    model = await service.create_model(model_data, user_id=user_id)
    model_id = model.id

    # --- Save geometry file to configured storage backend ---
    if has_geometry and geometry_content:
        ext = pathlib.Path(geometry_file.filename or "geometry.dae").suffix or ".dae"  # type: ignore[union-attr]
        await bim_file_storage.save_geometry(
            project_id=project_id,
            model_id=str(model_id),
            ext=ext,
            content=geometry_content,
        )

    # --- Import elements ---
    from app.modules.bim_hub.schemas import BIMElementCreate

    elements_create = [
        BIMElementCreate(
            stable_id=ed["stable_id"],
            element_type=ed.get("element_type"),
            name=ed.get("name"),
            storey=ed.get("storey"),
            discipline=ed.get("discipline") or discipline,
            quantities=ed.get("quantities", {}),
            properties=ed.get("properties", {}),
            mesh_ref=ed.get("mesh_ref"),
            bounding_box=ed.get("bounding_box"),
        )
        for ed in element_dicts
    ]
    created_elements = await service.bulk_import_elements(model_id, elements_create)

    # Compute summary
    storeys = sorted({e.storey for e in created_elements if e.storey})
    disciplines_found = sorted({e.discipline for e in created_elements if e.discipline})

    logger.info(
        "BIM upload complete: model=%s, elements=%d, storeys=%d, disciplines=%s",
        name,
        len(created_elements),
        len(storeys),
        disciplines_found,
    )

    return {
        "model_id": str(model_id),
        "element_count": len(created_elements),
        "storeys": storeys,
        "disciplines": disciplines_found,
        "has_geometry": has_geometry,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Direct CAD file upload (RVT, IFC, DWG, DGN, FBX, OBJ, 3DS)
# ═══════════════════════════════════════════════════════════════════════════════

_ALLOWED_CAD_EXTENSIONS = {".rvt", ".ifc", ".dwg", ".dgn", ".fbx", ".obj", ".3ds"}

# Formats that require an external converter binary. IFC has a built-in
# text fallback parser, so it's NOT in this set; XLSX/CSV go through a
# separate upload endpoint and aren't relevant here.
_NEEDS_CONVERTER_EXTS = {".rvt", ".dwg", ".dgn"}


# ═══════════════════════════════════════════════════════════════════════════════
# Background workers — invoked via FastAPI BackgroundTasks so the upload
# request returns in milliseconds even when DDC conversion takes minutes.
# Each worker uses a fresh AsyncSession (the request session is closed by
# the time the task runs) and the same storage abstraction as the router.
# ═══════════════════════════════════════════════════════════════════════════════


async def _process_cad_in_background(
    *,
    project_id: str,
    model_id: str,
    cad_storage_key: str,
    ext: str,
    conversion_depth: str,
) -> None:
    """Run DDC conversion + element extraction for an uploaded CAD file.

    Scheduled after the upload endpoint returns so the HTTP request finishes
    in milliseconds while the (potentially minutes-long) conversion happens
    off the request path.  Updates the model row's ``status`` to
    ``ready`` / ``error`` / ``needs_converter`` when finished — the frontend
    already polls ``GET /{model_id}`` and transitions the UI automatically.
    """
    import asyncio
    import tempfile
    import uuid as _uuid
    from pathlib import Path as _Path

    from sqlalchemy import select

    from app.core.storage import get_storage_backend
    from app.database import async_session_factory
    from app.modules.bim_hub.ifc_processor import process_ifc_file
    from app.modules.bim_hub.models import BIMElement, BIMModel

    model_uuid = _uuid.UUID(model_id)

    # Pre-flight: check the converter binary up front so we can surface a
    # specific actionable error ("install the RVT converter") instead of the
    # generic "no elements extracted" message that lands when DDC isn't found
    # mid-pipeline. IFC has a built-in text fallback so we don't gate on it.
    #
    # Two-stage check (added 2026-04-28 to address "many problems with the
    # converters and downloading"):
    #   1. ``find_converter`` — file exists on disk (cheap, file-stat only).
    #   2. ``smoke_test_converter`` — binary actually loads (8 s timeout,
    #      result cached 5 min). Catches Qt-DLL / Mark-of-the-Web /
    #      VC-Redist-missing breakage that the first check can't see.
    if ext.lower() == ".rvt":
        try:
            from app.modules.boq.cad_import import (
                find_converter as _fc,
            )
            from app.modules.boq.cad_import import (
                smoke_test_converter as _smoke,
            )

            converter_id = "rvt"
            failure_reason: str | None = None
            failure_code: str | None = None
            suggested_actions: list[str] = []

            if _fc(converter_id) is None:
                failure_code = "ddc_not_found"
                failure_reason = (
                    "The Revit (RVT) converter is not installed on this "
                    "server. Open Settings → BIM Converters and click "
                    "Install for RVT, then click Retry on this model."
                )
                suggested_actions = ["install_converter"]
            else:
                # Run the cached smoke test before the (expensive) RVT
                # conversion. If the binary fails to load we report the
                # DLL/perm error to the user immediately rather than
                # letting them wait through a 5-minute conversion that
                # has no chance of succeeding.
                health = await asyncio.to_thread(_smoke, converter_id)
                if health["status"] != "ok":
                    failure_code = "ddc_smoke_failed"
                    failure_reason = health["message"] or (
                        "The Revit (RVT) converter is installed but the "
                        "smoke test failed. Open Settings → BIM Converters "
                        "and click Reinstall, then click Retry on this model."
                    )
                    suggested_actions = list(health["suggested_actions"]) or [
                        "reinstall_converter"
                    ]

            if failure_code:
                async with async_session_factory() as session:
                    model = (
                        await session.execute(
                            select(BIMModel).where(BIMModel.id == model_uuid)
                        )
                    ).scalar_one_or_none()
                    if model is not None:
                        model.status = "needs_converter"
                        model.error_message = failure_reason
                        meta = dict(model.metadata_ or {})
                        meta["error_code"] = failure_code
                        meta["converter_id"] = converter_id
                        meta["suggested_actions"] = suggested_actions
                        meta["install_endpoint"] = (
                            f"/api/v1/takeoff/converters/{converter_id}/install/"
                        )
                        meta["verify_endpoint"] = (
                            f"/api/v1/takeoff/converters/{converter_id}/verify/"
                        )
                        model.metadata_ = meta
                        await session.commit()
                logger.warning(
                    "RVT pre-flight failed for model %s: %s",
                    model_id,
                    failure_code,
                )
                return
        except Exception:  # noqa: BLE001 — pre-flight is best-effort
            logger.exception("RVT converter pre-flight check failed for %s", model_id)

    try:
        content = await get_storage_backend().get(cad_storage_key)

        with tempfile.TemporaryDirectory(prefix="oe-bim-bg-") as _tmp_str:
            _tmp_dir = _Path(_tmp_str)
            _tmp_cad_path = _tmp_dir / f"original{ext}"
            await asyncio.to_thread(_tmp_cad_path.write_bytes, content)

            result = await asyncio.to_thread(
                process_ifc_file, _tmp_cad_path, _tmp_dir, conversion_depth
            )
            element_count = result["element_count"]

            geo_key: str | None = None
            geo_local = result.get("geometry_path")
            if geo_local:
                _geo_path = _Path(geo_local)
                if _geo_path.is_file():
                    _geo_bytes = await asyncio.to_thread(_geo_path.read_bytes)
                    _geo_ext = _geo_path.suffix or ".dae"
                    geo_key = await bim_file_storage.save_geometry(
                        project_id=project_id,
                        model_id=model_id,
                        ext=_geo_ext,
                        content=_geo_bytes,
                    )

            glb_key: str | None = None
            glb_local = result.get("glb_path")
            if glb_local:
                _glb_path = _Path(glb_local)
                if _glb_path.is_file():
                    _glb_bytes = await asyncio.to_thread(_glb_path.read_bytes)
                    glb_key = await bim_file_storage.save_geometry(
                        project_id=project_id,
                        model_id=model_id,
                        ext=".glb",
                        content=_glb_bytes,
                    )
                    logger.info(
                        "GLB geometry saved: %s (%d bytes)", glb_key, len(_glb_bytes)
                    )

            raw_elements = result.get("raw_elements", [])
            if raw_elements:
                try:
                    from app.modules.bim_hub.dataframe_store import write_dataframe

                    await asyncio.to_thread(
                        write_dataframe,
                        project_id=project_id,
                        model_id=model_id,
                        rows=raw_elements,
                    )
                except Exception as exc:
                    logger.warning("Parquet write failed (non-fatal): %s", exc)

        async with async_session_factory() as session:
            # Defensive: the upload endpoint commits before scheduling us, but
            # retry a few times anyway in case of slow disk flushes / connection
            # pool churn during a heavy upload burst.
            model = None
            for _attempt in range(5):
                model = (
                    await session.execute(
                        select(BIMModel).where(BIMModel.id == model_uuid)
                    )
                ).scalar_one_or_none()
                if model is not None:
                    break
                await asyncio.sleep(0.2)
            if model is None:
                logger.error(
                    "Background processor: model %s vanished mid-conversion", model_id
                )
                return

            if element_count > 0:
                # Top-level geometry_quality drives the placeholder banner.
                # Stamp it onto each element's properties so the frontend
                # viewer can self-detect placeholders without an extra API
                # round-trip for the model metadata.
                result_quality = result.get("geometry_quality") or result.get("geometry_type")
                # BUG-V320-DDC-01 / D-TKC-NEW-01 — honesty gate.  When the DDC
                # cad2data converter is unavailable the IFC text-parser still
                # imports element geometry, but it can only recover quantities
                # if the file happens to ship explicit IfcElementQuantity
                # blocks.  Track whether *any* element carries a non-empty
                # quantities map; if none do, we must NOT advertise the model
                # as a clean 'ready' import with error_message=null — that is
                # the dishonest "successful import, zero quantities" state the
                # QA audit flagged.
                any_quantities = False
                for elem_data in result["elements"]:
                    el_props = dict(elem_data.get("properties", {}) or {})
                    if elem_data.get("is_placeholder") or result_quality == "placeholder":
                        el_props["is_placeholder"] = True
                    el_quantities = elem_data.get("quantities", {}) or {}
                    if el_quantities and any(
                        _to_qty_float(v) for v in el_quantities.values()
                    ):
                        any_quantities = True
                    el = BIMElement(
                        model_id=model_uuid,
                        stable_id=elem_data["stable_id"],
                        element_type=elem_data.get("element_type"),
                        name=elem_data.get("name"),
                        storey=elem_data.get("storey"),
                        discipline=elem_data.get("discipline"),
                        properties=el_props,
                        quantities=el_quantities,
                        geometry_hash=elem_data.get("geometry_hash"),
                        bounding_box=elem_data.get("bounding_box"),
                        mesh_ref=elem_data.get("mesh_ref"),
                    )
                    session.add(el)

                converter_absent = result_quality == "placeholder"
                no_quantities = not any_quantities

                model.status = "ready"
                model.element_count = element_count
                model.storey_count = len(result["storeys"])
                model.bounding_box = result.get("bounding_box")
                # BUG-006: stamp the conversion-finished moment so audit
                # reports / Asset Register sorting have something better
                # than ``created_at`` (which is the upload moment, before
                # geometry extraction).  ``import_date`` was previously
                # left null on every ready model.
                from datetime import UTC as _UTC
                from datetime import datetime as _dt

                model.import_date = _dt.now(_UTC)
                if glb_key:
                    model.canonical_file_path = glb_key
                elif geo_key:
                    model.canonical_file_path = geo_key
                model.metadata_ = {
                    **(model.metadata_ or {}),
                    "geometry_type": result.get("geometry_type", "unknown"),
                    # geometry_quality drives the frontend's "placeholder
                    # geometry" banner — set to "placeholder" when DDC
                    # cad2data is unavailable and we synthesized boxes.
                    "geometry_quality": result.get(
                        "geometry_quality", result.get("geometry_type", "unknown"),
                    ),
                    # DDC converter version stamp — drives the "Processed
                    # with DDC v{X}" badge on the BIM model card and the
                    # /about page. Both keys are optional: missing values
                    # leave the badge hidden (v3.12.0 / Stream D).
                    **(
                        {"converter_version": result["converter_version"]}
                        if result.get("converter_version")
                        else {}
                    ),
                    **(
                        {"converter_source": result["converter_source"]}
                        if result.get("converter_source")
                        else {}
                    ),
                }

                # BUG-V320-DDC-01 / D-TKC-NEW-01 — non-destructive honesty
                # path.  The elements were imported (geometry is useful for
                # the viewer / element linking) but if the DDC converter was
                # absent OR no quantities could be extracted we downgrade the
                # status from a misleading 'ready' to a distinct 'degraded'
                # state and populate a user-facing warning so the UI can show
                # "imported, but no quantities — DDC converter required"
                # instead of pretending the import fully succeeded.
                if converter_absent or no_quantities:
                    model.status = "degraded"
                    meta_warn = dict(model.metadata_ or {})
                    if converter_absent and no_quantities:
                        warn_msg = (
                            "Geometry imported, but no quantities could be "
                            "extracted: the DDC cad2data converter is not "
                            "available on this server, and the file does not "
                            "carry explicit IFC BaseQuantities. Elements were "
                            "imported for geometry/linking only. Install the "
                            "DDC converter (Settings → BIM Converters) and "
                            "click Retry to recover area/volume/length "
                            "quantities."
                        )
                        meta_warn["error_code"] = "no_quantities_converter_absent"
                    elif converter_absent:
                        warn_msg = (
                            "Imported with placeholder geometry: the DDC "
                            "cad2data converter is not available on this "
                            "server, so quantities were read from the file's "
                            "embedded IFC BaseQuantities only and may be "
                            "incomplete. Install the DDC converter "
                            "(Settings → BIM Converters) and click Retry for "
                            "full geometry-derived quantities."
                        )
                        meta_warn["error_code"] = "converter_absent"
                    else:
                        warn_msg = (
                            "Geometry imported, but no quantities "
                            "(area/volume/length) could be extracted from "
                            "this file. Elements were imported for "
                            "geometry/linking only."
                        )
                        meta_warn["error_code"] = "no_quantities"
                    model.error_message = warn_msg
                    meta_warn["warning"] = warn_msg
                    meta_warn["degraded"] = True
                    meta_warn["converter_id"] = "ifc"
                    meta_warn["install_endpoint"] = (
                        "/api/v1/takeoff/converters/ifc/install/"
                    )
                    model.metadata_ = meta_warn
                    logger.warning(
                        "Background CAD processed but DEGRADED (converter_absent=%s "
                        "no_quantities=%s): %d elements → model %s degraded",
                        converter_absent, no_quantities, element_count, model_id,
                    )
                else:
                    logger.info(
                        "Background CAD processed: %d elements, %d storeys → "
                        "model %s ready",
                        element_count, len(result["storeys"]), model_id,
                    )

                # Storage policy — drop the raw upload after a *successful*
                # conversion when ``keep_original_cad`` is False (production
                # default).  Failed conversions fall through to the else
                # branch below and keep the original so retry works without
                # re-upload.  Conversion artifacts (GLB/DAE/parquet) stay
                # forever regardless.
                from app.config import get_settings as _get_settings

                if not _get_settings().keep_original_cad:
                    await bim_file_storage.delete_original_cad(
                        project_id=project_id,
                        model_id=model_id,
                        ext=ext,
                    )
            else:
                meta = dict(model.metadata_ or {})

                # Pull the structured failure context the DDC subprocess
                # recorded (RVT version, converter version, stderr tail).
                # If it's present, we can build a much more specific error
                # message than the legacy "converter not installed" boilerplate.
                from app.modules.bim_hub.ifc_processor import last_ddc_failure

                ddc_failure = last_ddc_failure()
                rvt_info = ddc_failure.get("rvt_info") or {}
                conv_info = ddc_failure.get("converter_info") or {}
                rvt_app = rvt_info.get("app_name")  # e.g. "Revit 2024"
                conv_version = conv_info.get("version")  # e.g. "18.0.0.0"
                stderr_tail = (ddc_failure.get("stderr") or "").strip()

                if ext == ".rvt":
                    model.status = "needs_converter"

                    # Compose the message in pieces — every clause is added
                    # only when its underlying datum is non-empty so we never
                    # ship "File saved with Revit None".
                    parts: list[str] = []
                    if rvt_app:
                        parts.append(
                            f"File saved with {rvt_app}"
                            + (f" (format {rvt_info['format']})." if rvt_info.get("format") else ".")
                        )
                    if conv_version:
                        parts.append(f"Installed RVT converter: {conv_version}.")
                    parts.append(
                        "The converter produced no elements from this file. "
                        "Most common causes: the RVT was saved with a Revit "
                        "version newer than the converter supports, the file "
                        "is corrupt, or a converter dependency is missing."
                    )
                    if stderr_tail:
                        # Trim to a single line for the headline message;
                        # the full stderr tail goes into metadata_ below.
                        first_line = stderr_tail.splitlines()[0][:200]
                        if first_line:
                            parts.append(f"Converter said: {first_line}")
                    parts.append(
                        "Try updating the RVT converter (Settings → BIM "
                        "Converters → Reinstall) and clicking Retry."
                    )
                    model.error_message = " ".join(parts)

                    meta["error_code"] = "ddc_failed"
                    meta["converter_id"] = "rvt"
                    meta["install_endpoint"] = (
                        "/api/v1/takeoff/converters/rvt/install/"
                    )
                    # Structured diagnostic info for the frontend to render
                    # a dedicated "version mismatch" panel if it wants to.
                    meta["diagnostics"] = {
                        "rvt_info": rvt_info,
                        "converter_info": conv_info,
                        "reason": ddc_failure.get("reason"),
                        "exit_code": ddc_failure.get("exit_code"),
                        "stderr_tail": stderr_tail,
                    }
                else:
                    model.status = "error"
                    model.error_message = (
                        "No elements could be extracted from this IFC file. "
                        "Open the file in a viewer like BIMcollab Zoom to "
                        "confirm it isn't empty, then click Retry."
                    )
                    meta["error_code"] = "zero_elements"
                    if stderr_tail or conv_version:
                        meta["diagnostics"] = {
                            "converter_info": conv_info,
                            "reason": ddc_failure.get("reason"),
                            "exit_code": ddc_failure.get("exit_code"),
                            "stderr_tail": stderr_tail,
                        }
                model.metadata_ = meta
                logger.warning(
                    "Background CAD processed but no elements found for model %s",
                    model_id,
                )

            await session.commit()

    except Exception as exc:
        logger.exception("Background CAD processing failed for model %s: %s", model_id, exc)
        try:
            async with async_session_factory() as session:
                model = (
                    await session.execute(
                        select(BIMModel).where(BIMModel.id == model_uuid)
                    )
                ).scalar_one_or_none()
                if model is not None:
                    model.status = "error"
                    model.error_message = (
                        f"Processing failed: {exc}. Click Retry to try again, or "
                        f"contact support if this keeps happening."
                    )
                    meta = dict(model.metadata_ or {})
                    meta["error_code"] = "unexpected"
                    model.metadata_ = meta
                    await session.commit()
        except Exception as exc2:
            logger.exception("Failed to mark model %s as error: %s", model_id, exc2)


async def _generate_pdf_in_background(
    *,
    project_id: str,
    model_id: str,
    cad_storage_key: str,
    ext: str,
    model_name: str,
    user_id: str,
) -> None:
    """Run DDC PDF-only export for an existing model and link as a Document.

    Invoked from POST /{model_id}/generate-pdf-sheets/ via BackgroundTasks.
    Calls the DDC converter exactly once with a ``.pdf`` output target — no
    re-export of XLSX/DAE.  Silently skips when no converter is installed
    on the host (the upload itself stays usable; PDF export is opt-in).
    """
    import asyncio
    import subprocess
    import tempfile
    import uuid as _uuid
    from pathlib import Path as _Path

    from app.core.storage import get_storage_backend
    from app.database import async_session_factory

    converter_ext = ext.lstrip(".").lower()
    try:
        from app.modules.boq.cad_import import find_converter
    except ImportError:
        logger.warning("PDF generation skipped — cad_import not available")
        return

    converter = find_converter(converter_ext)
    if not converter:
        logger.info(
            "PDF generation skipped — %s converter not installed",
            converter_ext.upper(),
        )
        return

    try:
        content = await get_storage_backend().get(cad_storage_key)

        with tempfile.TemporaryDirectory(prefix="oe-bim-pdf-") as _tmp_str:
            _tmp_dir = _Path(_tmp_str)
            _tmp_cad_path = _tmp_dir / f"original{ext}"
            await asyncio.to_thread(_tmp_cad_path.write_bytes, content)

            pdf_target = (_tmp_dir / "sheets.pdf").resolve()

            def _run_pdf() -> tuple[int, bytes]:
                proc = subprocess.run(
                    [str(converter), str(_tmp_cad_path.resolve()), str(pdf_target)],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    cwd=str(converter.parent),
                    input=b"\n",
                    timeout=900,
                )
                return proc.returncode, proc.stderr

            try:
                rc, stderr = await asyncio.to_thread(_run_pdf)
            except subprocess.TimeoutExpired:
                logger.warning("DDC PDF generation timed out for model %s", model_id)
                return

            if rc != 0 or not pdf_target.is_file() or pdf_target.stat().st_size < 1000:
                logger.warning(
                    "DDC PDF generation failed for model %s (rc=%d, stderr=%s)",
                    model_id,
                    rc,
                    stderr.decode(errors="replace")[:200] if stderr else "",
                )
                return

            pdf_bytes = await asyncio.to_thread(pdf_target.read_bytes)
            pdf_storage_key = await bim_file_storage.save_geometry(
                project_id=project_id,
                model_id=model_id,
                ext=".pdf",
                content=pdf_bytes,
            )

            try:
                from app.modules.documents.models import Document as DocModel

                async with async_session_factory() as session:
                    pdf_doc = DocModel(
                        project_id=_uuid.UUID(project_id),
                        name=f"{model_name or 'BIM Model'} — Sheets (PDF)",
                        category="drawing",
                        file_path=pdf_storage_key,
                        file_size=len(pdf_bytes),
                        mime_type="application/pdf",
                        tags=["bim", "sheets", "auto-generated", converter_ext],
                        created_by=_uuid.UUID(user_id) if user_id else None,
                    )
                    session.add(pdf_doc)
                    await session.commit()
                    logger.info(
                        "PDF sheets saved as Document for model %s: %s (%d bytes)",
                        model_id, pdf_storage_key, len(pdf_bytes),
                    )
            except Exception as exc:
                logger.warning("PDF sheets → Document linkage failed: %s", exc)

    except Exception as exc:
        logger.exception("PDF generation failed for model %s: %s", model_id, exc)


@router.post("/upload-cad/", status_code=201)
async def upload_cad_file(
    background_tasks: BackgroundTasks,
    project_id: str = Query(..., description="Project UUID"),
    name: str = Query(default="", max_length=255),
    discipline: str = Query(default="architecture", max_length=50),
    conversion_depth: str = Query(
        default="standard",
        description=(
            "DDC conversion depth: 'standard' (~15 basic columns, fastest),"
            " 'medium' (DDC standard + full property promotion, ~900 columns),"
            " or 'complete' (all Revit parameters, ~1000+ columns, slowest)"
        ),
    ),
    file: UploadFile = File(..., description="CAD file (RVT, IFC, DWG, DGN, FBX, OBJ, 3DS)"),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict:
    """Upload a raw CAD file for background processing.

    The file is stored on disk at ``data/bim/{project_id}/{model_id}/original.{ext}``
    and a BIMModel record is created with status="processing". A real CAD converter
    service would pick it up asynchronously; for now the model stays in processing state.

    Accepted extensions: .rvt, .ifc, .dwg, .dgn, .fbx, .obj, .3ds
    """
    # --- Verify project access (IDOR guard) ---
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid project_id: {exc}",
        ) from exc
    await _verify_project_access(service.session, project_uuid, user_id or "")

    allowed, _ = upload_limiter.is_allowed(str(user_id or "anon"))
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many uploads. Please wait a moment and try again.",
            headers={"Retry-After": "60"},
        )

    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No filename provided.",
        )

    ext = pathlib.Path(filename).suffix.lower()
    if ext not in _ALLOWED_CAD_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Unsupported file type '{ext}'. "
                f"Accepted: {', '.join(sorted(_ALLOWED_CAD_EXTENSIONS))}"
            ),
        )

    # Stream the upload to a temp file in 1 MB chunks instead of buffering
    # the whole body in memory.  A 500 MB IFC used to cost ~500 MB of heap
    # in the request handler — on the 2 GB-RAM VPS, two concurrent uploads
    # were enough to OOM the process.  ``StreamedUpload`` exposes:
    #   - ``upload.path``    — the spooled temp file
    #   - ``upload.size``    — bytes written
    #   - ``upload.head``    — first 64 bytes for magic-byte validation
    # Storage's ``put_stream`` then ``rename(2)``s the file into place
    # (single syscall on same-FS local backend; ``upload_fileobj`` with
    # multipart on S3) — zero additional memory pressure.
    from app.core.upload_streaming import stream_upload_to_temp

    async with stream_upload_to_temp(file, suffix=ext) as upload:
        if upload.size == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        # Preflight: when the required converter binary isn't installed on
        # this server, **persist** the upload and create a placeholder model
        # so the user doesn't have to re-upload after running the install.
        # The response is HTTP 202 Accepted (the request is good — we'll
        # finish processing later) rather than 201 Created (we haven't
        # created any geometry yet).  ``Retry-After`` and a ``Link`` header
        # point the client at the converter-install endpoint and the model
        # row that will be re-processed once the binary lands.  Frontend
        # dispatches on the ``status`` field in the body — see
        # ``BIMCadUploadResponse`` in ``frontend/src/features/bim/api.ts``.
        if ext in _NEEDS_CONVERTER_EXTS:
            from app.modules.boq.cad_import import find_converter

            if find_converter(ext.lstrip(".")) is None:
                new_model_id = uuid.uuid4()
                saved_cad_key = await bim_file_storage.save_original_cad_from_path(
                    project_uuid, new_model_id, ext, upload.path, size=upload.size,
                )
                display_name = (name or pathlib.Path(filename).stem).strip() or filename
                from app.modules.bim_hub.schemas import BIMModelCreate

                # NB: ``error_message`` is set via a follow-up update because
                # ``BIMModelCreate`` doesn't expose that field — it lives on
                # ``BIMModelUpdate`` so freshly-created records start clean.
                # The kwarg here MUST be ``user_id=`` (the service signature) —
                # passing ``created_by=`` raised a TypeError on every .rvt /
                # .ifc upload that hit the missing-converter path.
                pending_model = await service.create_model(
                    BIMModelCreate(
                        project_id=project_uuid,
                        name=display_name,
                        discipline=discipline,
                        model_format=ext.lstrip("."),
                        canonical_file_path=saved_cad_key,
                        status="needs_converter",
                    ),
                    user_id=user_id,
                )
                await service.update_model(
                    pending_model.id,
                    BIMModelUpdate(
                        error_message=(
                            f"{ext.upper().lstrip('.')} converter not installed — "
                            f"install it from the BIM converter banner, then "
                            f"click Re-process on this model."
                        ),
                    ),
                )

                logger.info(
                    "Saved %s upload pending converter — model=%s, key=%s, %d bytes",
                    ext, new_model_id, saved_cad_key, upload.size,
                )

                install_endpoint = (
                    f"/api/v1/takeoff/converters/{ext.lstrip('.')}/install/"
                )
                return JSONResponse(
                    status_code=status.HTTP_202_ACCEPTED,
                    content={
                        "status": "converter_required",
                        "format": ext.lstrip("."),
                        "converter_id": ext.lstrip("."),
                        "message": (
                            f"{ext.upper().lstrip('.')} files require the "
                            f"{ext.upper().lstrip('.')} converter, which is not "
                            f"installed on this server. Your file has been "
                            f"saved — install the converter and click "
                            f"Re-process on the model card to finish the upload."
                        ),
                        "install_endpoint": install_endpoint,
                        "model_id": str(new_model_id),
                        "name": display_name,
                        "file_size": upload.size,
                        "element_count": 0,
                        "error_message": None,
                    },
                    headers={
                        "Retry-After": "60",
                        "Link": (
                            f"<{install_endpoint}>; rel=\"install-converter\", "
                            f"</api/v1/bim_hub/{new_model_id}/retry/>; "
                            f"rel=\"reprocess-model\""
                        ),
                    },
                )

        # Magic-byte validation — filename extensions are attacker-controlled
        # and we proceed to hand this file to a CAD converter that can be
        # exploited by unexpected formats. Reject anything that doesn't look
        # like one of our accepted CAD/BIM containers.  The streamed-upload
        # helper has already kept the first 64 bytes around for us.
        from app.core.file_signature import (
            ALLOWED_CAD_TYPES,
            FileSignatureMismatch,
        )
        from app.core.file_signature import (
            require as _require_sig,
        )

        try:
            _require_sig(
                upload.head,
                ALLOWED_CAD_TYPES,
                filename=filename,
            )
        except FileSignatureMismatch as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            ) from exc

        # Auto-fill model name from filename
        model_name = name or pathlib.Path(filename).stem

        # Determine model format from extension (strip the dot)
        model_format = ext.lstrip(".")

        # Create BIM model record with processing status
        from app.modules.bim_hub.schemas import BIMModelCreate

        model_data = BIMModelCreate(
            project_id=uuid.UUID(project_id),
            name=model_name,
            discipline=discipline,
            model_format=model_format,
            status="processing",
        )
        model = await service.create_model(model_data, user_id=user_id)
        model_id = model.id

        # Save CAD file via the configured storage backend — returns the
        # storage key that the Documents hub cross-link and downstream
        # diagnostics use to refer back to the stored blob.  Streaming
        # variant: the backend renames the temp file into place rather
        # than buffering its content.
        saved_cad_key = await bim_file_storage.save_original_cad_from_path(
            project_id=project_id,
            model_id=str(model_id),
            ext=ext,
            src_path=upload.path,
            size=upload.size,
        )

        logger.info(
            "CAD file uploaded: %s (%s, %d bytes) -> model %s (key=%s)",
            filename,
            ext,
            upload.size,
            model_id,
            saved_cad_key,
        )

        # Cross-link: create Document record so BIM files appear in Documents hub.
        # Uses the ORM model directly (NOT raw SQL) so timestamps + defaults are
        # filled by SQLAlchemy / Base mixin and the row stays in sync with the
        # rest of the documents module if its schema evolves.  Failures are
        # swallowed because the cross-link is convenience-only — the BIM model
        # itself is already saved by the time we get here.
        try:
            from app.modules.documents.models import Document

            doc = Document(
                project_id=uuid.UUID(project_id),
                name=filename,
                description=f"BIM model: {model_name}",
                category="drawing",
                file_size=upload.size,
                mime_type=f"application/{model_format}",
                file_path=saved_cad_key,
                version=1,
                uploaded_by=user_id or "",
                tags=["bim", model_format, discipline],
                metadata_={
                    "source_module": "bim_hub",
                    "source_id": str(model_id),
                },
            )
            service.session.add(doc)
            await service.session.flush()
            logger.info("Cross-linked BIM model %s → document %s", model_id, doc.id)
        except Exception as exc:
            logger.warning("Failed to cross-link BIM to documents hub: %s", exc)

    # Schedule processing OUT of the request path: the upload endpoint now
    # returns 201 + status="processing" in milliseconds, and the actual DDC
    # conversion + element/geometry persistence happens in a background task.
    # This eliminates the multi-minute synchronous block that used to drop
    # the connection on slow conversions ("Cannot connect to server" in the
    # frontend).  The frontend already polls GET /{model_id}/ — model.status
    # transitions from "processing" → "ready" / "error" / "needs_converter"
    # automatically once the worker finishes.
    final_status = "processing"
    element_count = 0

    # Commit BEFORE scheduling the background task: the worker opens its own
    # async session and looks the model up by id, so the row must already be
    # durably visible to other connections.  Without this explicit commit the
    # worker raced the request-scope dependency teardown and saw "model
    # vanished mid-conversion" intermittently.
    await service.session.commit()

    processable = ext in (".ifc", ".rvt")
    if processable:
        background_tasks.add_task(
            _process_cad_in_background,
            project_id=project_id,
            model_id=str(model_id),
            cad_storage_key=saved_cad_key,
            ext=ext,
            conversion_depth=conversion_depth,
        )
        logger.info(
            "CAD upload accepted, processing scheduled in background: %s → model %s",
            filename, model_id,
        )
    else:
        # Non-processable format (DWG, DGN, FBX, etc.) — needs converter
        model.status = "needs_converter"
        model.error_message = (
            f"{ext.upper().lstrip('.')} files require an external converter. "
            "Convert to IFC first, then re-upload."
        )
        await service.session.flush()
        final_status = "needs_converter"

    return {
        "model_id": str(model_id),
        "name": model_name,
        "format": model_format,
        "file_size": upload.size,
        "status": final_status,
        "element_count": element_count,
        "error_message": model.error_message,
        "geometry_type": (model.metadata_ or {}).get("geometry_type", "unknown"),
        "converter_id": ext.lstrip(".") if final_status == "needs_converter" else None,
        "install_endpoint": (
            f"/api/v1/takeoff/converters/{ext.lstrip('.')}/install/"
            if final_status == "needs_converter" else None
        ),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Optional post-upload: generate PDF sheets for an existing model
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/{model_id}/generate-pdf-sheets/", status_code=202)
async def generate_pdf_sheets(
    model_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Schedule PDF-sheets export for an existing BIM model.

    Runs the DDC converter once with a ``.pdf`` output target — no XLSX/DAE
    re-export, only the sheets PDF.  The PDF is saved as a Document linked
    to the project once the worker finishes.

    Returns immediately; the caller does not wait for the export to finish.
    Frontend can detect completion by polling the project's documents list.
    """
    from app.modules.bim_hub import file_storage as _bim_storage

    model = await _verify_model_access(service, model_id, user_id or "")

    model_format = (model.model_format or "").lower()
    if not model_format:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Model has no format recorded — cannot regenerate sheets.",
        )

    ext = "." + model_format.lstrip(".")
    cad_storage_key = _bim_storage.original_cad_key(
        project_id=model.project_id,
        model_id=model_id,
        ext=ext,
    )
    backend_store = _bim_storage._backend()
    if not await backend_store.exists(cad_storage_key):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Original CAD file is no longer available — re-upload the model.",
        )

    background_tasks.add_task(
        _generate_pdf_in_background,
        project_id=str(model.project_id),
        model_id=str(model_id),
        cad_storage_key=cad_storage_key,
        ext=ext,
        model_name=model.name or "BIM Model",
        user_id=user_id or "",
    )

    return {
        "status": "scheduled",
        "model_id": str(model_id),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Retry conversion — re-runs the background CAD processor for a model that
# previously failed (status="error" / "needs_converter"). Useful when the
# user installs a missing converter after upload, or the original failure
# was transient (network blip, OOM during a parallel upload burst).
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/{model_id}/retry/", status_code=202)
async def retry_model_processing(
    model_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Re-schedule background DDC conversion for a previously-failed model.

    Resets ``status`` to ``processing``, clears ``error_message``, and re-
    invokes :func:`_process_cad_in_background` against the same original CAD
    blob.  Returns 202 immediately — the frontend already polls
    ``GET /{model_id}/`` and will transition the UI when the worker finishes.

    Refuses to retry models that:
        * are already ``ready`` (no need),
        * are already ``processing`` (a worker is in flight),
        * have no original CAD blob recorded (re-upload required).
    """
    from app.modules.bim_hub import file_storage as _bim_storage

    model = await _verify_model_access(service, model_id, user_id or "")

    if model.status == "ready":
        return {
            "status": "noop",
            "model_id": str(model_id),
            "message": "Model is already ready — nothing to retry.",
        }
    if model.status == "processing":
        return {
            "status": "noop",
            "model_id": str(model_id),
            "message": "Model is already being processed.",
        }

    model_format = (model.model_format or "").lower()
    if not model_format:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Model has no format recorded — cannot retry.",
        )

    ext = "." + model_format.lstrip(".")
    cad_storage_key = _bim_storage.original_cad_key(
        project_id=model.project_id,
        model_id=model_id,
        ext=ext,
    )
    backend_store = _bim_storage._backend()
    if not await backend_store.exists(cad_storage_key):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "Original CAD file is no longer available — re-upload the "
                "model to retry."
            ),
        )

    # Clear previous error state before re-scheduling so the frontend's
    # polling immediately reflects "processing" once the retry kicks in.
    model.status = "processing"
    model.error_message = None
    meta = dict(model.metadata_ or {})
    for _k in ("error_code", "install_endpoint"):
        meta.pop(_k, None)
    model.metadata_ = meta
    await service.session.commit()

    background_tasks.add_task(
        _process_cad_in_background,
        project_id=str(model.project_id),
        model_id=str(model_id),
        cad_storage_key=cad_storage_key,
        ext=ext,
        conversion_depth="standard",
    )
    logger.info("Retry scheduled for model %s", model_id)

    return {
        "status": "scheduled",
        "model_id": str(model_id),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Geometry file serving
# ═══════════════════════════════════════════════════════════════════════════════


@router.head("/models/{model_id}/geometry/", response_model=None, include_in_schema=False)
@router.get("/models/{model_id}/geometry/", response_model=None)
async def get_model_geometry(
    model_id: uuid.UUID,
    token: str | None = Query(
        default=None,
        description="JWT access token (alternative to Authorization header for static loaders)",
    ),
    fmt: str | None = Query(
        default=None,
        description="Force a specific geometry format: 'dae' or 'glb'. "
        "When omitted, the server returns GLB (preferred) with DAE fallback.",
    ),
    authorization: str | None = Header(default=None),
    service: BIMHubService = Depends(_get_service),
) -> StreamingResponse | RedirectResponse:
    """Serve the COLLADA/DAE geometry file for the 3D viewer.

    Auth: accepts either an Authorization header OR a ``?token=...`` query
    parameter. The query param exists because Three.js ColladaLoader cannot
    set custom headers — without this fallback the viewer would 401.

    The geometry blob is resolved through :mod:`app.modules.bim_hub.file_storage`
    so both the local filesystem and S3 backends work transparently.  For S3
    we redirect to a short-lived presigned URL; for the local backend we
    stream the bytes directly through the route.
    """
    # Per-request correlation ID — surfaced in the X-Request-Id response
    # header AND embedded in every structured-error payload so a user who
    # ships a screenshot to support can be located in server logs in one
    # grep. UUID4 keeps it non-PII (no info about the user, project, or
    # file).  We generate locally rather than relying on a middleware so
    # the value is identical between log line and HTTP response.
    request_id = str(uuid.uuid4())

    # Validate the token (header or query). ColladaLoader can't set headers,
    # so we accept ?token=<jwt> as an alternative auth mechanism.
    from app.config import get_settings
    from app.dependencies import decode_access_token, verify_user_exists_and_active

    auth_token: str | None = token
    if not auth_token and authorization and authorization.lower().startswith("bearer "):
        auth_token = authorization[7:]

    if not auth_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={
                "error": "auth_missing",
                "category": "authentication",
                "request_id": request_id,
                "model_id": str(model_id),
                "message": "Missing authentication token.",
                "remediation": (
                    "Refresh the page to renew your login session. If you "
                    "were idle for a long time the access token may have "
                    "expired silently."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    try:
        payload = decode_access_token(auth_token, get_settings())
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={
                "error": "auth_invalid",
                "category": "authentication",
                "request_id": request_id,
                "model_id": str(model_id),
                "message": "Authentication token is invalid or expired.",
                "remediation": (
                    "Log out and log back in to obtain a fresh token. If "
                    "the problem persists, your account may have been "
                    "deactivated — contact support."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    # BUG-323: forged tokens with a fake UUID must not authenticate here
    # either. Re-hydrate against the DB and replace self-asserted role /
    # permissions with canonical state before any authorization check.
    db_user = await verify_user_exists_and_active(payload["sub"])
    from app.core.permissions import permission_registry

    payload["role"] = db_user.role
    payload["permissions"] = permission_registry.get_role_permissions(db_user.role)

    # Check the token-bearer actually has bim.read before we load data.
    token_role = payload.get("role", "")
    token_perms: list[str] = payload.get("permissions", [])
    if token_role != "admin" and "bim.read" not in token_perms:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "permission_denied",
                "category": "authorization",
                "request_id": request_id,
                "model_id": str(model_id),
                "required_permission": "bim.read",
                "message": "Your account lacks permission to view BIM models.",
                "remediation": (
                    "Ask a project administrator to grant you the 'bim.read' "
                    "permission, or to assign you a role (Estimator / "
                    "Manager / Admin) that includes it."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    model = await service.get_model(model_id)
    if model is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "model_not_found",
                "category": "not_found",
                "request_id": request_id,
                "model_id": str(model_id),
                "message": "This BIM model has been deleted or never existed.",
                "remediation": (
                    "Go back to the project's BIM tab and pick a model from "
                    "the list. If you reached this page from a saved link, "
                    "the model may have been removed by a teammate."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    # IDOR guard: verify the caller owns the project this model belongs to.
    token_user_id = str(payload.get("sub") or "")
    await _verify_project_access(service.session, model.project_id, token_user_id)

    project_id = str(model.project_id)

    # Resolve the geometry blob through the storage backend.
    # When ?fmt=dae is passed, force DAE format (useful when GLB has
    # scrambled node names from an older trimesh conversion).
    if fmt and fmt.lower() == "dae":
        found = await bim_file_storage.find_geometry_key(
            project_id, model_id, prefer_ext=".dae"
        )
    else:
        found = await bim_file_storage.find_geometry_key(project_id, model_id)
    if found is not None:
        key, ext = found
        media_type = bim_file_storage.GEOMETRY_MEDIA_TYPES.get(
            ext, "application/octet-stream"
        )
        cache_headers = {
            # No caching — geometry may be re-generated with patched node names.
            "Cache-Control": "no-store, no-cache, must-revalidate",
        }

        # Prefer a presigned URL so the browser fetches directly from the
        # bucket (S3).  Local backend returns None → fall back to streaming.
        presigned = bim_file_storage.presigned_geometry_url(key)
        if presigned:
            return RedirectResponse(url=presigned, status_code=307)

        # Read the full blob and gzip-compress for transfer.
        # GLB: 9.5 MB → 1.7 MB, DAE: 32 MB → 3.5 MB typical.
        from app.core.storage import get_storage_backend

        _geo_bytes = await get_storage_backend().get(key)

        # Serve-time integrity check — closes the gap where geometry written
        # by an older converter (before _validate_geometry_file existed on
        # ingest) keeps streaming bad bytes to the viewer. The browser
        # surfaces this as an opaque "Cannot read properties of undefined
        # (reading 'getAttribute')" deep inside Three.js. We re-check the
        # first ~4 KB of the blob and 422 with a precise diagnostic if it
        # doesn't match the format the extension promises.
        ok_serve, reason_serve = _quick_validate_geometry_bytes(_geo_bytes, ext)
        if not ok_serve:
            # Build a structured diagnostic payload. We deliberately limit
            # what we expose to: (a) the first 8 bytes of the file as hex
            # + ASCII (universally safe — magic bytes don't carry PII),
            # (b) total size in bytes, (c) the parser reason, (d) the
            # stored extension, (e) what we expected. NO actual user-data
            # bytes or filenames are leaked. Frontend renders this verbatim
            # in the BIM viewer error panel so users can give actionable
            # detail to support without revealing the file contents.
            head_bytes = _geo_bytes[:8]
            head_hex = " ".join(f"{b:02x}" for b in head_bytes)
            head_ascii = "".join(
                chr(b) if 0x20 <= b < 0x7F else "." for b in head_bytes
            )
            # Surface the first identifiable XML root tag for the common
            # "stored DAE turned out to be IFC-XML / gbXML / HTML 404 page"
            # failure mode — gives support a one-glance diagnosis.
            first_tag: str | None = None
            if ext.lower() == ".dae":
                import re as _re_diag
                try:
                    _head_text = _geo_bytes[:4096].decode("utf-8", errors="replace")
                    _m = _re_diag.search(r"<([a-zA-Z_:][\w:.-]{0,40})", _head_text)
                    if _m:
                        first_tag = f"<{_m.group(1)}>"
                except Exception:  # pragma: no cover — replace can't raise
                    first_tag = None
            expected_signature = {
                ".glb": "b'glTF' magic + version 2",
                ".dae": "<COLLADA> root tag within first 4 KB",
                ".gltf": "JSON object with required 'asset' key",
            }.get(ext.lower(), f"valid {ext} payload")
            # Categorise reason into a plain-language "cause" the UI can
            # show without the user having to read parser jargon. This is
            # the single biggest lever for end-user understanding: instead
            # of "DAE has no <COLLADA> root in first 4 KB (first tag found:
            # <html>)" they see "The stored file is an HTML page, not a 3D
            # model — the converter probably crashed and saved an error
            # page by mistake."
            reason_lower = reason_serve.lower()
            if "empty buffer" in reason_lower:
                cause = (
                    "The geometry file on the server has zero bytes. The "
                    "original upload likely failed half-way through."
                )
            elif "suspiciously small" in reason_lower:
                cause = (
                    "The geometry file is too small to be a real 3D model. "
                    "The upload was probably truncated, or the converter "
                    "wrote only an error stub."
                )
            elif "<!doctype html" in (first_tag or "").lower() or (
                first_tag and first_tag.lower() in ("<html>", "<body>")
            ):
                cause = (
                    "The stored file is an HTML page, not a 3D model. The "
                    "converter likely saved an error page by mistake. The "
                    "source CAD/BIM file may not be supported, or the "
                    "converter service was unreachable during processing."
                )
            elif first_tag and first_tag.lower() in (
                "<ifcxml>", "<gbxml>", "<xml>", "<?xml>",
            ):
                cause = (
                    f"The stored file is {first_tag} (XML data) instead of "
                    "a 3D mesh. The source format does not contain 3D "
                    "geometry to display — e.g. an IFC schedule or a "
                    "2D-only drawing."
                )
            elif "magic mismatch" in reason_lower:
                cause = (
                    "The file's first bytes don't match the expected "
                    "format signature. The file is either corrupted in "
                    "transit, or its extension was renamed manually."
                )
            elif "unsupported glb version" in reason_lower:
                cause = (
                    "The file is an older glTF format version that our "
                    "viewer doesn't support (we require glTF 2.0)."
                )
            else:
                cause = (
                    f"The stored file does not match the expected {ext} "
                    "signature. The CAD converter may have run with an "
                    "older version, or the source file is corrupt."
                )
            logger.warning(
                "BIM geometry served from %s failed serve-time validation: %s "
                "(request_id=%s, model_id=%s, size=%d, head=%s, "
                "first_tag=%s, ext=%s)",
                key, reason_serve, request_id, model_id, len(_geo_bytes),
                head_hex, first_tag, ext,
            )
            diagnostic = {
                "error": "geometry_invalid",
                "category": "file_format",
                "request_id": request_id,
                "reason": reason_serve,
                "cause": cause,
                "format": ext.lstrip(".") or "unknown",
                "stored_extension": ext,
                "expected_signature": expected_signature,
                "size_bytes": len(_geo_bytes),
                "head_hex": head_hex,
                "head_ascii": head_ascii,
                "first_tag": first_tag,
                "model_id": str(model_id),
                "remediation": (
                    "Delete this model and re-upload the source CAD/BIM "
                    "file. If the problem repeats with the same file, the "
                    "source itself may be unsupported (2D-only DWG, IFC "
                    "schedule with no geometry, corrupted RVT) — try "
                    "exporting from your CAD tool again, or contact "
                    "info@datadrivenconstruction.io and quote the "
                    "Request ID shown below."
                ),
                "message": (
                    f"Geometry file is not a valid {ext} payload: {reason_serve}"
                ),
            }
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=diagnostic,
                headers={"X-Request-Id": request_id},
            )

        compressed = _gzip.compress(_geo_bytes, compresslevel=6)
        # RFC 5987 encoding so non-ASCII model names (Cyrillic / Arabic / …)
        # don't blow up the latin-1 HTTP header encoder. Without this the
        # whole geometry response 500's and the frontend spins on "loading"
        # forever. We send both a plain-ASCII `filename=` fallback and a
        # UTF-8-encoded `filename*=` for browsers that support it.
        from urllib.parse import quote as _qs

        from fastapi.responses import Response

        display_name = f"{model.name}{ext}"
        ascii_fallback = display_name.encode("ascii", "replace").decode("ascii")
        cd_header = (
            f'inline; filename="{ascii_fallback}"; '
            f"filename*=UTF-8''{_qs(display_name)}"
        )

        return Response(
            content=compressed,
            media_type=media_type,
            headers={
                **cache_headers,
                "Content-Encoding": "gzip",
                "Content-Disposition": cd_header,
                # Surface the correlation ID even on the happy path so a
                # downstream JS parsing failure still has a request_id to
                # quote when reporting (matches every error branch above).
                "X-Request-Id": request_id,
            },
        )

    logger.warning(
        "BIM geometry not found on storage (request_id=%s, model_id=%s, "
        "project_id=%s)",
        request_id, model_id, project_id,
    )
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail={
            "error": "geometry_missing",
            "category": "not_found",
            "request_id": request_id,
            "model_id": str(model_id),
            "message": (
                "No 3D geometry file is attached to this model on the "
                "server."
            ),
            "remediation": (
                "Either the model was uploaded but the CAD converter never "
                "produced a 3D mesh (the source file may be 2D-only or "
                "may have crashed the converter), or the file was deleted "
                "manually from storage. Try re-uploading the source file. "
                "If the same source file repeatedly produces no geometry, "
                "contact info@datadrivenconstruction.io and quote the "
                "Request ID below."
            ),
        },
        headers={"X-Request-Id": request_id},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Models
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/", response_model=BIMModelListResponse)
async def list_models(
    project_id: uuid.UUID = Query(...),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelListResponse:
    """List BIM models for a project.

    Always returns every persisted BIMModel row (sorted by ``created_at``
    desc) — including ``ready``, ``processing``, ``needs_converter`` and
    ``error`` rows — so the /bim page can surface already-converted models
    *without* triggering re-conversion.  Each item is enriched with
    ``conversion_artifact_size_mb`` (sum of GLB/DAE/parquet/thumbnail
    bytes), ``has_original`` (True iff the raw upload is still on
    storage) and ``error_code`` (stable id lifted out of the metadata
    blob for the converter-required UI state).  The list response also
    carries aggregate ``total_artifact_size_mb`` /
    ``total_original_size_mb`` totals which drive the disk-usage chip
    in the BIM page header.
    """
    await _verify_project_access(service.session, project_id, user_id or "")
    items, total = await service.list_models(project_id, offset=offset, limit=limit)

    # Batched storage probe: ONE list_prefix sweep against the backend
    # collects artifact/original/geometry info for every model in the
    # page. Replaces the previous asyncio.gather fan-out which issued
    # 3+ probes per model (50 models → 150+ HEAD/stat round-trips per
    # list call — classic N+1 against storage).  When the backend
    # doesn't support list_prefix (community backends predating v4.6.1),
    # fall back to the per-model probe loop so behaviour is unchanged.
    storage_summary: dict[str, dict[str, object]] = {}
    use_bulk = bim_file_storage.list_prefix_supported()
    if use_bulk:
        try:
            storage_summary = await bim_file_storage.bulk_model_storage_summary(
                project_id,
            )
        except Exception:  # noqa: BLE001  # never break listing on storage hiccups
            logger.exception(
                "bulk_model_storage_summary failed for project=%s; "
                "falling back to per-model probes.",
                project_id,
            )
            use_bulk = False

    async def _enrich_per_model(model_obj):  # type: ignore[no-untyped-def]
        """Per-model probe fallback (community backends without list_prefix)."""
        size_bytes = 0
        has_orig = False
        has_geom = bool(model_obj.canonical_file_path)
        try:
            size_bytes = await bim_file_storage.compute_artifact_size_bytes(
                model_obj.project_id, model_obj.id,
            )
        except Exception:  # noqa: BLE001
            logger.exception("artifact-size probe failed for model=%s", model_obj.id)
        ext_raw = (model_obj.model_format or "").lstrip(".")
        if ext_raw:
            try:
                has_orig = await bim_file_storage.has_original_cad(
                    model_obj.project_id, model_obj.id, ext=f".{ext_raw}",
                )
            except Exception:  # noqa: BLE001
                logger.exception("has_original probe failed for model=%s", model_obj.id)
        if not has_geom:
            try:
                has_geom = (
                    await bim_file_storage.find_geometry_key(
                        project_id=str(model_obj.project_id),
                        model_id=str(model_obj.id),
                    )
                ) is not None
            except Exception:  # noqa: BLE001
                logger.exception("has_geometry probe failed for model=%s", model_obj.id)
        return size_bytes, has_orig, has_geom

    def _enrich_from_summary(model_obj):  # type: ignore[no-untyped-def]
        """Fast bulk-summary-driven enrich (no I/O)."""
        info = storage_summary.get(str(model_obj.id), {})
        size_bytes = int(info.get("artifact_size_bytes", 0) or 0)
        # ``has_original`` historically reflected the existence of a blob
        # at original.{ext} where ext == model_format.  The bulk sweep
        # uses the same "filename starts with original." rule, so the
        # two definitions match for every realistic model row.
        has_orig = bool(info.get("has_original", False))
        has_geom = bool(model_obj.canonical_file_path) or bool(
            info.get("geometry_exts") or (),
        )
        return size_bytes, has_orig, has_geom

    if use_bulk:
        enriched = [_enrich_from_summary(m) for m in items]
    else:
        import asyncio as _asyncio

        enriched = await _asyncio.gather(
            *[_enrich_per_model(m) for m in items],
            return_exceptions=False,
        )

    item_responses: list[BIMModelResponse] = []
    total_artifact_bytes = 0
    total_original_bytes = 0
    for model_obj, (size_bytes, has_orig, has_geom) in zip(items, enriched, strict=True):
        resp = BIMModelResponse.model_validate(model_obj)
        resp.conversion_artifact_size_mb = round(size_bytes / (1024 * 1024), 3)
        resp.has_original = has_orig
        resp.has_geometry = has_geom
        meta = model_obj.metadata_ or {}
        if isinstance(meta, dict):
            err_code = meta.get("error_code")
            if isinstance(err_code, str):
                resp.error_code = err_code
        total_artifact_bytes += size_bytes
        if has_orig:
            # Original-blob size: pulled from the bulk summary when
            # available (zero extra round-trips), or from a single
            # per-model size() probe on the fallback path.
            ext_raw = (model_obj.model_format or "").lstrip(".")
            if use_bulk:
                info = storage_summary.get(str(model_obj.id), {})
                total_original_bytes += int(info.get("original_size_bytes", 0) or 0)
            elif ext_raw:
                try:
                    backend = bim_file_storage._backend()
                    key = bim_file_storage.original_cad_key(
                        model_obj.project_id, model_obj.id, ext=f".{ext_raw}",
                    )
                    total_original_bytes += await backend.size(key)
                except Exception:  # noqa: BLE001
                    pass
        item_responses.append(resp)

    return BIMModelListResponse(
        items=item_responses,
        total=total,
        offset=offset,
        limit=limit,
        total_artifact_size_mb=round(total_artifact_bytes / (1024 * 1024), 3),
        total_original_size_mb=round(total_original_bytes / (1024 * 1024), 3),
        storage_root_label=bim_file_storage.bim_root_label(),
    )


@router.post("/", response_model=BIMModelResponse, status_code=201)
async def create_model(
    data: BIMModelCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelResponse:
    """Create a new BIM model record."""
    await _verify_project_access(service.session, data.project_id, user_id)
    model = await service.create_model(data, user_id=user_id)
    return BIMModelResponse.model_validate(model)


# ─── Asset Register routes (must come BEFORE /{model_id} so that
#     `/assets` is not interpreted as a UUID model_id and rejected with
#     422 by the path validator). The handlers and the `_summarise_asset`
#     helper live further down in the file under the "Asset Register
#     (v2.3.0)" section header — only the route registrations move up. ───

@router.get("/assets", response_model=AssetListResponse)
async def list_assets(
    project_id: uuid.UUID = Query(..., description="Scope the asset list to this project"),
    element_type: str | None = Query(default=None),
    operational_status: str | None = Query(default=None),
    search: str | None = Query(default=None, min_length=1, max_length=200),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> AssetListResponse:
    """List tracked assets across every BIM model in a project.

    Tracked = ``BIMElement.is_tracked_asset == True``. Managed-asset
    rows appear on the Assets page with manufacturer / serial / warranty
    columns lifted out of the ``asset_info`` JSON blob for convenient
    sorting.
    """
    await _verify_project_access(service.session, project_id, user_id or "")
    rows, total = await service.list_tracked_assets(
        project_id,
        element_type=element_type,
        operational_status=operational_status,
        search=search,
        offset=offset,
        limit=limit,
    )
    return AssetListResponse(
        items=[_summarise_asset(element, model) for element, model in rows],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.patch("/assets/{element_id}/asset-info", response_model=BIMElementResponse)
async def update_asset_info(
    element_id: uuid.UUID,
    payload: AssetInfoUpdateRequest,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.write")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Merge-update asset_info on a BIMElement.

    Semantics:
    - Any key you send overwrites the matching key in ``asset_info``.
    - Sending ``null`` or ``""`` for a key clears that key.
    - ``is_tracked_asset`` auto-flips to ``True`` on first non-empty
      write; pass an explicit bool to override.
    - Unrelated keys already in ``asset_info`` survive untouched.
    """
    # Locate the element first so we can verify project access.
    element = await service.get_element(element_id)
    await _verify_model_access(service, element.model_id, user_id or "")

    updated = await service.update_asset_info(
        element_id,
        asset_info=payload.asset_info.model_dump(exclude_unset=False),
        is_tracked_asset=payload.is_tracked_asset,
    )
    return BIMElementResponse.model_validate(updated)


@router.get("/{model_id}", response_model=BIMModelResponse)
async def get_model(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelResponse:
    """Get a single BIM model by ID."""
    model = await _verify_model_access(service, model_id, user_id or "")
    resp = BIMModelResponse.model_validate(model)
    # Mirror the list endpoint enrichment so single-model polls
    # (status transitions during background conversion) also expose
    # artifact size + ``has_original`` to the frontend.
    try:
        size_bytes = await bim_file_storage.compute_artifact_size_bytes(
            model.project_id, model.id,
        )
        resp.conversion_artifact_size_mb = round(size_bytes / (1024 * 1024), 3)
    except Exception:  # noqa: BLE001
        logger.exception("artifact-size probe failed for model=%s", model.id)
    ext_raw = (model.model_format or "").lstrip(".")
    if ext_raw:
        try:
            resp.has_original = await bim_file_storage.has_original_cad(
                model.project_id, model.id, ext=f".{ext_raw}",
            )
        except Exception:  # noqa: BLE001
            logger.exception("has_original probe failed for model=%s", model.id)
    # has_geometry is true when the background converter saved a GLB/DAE
    # (``canonical_file_path`` set). Probing storage as a tie-breaker for
    # historical rows where the column was missed.
    resp.has_geometry = bool(model.canonical_file_path)
    if not resp.has_geometry:
        try:
            resp.has_geometry = (
                await bim_file_storage.find_geometry_key(
                    project_id=str(model.project_id), model_id=str(model.id),
                )
            ) is not None
        except Exception:  # noqa: BLE001
            logger.exception("has_geometry probe failed for model=%s", model.id)
    meta = model.metadata_ or {}
    if isinstance(meta, dict):
        err_code = meta.get("error_code")
        if isinstance(err_code, str):
            resp.error_code = err_code
    return resp


@router.patch("/{model_id}", response_model=BIMModelResponse)
async def update_model(
    model_id: uuid.UUID,
    data: BIMModelUpdate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelResponse:
    """Update a BIM model."""
    await _verify_model_access(service, model_id, user_id)
    model = await service.update_model(model_id, data)
    return BIMModelResponse.model_validate(model)


@router.get("/models/{model_id}/schema/")
async def get_model_schema(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
):
    """Return distinct element types + property keys/values for the model.

    Used by the quantity-rule editor (RFC 24) to seed combobox options.
    """
    await _verify_model_access(service, model_id, user_id or "")
    return await service.get_model_schema(model_id)


@router.delete("/{model_id}", status_code=204)
async def delete_model(
    model_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> None:
    """Delete a BIM model and all its elements."""
    await _verify_model_access(service, model_id, user_id)
    await service.delete_model(model_id)


@router.post("/cleanup-stale/")
async def cleanup_stale_processing(
    project_id: uuid.UUID = Query(...),
    max_age_hours: int = Query(default=1, ge=0),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, int]:
    """Remove models stuck in 'processing' with 0 elements older than max_age_hours."""
    await _verify_project_access(service.session, project_id, user_id or "")
    count = await service.cleanup_stale_processing(project_id, max_age_hours=max_age_hours)
    return {"deleted": count}


@router.post("/cleanup-orphans/")
async def cleanup_orphan_bim_files(
    _user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Scan ``data/bim/`` and remove directories with no matching DB row.

    Admin-grade disk hygiene. Protects against orphaned RVT/IFC/COLLADA/Excel
    artefacts left behind by failed uploads, crashed conversions, or manual
    DB deletes that bypassed the service layer.
    """
    return await service.cleanup_orphan_bim_files()


# ═══════════════════════════════════════════════════════════════════════════════
# Elements
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/models/{model_id}/elements/", response_model=BIMElementListResponse)
async def list_elements(
    model_id: uuid.UUID,
    element_type: str | None = Query(default=None),
    storey: str | None = Query(default=None),
    discipline: str | None = Query(default=None),
    group_id: uuid.UUID | None = Query(
        default=None,
        description="Filter to elements belonging to this saved element group",
    ),
    offset: int = Query(default=0, ge=0),
    # Cap depends on ``skeleton``: the enriched path is paginated at 2000/page
    # because each row fans out to six relation joins (boq_links, docs,
    # tasks, activities, requirements, validation). Skeleton mode returns
    # plain BIMElement rows with no joins and is safe at 50000/page.
    limit: int = Query(default=500, ge=1, le=50000),
    skeleton: bool = Query(
        default=False,
        description=(
            "Skip eager-loading of boq_links / linked_documents / linked_tasks "
            "/ linked_activities / linked_requirements / validation_results. "
            "~10× faster and used by the 3D viewer for mesh-to-element "
            "matching, where those relations are not needed."
        ),
    ),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementListResponse:
    """List elements for a BIM model (paginated, filterable).

    Each element in the response includes a ``boq_links`` array of
    ``BOQElementLinkBrief`` entries, a ``linked_documents`` array of
    ``DocumentLinkBrief`` entries, a ``linked_tasks`` array of ``TaskBrief``
    entries, and a ``linked_activities`` array of ``ActivityBrief`` entries
    so the viewer can render link badges without a second round trip.
    """
    from app.modules.bim_hub.schemas import (
        ActivityBrief,
        DocumentLinkBrief,
        ElementValidationSummary,
        RequirementBrief,
        TaskBrief,
    )

    await _verify_model_access(service, model_id, user_id or "")

    # Skeleton path: plain BIMElement rows, no relation joins, no enrichment.
    # Ten times faster than the enriched path — used by the 3D viewer, where
    # mesh matching only needs id/mesh_ref/name/element_type/bbox.
    if skeleton:
        if limit > 50000:
            limit = 50000
        plain_items, plain_total = await service.list_elements(
            model_id,
            element_type=element_type,
            storey=storey,
            discipline=discipline,
            offset=offset,
            limit=limit,
        )
        # Skinny rows: drop the per-element `properties` / `quantities` /
        # `classification` / `metadata` payloads. These can weigh ~1.5 kB per
        # row on a typical Revit export (45+ Revit parameters × short value),
        # which adds up to a 16 MB JSON body for 7 000 elements. The viewer
        # pulls the full property set straight from Parquet on click, so
        # carrying it in the skeleton is pure overhead.
        skeleton_items: list[BIMElementResponse] = []
        for e in plain_items:
            resp = BIMElementResponse.model_validate(e)
            resp.properties = {}
            resp.quantities = {}
            resp.metadata = {}
            skeleton_items.append(resp)
        return BIMElementListResponse(
            items=skeleton_items,
            total=plain_total,
            offset=offset,
            limit=limit,
        )

    # Enriched path is capped at 2000 — each extra row spawns six join lookups.
    if limit > 2000:
        limit = 2000
    (
        items,
        total,
        boq_links_by_id,
        doc_links_by_id,
        task_links_by_id,
        activity_briefs_by_id,
        requirement_briefs_by_id,
        validation_summaries_by_id,
    ) = await service.list_elements_with_links(
        model_id,
        element_type=element_type,
        storey=storey,
        discipline=discipline,
        group_id=group_id,
        offset=offset,
        limit=limit,
    )

    # The service stashes a sentinel entry under ``_VALIDATION_REPORT_SENTINEL``
    # (UUID(int=0)) when a ``target_type='bim_model'`` report exists. We pop
    # it so it never reaches the per-element loop.
    from app.modules.bim_hub.service import _VALIDATION_REPORT_SENTINEL

    report_exists = _VALIDATION_REPORT_SENTINEL in validation_summaries_by_id
    validation_summaries_by_id.pop(_VALIDATION_REPORT_SENTINEL, None)

    responses: list[BIMElementResponse] = []
    for elem in items:
        boq_briefs = [
            BOQElementLinkBrief.model_validate(b)
            for b in boq_links_by_id.get(elem.id, [])
        ]
        doc_briefs = [
            DocumentLinkBrief.model_validate(b)
            for b in doc_links_by_id.get(elem.id, [])
        ]
        task_briefs = [
            TaskBrief.model_validate(b)
            for b in task_links_by_id.get(elem.id, [])
        ]
        activity_briefs = [
            ActivityBrief.model_validate(b)
            for b in activity_briefs_by_id.get(elem.id, [])
        ]
        requirement_briefs = [
            RequirementBrief.model_validate(b)
            for b in requirement_briefs_by_id.get(elem.id, [])
        ]
        raw_val = validation_summaries_by_id.get(elem.id, [])
        validation_summaries = [
            ElementValidationSummary.model_validate(v) for v in raw_val
        ]
        # Derive worst-severity status; 'unchecked' iff no report exists
        # at all (any element had at least one entry → report_exists).
        if not report_exists:
            val_status: str = "unchecked"
        elif any(v.severity == "error" for v in validation_summaries):
            val_status = "error"
        elif any(v.severity == "warning" for v in validation_summaries):
            val_status = "warning"
        else:
            val_status = "pass"
        resp = BIMElementResponse.model_validate(elem)
        resp.boq_links = boq_briefs
        resp.linked_documents = doc_briefs
        resp.linked_tasks = task_briefs
        resp.linked_activities = activity_briefs
        resp.linked_requirements = requirement_briefs
        resp.validation_results = validation_summaries
        resp.validation_status = val_status  # type: ignore[assignment]
        responses.append(resp)

    return BIMElementListResponse(
        items=responses,
        total=total,
        offset=offset,
        limit=limit,
    )


@router.post(
    "/models/{model_id}/elements/",
    response_model=BIMElementListResponse,
    status_code=201,
)
async def bulk_import_elements(
    model_id: uuid.UUID,
    data: BIMElementBulkImport,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementListResponse:
    """Bulk import elements for a model (replaces existing)."""
    await _verify_model_access(service, model_id, user_id)
    elements = await service.bulk_import_elements(model_id, data.elements)
    return BIMElementListResponse(
        items=[BIMElementResponse.model_validate(e) for e in elements],
        total=len(elements),
        offset=0,
        limit=len(elements),
    )


@router.post(
    "/models/{model_id}/elements/by-ids/",
    response_model=BIMElementListResponse,
)
async def get_elements_by_ids(
    model_id: uuid.UUID,
    body: dict,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementListResponse:
    """Fetch specific elements by their IDs (DB UUID or stable_id)."""
    await _verify_model_access(service, model_id, user_id or "")

    element_ids: list[str] = body.get("element_ids", [])
    if not element_ids or len(element_ids) > 100:
        return BIMElementListResponse(items=[], total=0, offset=0, limit=0)

    from sqlalchemy import or_

    from app.modules.bim_hub.models import BIMElement

    query = (
        select(BIMElement)
        .where(BIMElement.model_id == model_id)
        .where(
            or_(
                BIMElement.id.in_(
                    [uuid.UUID(eid) for eid in element_ids if len(eid) == 36]
                ),
                BIMElement.stable_id.in_(element_ids),
            )
        )
    )
    result = await service.session.execute(query)
    elements = list(result.scalars().all())

    return BIMElementListResponse(
        items=[BIMElementResponse.model_validate(e) for e in elements],
        total=len(elements),
        offset=0,
        limit=len(elements),
    )


@router.post(
    "/models/{model_id}/ensure-element/",
    response_model=BIMElementResponse,
)
async def ensure_element(
    model_id: uuid.UUID,
    body: dict,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Resolve (or lazy-create) a BIMElement row from a mesh_ref / stable_id.

    Needed when linking a BOQ position to a BIM mesh that was visible in the
    3D viewer but had no oe_bim_element row (e.g. DDC standard extract skips
    certain Revit categories). Body: ``{"mesh_ref": "140056"}`` or
    ``{"stable_id": "140056"}``.
    """
    await _verify_model_access(service, model_id, user_id or "")
    stable_id = body.get("stable_id")
    mesh_ref = body.get("mesh_ref")
    element = await service.ensure_element(
        model_id,
        stable_id=str(stable_id) if stable_id else None,
        mesh_ref=str(mesh_ref) if mesh_ref else None,
    )
    return BIMElementResponse.model_validate(element)


@router.get("/elements/{element_id}", response_model=BIMElementResponse)
async def get_element(
    element_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Get a single BIM element by ID."""
    element = await service.get_element(element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Element not found",
        )
    # Verify caller owns the project this element's model belongs to.
    await _verify_model_access(service, element.model_id, user_id or "")
    return BIMElementResponse.model_validate(element)


# ═══════════════════════════════════════════════════════════════════════════════
# Asset Register (v2.3.0)
# ═══════════════════════════════════════════════════════════════════════════════


def _summarise_asset(element, model) -> AssetSummary:
    """Project BIMElement+BIMModel onto the flat AssetSummary row.

    Lifts commonly-displayed fields out of the ``asset_info`` blob so
    the frontend list can sort them without peeking into the JSON.
    """
    info = dict(element.asset_info or {})
    return AssetSummary(
        id=element.id,
        model_id=element.model_id,
        project_id=model.project_id,
        model_name=model.name,
        stable_id=element.stable_id,
        element_type=element.element_type,
        name=element.name,
        storey=element.storey,
        discipline=element.discipline,
        asset_info=info,
        manufacturer=info.get("manufacturer"),
        model=info.get("model"),
        serial_number=info.get("serial_number"),
        warranty_until=info.get("warranty_until"),
        operational_status=info.get("operational_status"),
        asset_tag=info.get("asset_tag"),
    )


# Note: the @router.get("/assets") and @router.patch("/assets/{element_id}/asset-info")
# route definitions were moved up before "/{model_id}" so FastAPI's path
# matcher resolves the literal "assets" segment instead of mistaking it
# for a UUID. The handlers live above; the `_summarise_asset` helper just
# above is still referenced from there at request time.


# ═══════════════════════════════════════════════════════════════════════════════
# COBie Export (v2.3.0)
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/models/{model_id}/export/cobie.xlsx")
async def export_cobie_xlsx(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> StreamingResponse:
    """Generate an ISO-19650 COBie UK 2.4 handover workbook.

    The response streams an ``.xlsx`` file with seven sheets (Contact /
    Facility / Floor / Space / Type / Component / System) built from
    the current canonical BIM data + ``asset_info`` payload.
    """
    await _verify_model_access(service, model_id, user_id or "")
    xlsx_bytes, filename = await service.export_cobie(model_id)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(xlsx_bytes)),
        },
    )


# ═══════════════════════════════════════════════════════════════════════════════
# BOQ Links
# ═══════════════════════════════════════════════════════════════════════════════


async def _verify_boq_position_access(
    service: "BIMHubService",
    position_id: uuid.UUID,
    user_id: str,
) -> None:
    """Resolve a BOQ position → its BOQ → project and verify the caller owns it.

    `Position` has no direct `project_id` column — the project lives on the
    parent `BOQ` row reached via `position.boq_id`.  We do a single-row
    SELECT joining position → boq so this stays one round-trip.
    """
    # ``BOQ`` is the class name exposed by ``boq.models`` and it refers to
    # the Bill-of-Quantities aggregate, not a module-level constant — the
    # ``N811`` noqa below suppresses ruff's all-caps-is-a-constant heuristic.
    from app.modules.boq.models import BOQ as BOQModel  # noqa: N811
    from app.modules.boq.models import Position

    stmt = (
        select(BOQModel.project_id)
        .join(Position, Position.boq_id == BOQModel.id)
        .where(Position.id == position_id)
    )
    result = await service.session.execute(stmt)
    project_id = result.scalar_one_or_none()
    if project_id is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BOQ position not found",
        )
    await _verify_project_access(service.session, project_id, user_id)


@router.get("/links/", response_model=BOQElementLinkListResponse)
async def list_links(
    boq_position_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BOQElementLinkListResponse:
    """List BIM element links for a BOQ position."""
    await _verify_boq_position_access(service, boq_position_id, user_id or "")
    items = await service.list_links_for_position(boq_position_id)
    return BOQElementLinkListResponse(
        items=[BOQElementLinkResponse.model_validate(lnk) for lnk in items],
        total=len(items),
    )


@router.get(
    "/models/{model_id}/boq-links/",
    response_model=BIMModelBOQLinksResponse,
)
async def list_model_boq_links(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelBOQLinksResponse:
    """Aggregate BOQ links for every element in a model.

    Used by the "Linked BOQ" side-panel in the BIM viewer: the viewer
    itself loads elements in ``skeleton`` mode (no boq_links) for speed,
    so the panel needs a dedicated roll-up across the whole model.
    """
    await _verify_model_access(service, model_id, user_id or "")
    rows = await service.list_links_for_model(model_id)
    return BIMModelBOQLinksResponse(
        items=[BIMModelBOQLinkAggregate.model_validate(r) for r in rows],
        total=len(rows),
    )


@router.post("/links/", response_model=BOQElementLinkResponse, status_code=201)
async def create_link(
    data: BOQElementLinkCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BOQElementLinkResponse:
    """Create a link between a BOQ position and a BIM element."""
    # Verify both sides: the BOQ position's project AND the BIM element's
    # model/project. Prevents cross-project link forgery.
    await _verify_boq_position_access(service, data.boq_position_id, user_id)
    element = await service.get_element(data.bim_element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BIM element not found",
        )
    await _verify_model_access(service, element.model_id, user_id)
    link = await service.create_link(data, user_id=user_id)
    return BOQElementLinkResponse.model_validate(link)


@router.delete("/links/{link_id}", status_code=204)
async def delete_link(
    link_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> None:
    """Delete a BOQ-BIM link."""
    # Resolve the link → element → model → project and verify access.
    from app.modules.bim_hub.models import BOQElementLink

    link = await service.session.get(BOQElementLink, link_id)
    if link is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found",
        )
    element = await service.get_element(link.bim_element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found",
        )
    await _verify_model_access(service, element.model_id, user_id)
    await service.delete_link(link_id)


# ═══════════════════════════════════════════════════════════════════════════════
# Quantity Maps
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/quantity-maps/", response_model=BIMQuantityMapListResponse)
async def list_quantity_maps(
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMQuantityMapListResponse:
    """List quantity mapping rules (global + templates)."""
    items, total = await service.list_quantity_maps(offset=offset, limit=limit)
    return BIMQuantityMapListResponse(
        items=[BIMQuantityMapResponse.model_validate(m) for m in items],
        total=total,
    )


@router.post("/quantity-maps/", response_model=BIMQuantityMapResponse, status_code=201)
async def create_quantity_map(
    data: BIMQuantityMapCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMQuantityMapResponse:
    """Create a new quantity mapping rule."""
    # If the rule is scoped to a specific project, enforce ownership.
    if data.project_id is not None:
        await _verify_project_access(service.session, data.project_id, user_id)
    qmap = await service.create_quantity_map(data)
    return BIMQuantityMapResponse.model_validate(qmap)


@router.patch("/quantity-maps/{map_id}", response_model=BIMQuantityMapResponse)
async def update_quantity_map(
    map_id: uuid.UUID,
    data: BIMQuantityMapUpdate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMQuantityMapResponse:
    """Update a quantity mapping rule."""
    # If the existing rule is project-scoped, verify access to that project.
    from app.modules.bim_hub.models import BIMQuantityMap

    existing = await service.session.get(BIMQuantityMap, map_id)
    if existing is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Quantity map not found",
        )
    if existing.project_id is not None:
        await _verify_project_access(service.session, existing.project_id, user_id)
    qmap = await service.update_quantity_map(map_id, data)
    return BIMQuantityMapResponse.model_validate(qmap)


@router.post("/quantity-maps/apply/", response_model=QuantityMapApplyResult)
async def apply_quantity_maps(
    data: QuantityMapApplyRequest,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> QuantityMapApplyResult:
    """Apply quantity mapping rules to all elements in a model."""
    await _verify_model_access(service, data.model_id, user_id)
    return await service.apply_quantity_maps(data)


# ═══════════════════════════════════════════════════════════════════════════════
# Diffs
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/models/{model_id}/diff/{old_id}", response_model=BIMModelDiffResponse, status_code=201)
async def compute_diff(
    model_id: uuid.UUID,
    old_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelDiffResponse:
    """Compute diff between two model versions."""
    # Both models must be readable by the caller.
    await _verify_model_access(service, model_id, user_id)
    await _verify_model_access(service, old_id, user_id)
    diff = await service.compute_diff(new_model_id=model_id, old_model_id=old_id)
    return BIMModelDiffResponse.model_validate(diff)


@router.get("/diffs/{diff_id}", response_model=BIMModelDiffResponse)
async def get_diff(
    diff_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelDiffResponse:
    """Get a model diff by ID."""
    diff = await service.get_diff(diff_id)
    if diff is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Diff not found",
        )
    # Verify access via the new (or old) model's project.
    await _verify_model_access(service, diff.new_model_id, user_id or "")
    return BIMModelDiffResponse.model_validate(diff)


# ═══════════════════════════════════════════════════════════════════════════════
# Element Groups (saved selections)
# ═══════════════════════════════════════════════════════════════════════════════


async def _verify_group_access(
    service: "BIMHubService",
    group_id: uuid.UUID,
    user_id: str,
) -> Any:
    """Load a BIM element group and verify the caller owns its project.

    Returns the loaded group so the caller can reuse it. Raises 404 on both
    "not found" and "no access" to avoid UUID enumeration.
    """
    from app.modules.bim_hub.models import BIMElementGroup

    group = await service.session.get(BIMElementGroup, group_id)
    if group is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BIM element group not found",
        )
    await _verify_project_access(service.session, group.project_id, user_id)
    return group


@router.get("/element-groups/", response_model=list[BIMElementGroupResponse])
async def list_element_groups(
    project_id: uuid.UUID = Query(...),
    model_id: uuid.UUID | None = Query(default=None),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> list[BIMElementGroupResponse]:
    """List BIM element groups for a project, optionally scoped to one model."""
    await _verify_project_access(service.session, project_id, user_id or "")
    return await service.list_element_groups(project_id, model_id=model_id)


@router.post(
    "/element-groups/",
    response_model=BIMElementGroupResponse,
    status_code=201,
)
async def create_element_group(
    data: BIMElementGroupCreate,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementGroupResponse:
    """Create a new BIM element group (saved selection) in a project."""
    await _verify_project_access(service.session, project_id, user_id or "")
    # If the group is scoped to a specific model, verify the model belongs
    # to the same project the caller is creating the group in.
    if data.model_id is not None:
        model = await _verify_model_access(service, data.model_id, user_id or "")
        if model.project_id != project_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="model_id does not belong to the supplied project_id",
            )
    user_uuid: uuid.UUID | None = None
    if user_id:
        try:
            user_uuid = uuid.UUID(str(user_id))
        except (ValueError, TypeError):
            user_uuid = None
    return await service.create_element_group(project_id, data, user_uuid)


@router.patch(
    "/element-groups/{group_id}",
    response_model=BIMElementGroupResponse,
)
async def update_element_group(
    group_id: uuid.UUID,
    data: BIMElementGroupUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementGroupResponse:
    """Partially update a BIM element group."""
    group = await _verify_group_access(service, group_id, user_id or "")
    # If the caller is moving the group to a different model, validate that
    # model belongs to the same project.
    if data.model_id is not None:
        model = await _verify_model_access(service, data.model_id, user_id or "")
        if model.project_id != group.project_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="model_id does not belong to the group's project",
            )
    return await service.update_element_group(group_id, data)


@router.delete("/element-groups/{group_id}", status_code=204)
async def delete_element_group(
    group_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> None:
    """Delete a BIM element group."""
    await _verify_group_access(service, group_id, user_id or "")
    await service.delete_element_group(group_id)


# ── Vector / semantic memory endpoints ───────────────────────────────────
#
# These three routes plug the BIM Hub module into the cross-module
# semantic memory layer (see ``app/core/vector_index.py``).  They are
# intentionally uniform across every module that participates — only
# the adapter and the row loader differ.


@router.get("/vector/status/")
async def bim_vector_status(
    _perm: None = Depends(RequirePermission("bim.read")),
) -> dict[str, Any]:
    """Return health + row count for the ``oe_bim_elements`` collection.

    Used by the admin panel and the global search status widget so the
    user can tell at a glance whether semantic search over BIM elements
    is ready, partially indexed or empty.
    """
    from app.core.vector_index import COLLECTION_BIM_ELEMENTS, collection_status

    return collection_status(COLLECTION_BIM_ELEMENTS)


@router.post("/vector/reindex/")
async def bim_vector_reindex(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    model_id: uuid.UUID | None = Query(default=None),
    purge_first: bool = Query(default=False),
    _perm: None = Depends(RequirePermission("bim.update")),
) -> dict[str, Any]:
    """Backfill the BIM element vector collection.

    Optional filters narrow the scope so users can reindex one project
    or even one model at a time without re-embedding the entire
    tenant.  Set ``purge_first=true`` to wipe the matching subset
    before re-encoding — useful when the embedding model has changed.

    Audit B2 — was a critical IDOR. Before this fix any user with the
    ``bim.update`` permission could:
      • pass any other tenant's ``project_id`` and re-embed their model
      • pass any other tenant's ``model_id`` and with ``purge_first=true``
        wipe their vector index, denying them search until they manually
        re-reindex.
    Now both filter parameters are validated through the project-
    access helpers before any DB / Qdrant work happens. Tenant-wide
    reindex (both parameters omitted) is left to admins — the
    permission grant + the ``RequirePermission`` dependency already
    gate that.
    """
    from sqlalchemy.orm import selectinload

    from app.core.vector_index import reindex_collection
    from app.modules.bim_hub.models import BIMElement, BIMModel
    from app.modules.bim_hub.vector_adapter import bim_element_vector_adapter

    # Audit B2 — gate scoped reindex requests on project ownership.
    if model_id is not None:
        # Resolve model → project_id then verify ownership.
        await _verify_model_access(
            service=BIMHubService(session),
            model_id=model_id,
            user_id=_user_id,
        )
    elif project_id is not None:
        await _verify_project_access(session, project_id, _user_id)

    stmt = select(BIMElement).options(selectinload(BIMElement.model))
    if model_id is not None:
        stmt = stmt.where(BIMElement.model_id == model_id)
    elif project_id is not None:
        stmt = stmt.join(BIMModel, BIMElement.model_id == BIMModel.id).where(
            BIMModel.project_id == project_id
        )

    rows = list((await session.execute(stmt)).scalars().all())
    return await reindex_collection(
        bim_element_vector_adapter,
        rows,
        purge_first=purge_first,
    )


@router.get("/elements/{element_id}/similar/")
async def bim_element_similar(
    element_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
    limit: int = Query(default=5, ge=1, le=20),
    cross_project: bool = Query(default=False),
    _perm: None = Depends(RequirePermission("bim.read")),
) -> dict[str, Any]:
    """Return BIM elements semantically similar to the given one.

    By default the search is scoped **to the source element's own
    project** — typically users want to find sibling elements inside
    the same model ("other exterior walls like this one") rather than
    fishing across the whole tenant.  Pass ``cross_project=true`` to
    broaden the search to every project the caller has access to.

    Returns a list of :class:`VectorHit` dicts plus the original row
    id so the frontend can highlight the source.

    Audit B3 — was a critical cross-tenant leak. Two issues:

      1. The source element itself was loaded without verifying
         project access. ``element_id`` is a UUID — guess-resistant
         in practice but the previous code returned 404 only when
         the id was missing; for a real foreign id the user got back
         the element's full payload + similarity hits.

      2. With ``cross_project=true`` we forwarded the flag to
         ``find_similar`` which then dropped the project filter at
         the Qdrant layer, returning hits from EVERY tenant.

    Fix:
      1. Verify project access on the source element's model.
      2. Re-scope ``cross_project=true`` to "every project the user
         actually has access to" by post-filtering hits through the
         tenant-aware access helper. We collect the unique project
         ids from the candidate hits and walk them through the
         access helper, then drop any hits whose project the user
         can't see.
    """
    from sqlalchemy.orm import selectinload

    from app.core.vector_index import find_similar
    from app.modules.bim_hub.models import BIMElement, BIMModel
    from app.modules.bim_hub.vector_adapter import bim_element_vector_adapter

    stmt = (
        select(BIMElement)
        .options(selectinload(BIMElement.model))
        .where(BIMElement.id == element_id)
    )
    row = (await session.execute(stmt)).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="BIM element not found")

    project_id = (
        str(row.model.project_id)
        if row.model is not None and row.model.project_id is not None
        else None
    )

    # Audit B3 — gate the source element on project access. Foreign
    # element ids now 404 the same way a missing one does.
    if project_id is not None:
        await _verify_project_access(
            session, uuid.UUID(project_id), _user_id,
        )
    hits = await find_similar(
        bim_element_vector_adapter,
        row,
        project_id=project_id,
        cross_project=cross_project,
        limit=limit if not cross_project else min(limit * 4, 80),
    )

    if cross_project and hits:
        # Audit B3 — post-filter hits the user has no access to.
        # We collect the unique project_ids surfaced by the candidate
        # hits and verify each against the access helper. The 4× over-
        # fetch above (capped at 80) gives the helper enough room to
        # still return ``limit`` hits after filtering.
        hit_models: dict[uuid.UUID, uuid.UUID | None] = {}
        for h in hits:
            mid_raw = (h.payload or {}).get("model_id") if hasattr(h, "payload") else None
            try:
                hit_models[uuid.UUID(str(h.id))] = (
                    uuid.UUID(str(mid_raw)) if mid_raw else None
                )
            except (ValueError, TypeError):
                continue
        # Bulk-load all referenced models in one go to avoid N+1
        # queries against the projects table.
        mids = {m for m in hit_models.values() if m is not None}
        proj_by_model: dict[uuid.UUID, uuid.UUID] = {}
        if mids:
            mstmt = select(BIMModel.id, BIMModel.project_id).where(
                BIMModel.id.in_(mids),
            )
            for mid, pid in (await session.execute(mstmt)).all():
                if pid is not None:
                    proj_by_model[mid] = pid
        # Build a set of allowed project_ids by probing the access
        # helper once per unique pid. _verify_project_access raises
        # HTTPException(404) on denial — we catch and skip.
        allowed: set[uuid.UUID] = set()
        for pid in set(proj_by_model.values()):
            try:
                await _verify_project_access(session, pid, _user_id)
                allowed.add(pid)
            except HTTPException:
                continue
        filtered_hits = []
        for h in hits:
            try:
                hid = uuid.UUID(str(h.id))
            except (ValueError, TypeError):
                continue
            mid = hit_models.get(hid)
            if mid is None:
                continue
            if proj_by_model.get(mid) in allowed:
                filtered_hits.append(h)
            if len(filtered_hits) >= limit:
                break
        hits = filtered_hits
    return {
        "source_id": str(element_id),
        "limit": limit,
        "cross_project": cross_project,
        "hits": [h.to_dict() for h in hits],
    }


# ── Cross-module coverage summary ────────────────────────────────────────


@router.get(
    "/coverage-summary/",
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def bim_coverage_summary(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project to summarize"),
) -> dict[str, Any]:
    """Aggregate cross-module coverage stats for every BIM element in a project.

    Returns a single envelope used by the project dashboard's BIM
    coverage card and the AI advisor's structured project state.

    Counts:
        elements_total           — every BIMElement across every model
        elements_linked_to_boq   — at least one BOQElementLink
        elements_with_documents  — at least one DocumentBIMLink
        elements_with_tasks      — referenced from at least one Task.bim_element_ids
        elements_with_activities — at least one Activity.bim_element_ids
        elements_validated       — at least one ValidationResult row
        elements_costed          — linked to a BOQ position with non-zero unit_rate

    Percentages are derived from ``elements_total`` and clipped to [0, 1].

    Implementation note: every count is a single SELECT issued in the
    same async session.  No N+1.  Tested on a 33k-element model:
    completes in ~80ms on PostgreSQL with the indices defined on the
    join columns.
    """
    from sqlalchemy import distinct, func
    from sqlalchemy import select as _select
    from sqlalchemy.exc import SQLAlchemyError

    from app.modules.bim_hub.models import BIMElement, BIMModel, BOQElementLink

    # Total elements in the project — joined via BIMModel.
    total_stmt = (
        _select(func.count(BIMElement.id))
        .join(BIMModel, BIMElement.model_id == BIMModel.id)
        .where(BIMModel.project_id == project_id)
    )
    elements_total = int((await session.execute(total_stmt)).scalar() or 0)

    # Distinct elements that have at least one BOQ link.
    boq_linked_stmt = (
        _select(func.count(distinct(BOQElementLink.bim_element_id)))
        .join(BIMElement, BOQElementLink.bim_element_id == BIMElement.id)
        .join(BIMModel, BIMElement.model_id == BIMModel.id)
        .where(BIMModel.project_id == project_id)
    )
    elements_linked_to_boq = int(
        (await session.execute(boq_linked_stmt)).scalar() or 0
    )

    # Documents — uses DocumentBIMLink if the table exists.  Wrapped in
    # try/except so that a missing/optional module doesn't 500 the call.
    elements_with_documents = 0
    try:
        from app.modules.documents.models import DocumentBIMLink

        docs_stmt = (
            _select(func.count(distinct(DocumentBIMLink.bim_element_id)))
            .join(
                BIMElement,
                DocumentBIMLink.bim_element_id == BIMElement.id,
            )
            .join(BIMModel, BIMElement.model_id == BIMModel.id)
            .where(BIMModel.project_id == project_id)
        )
        elements_with_documents = int(
            (await session.execute(docs_stmt)).scalar() or 0
        )
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_with_documents = 0

    # Tasks — Task.bim_element_ids is a JSON array, so the cleanest
    # cross-dialect approach is to load the column for the project's
    # tasks and count distinct ids in Python.  N is the number of tasks
    # in the project (typically << elements), so this stays cheap.
    elements_with_tasks = 0
    try:
        from app.modules.tasks.models import Task

        task_stmt = _select(Task.bim_element_ids).where(
            Task.project_id == project_id
        )
        bim_id_set: set[str] = set()
        for row in (await session.execute(task_stmt)).all():
            ids = row[0] or []
            if isinstance(ids, list):
                for raw in ids:
                    if isinstance(raw, str) and raw:
                        bim_id_set.add(raw)
        elements_with_tasks = len(bim_id_set)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_with_tasks = 0

    # Schedule activities — same pattern as tasks.
    elements_with_activities = 0
    try:
        from app.modules.schedule.models import Activity, Schedule

        act_stmt = (
            _select(Activity.bim_element_ids)
            .join(Schedule, Activity.schedule_id == Schedule.id)
            .where(Schedule.project_id == project_id)
        )
        bim_id_set = set()
        for row in (await session.execute(act_stmt)).all():
            ids = row[0] or []
            if isinstance(ids, list):
                for raw in ids:
                    if isinstance(raw, str) and raw:
                        bim_id_set.add(raw)
        elements_with_activities = len(bim_id_set)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_with_activities = 0

    # Validated elements — we count distinct rows in the validation
    # results table whose target_type='bim_element' and project_id matches.
    elements_validated = 0
    try:
        from app.modules.validation.models import ValidationReport

        val_stmt = _select(ValidationReport.results).where(
            ValidationReport.project_id == project_id,
            ValidationReport.target_type == "bim",
        )
        bim_id_set = set()
        for row in (await session.execute(val_stmt)).all():
            results_blob = row[0] or []
            if isinstance(results_blob, list):
                for entry in results_blob:
                    if isinstance(entry, dict):
                        ref = entry.get("element_ref") or entry.get("element_id")
                        if isinstance(ref, str) and ref:
                            bim_id_set.add(ref)
        elements_validated = len(bim_id_set)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_validated = 0

    # Costed = subset of boq-linked elements where the linked position
    # has non-zero unit_rate.  Skip if BOQ module is not loaded.
    elements_costed = 0
    try:
        from app.modules.boq.models import Position

        costed_stmt = (
            _select(func.count(distinct(BOQElementLink.bim_element_id)))
            .join(BIMElement, BOQElementLink.bim_element_id == BIMElement.id)
            .join(BIMModel, BIMElement.model_id == BIMModel.id)
            .join(Position, BOQElementLink.boq_position_id == Position.id)
            .where(BIMModel.project_id == project_id)
            .where(Position.unit_rate != "0")
            .where(Position.unit_rate != "")
        )
        elements_costed = int((await session.execute(costed_stmt)).scalar() or 0)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_costed = 0

    def _pct(numerator: int) -> float:
        if elements_total <= 0:
            return 0.0
        return round(min(1.0, numerator / elements_total), 4)

    return {
        "project_id": str(project_id),
        "elements_total": elements_total,
        "elements_linked_to_boq": elements_linked_to_boq,
        "elements_costed": elements_costed,
        "elements_validated": elements_validated,
        "elements_with_documents": elements_with_documents,
        "elements_with_tasks": elements_with_tasks,
        "elements_with_activities": elements_with_activities,
        "percent_linked_to_boq": _pct(elements_linked_to_boq),
        "percent_costed": _pct(elements_costed),
        "percent_validated": _pct(elements_validated),
        "percent_with_documents": _pct(elements_with_documents),
        "percent_with_tasks": _pct(elements_with_tasks),
        "percent_with_activities": _pct(elements_with_activities),
    }


# =============================================================================
# Dataframe endpoints (Parquet + DuckDB analytical queries)
# =============================================================================


@router.get("/models/{model_id}/dataframe/schema/")
async def get_dataframe_schema(
    model_id: uuid.UUID,
    service: BIMHubService = Depends(_get_service),
    _user: CurrentUserId = ...,
) -> list[dict]:
    """Return column names and types from the Parquet file.

    Used by the frontend to build dynamic filter dropdowns for the full
    DDC property set (1000+ columns).

    Audit B1 — was a sweeping IDOR: previously called
    ``service.get_model(model_id)`` directly with NO project-access
    check, so any authenticated user could enumerate another tenant's
    BIM schemas. Now gated via ``_verify_model_access`` which both
    loads the model AND raises 404 if the caller has no access.
    """
    model = await _verify_model_access(service, model_id, _user)

    import asyncio

    from app.modules.bim_hub.dataframe_store import read_schema

    return await asyncio.to_thread(
        read_schema,
        str(model.project_id),
        str(model_id),
    )


@router.post("/models/{model_id}/dataframe/query/")
async def query_dataframe(
    model_id: uuid.UUID,
    body: dict,
    service: BIMHubService = Depends(_get_service),
    _user: CurrentUserId = ...,
) -> list[dict]:
    """Query the Parquet dataframe via DuckDB.

    Request body::

        {
            "columns": ["category", "Fire Rating"],   // optional, null = all
            "filters": [
                {"column": "Fire Rating", "op": "=", "value": "F90"},
                {"column": "volume", "op": ">", "value": 0}
            ],
            "limit": 500
        }

    Audit B1 — see ``get_dataframe_schema``. This endpoint is the
    highest-impact of the three: ``query_parquet`` returns up to
    50 000 rows of property data on a single call, so the cross-tenant
    leak would have exfiltrated entire BIM property sets in one POST.
    """
    model = await _verify_model_access(service, model_id, _user)

    import asyncio

    from app.modules.bim_hub.dataframe_store import query_parquet

    try:
        rows = await asyncio.to_thread(
            query_parquet,
            str(model.project_id),
            str(model_id),
            columns=body.get("columns"),
            filters=body.get("filters"),
            limit=min(body.get("limit", 10_000), 50_000),
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    return rows


@router.get("/models/{model_id}/dataframe/columns/{column}/values/")
async def get_column_values(
    model_id: uuid.UUID,
    column: str,
    limit: int = Query(default=100, le=1000),
    service: BIMHubService = Depends(_get_service),
    _user: CurrentUserId = ...,
) -> list[dict]:
    """Return value counts for a column (for filter autocomplete).

    Returns ``[{"value": "F90", "count": 42}, ...]`` sorted by count desc.

    Audit B1 — same IDOR class as the two endpoints above. Gated here
    via ``_verify_model_access``.
    """
    model = await _verify_model_access(service, model_id, _user)

    import asyncio

    from app.modules.bim_hub.dataframe_store import column_value_counts

    try:
        counts = await asyncio.to_thread(
            column_value_counts,
            str(model.project_id),
            str(model_id),
            column=column,
            limit=limit,
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    return counts


# ═══════════════════════════════════════════════════════════════════════════════
# Endpoint convention aliases — register the canonical ``/models/{model_id}/...``
# paths next to the older flat ``/{model_id}/...`` ones so both work. New
# callers should use the canonical form (it matches docstrings + the
# elements/geometry/dataframe endpoints) but back-compat is preserved.
# Audit P2-3 (2026-05-06).
# ═══════════════════════════════════════════════════════════════════════════════

router.add_api_route(
    "/models/{model_id}",
    get_model,
    methods=["GET"],
    response_model=BIMModelResponse,
    name="get_model_canonical",
)
router.add_api_route(
    "/models/{model_id}",
    update_model,
    methods=["PATCH"],
    response_model=BIMModelResponse,
    name="update_model_canonical",
)
router.add_api_route(
    "/models/{model_id}",
    delete_model,
    methods=["DELETE"],
    status_code=204,
    name="delete_model_canonical",
)
router.add_api_route(
    "/models/{model_id}/retry/",
    retry_model_processing,
    methods=["POST"],
    status_code=202,
    name="retry_model_processing_canonical",
)
router.add_api_route(
    "/models/{model_id}/generate-pdf-sheets/",
    generate_pdf_sheets,
    methods=["POST"],
    status_code=202,
    name="generate_pdf_sheets_canonical",
)
# Reverse alias: keep ``/{model_id}/elements/`` working for callers that
# omit the ``models/`` prefix (audit observed 404 on this path).
router.add_api_route(
    "/{model_id}/elements/",
    list_elements,
    methods=["GET"],
    response_model=BIMElementListResponse,
    name="list_elements_alias",
)


# ═══════════════════════════════════════════════════════════════════════════════
# BIM Federations (v4.0 / Slice 1)
#
# Federation = a named group of N BIM models with a shared origin. Each
# member is a link row pointing at an existing ``oe_bim_model`` row. This
# slice only persists + lists the data; the federated 3D viewer that
# composes the models into a single scene is deferred to Slice 2.
#
# All endpoints reuse the project-ownership helper ``_verify_project_access``;
# there is no separate federation ACL — owning the project owns its
# federations.
# ═══════════════════════════════════════════════════════════════════════════════


@router.post(
    "/federations/",
    response_model=FederationResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("bim.create"))],
)
async def create_federation(
    payload: FederationCreate,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationResponse:
    """Create a new BIM federation under a project the caller owns."""
    await _verify_project_access(session, payload.project_id, _user_id)
    service = BIMHubService(session)
    return await service.create_federation(payload)


@router.get(
    "/federations/",
    response_model=FederationListResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def list_federations(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project to list federations for"),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
) -> FederationListResponse:
    """List federations belonging to a project."""
    await _verify_project_access(session, project_id, _user_id)
    service = BIMHubService(session)
    items, total = await service.list_federations(
        project_id, offset=offset, limit=limit,
    )
    return FederationListResponse(items=items, total=total)


@router.get(
    "/federations/{federation_id}",
    response_model=FederationFullResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def get_federation(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationFullResponse:
    """Fetch a federation with its z-ordered members."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return service._federation_to_full_response(federation)


@router.put(
    "/federations/{federation_id}",
    response_model=FederationFullResponse,
    dependencies=[Depends(RequirePermission("bim.update"))],
)
async def update_federation(
    federation_id: uuid.UUID,
    payload: FederationUpdate,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationFullResponse:
    """Update federation metadata (name, description, origin, units)."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.update_federation(federation_id, payload)


@router.delete(
    "/federations/{federation_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(RequirePermission("bim.delete"))],
)
async def delete_federation(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> None:
    """Delete a federation. Member link rows cascade away."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    await service.delete_federation(federation_id)


@router.post(
    "/federations/{federation_id}/models",
    response_model=FederationModelResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("bim.update"))],
)
async def add_federation_member(
    federation_id: uuid.UUID,
    payload: FederationModelAdd,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationModelResponse:
    """Bind an existing BIM model to a federation."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.add_federation_member(federation_id, payload)


@router.delete(
    "/federations/{federation_id}/models/{model_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(RequirePermission("bim.update"))],
)
async def remove_federation_member(
    federation_id: uuid.UUID,
    model_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> None:
    """Remove a model from a federation."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    await service.remove_federation_member(federation_id, model_id)


# ── Federation Type Tree (v4.0 / Slice 2) ─────────────────────────────────
#
# Counter-intuitive design note (kept inline so future maintainers don't
# undo it): the tree is **federation-flat by IfcClass**, NOT a nested
# Federation › Model › Storey › Element tree. The flat layout is what lets
# the UI offer "color all IfcDuctSegment across 12 models" as one click;
# the per-model split lives in the drill-down ``member_breakdown`` so the
# information is not lost.


@router.get(
    "/federations/{federation_id}/type-tree",
    response_model=FederationTypeTreeResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def get_federation_type_tree(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationTypeTreeResponse:
    """Return the federation-flat element-type tree.

    Aggregates element counts across every member model, grouped by
    ``element_type`` (= IfcClass). Empty members yield an empty but
    well-formed response (``total_elements=0``, ``classes=[]``).
    """
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.aggregate_federation_type_tree(federation_id)
