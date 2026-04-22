"""COBie (UK 2.4) workbook builder.

COBie — Construction Operations Building Information Exchange — is the
ISO-19650 handover format. An XLSX with one sheet per concept, the
receiving FM / CAFM system (Archibus, Maximo, Planon, etc.) imports
into its own asset register.

This module builds the seven most-commonly-requested sheets in
``COBie.UK.2.4`` shape:

    Contact · Facility · Floor · Space · Type · Component · System

We intentionally leave out the remaining spec sheets (Zone, Job,
Resource, Spare, Impact, Document, Attribute, Coordinate, Issue,
Connection, Assembly) on first iteration — they're either rarely
consumed by real-world CAFM imports or require data we don't yet store
in the canonical model. Later revisions can add them without breaking
the signature.

Source of truth
---------------
The workbook is a PROJECTION of data already held in our models:

    Facility ── BIMModel (name, project metadata)
    Floor    ── BIMElement.storey (unique values)
    Space    ── BIMElement where element_type ∈ {Room, Space, IfcSpace}
    Type     ── BIMElement where element_type matches equipment types
    Component── BIMElement WHERE is_tracked_asset=True (asset register)
    System   ── distinct asset_info.parent_system values
    Contact  ── minimum required "createdBy" contact per row

No SQL is issued from this module — it takes pre-fetched ORM objects
and renders them. The router is responsible for loading the data.

Determinism
-----------
Every row is sorted by a stable key so ``build_cobie_workbook`` is
byte-reproducible for a given input. Snapshot tests compare against a
fixture blob; if sort order were non-deterministic the tests would be
flaky.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# ── COBie column schemas ─────────────────────────────────────────────────
# Columns follow the BuildingSMART/BS 1192-4 COBie UK 2.4 template.
# "Required" fields are marked with * in the spec; we fall back to "n/a"
# when a field isn't available so importers don't reject rows for
# missing data.

CONTACT_COLUMNS = [
    "Email*", "CreatedBy", "CreatedOn", "Category", "Company",
    "Phone", "ExtSystem", "ExtObject", "ExtIdentifier",
    "Department", "OrganizationCode", "GivenName", "FamilyName",
    "Street", "PostalBox", "Town", "StateRegion", "PostalCode",
    "Country",
]

FACILITY_COLUMNS = [
    "Name*", "CreatedBy", "CreatedOn", "Category", "ProjectName",
    "SiteName", "LinearUnits", "AreaUnits", "VolumeUnits", "CurrencyUnit",
    "AreaMeasurement", "ExternalSystem", "ExternalProjectObject",
    "ExternalProjectIdentifier", "ExternalSiteObject",
    "ExternalSiteIdentifier", "ExternalFacilityObject",
    "ExternalFacilityIdentifier", "Description", "ProjectDescription",
    "SiteDescription", "Phase",
]

FLOOR_COLUMNS = [
    "Name*", "CreatedBy", "CreatedOn", "Category", "ExtSystem",
    "ExtObject", "ExtIdentifier", "Description", "Elevation", "Height",
]

SPACE_COLUMNS = [
    "Name*", "CreatedBy", "CreatedOn", "Category", "FloorName",
    "Description", "ExtSystem", "ExtObject", "ExtIdentifier",
    "RoomTag", "UsableHeight", "GrossArea", "NetArea",
]

TYPE_COLUMNS = [
    "Name*", "CreatedBy", "CreatedOn", "Category", "Description",
    "AssetType", "Manufacturer", "ModelNumber", "WarrantyGuarantorParts",
    "WarrantyDurationParts", "WarrantyGuarantorLabor",
    "WarrantyDurationLabor", "WarrantyDurationUnit", "ExtSystem",
    "ExtObject", "ExtIdentifier", "ReplacementCost", "ExpectedLife",
    "DurationUnit", "NominalLength", "NominalWidth", "NominalHeight",
    "ModelReference", "Shape", "Size", "Color", "Finish", "Grade",
    "Material", "Constituents", "Features", "AccessibilityPerformance",
    "CodePerformance", "SustainabilityPerformance",
]

COMPONENT_COLUMNS = [
    "Name*", "CreatedBy", "CreatedOn", "TypeName", "Space",
    "Description", "ExtSystem", "ExtObject", "ExtIdentifier",
    "SerialNumber", "InstallationDate", "WarrantyStartDate", "TagNumber",
    "BarCode", "AssetIdentifier",
]

SYSTEM_COLUMNS = [
    "Name*", "CreatedBy", "CreatedOn", "Category", "ComponentNames",
    "ExtSystem", "ExtObject", "ExtIdentifier", "Description",
]


# Default "unknown" contact — every COBie row needs a creator email,
# even when we don't know one. Real deployments override this via
# ``CobieOptions.default_contact_email``.
DEFAULT_CONTACT_EMAIL = "handover@openconstructionerp.local"
DEFAULT_CREATED_BY = DEFAULT_CONTACT_EMAIL
SHEET_NAME_OVERRIDES = {
    "Facility": "Facility",
    "Floor": "Floor",
    "Space": "Space",
    "Type": "Type",
    "Component": "Component",
    "System": "System",
    "Contact": "Contact",
}

# Elements whose ``element_type`` hints that they ARE a Space (not an
# equipment asset). Case-insensitive suffix match so IfcSpace / Room /
# Rooms / Space all land on the Space sheet.
SPACE_TYPE_TOKENS = {"space", "room", "ifcspace", "rooms"}


@dataclass(frozen=True)
class CobieOptions:
    """Builder configuration. Immutable so we can hash it into cache keys."""

    spec: str = "UK_2.4"
    default_contact_email: str = DEFAULT_CONTACT_EMAIL
    project_name: str | None = None
    site_name: str | None = None
    phase: str = "Handover"
    linear_units: str = "meters"
    area_units: str = "square meters"
    volume_units: str = "cubic meters"
    currency_unit: str = "EUR"
    # Pin the CreatedOn timestamp so snapshot tests stay deterministic;
    # callers pass ``frozen_now`` when exporting and None in production.
    frozen_now: datetime | None = None


# ── Public API ──────────────────────────────────────────────────────────


def build_cobie_workbook(
    model: Any,
    elements: list[Any],
    documents: list[Any] | None = None,
    options: CobieOptions | None = None,
) -> bytes:
    """Build a COBie.UK.2.4 workbook for the supplied model.

    Args:
        model: BIMModel-like object. Must expose ``name``, ``project_id``,
            ``discipline`` (optional). Attribute access only — tests can
            pass a dataclass / SimpleNamespace without going through ORM.
        elements: List of BIMElement-like objects. Each must expose
            ``stable_id``, ``element_type``, ``name``, ``storey``,
            ``discipline``, ``asset_info`` (dict), ``is_tracked_asset``,
            ``quantities`` (dict), ``properties`` (dict).
        documents: Optional list of Document-like objects linked to the
            model. Not yet consumed by any sheet — reserved for the
            future ``Document`` sheet.
        options: Builder configuration. Defaults used when ``None``.

    Returns:
        XLSX file bytes.
    """
    opts = options or CobieOptions()
    wb = Workbook()
    # openpyxl creates a default sheet — we'll rename it to the first
    # COBie sheet (Contact) to avoid an orphan "Sheet" tab.
    first_ws = wb.active
    if first_ws is not None:
        first_ws.title = "Contact"

    created_on = _format_cobie_datetime(opts.frozen_now or datetime.now(UTC))
    contact_ws = wb["Contact"]
    _write_contact_sheet(contact_ws, created_on, opts)

    _write_facility_sheet(wb.create_sheet("Facility"), model, created_on, opts)
    _write_floor_sheet(wb.create_sheet("Floor"), elements, created_on, opts)
    _write_space_sheet(wb.create_sheet("Space"), elements, created_on, opts)
    _write_type_sheet(wb.create_sheet("Type"), elements, created_on, opts)
    _write_component_sheet(wb.create_sheet("Component"), elements, created_on, opts)
    _write_system_sheet(wb.create_sheet("System"), elements, created_on, opts)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Sheet renderers ─────────────────────────────────────────────────────


def _write_header_row(ws: Worksheet, columns: list[str]) -> None:
    """Bold, background-filled COBie header row."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
        cell.fill = PatternFill(
            start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid"
        )


def _write_contact_sheet(ws: Worksheet, created_on: str, opts: CobieOptions) -> None:
    """One placeholder contact row — many CAFM imports reject a blank
    sheet, so we always emit at least the default handover contact."""
    _write_header_row(ws, CONTACT_COLUMNS)
    row = {
        "Email*": opts.default_contact_email,
        "CreatedBy": opts.default_contact_email,
        "CreatedOn": created_on,
        "Category": "Contact",
        "Company": "OpenConstructionERP",
    }
    _append_row(ws, CONTACT_COLUMNS, row)


def _write_facility_sheet(
    ws: Worksheet, model: Any, created_on: str, opts: CobieOptions
) -> None:
    _write_header_row(ws, FACILITY_COLUMNS)
    row = {
        "Name*": getattr(model, "name", "Facility"),
        "CreatedBy": opts.default_contact_email,
        "CreatedOn": created_on,
        "Category": "Facility",
        "ProjectName": opts.project_name or getattr(model, "name", "Project"),
        "SiteName": opts.site_name or "Main Site",
        "LinearUnits": opts.linear_units,
        "AreaUnits": opts.area_units,
        "VolumeUnits": opts.volume_units,
        "CurrencyUnit": opts.currency_unit,
        "AreaMeasurement": "gross",
        "ExternalSystem": "OpenConstructionERP",
        "ExternalProjectObject": "Project",
        "ExternalProjectIdentifier": str(getattr(model, "project_id", "")),
        "ExternalSiteObject": "Site",
        "ExternalSiteIdentifier": "1",
        "ExternalFacilityObject": "BIMModel",
        "ExternalFacilityIdentifier": str(getattr(model, "id", "")),
        "Description": opts.project_name or getattr(model, "name", ""),
        "ProjectDescription": opts.project_name or "",
        "SiteDescription": opts.site_name or "",
        "Phase": opts.phase,
    }
    _append_row(ws, FACILITY_COLUMNS, row)


def _write_floor_sheet(
    ws: Worksheet, elements: list[Any], created_on: str, opts: CobieOptions
) -> None:
    _write_header_row(ws, FLOOR_COLUMNS)
    storeys = _distinct_sorted(
        [getattr(e, "storey", None) for e in elements if getattr(e, "storey", None)]
    )
    if not storeys:
        # Some models have no storey data — emit at least one floor so
        # spaces/components can dangle reference "Floor1".
        storeys = ["Floor 1"]
    for storey in storeys:
        row = {
            "Name*": storey,
            "CreatedBy": opts.default_contact_email,
            "CreatedOn": created_on,
            "Category": "Floor",
            "ExtSystem": "OpenConstructionERP",
            "ExtObject": "BIMElement.storey",
            "ExtIdentifier": storey,
            "Description": storey,
            "Elevation": "0",
            "Height": "0",
        }
        _append_row(ws, FLOOR_COLUMNS, row)


def _is_space_element(element: Any) -> bool:
    et = (getattr(element, "element_type", None) or "").lower()
    return any(token in et for token in SPACE_TYPE_TOKENS)


def _write_space_sheet(
    ws: Worksheet, elements: list[Any], created_on: str, opts: CobieOptions
) -> None:
    _write_header_row(ws, SPACE_COLUMNS)
    spaces = [e for e in elements if _is_space_element(e)]
    spaces.sort(key=lambda e: (getattr(e, "storey", "") or "", getattr(e, "stable_id", "")))
    for sp in spaces:
        q = getattr(sp, "quantities", {}) or {}
        row = {
            "Name*": getattr(sp, "name", None) or getattr(sp, "stable_id", ""),
            "CreatedBy": opts.default_contact_email,
            "CreatedOn": created_on,
            "Category": getattr(sp, "element_type", None) or "Space",
            "FloorName": getattr(sp, "storey", "") or "Floor 1",
            "Description": getattr(sp, "name", None) or getattr(sp, "stable_id", ""),
            "ExtSystem": "OpenConstructionERP",
            "ExtObject": "BIMElement",
            "ExtIdentifier": getattr(sp, "stable_id", ""),
            "RoomTag": getattr(sp, "stable_id", ""),
            "UsableHeight": str(q.get("height", "0")),
            "GrossArea": str(q.get("area", "0")),
            "NetArea": str(q.get("area", "0")),
        }
        _append_row(ws, SPACE_COLUMNS, row)


def _write_type_sheet(
    ws: Worksheet, elements: list[Any], created_on: str, opts: CobieOptions
) -> None:
    """Aggregate tracked assets by (element_type, manufacturer, model)
    → one Type row. Components reference the Type by its Name."""
    _write_header_row(ws, TYPE_COLUMNS)
    type_map: dict[str, dict[str, Any]] = {}
    for element in elements:
        if not getattr(element, "is_tracked_asset", False):
            continue
        ai = getattr(element, "asset_info", {}) or {}
        et = getattr(element, "element_type", None) or "Unknown"
        mfr = ai.get("manufacturer") or "Unknown"
        mdl = ai.get("model") or "Unknown"
        key = _type_key(et, mfr, mdl)
        if key not in type_map:
            type_map[key] = {
                "name": key,
                "category": et,
                "manufacturer": mfr,
                "model": mdl,
                "description": f"{et} · {mfr} {mdl}",
            }
    for key in sorted(type_map.keys()):
        t = type_map[key]
        row = {
            "Name*": t["name"],
            "CreatedBy": opts.default_contact_email,
            "CreatedOn": created_on,
            "Category": t["category"],
            "Description": t["description"],
            "AssetType": "Movable",  # conservative default — CAFM overrides
            "Manufacturer": t["manufacturer"],
            "ModelNumber": t["model"],
            "WarrantyGuarantorParts": t["manufacturer"],
            "WarrantyDurationParts": "0",
            "WarrantyGuarantorLabor": t["manufacturer"],
            "WarrantyDurationLabor": "0",
            "WarrantyDurationUnit": "Years",
            "ExtSystem": "OpenConstructionERP",
            "ExtObject": "BIMElement.element_type",
            "ExtIdentifier": t["name"],
            "ReplacementCost": "0",
            "ExpectedLife": "0",
            "DurationUnit": "Years",
            "NominalLength": "0",
            "NominalWidth": "0",
            "NominalHeight": "0",
            "ModelReference": t["model"],
            "Shape": "n/a",
            "Size": "n/a",
            "Color": "n/a",
            "Finish": "n/a",
            "Grade": "n/a",
            "Material": "n/a",
            "Constituents": "n/a",
            "Features": "n/a",
            "AccessibilityPerformance": "n/a",
            "CodePerformance": "n/a",
            "SustainabilityPerformance": "n/a",
        }
        _append_row(ws, TYPE_COLUMNS, row)


def _write_component_sheet(
    ws: Worksheet, elements: list[Any], created_on: str, opts: CobieOptions
) -> None:
    """One row per tracked asset. Name = stable_id or asset_tag, TypeName
    back-references the Type sheet via (element_type + manufacturer + model)."""
    _write_header_row(ws, COMPONENT_COLUMNS)
    components = [e for e in elements if getattr(e, "is_tracked_asset", False)]
    components.sort(key=lambda e: getattr(e, "stable_id", ""))
    for comp in components:
        ai = getattr(comp, "asset_info", {}) or {}
        et = getattr(comp, "element_type", None) or "Unknown"
        mfr = ai.get("manufacturer") or "Unknown"
        mdl = ai.get("model") or "Unknown"
        type_name = _type_key(et, mfr, mdl)
        row = {
            "Name*": ai.get("asset_tag")
            or getattr(comp, "name", None)
            or getattr(comp, "stable_id", ""),
            "CreatedBy": opts.default_contact_email,
            "CreatedOn": created_on,
            "TypeName": type_name,
            "Space": getattr(comp, "storey", "") or "Floor 1",
            "Description": getattr(comp, "name", "") or getattr(comp, "stable_id", ""),
            "ExtSystem": "OpenConstructionERP",
            "ExtObject": "BIMElement",
            "ExtIdentifier": getattr(comp, "stable_id", ""),
            "SerialNumber": ai.get("serial_number") or "n/a",
            "InstallationDate": ai.get("commissioned_at") or "n/a",
            "WarrantyStartDate": ai.get("commissioned_at") or "n/a",
            "TagNumber": ai.get("asset_tag") or getattr(comp, "stable_id", ""),
            "BarCode": ai.get("asset_tag") or "n/a",
            "AssetIdentifier": ai.get("asset_tag") or getattr(comp, "stable_id", ""),
        }
        _append_row(ws, COMPONENT_COLUMNS, row)


def _write_system_sheet(
    ws: Worksheet, elements: list[Any], created_on: str, opts: CobieOptions
) -> None:
    """Aggregate components by ``asset_info.parent_system``. One row per
    system, ``ComponentNames`` joins the member stable_ids with ``,``."""
    _write_header_row(ws, SYSTEM_COLUMNS)
    system_map: dict[str, list[str]] = {}
    for element in elements:
        if not getattr(element, "is_tracked_asset", False):
            continue
        ai = getattr(element, "asset_info", {}) or {}
        sys_name = ai.get("parent_system")
        if not sys_name:
            continue
        members = system_map.setdefault(sys_name, [])
        members.append(
            ai.get("asset_tag")
            or getattr(element, "name", None)
            or getattr(element, "stable_id", "")
        )
    for sys_name in sorted(system_map.keys()):
        members = sorted(system_map[sys_name])
        row = {
            "Name*": sys_name,
            "CreatedBy": opts.default_contact_email,
            "CreatedOn": created_on,
            "Category": "System",
            "ComponentNames": ",".join(members),
            "ExtSystem": "OpenConstructionERP",
            "ExtObject": "asset_info.parent_system",
            "ExtIdentifier": sys_name,
            "Description": sys_name,
        }
        _append_row(ws, SYSTEM_COLUMNS, row)


# ── Helpers ──────────────────────────────────────────────────────────────


def _append_row(ws: Worksheet, columns: list[str], row: dict[str, Any]) -> None:
    """Append a row in column order, defaulting missing keys to "n/a"
    (COBie expects a value in every cell; blank cells fail strict
    importers)."""
    values = [row.get(col, "n/a") for col in columns]
    # Convert None → "n/a" for type-safety — cells with None become
    # empty in Excel and some CAFM parsers choke.
    values = ["n/a" if v is None else v for v in values]
    ws.append(values)


def _distinct_sorted(items: list[Any]) -> list[str]:
    """Drop duplicates, coerce to string, sort for stability."""
    return sorted({str(x) for x in items if x})


def _type_key(element_type: str, manufacturer: str, model: str) -> str:
    """Canonical Type name: ``"{element_type} - {manufacturer} {model}"``.

    Kept in a single helper so Component rows can back-reference the
    Type row without duplicating the format string (and drifting)."""
    return f"{element_type} - {manufacturer} {model}"


def _format_cobie_datetime(dt: datetime) -> str:
    """COBie wants ISO-8601 without timezone offset — so we strip the
    tzinfo after converting to UTC. Example: ``2026-04-22T09:15:42``."""
    if dt.tzinfo is not None:
        dt = dt.astimezone(UTC).replace(tzinfo=None)
    return dt.strftime("%Y-%m-%dT%H:%M:%S")
