"""Excel + CSV import/export for EAC requirements.

The EAC schema (entity, attribute, constraint_type, constraint_value,
unit, category, priority, ...) maps cleanly to a flat sheet, so this
module sticks to one workbook with one row per requirement. The
template doubles as a working example: row 1 = headers, row 2 = a
filled-in sample, row 3 onwards = blank.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from app.modules.requirements.evaluator import OPERATORS

# Column order is the contract — keep it stable across template,
# import, and export so downstream tooling doesn't break.
COLUMNS: tuple[str, ...] = (
    "entity",
    "attribute",
    "constraint_type",
    "constraint_value",
    "unit",
    "category",
    "priority",
    "source_ref",
    "notes",
)

REQUIRED_COLUMNS: frozenset[str] = frozenset({"entity", "attribute", "constraint_type"})

# Friendly per-column hints rendered into row 2 of the template so the
# user never has to guess what the field expects.
_TEMPLATE_SAMPLE: dict[str, str] = {
    "entity": "Walls",
    "attribute": "FireRating",
    "constraint_type": "regex",
    "constraint_value": "^F\\d{2,3}$",
    "unit": "",
    "category": "fire_safety",
    "priority": "must",
    "source_ref": "DIN 4102",
    "notes": "Required by local building code",
}

_TEMPLATE_HINT: dict[str, str] = {
    "entity": "Element type — Revit category or IFC class (Walls, IfcWall*)",
    "attribute": "Property name — flat or 'Group.Name' (Pset_WallCommon.FireRating)",
    "constraint_type": " | ".join(OPERATORS),
    "constraint_value": "Match value, regex, threshold, or 'min..max'",
    "unit": "Optional — m, m2, W/m2K, F90, …",
    "category": "structural | fire_safety | thermal | acoustic | ...",
    "priority": "must | should | may",
    "source_ref": "Standard, drawing, or spec reference",
    "notes": "Free text",
}


def build_template_xlsx() -> bytes:
    """Generate a friendly Excel template — headers, a sample, and hints."""
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sample_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    sample_font = Font(italic=True, color="555555")

    for idx, col in enumerate(COLUMNS, start=1):
        letter = get_column_letter(idx)
        cell = ws[f"{letter}1"]
        cell.value = col
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.comment = Comment(_TEMPLATE_HINT[col], "OpenConstructionERP")
        ws.column_dimensions[letter].width = max(16, len(col) + 6)

        sample_cell = ws[f"{letter}2"]
        sample_cell.value = _TEMPLATE_SAMPLE.get(col, "")
        sample_cell.fill = sample_fill
        sample_cell.font = sample_font

    # Row 3+ left blank for the user to fill in.
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Legend sheet — operator cheat-sheet so users can pick without docs.
    legend = wb.create_sheet("Operators")
    legend["A1"] = "Operator"
    legend["B1"] = "constraint_value example"
    legend["C1"] = "Meaning"
    for cell_ref in ("A1", "B1", "C1"):
        legend[cell_ref].fill = header_fill
        legend[cell_ref].font = header_font
    legend.column_dimensions["A"].width = 18
    legend.column_dimensions["B"].width = 28
    legend.column_dimensions["C"].width = 50

    legend_rows: list[tuple[str, str, str]] = [
        ("equals", "Concrete C30/37", "actual must equal value (case-insensitive)"),
        ("not_equals", "Steel", "actual must not equal value"),
        ("contains", "fire", "value must appear in actual (substring)"),
        ("not_contains", "draft", "value must NOT appear in actual"),
        ("min", "0.24", "numeric: actual >= value"),
        ("max", "5.0", "numeric: actual <= value"),
        ("range", "200..400", "numeric: value lo..hi (also '-', ',', ';')"),
        ("regex", r"^F\d{2,3}$", "actual must match the regex"),
        ("exists", "(blank)", "property must be present and non-empty"),
        ("not_exists", "(blank)", "property must be missing or empty"),
    ]
    for i, row in enumerate(legend_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            legend.cell(row=i, column=col_idx, value=value)
    legend.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalise_row(row: dict[str, Any]) -> dict[str, Any]:
    """Coerce one parsed row into a RequirementCreate-friendly dict."""
    out: dict[str, Any] = {}
    for col in COLUMNS:
        raw = row.get(col, "")
        if raw is None:
            out[col] = ""
            continue
        out[col] = str(raw).strip() if not isinstance(raw, str) else raw.strip()

    op = out["constraint_type"].lower() or "equals"
    out["constraint_type"] = op
    if not out.get("priority"):
        out["priority"] = "must"
    if not out.get("category"):
        out["category"] = "general"
    return out


def parse_xlsx(payload: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse an Excel workbook into rows + a list of warnings.

    The parser is forgiving: header row may be in any column order,
    extra columns are ignored, and missing rows are skipped silently.
    Empty rows (no entity AND no attribute) are dropped.
    """
    import openpyxl

    warnings: list[str] = []
    rows: list[dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl raises a variety
        warnings.append(f"Could not open workbook: {exc}")
        return rows, warnings

    sheet = wb.active
    if sheet is None:
        warnings.append("Workbook has no active sheet")
        return rows, warnings

    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        warnings.append("Workbook is empty")
        return rows, warnings

    headers = [str(c).strip().lower() if c is not None else "" for c in header_row]
    col_index: dict[str, int] = {}
    for idx, name in enumerate(headers):
        if name in COLUMNS and name not in col_index:
            col_index[name] = idx

    missing_required = REQUIRED_COLUMNS - col_index.keys()
    if missing_required:
        warnings.append(
            "Missing required columns: " + ", ".join(sorted(missing_required))
        )
        return rows, warnings

    for line_no, raw_row in enumerate(iterator, start=2):
        if raw_row is None:
            continue
        entity = (
            str(raw_row[col_index["entity"]]).strip()
            if col_index.get("entity") is not None
            and raw_row[col_index["entity"]] is not None
            else ""
        )
        attribute = (
            str(raw_row[col_index["attribute"]]).strip()
            if col_index.get("attribute") is not None
            and raw_row[col_index["attribute"]] is not None
            else ""
        )
        if not entity and not attribute:
            continue

        record: dict[str, Any] = {col: "" for col in COLUMNS}
        for col, idx in col_index.items():
            if idx >= len(raw_row):
                continue
            cell = raw_row[idx]
            record[col] = "" if cell is None else cell

        normalised = _normalise_row(record)
        if normalised["constraint_type"] not in OPERATORS:
            warnings.append(
                f"Row {line_no}: unknown operator "
                f"'{normalised['constraint_type']}', defaulting to equals"
            )
            normalised["constraint_type"] = "equals"
        rows.append(normalised)

    return rows, warnings


def parse_csv(payload: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse a CSV file into rows + warnings (same contract as xlsx)."""
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []

    text = payload.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        warnings.append("CSV is empty")
        return rows, warnings

    headers = [c.strip().lower() for c in header]
    col_index = {name: idx for idx, name in enumerate(headers) if name in COLUMNS}

    missing_required = REQUIRED_COLUMNS - col_index.keys()
    if missing_required:
        warnings.append("Missing required columns: " + ", ".join(sorted(missing_required)))
        return rows, warnings

    for line_no, raw_row in enumerate(reader, start=2):
        if not raw_row or all(not cell.strip() for cell in raw_row):
            continue
        record = {col: "" for col in COLUMNS}
        for col, idx in col_index.items():
            if idx < len(raw_row):
                record[col] = raw_row[idx]
        normalised = _normalise_row(record)
        if not normalised["entity"] and not normalised["attribute"]:
            continue
        if normalised["constraint_type"] not in OPERATORS:
            warnings.append(
                f"Row {line_no}: unknown operator "
                f"'{normalised['constraint_type']}', defaulting to equals"
            )
            normalised["constraint_type"] = "equals"
        rows.append(normalised)

    return rows, warnings


def export_xlsx(rows: list[dict[str, Any]], title: str = "Requirements") -> bytes:
    """Build a formatted Excel workbook from already-fetched rows."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Requirements"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for idx, col in enumerate(COLUMNS, start=1):
        letter = get_column_letter(idx)
        cell = ws[f"{letter}1"]
        cell.value = col
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[letter].width = max(14, len(col) + 6)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col, "") or "")

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
