"""Field Reports API routes.

Endpoints:
    POST   /reports                        - Create field report
    GET    /reports?project_id=X           - List with filters
    GET    /reports/{id}                   - Get single
    PATCH  /reports/{id}                   - Update
    DELETE /reports/{id}                   - Delete
    POST   /reports/{id}/submit            - Submit for approval
    POST   /reports/{id}/approve           - Approve
    GET    /reports/summary?project_id=X   - Aggregated stats
    GET    /reports/{id}/export/pdf         - Download PDF
    GET    /reports/calendar?project_id=X  - Reports by month for calendar
    GET    /reports/template               - Download import template
    POST   /reports/import/file?project_id=X - Import from CSV/Excel
    GET    /reports/export?project_id=X    - Export all as Excel
"""

import csv
import io
import logging
import uuid
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response, StreamingResponse

from app.core.upload_guards import reject_if_xlsx_bomb
from app.dependencies import CurrentUserId, RequirePermission, SessionDep, verify_project_access
from app.modules.fieldreports.schemas import (
    FieldReportCreate,
    FieldReportResponse,
    FieldReportSummary,
    FieldReportUpdate,
    LinkDocumentsRequest,
    LinkedDocumentResponse,
    SiteEquipmentLogCreate,
    SiteEquipmentLogResponse,
    SiteEquipmentLogUpdate,
    SiteWorkforceLogCreate,
    SiteWorkforceLogResponse,
    SiteWorkforceLogUpdate,
)
from app.modules.fieldreports.service import FieldReportService

router = APIRouter()
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> FieldReportService:
    return FieldReportService(session)


def _report_to_response(report: object) -> FieldReportResponse:
    """Build a FieldReportResponse from a FieldReport ORM object."""
    return FieldReportResponse(
        id=report.id,  # type: ignore[attr-defined]
        project_id=report.project_id,  # type: ignore[attr-defined]
        report_date=report.report_date,  # type: ignore[attr-defined]
        report_type=report.report_type,  # type: ignore[attr-defined]
        weather_condition=report.weather_condition,  # type: ignore[attr-defined]
        temperature_c=report.temperature_c,  # type: ignore[attr-defined]
        wind_speed=report.wind_speed,  # type: ignore[attr-defined]
        precipitation=report.precipitation,  # type: ignore[attr-defined]
        humidity=report.humidity,  # type: ignore[attr-defined]
        workforce=report.workforce or [],  # type: ignore[attr-defined]
        equipment_on_site=report.equipment_on_site or [],  # type: ignore[attr-defined]
        work_performed=report.work_performed,  # type: ignore[attr-defined]
        delays=report.delays,  # type: ignore[attr-defined]
        delay_hours=report.delay_hours,  # type: ignore[attr-defined]
        visitors=report.visitors,  # type: ignore[attr-defined]
        deliveries=report.deliveries,  # type: ignore[attr-defined]
        safety_incidents=report.safety_incidents,  # type: ignore[attr-defined]
        materials_used=report.materials_used or [],  # type: ignore[attr-defined]
        photos=report.photos or [],  # type: ignore[attr-defined]
        notes=report.notes,  # type: ignore[attr-defined]
        signature_by=report.signature_by,  # type: ignore[attr-defined]
        signature_data=report.signature_data,  # type: ignore[attr-defined]
        status=report.status,  # type: ignore[attr-defined]
        approved_by=report.approved_by,  # type: ignore[attr-defined]
        approved_at=report.approved_at,  # type: ignore[attr-defined]
        document_ids=report.document_ids or [],  # type: ignore[attr-defined]
        created_by=report.created_by,  # type: ignore[attr-defined]
        metadata=getattr(report, "metadata_", {}),  # type: ignore[attr-defined]
        created_at=report.created_at,  # type: ignore[attr-defined]
        updated_at=report.updated_at,  # type: ignore[attr-defined]
    )


# ── Summary ──────────────────────────────────────────────────────────────────


@router.get("/reports/summary/", response_model=FieldReportSummary)
async def get_summary(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FieldReportService = Depends(_get_service),
) -> FieldReportSummary:
    """Aggregated field report stats for a project."""
    await verify_project_access(project_id, user_id, session)
    data = await service.get_summary(project_id)
    return FieldReportSummary(**data)


# ── Calendar ─────────────────────────────────────────────────────────────────


@router.get("/reports/calendar/", response_model=list[FieldReportResponse])
async def get_calendar(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    month: str = Query(..., pattern=r"^\d{4}-\d{2}$"),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FieldReportService = Depends(_get_service),
) -> list[FieldReportResponse]:
    """Get reports for a month (calendar view). Month format: YYYY-MM."""
    await verify_project_access(project_id, user_id, session)
    parts = month.split("-")
    year, mon = int(parts[0]), int(parts[1])
    reports = await service.get_calendar(project_id, year, mon)
    return [_report_to_response(r) for r in reports]


# ── Weather ─────────────────────────────────────────────────────────────────


@router.get("/weather/")
async def get_current_weather(
    lat: float = Query(..., description="Latitude"),
    lon: float = Query(..., description="Longitude"),
    _user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict:
    """Fetch current weather for a location (optional, requires OPENWEATHERMAP_API_KEY)."""
    from app.config import get_settings
    from app.modules.fieldreports.weather import fetch_weather

    settings = get_settings()
    api_key = settings.openweathermap_api_key
    if not api_key:
        return {"available": False, "error": "OpenWeatherMap API key not configured"}

    result = await fetch_weather(lat, lon, api_key=api_key)
    if result is None:
        return {"available": False, "error": "Weather fetch failed"}

    return {"available": True, **result}


# ── Import template ─────────────────────────────────────────────────────────


@router.get("/reports/template/")
async def download_field_reports_template(
    _user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> StreamingResponse:
    """Download an Excel template for importing field reports.

    Contains three sheets (Field Reports, Workforce Log, Equipment Log) with
    headers and one example row each, plus a Notes sheet explaining columns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()

    # ── Sheet 1: Field Reports ──────────────────────────────────────────
    ws_reports = wb.active
    ws_reports.title = "Field Reports"

    report_headers = [
        "Date",
        "Weather",
        "Temperature",
        "Wind",
        "Description",
        "Workforce Summary",
        "Equipment Summary",
        "Notes",
    ]
    for i, h in enumerate(report_headers, 1):
        cell = ws_reports.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    report_example = [
        "2026-04-07",
        "clear",
        "18",
        "light",
        "Foundation pour for zone A completed",
        "25 workers, 8h avg",
        "Crane x2, Excavator x1",
        "Minor delay due to material delivery",
    ]
    for i, val in enumerate(report_example, 1):
        ws_reports.cell(row=2, column=i, value=val)

    # ── Sheet 2: Workforce Log ──────────────────────────────────────────
    ws_workforce = wb.create_sheet("Workforce Log")

    workforce_headers = [
        "Worker Type",
        "Company",
        "Headcount",
        "Hours Worked",
        "Overtime Hours",
    ]
    for i, h in enumerate(workforce_headers, 1):
        cell = ws_workforce.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    workforce_example = ["Concrete", "Acme Construction GmbH", "8", "8.0", "1.5"]
    for i, val in enumerate(workforce_example, 1):
        ws_workforce.cell(row=2, column=i, value=val)

    # ── Sheet 3: Equipment Log ──────────────────────────────────────────
    ws_equipment = wb.create_sheet("Equipment Log")

    equipment_headers = [
        "Equipment Description",
        "Type",
        "Hours Operational",
        "Hours Standby",
        "Hours Breakdown",
        "Operator",
    ]
    for i, h in enumerate(equipment_headers, 1):
        cell = ws_equipment.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    equipment_example = ["Liebherr LTM 1100", "crane", "6.0", "1.5", "0.5", "Hans Mueller"]
    for i, val in enumerate(equipment_example, 1):
        ws_equipment.cell(row=2, column=i, value=val)

    # ── Sheet 4: Notes ──────────────────────────────────────────────────
    ws_notes = wb.create_sheet("Notes")
    ws_notes.cell(row=1, column=1, value="Column").font = Font(bold=True)
    ws_notes.cell(row=1, column=2, value="Description").font = Font(bold=True)

    notes_data = [
        ("Date", "Report date in YYYY-MM-DD format (e.g. 2026-04-07)"),
        ("Weather", "clear, cloudy, rain, snow, fog, or storm"),
        ("Temperature", "Temperature in Celsius (number)"),
        ("Wind", "Wind description: calm, light, moderate, strong, storm"),
        ("Description", "Free-text summary of work performed on site"),
        ("Workforce Summary", "Summary of workforce present (free text)"),
        ("Equipment Summary", "Summary of equipment on site (free text)"),
        ("Notes", "Additional notes, delays, incidents"),
        ("Worker Type", "Trade or role (e.g. Concrete, Electrical, General Labor)"),
        ("Company", "Subcontractor or company name"),
        ("Headcount", "Number of workers (integer)"),
        ("Hours Worked", "Regular hours per worker (decimal)"),
        ("Overtime Hours", "Overtime hours per worker (decimal)"),
        ("Equipment Description", "Equipment name/model"),
        ("Type", "Equipment category (crane, excavator, pump, etc.)"),
        ("Hours Operational", "Hours the equipment was in use (decimal)"),
        ("Hours Standby", "Hours the equipment was on standby (decimal)"),
        ("Hours Breakdown", "Hours the equipment was broken down (decimal)"),
        ("Operator", "Name of the equipment operator"),
    ]
    for row_idx, (col_name, desc) in enumerate(notes_data, 2):
        ws_notes.cell(row=row_idx, column=1, value=col_name)
        cell = ws_notes.cell(row=row_idx, column=2, value=desc)
        cell.alignment = Alignment(wrap_text=True)

    ws_notes.column_dimensions["A"].width = 25
    ws_notes.column_dimensions["B"].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="field_reports_import_template.xlsx"'
        },
    )


# ── Column alias maps for import ───────────────────────────────────────────

_REPORT_COLUMN_ALIASES: dict[str, list[str]] = {
    "report_date": [
        "date",
        "report_date",
        "report date",
        "datum",
        "berichtsdatum",
    ],
    "weather_condition": [
        "weather",
        "weather_condition",
        "weather condition",
        "wetter",
    ],
    "temperature_c": [
        "temperature",
        "temperature_c",
        "temp",
        "temperatur",
    ],
    "wind_speed": [
        "wind",
        "wind_speed",
        "wind speed",
    ],
    "work_performed": [
        "description",
        "work_performed",
        "work performed",
        "beschreibung",
        "arbeit",
    ],
    "workforce_summary": [
        "workforce summary",
        "workforce_summary",
        "workforce",
        "arbeitskraefte",
    ],
    "equipment_summary": [
        "equipment summary",
        "equipment_summary",
        "equipment",
        "geraete",
    ],
    "notes": [
        "notes",
        "note",
        "bemerkung",
        "anmerkung",
        "notizen",
    ],
}

_ALLOWED_WEATHER = {"clear", "cloudy", "rain", "snow", "fog", "storm"}


def _match_report_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map."""
    normalised = header.strip().lower()
    for canonical, aliases in _REPORT_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _parse_report_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for field report import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_report_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_report_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for field report import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_report_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


# ── Import from file ───────────────────────────────────────────────────────


@router.post("/reports/import/file/")
async def import_field_reports_file(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    _user_id: CurrentUserId = None,  # type: ignore[assignment]
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    service: FieldReportService = Depends(_get_service),
) -> dict[str, Any]:
    """Import field reports from an Excel or CSV file upload.

    Parses columns with flexible EN/DE aliases.  Returns a summary with
    counts of imported, skipped, and error details.
    """
    await verify_project_access(project_id, _user_id, session)
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File too large. Maximum size is 10 MB.",
        )

    # Zip-bomb guard: reject .xlsx whose uncompressed sheets exceed 50 MB.
    reject_if_xlsx_bomb(content)

    try:
        if filename.endswith((".xlsx", ".xls")):
            rows = _parse_report_rows_from_excel(content)
        else:
            rows = _parse_report_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing field report import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    imported_count = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            report_date_raw = str(row.get("report_date", "")).strip()
            if not report_date_raw:
                skipped += 1
                continue

            # Parse date
            try:
                report_date_val = date.fromisoformat(report_date_raw)
            except ValueError:
                errors.append({
                    "row": row_idx,
                    "error": f"Invalid date format: {report_date_raw}. Use YYYY-MM-DD.",
                    "data": {k: str(v)[:100] for k, v in row.items()},
                })
                continue

            # Weather
            weather = str(row.get("weather_condition", "clear")).strip().lower()
            if weather not in _ALLOWED_WEATHER:
                weather = "clear"

            # Temperature
            temp_raw = str(row.get("temperature_c", "")).strip()
            temperature_c: float | None = None
            if temp_raw:
                try:
                    temperature_c = float(temp_raw)
                except ValueError:
                    pass

            wind_speed = str(row.get("wind_speed", "")).strip() or None
            work_performed = str(row.get("work_performed", "")).strip()
            notes = str(row.get("notes", "")).strip() or None

            # Build workforce/equipment from summary text (stored in notes/metadata)
            workforce_summary = str(row.get("workforce_summary", "")).strip()
            equipment_summary = str(row.get("equipment_summary", "")).strip()

            # Create via service
            create_data = FieldReportCreate(
                project_id=project_id,
                report_date=report_date_val,
                weather_condition=weather,
                temperature_c=temperature_c,
                wind_speed=wind_speed,
                work_performed=work_performed,
                notes=notes,
                equipment_on_site=[equipment_summary] if equipment_summary else [],
                metadata={
                    "imported": True,
                    "workforce_summary": workforce_summary,
                    "equipment_summary": equipment_summary,
                },
            )
            await service.create_report(create_data, user_id=_user_id or "import")
            imported_count += 1

        except Exception as exc:
            errors.append({
                "row": row_idx,
                "error": str(exc),
                "data": {k: str(v)[:100] for k, v in row.items()},
            })
            logger.warning("Field report import error at row %d: %s", row_idx, exc)

    logger.info(
        "Field report file import complete: imported=%d, skipped=%d, errors=%d",
        imported_count,
        skipped,
        len(errors),
    )

    return {
        "imported": imported_count,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Export all reports as Excel ────────────────────────────────────────────


@router.get("/reports/export/")
async def export_field_reports(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    _user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FieldReportService = Depends(_get_service),
) -> StreamingResponse:
    """Export all field reports for a project as an Excel file."""
    await verify_project_access(project_id, _user_id, session)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    reports, _ = await service.list_reports(project_id, offset=0, limit=2000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Field Reports"

    headers = [
        "Date",
        "Weather",
        "Temperature",
        "Wind",
        "Description",
        "Workforce Count",
        "Equipment Count",
        "Notes",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, report in enumerate(reports, 2):
        workforce = report.workforce or []  # type: ignore[attr-defined]
        equipment = report.equipment_on_site or []  # type: ignore[attr-defined]
        workforce_count = sum(
            (e.get("count", 0) if isinstance(e, dict) else 0) for e in workforce
        )
        ws.cell(
            row=row_idx, column=1, value=str(report.report_date)  # type: ignore[attr-defined]
        )
        ws.cell(
            row=row_idx, column=2, value=report.weather_condition  # type: ignore[attr-defined]
        )
        ws.cell(
            row=row_idx, column=3, value=report.temperature_c  # type: ignore[attr-defined]
        )
        ws.cell(
            row=row_idx, column=4, value=report.wind_speed  # type: ignore[attr-defined]
        )
        ws.cell(
            row=row_idx, column=5, value=report.work_performed  # type: ignore[attr-defined]
        )
        ws.cell(row=row_idx, column=6, value=workforce_count)
        ws.cell(row=row_idx, column=7, value=len(equipment))
        ws.cell(row=row_idx, column=8, value=report.notes)  # type: ignore[attr-defined]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="field_reports_export.xlsx"'},
    )


# ── Create ───────────────────────────────────────────────────────────────────


@router.post("/reports/", response_model=FieldReportResponse, status_code=201)
async def create_report(
    data: FieldReportCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("fieldreports.create")),
    service: FieldReportService = Depends(_get_service),
) -> FieldReportResponse:
    """Create a new field report."""
    await verify_project_access(data.project_id, user_id, session)
    report = await service.create_report(data, user_id=user_id)
    return _report_to_response(report)


# ── List ─────────────────────────────────────────────────────────────────────


@router.get("/reports/", response_model=list[FieldReportResponse])
async def list_reports(
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    type_filter: str | None = Query(default=None, alias="type"),
    service: FieldReportService = Depends(_get_service),
) -> list[FieldReportResponse]:
    """List field reports for a project with optional filters."""
    await verify_project_access(project_id, user_id, session)
    reports, _ = await service.list_reports(
        project_id,
        offset=offset,
        limit=limit,
        date_from=date_from,
        date_to=date_to,
        report_type=type_filter,
        status_filter=status_filter,
    )
    return [_report_to_response(r) for r in reports]


# ── Get ──────────────────────────────────────────────────────────────────────


@router.get("/reports/{report_id}", response_model=FieldReportResponse)
async def get_report(
    report_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FieldReportService = Depends(_get_service),
) -> FieldReportResponse:
    """Get a single field report."""
    report = await service.get_report(report_id)
    return _report_to_response(report)


# ── Update ───────────────────────────────────────────────────────────────────


@router.patch("/reports/{report_id}", response_model=FieldReportResponse)
async def update_report(
    report_id: uuid.UUID,
    data: FieldReportUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
    service: FieldReportService = Depends(_get_service),
) -> FieldReportResponse:
    """Update a field report."""
    report = await service.update_report(report_id, data)
    return _report_to_response(report)


# ── Delete ───────────────────────────────────────────────────────────────────


@router.delete("/reports/{report_id}", status_code=204)
async def delete_report(
    report_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.delete")),
    service: FieldReportService = Depends(_get_service),
) -> None:
    """Delete a field report."""
    await service.delete_report(report_id)


# ── Submit ───────────────────────────────────────────────────────────────────


@router.post("/reports/{report_id}/submit/", response_model=FieldReportResponse)
async def submit_report(
    report_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
    service: FieldReportService = Depends(_get_service),
) -> FieldReportResponse:
    """Submit a draft report for approval."""
    report = await service.submit_report(report_id)
    return _report_to_response(report)


# ── Approve ──────────────────────────────────────────────────────────────────


@router.post("/reports/{report_id}/approve/", response_model=FieldReportResponse)
async def approve_report(
    report_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("fieldreports.approve")),
    service: FieldReportService = Depends(_get_service),
) -> FieldReportResponse:
    """Approve a submitted report."""
    report = await service.approve_report(report_id, user_id)
    return _report_to_response(report)


# ── Link documents ──────────────────────────────────────────────────────────


@router.post("/reports/{report_id}/link-documents/", response_model=FieldReportResponse)
async def link_documents(
    report_id: uuid.UUID,
    data: LinkDocumentsRequest,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
    service: FieldReportService = Depends(_get_service),
) -> FieldReportResponse:
    """Link one or more documents to a field report.

    Merges the provided document_ids with any already linked, avoiding
    duplicates.
    """
    report = await service.link_documents(report_id, data.document_ids)
    return _report_to_response(report)


@router.get("/reports/{report_id}/documents/", response_model=list[LinkedDocumentResponse])
async def get_linked_documents(
    report_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FieldReportService = Depends(_get_service),
) -> list[LinkedDocumentResponse]:
    """Return the documents linked to a field report.

    Looks up each document_id in the documents module and returns basic
    metadata for each.
    """
    report = await service.get_report(report_id)
    doc_ids = report.document_ids or []

    if not doc_ids:
        return []

    from sqlalchemy import select

    from app.modules.documents.models import Document

    stmt = select(Document).where(Document.id.in_([uuid.UUID(d) for d in doc_ids]))
    result = await session.execute(stmt)
    docs = result.scalars().all()

    return [
        LinkedDocumentResponse(
            id=doc.id,  # type: ignore[attr-defined]
            name=doc.name,  # type: ignore[attr-defined]
            category=doc.category,  # type: ignore[attr-defined]
            file_size=doc.file_size,  # type: ignore[attr-defined]
            mime_type=doc.mime_type,  # type: ignore[attr-defined]
        )
        for doc in docs
    ]


# ── PDF Export ───────────────────────────────────────────────────────────────


@router.get("/reports/{report_id}/export/pdf/")
async def export_pdf(
    report_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: FieldReportService = Depends(_get_service),
) -> Response:
    """Export a field report as PDF."""
    pdf_bytes = await service.generate_pdf(report_id)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=field_report_{report_id}.pdf"},
    )


# ── Site Workforce Log CRUD ────────────────────────────────────────────────


@router.post(
    "/reports/{report_id}/workforce/",
    response_model=SiteWorkforceLogResponse,
    status_code=201,
)
async def create_workforce_log(
    report_id: uuid.UUID,
    data: SiteWorkforceLogCreate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
) -> SiteWorkforceLogResponse:
    """Add a workforce log entry to a field report."""
    from app.modules.fieldreports.models import SiteWorkforceLog

    entry = SiteWorkforceLog(
        field_report_id=report_id,
        worker_type=data.worker_type,
        company=data.company,
        headcount=data.headcount,
        hours_worked=data.hours_worked,
        overtime_hours=data.overtime_hours,
        wbs_id=data.wbs_id,
        cost_category=data.cost_category,
        metadata_=data.metadata,
    )
    session.add(entry)
    await session.flush()
    return SiteWorkforceLogResponse.model_validate(entry)


@router.get(
    "/reports/{report_id}/workforce/",
    response_model=list[SiteWorkforceLogResponse],
)
async def list_workforce_logs(
    report_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> list[SiteWorkforceLogResponse]:
    """List all workforce log entries for a field report."""
    from sqlalchemy import select

    from app.modules.fieldreports.models import SiteWorkforceLog

    stmt = select(SiteWorkforceLog).where(SiteWorkforceLog.field_report_id == report_id)
    result = await session.execute(stmt)
    entries = list(result.scalars().all())
    return [SiteWorkforceLogResponse.model_validate(e) for e in entries]


@router.patch(
    "/workforce/{entry_id}",
    response_model=SiteWorkforceLogResponse,
)
async def update_workforce_log(
    entry_id: uuid.UUID,
    data: SiteWorkforceLogUpdate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
) -> SiteWorkforceLogResponse:
    """Update a workforce log entry."""
    from fastapi import HTTPException
    from sqlalchemy import update

    from app.modules.fieldreports.models import SiteWorkforceLog

    entry = await session.get(SiteWorkforceLog, entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Workforce log entry not found")

    updates = data.model_dump(exclude_unset=True)
    if "metadata" in updates:
        updates["metadata_"] = updates.pop("metadata")
    if updates:
        stmt = update(SiteWorkforceLog).where(SiteWorkforceLog.id == entry_id).values(**updates)
        await session.execute(stmt)
        await session.flush()
        session.expire_all()
        entry = await session.get(SiteWorkforceLog, entry_id)
    return SiteWorkforceLogResponse.model_validate(entry)


@router.delete("/workforce/{entry_id}", status_code=204)
async def delete_workforce_log(
    entry_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.delete")),
) -> None:
    """Delete a workforce log entry."""
    from fastapi import HTTPException

    from app.modules.fieldreports.models import SiteWorkforceLog

    entry = await session.get(SiteWorkforceLog, entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Workforce log entry not found")
    await session.delete(entry)
    await session.flush()


# ── Site Equipment Log CRUD ────────────────────────────────────────────────


@router.post(
    "/reports/{report_id}/equipment/",
    response_model=SiteEquipmentLogResponse,
    status_code=201,
)
async def create_equipment_log(
    report_id: uuid.UUID,
    data: SiteEquipmentLogCreate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
) -> SiteEquipmentLogResponse:
    """Add an equipment log entry to a field report."""
    from app.modules.fieldreports.models import SiteEquipmentLog

    entry = SiteEquipmentLog(
        field_report_id=report_id,
        equipment_description=data.equipment_description,
        equipment_type=data.equipment_type,
        hours_operational=data.hours_operational,
        hours_standby=data.hours_standby,
        hours_breakdown=data.hours_breakdown,
        operator_name=data.operator_name,
        metadata_=data.metadata,
    )
    session.add(entry)
    await session.flush()
    return SiteEquipmentLogResponse.model_validate(entry)


@router.get(
    "/reports/{report_id}/equipment/",
    response_model=list[SiteEquipmentLogResponse],
)
async def list_equipment_logs(
    report_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> list[SiteEquipmentLogResponse]:
    """List all equipment log entries for a field report."""
    from sqlalchemy import select

    from app.modules.fieldreports.models import SiteEquipmentLog

    stmt = select(SiteEquipmentLog).where(SiteEquipmentLog.field_report_id == report_id)
    result = await session.execute(stmt)
    entries = list(result.scalars().all())
    return [SiteEquipmentLogResponse.model_validate(e) for e in entries]


@router.patch(
    "/equipment/{entry_id}",
    response_model=SiteEquipmentLogResponse,
)
async def update_equipment_log(
    entry_id: uuid.UUID,
    data: SiteEquipmentLogUpdate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.update")),
) -> SiteEquipmentLogResponse:
    """Update an equipment log entry."""
    from fastapi import HTTPException
    from sqlalchemy import update

    from app.modules.fieldreports.models import SiteEquipmentLog

    entry = await session.get(SiteEquipmentLog, entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Equipment log entry not found")

    updates = data.model_dump(exclude_unset=True)
    if "metadata" in updates:
        updates["metadata_"] = updates.pop("metadata")
    if updates:
        stmt = update(SiteEquipmentLog).where(SiteEquipmentLog.id == entry_id).values(**updates)
        await session.execute(stmt)
        await session.flush()
        session.expire_all()
        entry = await session.get(SiteEquipmentLog, entry_id)
    return SiteEquipmentLogResponse.model_validate(entry)


@router.delete("/equipment/{entry_id}", status_code=204)
async def delete_equipment_log(
    entry_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("fieldreports.delete")),
) -> None:
    """Delete an equipment log entry."""
    from fastapi import HTTPException

    from app.modules.fieldreports.models import SiteEquipmentLog

    entry = await session.get(SiteEquipmentLog, entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="Equipment log entry not found")
    await session.delete(entry)
    await session.flush()
