"""‌⁠‍Finance API routes.

Endpoints:
    GET    /                    — List invoices with filters
    POST   /                    — Create invoice (auth required)
    GET    /invoices/export      — Export invoices as Excel
    GET    /invoices/{id}/br-pdf — Brazilian-styled invoice PDF (RPS layout)
    GET    /payments             — List payments
    POST   /payments             — Create payment (auth required)
    GET    /budgets              — List budgets
    POST   /budgets              — Create budget (auth required)
    PATCH  /budgets/{id}         — Update budget (auth required)
    POST   /budgets/import/file  — Import budgets from Excel/CSV (auth required)
    GET    /budgets/export       — Export budgets as Excel
    GET    /evm                  — List EVM snapshots
    POST   /evm/snapshot         — Create EVM snapshot (auth required)
    GET    /{id}                — Get single invoice
    PATCH  /{id}                — Update invoice (auth required)
    POST   /{id}/approve        — Approve invoice (auth required)
    POST   /{id}/pay            — Mark invoice as paid (auth required)

NOTE: Fixed-path routes (/payments, /budgets, /evm, /invoices/export) are
registered BEFORE the parametric /{invoice_id} route so that FastAPI does not
try to parse those path segments as UUIDs.
"""

import csv
import io
import logging
import uuid
from collections.abc import Iterable
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.file_signature import (
    SIGNATURE_BYTES_REQUIRED,
    FileSignatureMismatch,
)
from app.core.file_signature import (
    require as require_signature,
)
from app.core.rate_limiter import approval_limiter
from app.core.upload_guards import reject_if_xlsx_bomb
from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.contacts.models import Contact
from app.modules.finance.models import EVMSnapshot, Invoice, Payment, ProjectBudget
from app.modules.finance.schemas import (
    BudgetCreate,
    BudgetListResponse,
    BudgetResponse,
    BudgetUpdate,
    EVMListResponse,
    EVMSnapshotCreate,
    EVMSnapshotResponse,
    InvoiceCreate,
    InvoiceListResponse,
    InvoiceResponse,
    InvoiceUpdate,
    PaymentCreate,
    PaymentListResponse,
    PaymentResponse,
)
from app.modules.finance.service import FinanceService

router = APIRouter(tags=["finance"])
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> FinanceService:
    return FinanceService(session)


# ── Counterparty enrichment ─────────────────────────────────────────────────


def _contact_display_name(c: Contact) -> str:
    """‌⁠‍Return the human-readable contact label (company > "first last" > email)."""
    if c.company_name:
        return c.company_name
    full = f"{c.first_name or ''} {c.last_name or ''}".strip()
    return full or c.email or ""


async def _fetch_counterparty_names(session: AsyncSession, contact_ids: Iterable[str | None]) -> dict[str, str]:
    """‌⁠‍Resolve Invoice.contact_id → display name in one round trip."""
    ids = {cid for cid in contact_ids if cid}
    if not ids:
        return {}
    rows = (await session.execute(select(Contact).where(Contact.id.in_(ids)))).scalars().all()
    return {str(c.id): _contact_display_name(c) for c in rows}


def _invoice_to_response(invoice: Invoice, names: dict[str, str]) -> InvoiceResponse:
    resp = InvoiceResponse.model_validate(invoice)
    if invoice.contact_id:
        resp.counterparty_name = names.get(invoice.contact_id)
    return resp


# ── IDOR protection helpers ─────────────────────────────────────────────────


async def _require_project_access(
    session: AsyncSession,
    project_id: uuid.UUID | None,
    user_id: str | None,
) -> None:
    """Verify the current user owns (or is admin on) the referenced project.

    Central choke-point for project-scoped finance endpoints — must be called
    before reading or writing invoices/budgets/payments/EVM snapshots that
    belong to a specific project. Mirrors the pattern used by
    ``erp_chat.tools._require_project_access`` and the shared
    :func:`app.dependencies.verify_project_access`.

    R7 hardening (2026-05-24): cross-tenant fetches now answer **404**
    rather than 403. A 403 leaks the existence of project UUIDs the
    caller is not allowed to see — the global R7 standard (and the
    shared ``verify_project_access`` helper) returns 404 on both
    "missing" and "access denied". Bringing finance in line closes the
    enumeration sidechannel for project IDs.

    A ``None`` ``project_id`` is treated as a no-op (dashboard/list
    fall-throughs that legitimately aggregate across the user's own
    projects scope it themselves in the service layer).
    """
    if project_id is None:
        return
    if user_id is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required",
        )

    try:
        from app.modules.projects.repository import ProjectRepository
        from app.modules.users.repository import UserRepository

        proj_repo = ProjectRepository(session)
        project = await proj_repo.get_by_id(project_id)
        if project is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Project not found",
            )

        # Admin bypass
        try:
            user_repo = UserRepository(session)
            user = await user_repo.get_by_id(user_id)
            if user is not None and getattr(user, "role", "") == "admin":
                return
        except Exception:  # noqa: BLE001 — best-effort admin check
            pass

        if str(getattr(project, "owner_id", "")) != str(user_id):
            # R7: 404 not 403 — never confirm a project UUID exists
            # for callers that don't own it. Mirrors the response shape
            # of the "project missing" branch above.
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Project not found",
            )
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.warning("Finance project access check failed for %s: %s", project_id, exc)
        # Generic auth failure stays as 404 too — anything else would
        # again let the caller distinguish "exists but I lack access"
        # from "doesn't exist".
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )


async def _require_invoice_access(
    session: AsyncSession,
    invoice_id: uuid.UUID,
    user_id: str | None,
) -> Invoice:
    """Load an invoice and verify the caller has access to its parent project."""
    invoice = await session.get(Invoice, invoice_id)
    if invoice is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Invoice {invoice_id} not found",
        )
    await _require_project_access(session, invoice.project_id, user_id)
    return invoice


async def _require_budget_access(
    session: AsyncSession,
    budget_id: uuid.UUID,
    user_id: str | None,
) -> ProjectBudget:
    """Load a budget and verify the caller has access to its parent project."""
    budget = await session.get(ProjectBudget, budget_id)
    if budget is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Budget {budget_id} not found",
        )
    await _require_project_access(session, budget.project_id, user_id)
    return budget


async def _require_payment_access(
    session: AsyncSession,
    payment_id: uuid.UUID,
    user_id: str | None,
) -> Payment:
    """Load a payment and verify caller has access via its parent invoice."""
    payment = await session.get(Payment, payment_id)
    if payment is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Payment {payment_id} not found",
        )
    await _require_invoice_access(session, payment.invoice_id, user_id)
    return payment


async def _require_evm_access(
    session: AsyncSession,
    snapshot_id: uuid.UUID,
    user_id: str | None,
) -> EVMSnapshot:
    """Load an EVM snapshot and verify caller has access to its project."""
    snapshot = await session.get(EVMSnapshot, snapshot_id)
    if snapshot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"EVM snapshot {snapshot_id} not found",
        )
    await _require_project_access(session, snapshot.project_id, user_id)
    return snapshot


# ── Invoices (list / create) ───────────────────────────────────────────────


@router.get(
    "/",
    response_model=InvoiceListResponse,
    summary="List invoices",
    description="Retrieve a paginated list of invoices with optional filters by project, "
    "direction (payable/receivable), and status.",
)
async def list_invoices(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    direction: str | None = Query(default=None),
    status: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceListResponse:
    """List invoices with optional filters."""
    await _require_project_access(session, project_id, user_id)
    items, total = await service.list_invoices(
        project_id=project_id,
        direction=direction,
        invoice_status=status,
        offset=offset,
        limit=limit,
    )
    names = await _fetch_counterparty_names(session, (i.contact_id for i in items))
    return InvoiceListResponse(
        items=[_invoice_to_response(i, names) for i in items],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.post(
    "/",
    response_model=InvoiceResponse,
    status_code=201,
    summary="Create invoice",
    description="Create a new invoice with optional line items. Set invoice_direction "
    "to 'payable' (vendor invoices) or 'receivable' (client invoices).",
)
async def create_invoice(
    data: InvoiceCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Create a new invoice."""
    await _require_project_access(session, data.project_id, user_id)
    invoice = await service.create_invoice(data, user_id=user_id)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


# ── /invoices alias collection (BUG-API12) ─────────────────────────────────
#
# Frontend and external clients expect ``/api/v1/finance/invoices`` (without
# trailing slash) and ``/api/v1/finance/invoices/`` to behave like the root
# ``/``. Without an explicit alias FastAPI matches the ``/{invoice_id}``
# parametric route below and returns 422 on the literal string "invoices"
# (UUID parse error). Declared BEFORE ``/{invoice_id}`` so the static path
# wins.


@router.get(
    "/invoices/",
    response_model=InvoiceListResponse,
    summary="List invoices (alias of GET /)",
    description="Alias of ``GET /api/v1/finance/`` — returns the same paginated "
    "list. Provided so that clients hitting ``/api/v1/finance/invoices`` get a "
    "sensible 200 instead of a UUID-parse 422.",
)
async def list_invoices_alias(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    direction: str | None = Query(default=None),
    status: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceListResponse:
    """List invoices (alias of ``GET /``).

    Replicates the behaviour of ``list_invoices`` inline (rather than calling
    it through Python) so that FastAPI dependency wiring resolves cleanly.
    """
    await _require_project_access(session, project_id, user_id)
    items, total = await service.list_invoices(
        project_id=project_id,
        direction=direction,
        invoice_status=status,
        offset=offset,
        limit=limit,
    )
    names = await _fetch_counterparty_names(session, (i.contact_id for i in items))
    return InvoiceListResponse(
        items=[_invoice_to_response(i, names) for i in items],
        total=total,
        offset=offset,
        limit=limit,
    )


# ── Export invoices as Excel ────────────────────────────────────────────────


@router.get(
    "/invoices/export/",
    summary="Export invoices as Excel",
    description="Download invoices for a project as an Excel (.xlsx) file. "
    "Optionally filter by direction (payable/receivable).",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_invoices(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    direction: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
) -> StreamingResponse:
    """Export invoices for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    await _require_project_access(session, project_id, _user_id)

    stmt = select(Invoice).where(Invoice.project_id == project_id)
    if direction:
        stmt = stmt.where(Invoice.invoice_direction == direction)
    stmt = stmt.limit(50000)

    result = await session.execute(stmt)
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice #",
        "Direction",
        "Date",
        "Due Date",
        "Vendor/Client",
        "Subtotal",
        "Tax",
        "Total",
        "Status",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    from decimal import Decimal, InvalidOperation

    def _safe_decimal(raw: Any) -> Decimal:
        """Coerce DB string to Decimal preserving precision; NaN/Inf → 0.

        openpyxl accepts Decimal natively and stores as number without the
        float-precision loss ``float()`` would introduce on large currency
        values (BUG-069/070).
        """
        if raw is None or raw == "":
            return Decimal("0")
        try:
            d = Decimal(str(raw).strip())
        except (InvalidOperation, ValueError, TypeError):
            return Decimal("0")
        return d if d.is_finite() else Decimal("0")

    for row_idx, inv in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=inv.invoice_number)
        ws.cell(row=row_idx, column=2, value=inv.invoice_direction)
        ws.cell(row=row_idx, column=3, value=inv.invoice_date)
        ws.cell(row=row_idx, column=4, value=inv.due_date)
        ws.cell(row=row_idx, column=5, value=inv.contact_id or "")
        ws.cell(row=row_idx, column=6, value=_safe_decimal(inv.amount_subtotal))
        ws.cell(row=row_idx, column=7, value=_safe_decimal(inv.tax_amount))
        ws.cell(row=row_idx, column=8, value=_safe_decimal(inv.amount_total))
        ws.cell(row=row_idx, column=9, value=inv.status)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoices_export.xlsx"'},
    )


# ── Brazilian-styled invoice PDF (Tier-1 — pre-NF-e bridge) ─────────────────
#
# Path lives under ``/invoices/{invoice_id}/br-pdf/`` so FastAPI's static
# prefix ``/invoices/`` wins over the bare ``/{invoice_id}`` parametric
# route. See ``br_invoice_pdf.py`` for the rendering logic and the
# disclaimer text explaining why this PDF is NOT a fiscal document
# (NF-e / NFS-e SEFAZ integration is Tier-2 — see
# ``__brazil_tier2_followups.md``).


@router.get(
    "/invoices/{invoice_id}/br-pdf/",
    summary="Export invoice as Brazil-styled PDF (RPS layout)",
    description=(
        "Render the invoice as a one-page PDF in the Brazilian RPS "
        "(Recibo Provisório de Serviços) layout, with CNPJ / IE / Razão "
        "Social / código de serviço / retenções fields. NOT a fiscal "
        "document — for full NF-e / NFS-e SEFAZ output see Tier-2 roadmap."
    ),
    response_description="application/pdf stream",
)
async def export_invoice_br_pdf(
    invoice_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> StreamingResponse:
    """Render a Brazilian-styled invoice PDF and stream it back."""
    from app.modules.finance.br_invoice_pdf import render_br_invoice_pdf

    invoice = await _require_invoice_access(session, invoice_id, user_id)
    fresh = await service.get_invoice(invoice_id)

    # Project context (best-effort — never block the PDF on project lookup)
    project_dict: dict[str, Any] = {}
    try:
        from app.modules.projects.repository import ProjectRepository

        proj = await ProjectRepository(session).get_by_id(fresh.project_id)
        if proj is not None:
            project_dict = {
                "name": getattr(proj, "name", "") or "",
                "code": getattr(proj, "code", "") or "",
            }
    except Exception:  # noqa: BLE001 — header is decorative
        logger.debug("BR invoice PDF: project lookup failed", exc_info=True)

    invoice_dict: dict[str, Any] = {
        "invoice_number": fresh.invoice_number,
        "invoice_direction": fresh.invoice_direction,
        "invoice_date": fresh.invoice_date,
        "due_date": fresh.due_date,
        "amount_subtotal": fresh.amount_subtotal,
        "tax_amount": fresh.tax_amount,
        "retention_amount": fresh.retention_amount,
        "amount_total": fresh.amount_total,
        "notes": fresh.notes,
        "metadata": dict(fresh.metadata_ or {}),
    }
    line_items: list[dict[str, Any]] = [
        {
            "description": li.description,
            "unit": li.unit,
            "quantity": li.quantity,
            "unit_rate": li.unit_rate,
            "amount": li.amount,
        }
        for li in (fresh.line_items or [])
    ]

    pdf_bytes = render_br_invoice_pdf(
        invoice=invoice_dict,
        line_items=line_items,
        project=project_dict or None,
    )

    # Sanitise invoice_number before embedding in a quoted Content-Disposition
    # header.  invoice_number is a user-controlled DB value — it can contain
    # characters that would break the RFC 6266 quoted-string or inject
    # additional headers (CRLF injection).  Strip every character that is not
    # ASCII printable, remove double-quotes (which terminate the quoted-string
    # token) and forward-slashes (already done historically), and cap length.
    _raw_num = (invoice.invoice_number or "invoice")
    _safe_num = (
        _raw_num
        .encode("ascii", errors="replace")  # non-ASCII → b'?'
        .decode("ascii")
        .replace("\r", "")
        .replace("\n", "")
        .replace('"', "'")
        .replace("/", "-")
        .strip()
    )[:80] or "invoice"
    filename = f"RPS_{_safe_num}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Payments (MUST be before /{invoice_id}) ─────────────────────────────────


@router.get(
    "/payments/",
    response_model=PaymentListResponse,
    summary="List payments",
    description="Retrieve a paginated list of payments, optionally filtered by invoice.",
)
async def list_payments(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    invoice_id: uuid.UUID | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> PaymentListResponse:
    """List payments with optional invoice filter."""
    if invoice_id is not None:
        await _require_invoice_access(session, invoice_id, user_id)
    items, total = await service.list_payments(invoice_id=invoice_id, limit=limit, offset=offset)
    return PaymentListResponse(
        items=[PaymentResponse.model_validate(p) for p in items],
        total=total,
    )


@router.post(
    "/payments/",
    response_model=PaymentResponse,
    status_code=201,
    summary="Create payment",
    description="Record a payment against an invoice. Updates the invoice's paid amount. "
    "MANAGER-only — recording a payment is a binding ledger entry, not a CRUD edit.",
)
async def create_payment(
    data: PaymentCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.record_payment")),
    service: FinanceService = Depends(_get_service),
) -> PaymentResponse:
    """Record a payment against an invoice.

    R7 (2026-05-24): pinned to ``finance.record_payment`` (MANAGER+).
    Recording a payment row is a financial commitment that affects the
    invoice's paid/outstanding state and downstream budget actuals — it
    cannot remain an EDITOR-level action.
    """
    await _require_invoice_access(session, data.invoice_id, user_id)
    payment = await service.create_payment(data, actor_id=str(user_id) if user_id else None)
    return PaymentResponse.model_validate(payment)


# ── Budgets (MUST be before /{invoice_id}) ──────────────────────────────────


@router.get(
    "/budgets/",
    response_model=BudgetListResponse,
    summary="List budgets",
    description="Retrieve project budget lines with optional filters by project and cost category.",
)
async def list_budgets(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    category: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> BudgetListResponse:
    """List project budgets."""
    await _require_project_access(session, project_id, user_id)
    items, total = await service.list_budgets(project_id=project_id, category=category)
    return BudgetListResponse(
        items=[BudgetResponse.model_validate(b) for b in items],
        total=total,
    )


@router.post(
    "/budgets/",
    response_model=BudgetResponse,
    status_code=201,
    summary="Create budget line",
    description="Create a project budget line for a specific WBS element and cost category.",
)
async def create_budget(
    data: BudgetCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> BudgetResponse:
    """Create a project budget line."""
    await _require_project_access(session, data.project_id, user_id)
    budget = await service.create_budget(data)
    return BudgetResponse.model_validate(budget)


# ── Budget import (CSV / Excel) ─────────────────────────────────────────────

_BUDGET_COLUMN_ALIASES: dict[str, list[str]] = {
    "wbs_id": [
        "wbs_id",
        "wbs code",
        "wbs",
        "code",
        "wbs_code",
    ],
    "category": [
        "category",
        "cost category",
        "kategorie",
        "type",
    ],
    "original_budget": [
        "original_budget",
        "original budget",
        "original",
        "budget",
        "amount",
    ],
    "notes": [
        "notes",
        "note",
        "remarks",
        "bemerkung",
    ],
}

_ALLOWED_BUDGET_CATEGORIES = {
    "labor",
    "material",
    "equipment",
    "subcontractor",
    "overhead",
    "contingency",
    "other",
}


def _match_budget_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map."""
    normalised = header.strip().lower()
    for canonical, aliases in _BUDGET_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _safe_decimal_str(value: Any, default: str = "0") -> str:
    """Parse a value to a decimal string, returning default on failure."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if not text:
        return default
    # Handle European-style numbers: "1.234,56" -> "1234.56"
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        float(text)  # validate
        return text
    except (ValueError, TypeError):
        return default


def _parse_budget_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for budget import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_budget_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_budget_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for budget import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_budget_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


@router.post(
    "/budgets/import/file/",
    summary="Import budgets from file",
    description="Upload an Excel (.xlsx) or CSV (.csv) file to bulk-import budget lines. "
    "Column headers are auto-detected using flexible aliases (EN/DE). "
    "Returns a summary with imported, skipped, and error counts per row.",
)
async def import_budgets_file(
    _user_id: CurrentUserId,
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> dict[str, Any]:
    """Import project budgets from an Excel or CSV file upload.

    Expected columns:
    - **WBS Code** -- work breakdown structure code
    - **Category** -- budget category (labor, material, equipment, etc.)
    - **Original Budget** -- original budget amount
    - **Notes** -- optional notes

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    await _require_project_access(session, project_id, _user_id)

    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # R7 (2026-05-24): magic-byte gate. The filename-extension check above
    # is necessary but trivially bypassable — an attacker who renames a
    # PE/ELF/script to ``payload.xlsx`` would pass the extension test and
    # land in the parser. ``require_signature`` rejects anything that
    # isn't a ZIP-container (xlsx/xls OLE) or plain-text/CSV (which has
    # no signature and surfaces as ``None``).
    #
    # CSV genuinely has no magic bytes so it returns ``None`` from
    # ``detect`` — only require the signature check for the spreadsheet
    # branches; plain CSV falls through to the parser as-is.
    fname_low = filename
    if fname_low.endswith((".xlsx", ".xls")):
        head = content[:SIGNATURE_BYTES_REQUIRED]
        try:
            require_signature(
                head,
                # xlsx → ZIP container; legacy .xls → OLE compound document.
                frozenset({"zip", "ole"}),
                filename=file.filename,
            )
        except FileSignatureMismatch as exc:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail=str(exc),
            )

    # Zip-bomb guard: reject .xlsx whose uncompressed sheets exceed 50 MB.
    reject_if_xlsx_bomb(content)

    # Parse rows based on file type
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_budget_rows_from_excel(content)
        else:
            rows = _parse_budget_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing budget import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # Convert rows to BudgetCreate objects and import
    imported_count = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            wbs_id = str(row.get("wbs_id", "")).strip() or None

            # Parse category
            category = str(row.get("category", "")).strip().lower() or None
            if category and category not in _ALLOWED_BUDGET_CATEGORIES:
                errors.append(
                    {
                        "row": row_idx,
                        "error": (
                            f"Invalid category: '{category}'. Allowed: {', '.join(sorted(_ALLOWED_BUDGET_CATEGORIES))}"
                        ),
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue

            # Parse amount
            original_budget = _safe_decimal_str(row.get("original_budget"))

            # Validate amount is a valid number
            try:
                float(original_budget)
            except (ValueError, TypeError):
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Invalid budget amount: {row.get('original_budget')}",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue

            # Skip rows with no data
            if not wbs_id and not category and original_budget == "0":
                skipped += 1
                continue

            data = BudgetCreate(
                project_id=project_id,
                wbs_id=wbs_id,
                category=category,
                original_budget=original_budget,
                revised_budget=original_budget,  # default revised = original
            )
            await service.create_budget(data)
            imported_count += 1

        except Exception as exc:
            errors.append(
                {
                    "row": row_idx,
                    "error": str(exc),
                    "data": {k: str(v)[:100] for k, v in row.items()},
                }
            )
            logger.warning("Budget import error at row %d: %s", row_idx, exc)

    logger.info(
        "Budget file import complete: imported=%d, skipped=%d, errors=%d",
        imported_count,
        skipped,
        len(errors),
    )

    return {
        "imported": imported_count,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Export budgets as Excel ──────────────────────────────────────────────────


@router.get(
    "/budgets/export/",
    summary="Export budgets as Excel",
    description="Download budgets for a project as an Excel (.xlsx) file with "
    "original, revised, committed, actual, forecast, and variance columns.",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_budgets(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    _perm: None = Depends(RequirePermission("finance.read")),
) -> StreamingResponse:
    """Export budgets for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    await _require_project_access(session, project_id, _user_id)

    result = await session.execute(select(ProjectBudget).where(ProjectBudget.project_id == project_id).limit(50000))
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Budgets"

    headers = [
        "WBS",
        "Category",
        "Original",
        "Revised",
        "Committed",
        "Actual",
        "Forecast",
        "Variance",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, b in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=b.wbs_id or "")
        ws.cell(row=row_idx, column=2, value=b.category or "")
        # BUG-069: use Decimal (not float) so large construction-budget values
        # (e.g. 123456789.99) don't suffer IEEE-754 rounding when Excel reads
        # them back — openpyxl stores Decimal natively as a NUMERIC cell.
        from decimal import Decimal as _Dec, InvalidOperation as _IOp

        def _bd(raw: Any) -> _Dec:
            if raw is None or raw == "":
                return _Dec("0")
            try:
                d = _Dec(str(raw).strip())
            except (_IOp, ValueError, TypeError):
                return _Dec("0")
            return d if d.is_finite() else _Dec("0")

        original = _bd(b.original_budget)
        revised = _bd(b.revised_budget)
        committed = _bd(b.committed)
        actual = _bd(b.actual)
        forecast = _bd(b.forecast_final)
        variance = revised - actual

        ws.cell(row=row_idx, column=3, value=original)
        ws.cell(row=row_idx, column=4, value=revised)
        ws.cell(row=row_idx, column=5, value=committed)
        ws.cell(row=row_idx, column=6, value=actual)
        ws.cell(row=row_idx, column=7, value=forecast)
        ws.cell(row=row_idx, column=8, value=variance)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="budgets_export.xlsx"'},
    )


@router.patch(
    "/budgets/{budget_id}",
    response_model=BudgetResponse,
    summary="Update budget line",
    description="Partially update a budget line. Only provided fields are modified.",
)
async def update_budget(
    budget_id: uuid.UUID,
    data: BudgetUpdate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.update")),
    service: FinanceService = Depends(_get_service),
) -> BudgetResponse:
    """Update a budget line."""
    await _require_budget_access(session, budget_id, user_id)
    budget = await service.update_budget(budget_id, data)
    return BudgetResponse.model_validate(budget)


# ── EVM (MUST be before /{invoice_id}) ──────────────────────────────────────


@router.get(
    "/evm/",
    response_model=EVMListResponse,
    summary="List EVM snapshots",
    description="List Earned Value Management snapshots for a project. "
    "Each snapshot captures PV, EV, AC, SPI, CPI, and EAC at a point in time.",
)
async def list_evm_snapshots(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> EVMListResponse:
    """List EVM snapshots for a project."""
    await _require_project_access(session, project_id, user_id)
    items, total = await service.list_evm_snapshots(project_id=project_id)
    return EVMListResponse(
        items=[EVMSnapshotResponse.model_validate(s) for s in items],
        total=total,
    )


@router.post(
    "/evm/snapshot/",
    response_model=EVMSnapshotResponse,
    status_code=201,
    summary="Create EVM snapshot",
    description="Capture a new Earned Value Management snapshot for a project. "
    "Records planned value (PV), earned value (EV), actual cost (AC), and derived indices.",
)
async def create_evm_snapshot(
    data: EVMSnapshotCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> EVMSnapshotResponse:
    """Create an EVM snapshot."""
    await _require_project_access(session, data.project_id, user_id)
    snapshot = await service.create_evm_snapshot(data)
    return EVMSnapshotResponse.model_validate(snapshot)


# ── Finance Dashboard ─────────────────────────────────────────────────────────


@router.get(
    "/dashboard/",
    summary="Get finance dashboard",
    description="Aggregated finance KPIs: payable, receivable, overdue totals, "
    "budget utilisation, cash flow overview, and budget warning level "
    "(normal / caution at 80%+ / critical at 95%+). "
    "Optionally scope to a single project via project_id query parameter.",
)
async def finance_dashboard(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> dict:
    """Aggregated finance KPIs: payable, receivable, overdue, budget, cash flow.

    Optionally scope to a single project via ``project_id`` query parameter.
    Returns budget warning level ("normal", "caution" at 80%+, "critical" at 95%+).
    """
    await _require_project_access(session, project_id, user_id)
    return await service.get_dashboard(project_id=project_id)


# ── Invoice by ID (parametric routes LAST) ──────────────────────────────────


@router.get(
    "/{invoice_id}",
    response_model=InvoiceResponse,
    summary="Get invoice",
    description="Retrieve a single invoice by its UUID, including line items and payment history.",
)
async def get_invoice(
    invoice_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Get a single invoice by ID."""
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.get_invoice(invoice_id)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.patch(
    "/{invoice_id}",
    response_model=InvoiceResponse,
    summary="Update invoice",
    description="Partially update an invoice. Only provided fields are modified.",
)
async def update_invoice(
    invoice_id: uuid.UUID,
    data: InvoiceUpdate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.update")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Update an invoice."""
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.update_invoice(invoice_id, data)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.post(
    "/{invoice_id}/approve/",
    response_model=InvoiceResponse,
    summary="Approve invoice",
    description="Transition an invoice to 'sent' (legacy alias 'approved') status. "
    "Only invoices in 'draft' or 'pending' status can be approved. "
    "MANAGER-only — invoice approval is the financial-commitment gate.",
)
async def approve_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.approve")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Approve an invoice.

    R7 (2026-05-24): pinned to ``finance.approve`` (MANAGER+).
    Previously this route used ``finance.update`` (EDITOR), which let
    any estimator commit a payable to the project.
    """
    allowed, _ = approval_limiter.is_allowed(str(user_id))
    if not allowed:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Rate limit exceeded. Try again later.")
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.approve_invoice(
        invoice_id,
        actor_id=str(user_id) if user_id else None,
    )
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.post(
    "/{invoice_id}/pay/",
    response_model=InvoiceResponse,
    summary="Mark invoice as paid",
    description="Transition an invoice to 'paid' status. Records the payment date. "
    "MANAGER-only — marking an invoice paid is a binding ledger action. "
    "Idempotent: a second call against an already-paid invoice returns 400, "
    "not a duplicate ledger write.",
)
async def pay_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.pay")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Mark invoice as paid.

    R7 (2026-05-24): pinned to ``finance.pay`` (MANAGER+). The FSM
    allowlist (``_INVOICE_STATUS_TRANSITIONS``) ensures a second click
    against an already-paid invoice cannot re-trigger budget-actual
    recompute — it 400s on the disallowed ``paid -> paid`` transition
    (idempotency by allowlist).
    """
    allowed, _ = approval_limiter.is_allowed(str(user_id))
    if not allowed:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Rate limit exceeded. Try again later.")
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.pay_invoice(
        invoice_id,
        actor_id=str(user_id) if user_id else None,
    )
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)
