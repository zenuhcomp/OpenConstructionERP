"""‌⁠‍Takeoff HTTP endpoints.

Routes:
    GET    /converters                          — list CAD/BIM converter status
    POST   /converters/{converter_id}/install   — download & install a converter
    POST   /converters/{converter_id}/uninstall — remove an installed converter
    POST   /documents/upload                    — upload a PDF for takeoff
    GET    /documents/                          — list uploaded documents
    GET    /documents/{doc_id}                  — get single document
    POST   /documents/{doc_id}/extract-tables   — extract tables from document
    POST   /documents/{doc_id}/analyze          — AI analysis of extracted text
    GET    /documents/{doc_id}/download          — download the stored PDF file
    DELETE /documents/{doc_id}                  — delete a document

    POST   /measurements                       — create measurement
    GET    /measurements                        — list measurements (filtered)
    GET    /measurements/summary                — stats by group/type
    GET    /measurements/export                 — export measurements as CSV/JSON
    POST   /measurements/bulk                   — bulk create measurements
    GET    /measurements/{id}                   — get single measurement
    PATCH  /measurements/{id}                   — update measurement
    DELETE /measurements/{id}                   — delete measurement
    POST   /measurements/{id}/link-to-boq       — link measurement to BOQ position

    POST   /cad-group/create-boq               — create BOQ from grouped CAD QTO
    GET    /cad-group/export                    — export grouped QTO as Excel

    POST   /cad-data/describe                   — DataFrame-like describe of CAD session
    POST   /cad-data/value-counts               — value counts for a single column
    GET    /cad-data/elements                    — paginated element table with sort/filter
    POST   /cad-data/aggregate                   — group-by aggregation on CAD elements
    GET    /cad-data/missingness                 — per-column fill-rate + row completeness
"""

import logging
import random as _random
import threading
import time as _time
import uuid as _uuid
from datetime import UTC, datetime, timedelta
from pathlib import Path
from statistics import mean as _mean
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import delete, select

from app.core.csv_safety import neutralise_formula
from app.core.rate_limiter import upload_limiter
from app.dependencies import CurrentUserId, RequirePermission, SessionDep, verify_project_access
from app.modules.takeoff.manifest_verifier import (
    InstallNotSupported,
    InstallSHAMismatch,
    ManifestError,
    ManifestSignatureInvalid,
    fetch_manifest,
    maybe_warn_disabled,
    resolve_install,
    verify_downloaded_file,
)
from app.modules.takeoff.models import CadExtractionSession
from app.modules.takeoff.schemas import (
    LinkToBoqRequest,
    TakeoffMeasurementBulkCreate,
    TakeoffMeasurementCreate,
    TakeoffMeasurementResponse,
    TakeoffMeasurementSummary,
    TakeoffMeasurementUpdate,
)
from app.modules.takeoff.service import TakeoffService

logger = logging.getLogger(__name__)

# Surface the manifest-bypass warning at import time so it shows up in
# server logs alongside the rest of the boot banner — operators who
# left the escape hatch on by accident notice it on the next restart.
maybe_warn_disabled()

router = APIRouter(tags=["takeoff"])


# ── Converter status ─────────────────────────────────────────────────────


#
# Converter metadata. Sizes reflect what is actually committed in the
# `datadrivenconstruction/cad2data-Revit-IFC-DWG-DGN` repository under
# `DDC_WINDOWS_Converters/DDC_CONVERTER_{FORMAT}/` — typically ~30-50 MB
# per format (the small `*Exporter.exe` plus the bundled Qt6 DLLs).
# These are NOT GitHub Releases; the binaries live directly on the
# default branch and we install them by walking the directory tree
# via the Contents API and downloading each file from
# `raw.githubusercontent.com`.
_CONVERTER_META: list[dict[str, Any]] = [
    {
        "id": "dwg",
        "name": "DWG/DXF Converter",
        "description": "Import AutoCAD DWG and DXF files. Extracts geometry, layers, blocks, and properties into structured element tables for cost estimation.",
        "engine": "DDC Community",
        "extensions": [".dwg", ".dxf"],
        "exe": "DwgExporter.exe",
        "version": "1.0.0",
        "size_mb": 218.0,
    },
    {
        "id": "rvt",
        "name": "Revit (RVT) Parser",
        "description": "Native Revit file parser. Supports Revit 2015-2026. Extracts families, parameters, quantities, and spatial structure without an Autodesk license. Bundles format readers for every Revit version 2011-2026, which is why the download is large.",
        "engine": "DDC Community",
        "extensions": [".rvt", ".rfa"],
        "exe": "RvtExporter.exe",
        "version": "1.0.0",
        "size_mb": 598.0,
    },
    {
        "id": "ifc",
        "name": "IFC Import",
        "description": "Import IFC 2x3 and IFC4 files. Maps IFC entities to structured element tables with full property set extraction.",
        "engine": "DDC Community",
        "extensions": [".ifc", ".ifczip"],
        "exe": "IfcExporter.exe",
        "version": "1.0.0",
        "size_mb": 241.0,
    },
    {
        "id": "dgn",
        "name": "DGN Converter",
        "description": "Import MicroStation DGN files. Extracts elements, levels, properties, and 3D geometry into structured tables.",
        "engine": "DDC Community",
        "extensions": [".dgn"],
        "exe": "DgnExporter.exe",
        "version": "1.0.0",
        "size_mb": 217.0,
    },
]


@router.get("/converters/")
async def list_converters(verify: bool = False) -> dict[str, Any]:
    """‌⁠‍Return the status of all known CAD/BIM converters.

    Scans standard install paths and returns which converters are found.
    No authentication required — this is a public status check.

    Args:
        verify: When ``true``, also runs a quick smoke test (~8 s timeout
            per installed converter) to confirm the binary actually
            loads. Result is cached for 5 minutes so repeated calls are
            cheap. The default is ``false`` so the page-load list call
            stays fast (<50 ms); the BIM page polls with ``verify=true``
            after install completes.
    """
    import asyncio

    from app.modules.boq.cad_import import find_converter, smoke_test_converter

    # Phase 1: cheap file-stat lookup for every converter (synchronous,
    # bounded by ~4 disk reads — sub-millisecond on a warm cache).
    paths: list[Path | None] = [find_converter(m["id"]) for m in _CONVERTER_META]

    # Phase 2: when ``verify=true``, run ALL installed-converter smoke
    # tests CONCURRENTLY with ``asyncio.gather`` so the wall-clock cost
    # is bounded by the slowest one (8 s timeout) instead of the sum of
    # all of them. Without this the request was up to 32 s for four
    # installed converters — long enough that React Query stayed in
    # ``isFetching`` and the user saw a stale "0/4 verified" view with
    # no health pills.
    health_results: list[dict[str, Any] | None] = [None] * len(_CONVERTER_META)
    if verify:
        smoke_tasks: list[Any] = []
        smoke_indices: list[int] = []
        for idx, (meta, path) in enumerate(zip(_CONVERTER_META, paths, strict=True)):
            if path is not None:
                smoke_tasks.append(
                    asyncio.to_thread(smoke_test_converter, meta["id"])
                )
                smoke_indices.append(idx)
        if smoke_tasks:
            results = await asyncio.gather(*smoke_tasks, return_exceptions=True)
            for slot, result in zip(smoke_indices, results, strict=True):
                if isinstance(result, BaseException):
                    health_results[slot] = {
                        "status": "unknown",
                        "message": f"Smoke test errored: {result}",
                        "suggested_actions": [],
                        "checked_at": 0.0,
                    }
                else:
                    health_results[slot] = result  # type: ignore[assignment]

    converters: list[dict[str, Any]] = []
    for idx, meta in enumerate(_CONVERTER_META):
        path = paths[idx]
        installed = path is not None
        entry: dict[str, Any] = {
            **meta,
            "installed": installed,
            "path": str(path) if path else None,
        }
        if not installed:
            entry["health"] = "not_installed"
            entry["health_message"] = ""
            entry["suggested_actions"] = ["install_converter"]
        elif verify and health_results[idx] is not None:
            h = health_results[idx]
            assert h is not None
            entry["health"] = h["status"]
            entry["health_message"] = h["message"]
            entry["suggested_actions"] = h["suggested_actions"]
        else:
            entry["health"] = "unknown"
            entry["health_message"] = ""
            entry["suggested_actions"] = []
        converters.append(entry)

    installed_count = sum(1 for c in converters if c["installed"])
    healthy_count = sum(1 for c in converters if c["health"] == "ok")
    return {
        "converters": converters,
        "installed_count": installed_count,
        "healthy_count": healthy_count,
        "total_count": len(converters),
    }


@router.post("/converters/{converter_id}/verify/")
async def verify_converter(converter_id: str) -> dict[str, Any]:
    """‌⁠‍Force-run the smoke test for one converter and return health.

    Bypasses the 5-minute cache. Used by the BIM page's "Re-check" button
    so the user can re-verify after manually fixing a broken install
    (e.g. installing VC++ Redistributable, unblocking files, or running
    the converter exe once as administrator).
    """
    import asyncio

    from app.modules.boq.cad_import import find_converter, smoke_test_converter

    if converter_id not in {m["id"] for m in _CONVERTER_META}:
        raise HTTPException(
            status_code=404,
            detail=f"Unknown converter id: {converter_id}",
        )

    path = find_converter(converter_id)
    if path is None:
        return {
            "converter_id": converter_id,
            "installed": False,
            "path": None,
            "health": "not_installed",
            "health_message": (
                f"The .{converter_id.upper()} converter is not installed. "
                f"Use the Install button on the BIM page to download it."
            ),
            "suggested_actions": ["install_converter"],
        }

    health = await asyncio.to_thread(
        smoke_test_converter, converter_id, True
    )
    return {
        "converter_id": converter_id,
        "installed": True,
        "path": str(path),
        "health": health["status"],
        "health_message": health["message"],
        "suggested_actions": health["suggested_actions"],
    }


# ── Converter install / uninstall ────────────────────────────────────────


#
# Source repository for DDC Community converters. The binaries are NOT
# published as GitHub Releases — they live committed on the default
# branch under `DDC_WINDOWS_Converters/DDC_CONVERTER_{FORMAT}/`. Linux
# users get separate `.deb` packages from the apt source maintained at
# `pkg.datadrivenconstruction.io` (handled separately below).
_DDC_REPO = "datadrivenconstruction/cad2data-Revit-IFC-DWG-DGN"
_DDC_BRANCH = "main"

# Per-format directory inside the repo for Windows binaries. Each
# directory contains the small `*Exporter.exe`, the matching
# `DDC_Community_*_converter.exe` GUI shell, the bundled Qt6 DLLs, and
# `platforms/`, `styles/`, `datadrivenlibs/` subfolders.
_WINDOWS_CONVERTER_DIRS: dict[str, str] = {
    "rvt": "DDC_WINDOWS_Converters/DDC_CONVERTER_REVIT",
    "ifc": "DDC_WINDOWS_Converters/DDC_CONVERTER_IFC",
    "dwg": "DDC_WINDOWS_Converters/DDC_CONVERTER_DWG",
    "dgn": "DDC_WINDOWS_Converters/DDC_CONVERTER_DGN",
}

# Linux apt package names. We don't auto-install these (would need
# `sudo` and an apt source rewrite of `/etc/apt/sources.list.d/`),
# but we surface them in the install endpoint's error response so
# the user can run the command themselves.
_LINUX_APT_PACKAGES: dict[str, str] = {
    "rvt": "ddc-rvtconverter",
    "ifc": "ddc-ifcconverter",
    "dwg": "ddc-dwgconverter",
    "dgn": "ddc-dgnconverter",
}

_CONVERTER_CACHE_DIR = Path.home() / ".openestimator" / "cache" / "converters"
_CONVERTER_INSTALL_DIR = Path.home() / ".openestimator" / "converters"

_META_BY_ID: dict[str, dict[str, Any]] = {m["id"]: m for m in _CONVERTER_META}

# ── Live install progress (in-memory, per-process) ───────────────────────────
# The Windows installer downloads ~30-175 files (RVT alone is 598 MB) inside
# a thread pool. Without a progress feed the UI shows only a spinner for
# 30-90 s and users think it hung. We publish per-file progress into this
# dict from the worker thread; the frontend polls /install-progress/ every
# 500 ms. Lock is required because the ThreadPoolExecutor workers write
# concurrently while the FastAPI handler may read at any moment.
_INSTALL_PROGRESS: dict[str, dict[str, Any]] = {}
_INSTALL_PROGRESS_LOCK = threading.Lock()


def _set_install_progress(converter_id: str, **fields: Any) -> None:
    """Threadsafe merge of progress fields for a converter install."""
    with _INSTALL_PROGRESS_LOCK:
        slot = _INSTALL_PROGRESS.setdefault(converter_id, {})
        slot.update(fields)


def _clear_install_progress(converter_id: str) -> None:
    with _INSTALL_PROGRESS_LOCK:
        _INSTALL_PROGRESS.pop(converter_id, None)


def _get_install_progress(converter_id: str) -> dict[str, Any] | None:
    with _INSTALL_PROGRESS_LOCK:
        slot = _INSTALL_PROGRESS.get(converter_id)
        return dict(slot) if slot else None

# ── Audit A2 / A11: converter download hardening ─────────────────────────────

# Allowed hosts for converter file downloads. We hard-code this against
# GitHub's raw.githubusercontent.com (where the Contents API redirects
# blob downloads) so an attacker who tampers with the API response —
# substituting download_url with an attacker-controlled CDN — can't
# trick the installer into fetching a poisoned exe. The Contents API
# itself returns absolute URLs but FastAPI never trusts user-supplied
# URLs, so this matters only if GitHub itself is compromised OR if
# an upstream MITM rewrites the JSON in transit (which TLS already
# prevents, but defence in depth).
_ALLOWED_DOWNLOAD_HOSTS = frozenset({
    "raw.githubusercontent.com",
    "github.com",
    "objects.githubusercontent.com",  # GitHub's blob CDN, used for >5 MB
})

# Hard size cap per file. The largest single file in the DDC converter
# repo today is the ~140 MB IfcExporter.exe; we add ~3x headroom for
# future versions and Teigha format readers. Anything above this
# threshold is almost certainly a substitution attack or a pathological
# upstream change — refuse rather than waste disk and download time.
_MAX_DOWNLOAD_BYTES = 512 * 1024 * 1024  # 512 MB

# Total cap per converter install. RVT today is ~600 MB; we set 1.5 GB
# so existing installs continue to work but a runaway listing can't
# eat all the disk.
_MAX_INSTALL_BYTES = 1536 * 1024 * 1024  # 1.5 GB


def _check_download_url_allowed(url: str) -> None:
    """Reject converter download URLs whose host isn't on the allow-list.

    Audit A2 — without this check, a tampered GitHub Contents API
    response (or a future bug that lets a malicious value leak into
    ``download_url``) could redirect the installer to an attacker-
    controlled CDN. Allow-listing the three GitHub hosts that
    legitimately serve blob downloads closes that vector.
    """
    from urllib.parse import urlparse

    parsed = urlparse(url)
    if parsed.scheme not in {"https", "http"}:
        raise RuntimeError(
            f"Refused to download {url!r} — non-HTTP(S) scheme"
        )
    host = (parsed.hostname or "").lower()
    if host not in _ALLOWED_DOWNLOAD_HOSTS:
        raise RuntimeError(
            f"Refused to download {url!r} — host {host!r} is not on "
            f"the converter allow-list {sorted(_ALLOWED_DOWNLOAD_HOSTS)}"
        )


def _github_list_directory(repo_path: str) -> list[dict[str, Any]]:
    """Recursively list every file in a GitHub repo directory.

    Walks the GitHub Contents API at
    ``https://api.github.com/repos/{_DDC_REPO}/contents/{repo_path}``
    and follows nested directories. Returns a flat list of file
    descriptors with at least ``path``, ``download_url``, and ``size``
    fields. Directories themselves are not included; only their files.

    Raises ``RuntimeError`` if the API call fails so the caller can
    surface a useful error to the user instead of silently installing
    nothing.
    """
    import json
    import urllib.error
    import urllib.request

    api_url = (
        f"https://api.github.com/repos/{_DDC_REPO}/contents/{repo_path}"
        f"?ref={_DDC_BRANCH}"
    )
    req = urllib.request.Request(
        api_url,
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": "OpenConstructionERP-converter-installer",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        raise RuntimeError(
            f"GitHub Contents API returned {exc.code} for {repo_path}: {exc.reason}"
        ) from exc
    except (urllib.error.URLError, TimeoutError) as exc:
        raise RuntimeError(
            f"Could not reach GitHub Contents API: {exc}"
        ) from exc

    if not isinstance(payload, list):
        raise RuntimeError(
            f"GitHub Contents API returned a non-list for {repo_path} — "
            f"is the path correct?"
        )

    files: list[dict[str, Any]] = []
    for item in payload:
        item_type = item.get("type")
        if item_type == "file":
            files.append(item)
        elif item_type == "dir":
            # Recurse into subdirectories. The Qt-based converters keep
            # their plugins under platforms/ / styles/ / datadrivenlibs/
            # so we MUST recurse — flat downloads would miss the DLLs
            # the .exe needs at runtime.
            files.extend(_github_list_directory(item["path"]))
    return files


def _resolve_target_path(
    repo_path: str,
    src_prefix: str,
    dest_root: Path,
    install_dir_resolved: Path,
) -> Path:
    """Compute the on-disk target for a file from the GitHub listing.

    Strips the converter's source-dir prefix so we mirror the inner
    tree (datadrivenlibs/, platforms/, styles/) into ``dest_root``.
    Rejects path-traversal attempts up-front so an attacker who
    compromised the upstream repo can not write outside the install
    directory.
    """
    if repo_path.startswith(src_prefix):
        rel = repo_path[len(src_prefix):]
    else:
        rel = Path(repo_path).name
    if Path(rel).is_absolute() or ".." in Path(rel).parts:
        raise RuntimeError(
            f"Refused to write to suspicious path {rel!r} from "
            f"GitHub Contents response (path-traversal attempt)"
        )
    target = (dest_root / rel).resolve()
    try:
        target.relative_to(install_dir_resolved)
    except ValueError as exc:
        raise RuntimeError(
            f"Refused to write {target} — escapes install directory"
        ) from exc
    return target


def _download_one_file(download_url: str, target: Path) -> int:
    """Download a single file. Returns bytes written. Used by the pool.

    Audit A2 / A9 / A11 — three hardenings:

    1. **Host allow-list** (A2): refuse any URL whose hostname isn't
       in ``_ALLOWED_DOWNLOAD_HOSTS``. Closes a poisoned-redirect
       attack where the Contents API response (or a MITM) substitutes
       ``download_url`` for an attacker-controlled CDN.

    2. **Size cap during stream** (A11): we used to call
       ``urlretrieve`` which has no size cap — a hostile (or
       runaway-grown) blob could fill the disk before failing. We
       now stream the body chunk-by-chunk and abort the moment we
       exceed ``_MAX_DOWNLOAD_BYTES``.

    3. **Symlink/TOCTOU guard** (A9): if a malicious file existed
       at ``target`` (e.g. created by another local user between
       ``_resolve_target_path`` and this call), ``urlretrieve``
       would happily follow the symlink and overwrite whatever it
       pointed at. We delete any existing symlink before opening
       the destination and use ``O_NOFOLLOW`` on POSIX so a
       race-replaced symlink is rejected at open() time.
    """
    import os
    import urllib.request

    _check_download_url_allowed(download_url)

    target.parent.mkdir(parents=True, exist_ok=True)

    # A9 — drop any pre-existing symlink at the target. ``is_symlink``
    # uses lstat so we don't follow the link to check its target.
    if target.is_symlink():
        target.unlink()

    req = urllib.request.Request(
        download_url,
        headers={"User-Agent": "OpenConstructionERP-converter-installer"},
    )
    bytes_written = 0
    # POSIX gets O_NOFOLLOW. Windows doesn't expose it but `is_symlink`
    # + the unlink-first pass already covers the common case there
    # (Windows requires elevated rights to create symlinks).
    open_flags = os.O_WRONLY | os.O_CREAT | os.O_TRUNC
    # Windows: the MS C runtime defaults an os.open() fd to *text* mode,
    # so os.write() would translate every 0x0A into 0x0D 0x0A and shred
    # the binary (shifted PE header → WinError 216 on launch). O_BINARY
    # only exists on Windows; this is a no-op on POSIX.
    if hasattr(os, "O_BINARY"):
        open_flags |= os.O_BINARY  # type: ignore[attr-defined]
    if hasattr(os, "O_NOFOLLOW"):
        open_flags |= os.O_NOFOLLOW  # type: ignore[attr-defined]
    try:
        fd = os.open(str(target), open_flags, 0o644)
    except OSError as exc:
        # ELOOP on Linux when O_NOFOLLOW hits a symlink — surface as
        # a clear refusal rather than a generic OSError.
        raise RuntimeError(
            f"Refused to open {target} for writing: {exc}"
        ) from exc

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            # Servers may advertise size up-front; reject grossly large
            # files before reading a single byte.
            content_length = resp.headers.get("Content-Length")
            if content_length is not None:
                try:
                    declared = int(content_length)
                    if declared > _MAX_DOWNLOAD_BYTES:
                        raise RuntimeError(
                            f"Refused to download {download_url!r} — declared size "
                            f"{declared} bytes exceeds the per-file cap of "
                            f"{_MAX_DOWNLOAD_BYTES} bytes"
                        )
                except (ValueError, TypeError):
                    pass  # Bogus header — fall through to streaming cap.

            while True:
                chunk = resp.read(65536)
                if not chunk:
                    break
                bytes_written += len(chunk)
                if bytes_written > _MAX_DOWNLOAD_BYTES:
                    raise RuntimeError(
                        f"Aborted download of {download_url!r} — body exceeded "
                        f"the per-file cap of {_MAX_DOWNLOAD_BYTES} bytes "
                        f"(possible substitution attack)"
                    )
                os.write(fd, chunk)
    finally:
        os.close(fd)
    return bytes_written


def _download_converter_files_windows(converter_id: str) -> Path:
    """Download every file of a Windows converter into the install dir.

    Mirrors the per-format directory tree from
    `DDC_WINDOWS_Converters/DDC_CONVERTER_{FORMAT}/` into
    `~/.openestimator/converters/{format}_windows/` so multiple
    converters can coexist without overwriting each other's Qt DLLs
    or format readers.

    The RVT converter alone is ~600 MB across ~175 files (most of
    which are the bundled Teigha format readers for every Revit
    version 2011-2026), so we use a small ThreadPoolExecutor to
    download files in parallel — sequential `urlretrieve` calls
    against `raw.githubusercontent.com` would take 5+ minutes from
    a typical home connection. Eight workers gets RVT down to
    ~30-60 seconds without tripping any GitHub abuse detection.

    Returns the path to the installed `*Exporter.exe`. Raises
    ``RuntimeError`` if the listing fails, any download fails, or
    the expected exe is missing after download.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _set_install_progress(
        converter_id,
        stage="listing",
        current=0,
        total=0,
        bytes_done=0,
        file=None,
        started_at=_time.time(),
    )
    src_dir = _WINDOWS_CONVERTER_DIRS[converter_id]
    files = _github_list_directory(src_dir)
    if not files:
        _clear_install_progress(converter_id)
        raise RuntimeError(
            f"GitHub directory {src_dir!r} contains no files — "
            f"the DDC converter repo layout may have changed."
        )

    # Per-format install root keeps Qt DLLs and Teigha readers from
    # clobbering each other when multiple formats are installed.
    dest_root = (_CONVERTER_INSTALL_DIR / f"{converter_id}_windows").resolve()
    dest_root.mkdir(parents=True, exist_ok=True)
    install_dir_resolved = _CONVERTER_INSTALL_DIR.resolve()
    src_prefix = src_dir.rstrip("/") + "/"

    # Pre-resolve every target path so the security checks happen
    # BEFORE any network IO — we want to fail fast on a hostile
    # listing rather than partway through a 600 MB download.
    download_jobs: list[tuple[str, Path]] = []
    for entry in files:
        download_url = entry.get("download_url")
        if not download_url:
            continue  # submodules / symlinks — skip
        target = _resolve_target_path(
            entry["path"], src_prefix, dest_root, install_dir_resolved,
        )
        download_jobs.append((download_url, target))

    if not download_jobs:
        _clear_install_progress(converter_id)
        raise RuntimeError(
            f"GitHub listing for {src_dir} contained no downloadable files."
        )

    total_bytes = 0
    file_count = 0
    failures: list[str] = []
    _set_install_progress(
        converter_id,
        stage="downloading",
        current=0,
        total=len(download_jobs),
        bytes_done=0,
    )

    # Eight workers is a sweet spot — enough parallelism to saturate
    # most home links without triggering GitHub's anti-abuse limiter
    # on raw.githubusercontent.com (we've measured no 429s up to 16
    # workers in practice but 8 leaves headroom).
    #
    # Audit A11 — cumulative cap. Per-file caps don't help if the
    # repo listing itself grows hostile (1000 files * 500 MB each).
    # We abort the install when the running total exceeds
    # ``_MAX_INSTALL_BYTES`` and clean up the partial download.
    with ThreadPoolExecutor(max_workers=8) as pool:
        future_to_path = {
            pool.submit(_download_one_file, url, target): (url, target)
            for url, target in download_jobs
        }
        for fut in as_completed(future_to_path):
            url, target = future_to_path[fut]
            try:
                size = fut.result()
            except Exception as exc:
                failures.append(f"{url}: {exc}")
                continue
            total_bytes += size
            file_count += 1
            _set_install_progress(
                converter_id,
                current=file_count,
                bytes_done=total_bytes,
                file=Path(target).name,
            )
            if total_bytes > _MAX_INSTALL_BYTES:
                failures.append(
                    f"cumulative install size {total_bytes} bytes exceeded "
                    f"cap of {_MAX_INSTALL_BYTES} bytes — aborting"
                )
                # Cancel anything still in flight; the partial directory
                # gets cleaned up by the failure-handler below.
                for pending in future_to_path:
                    pending.cancel()
                break
            if file_count % 25 == 0:
                logger.info(
                    "Converter %s: downloaded %d/%d files (%.1f MB)",
                    converter_id, file_count, len(download_jobs),
                    total_bytes / 1024 / 1024,
                )

    if failures:
        # Roll back the partial download — we don't want a half-
        # installed converter that find_converter() will then
        # discover and try to use.
        import shutil as _shutil
        _shutil.rmtree(dest_root, ignore_errors=True)
        _clear_install_progress(converter_id)
        raise RuntimeError(
            f"{len(failures)} of {len(download_jobs)} downloads failed; "
            f"first error: {failures[0]}"
        )

    exe_name: str = _META_BY_ID[converter_id]["exe"]
    exe_path = dest_root / exe_name
    if not exe_path.exists():
        _clear_install_progress(converter_id)
        raise RuntimeError(
            f"Installed {file_count} files ({total_bytes} bytes) for "
            f"{converter_id} but {exe_name} is missing at {exe_path}"
        )

    _set_install_progress(
        converter_id,
        stage="verifying",
        current=len(download_jobs),
        bytes_done=total_bytes,
        file=None,
    )
    logger.info(
        "Installed %s converter: %d files, %.1f MB -> %s",
        converter_id, file_count, total_bytes / 1024 / 1024, exe_path,
    )
    return exe_path


@router.get(
    "/converters/{converter_id}/install-progress/",
    include_in_schema=True,
)
async def get_install_progress(converter_id: str) -> dict[str, Any]:
    """Lightweight progress poll for an in-flight converter install.

    The Windows installer downloads 30-175 files (~600 MB for RVT) inside
    a thread pool. ``install_converter`` doesn't return until the smoke
    test has finished — without this endpoint the frontend can only show
    a spinner for the full 30-90 s. We poll every 500 ms while the install
    mutation is pending and render `<progress>` + microcopy.

    Response shape:
      * ``{"active": False}`` — no install currently in flight (default,
        and what every fresh page-load returns)
      * ``{"active": True, "stage": "downloading", "current": 12,
            "total": 175, "bytes_done": 41943040, "file": "Qt6Core.dll",
            "started_at": 1715814723.49}`` — install in progress

    The dict is in-memory only — restarting the backend wipes it. That's
    fine because the install mutation will fail with a connection error
    on restart and the frontend will surface the error toast as the
    progress poll falls silent.
    """
    if converter_id not in _META_BY_ID:
        raise HTTPException(
            status_code=404,
            detail=f"Unknown converter: '{converter_id}'",
        )
    progress = _get_install_progress(converter_id)
    if progress is None:
        return {"active": False}
    return {"active": True, **progress}


@router.post(
    "/converters/{converter_id}/install/",
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def install_converter(
    converter_id: str,
    request: Request,
    _user_id: CurrentUserId,
    force: bool = Query(
        default=False,
        description=(
            "Bypass the 'already installed' short-circuit and re-download "
            "even when a binary is already present. Used by the 'Update' "
            "action when the version-check banner reports an outdated SHA."
        ),
    ),
) -> dict[str, Any]:
    """Download and install a DDC CAD/BIM converter.

    On Windows: walks the
    ``DDC_WINDOWS_Converters/DDC_CONVERTER_{FORMAT}/`` directory in
    the upstream repo via the GitHub Contents API and downloads each
    file (binary, Qt6 DLLs, plugins) into a per-format folder under
    ``~/.openestimator/converters/{format}_windows/``.

    On Linux: returns ``platform_unsupported`` with the apt-get
    command the user should run themselves. We deliberately do NOT
    auto-shell-out to ``apt`` here because that would require root
    + writing to ``/etc/apt/sources.list.d/`` and silently elevating
    privileges from a web request is exactly the wrong default.

    Returns 200 with ``installed: true`` on success, or 200 with
    ``installed: false`` + ``platform_unsupported`` on Linux/Mac. Only
    truly unrecoverable failures (network down, repo layout changed)
    raise HTTPException 502 — and even then, the response body carries
    the real error message so the user can act on it.

    Hardening note (v2.6.22): the function body is wrapped in a top-
    level try/except so an unexpected exception class (anything that
    isn't ``RuntimeError``) translates to a 502 with the underlying
    error class + message in ``detail`` instead of leaking a generic
    "Internal server error" 500.  Without this guard the install banner
    showed a useless toast when GitHub rate-limited or returned an
    unexpected payload — Hans's reproducible DWG/DGN failure case.
    """
    import asyncio
    import sys

    from app.modules.boq.cad_import import find_converter

    meta = _META_BY_ID.get(converter_id)
    if not meta:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Unknown converter: '{converter_id}'. "
                f"Available: {list(_META_BY_ID.keys())}"
            ),
        )

    try:
        # Already installed?
        existing = find_converter(converter_id)
        if existing and not force:
            # Skip the short-circuit when ``force=true`` so the "Update"
            # button can re-download a binary whose blob SHA no longer
            # matches the upstream GitHub Contents API.
            return {
                "converter_id": converter_id,
                "installed": True,
                "path": str(existing),
                "already_installed": True,
                "message": f"{meta['name']} is already installed at {existing}",
            }

        platform = sys.platform
        if platform == "win32":
            # Windows: download files from the upstream GitHub repo. The
            # download walks ~30-50 small-to-medium files (~30-50 MB total)
            # so we offload to a thread to keep the FastAPI event loop
            # responsive.
            try:
                exe_path = await asyncio.to_thread(
                    _download_converter_files_windows, converter_id,
                )
            except RuntimeError as exc:
                logger.warning(
                    "Windows converter install failed for %s: %s",
                    converter_id, exc,
                )
                raise HTTPException(
                    status_code=502,
                    detail=(
                        f"Could not install {meta['name']}: {exc}. "
                        f"You can install it manually from "
                        f"https://github.com/{_DDC_REPO}/tree/{_DDC_BRANCH}/"
                        f"{_WINDOWS_CONVERTER_DIRS[converter_id]}"
                    ),
                ) from exc

            # Post-install smoke test: launch the binary with a non-existent
            # input + output so it exits quickly. We're only checking that the
            # OS can load the exe + its Qt6 DLLs — a "missing DLL" error here
            # means the install is broken and would fail silently on the next
            # CAD upload, leaving the user staring at a "needs_converter"
            # banner with no clue why. Exit codes and stderr from valid CLI
            # error paths are fine; the only failure we care about is the
            # Windows loader emitting WinError 3221225781 (0xc0000135) —
            # "DLL not found" — which usually surfaces as a non-zero exit
            # AND empty stdout/stderr on stderr=PIPE.
            smoke_ok = True
            smoke_message: str | None = None
            try:
                import subprocess

                def _smoke() -> tuple[int, bytes, bytes]:
                    proc = subprocess.run(
                        [str(exe_path)],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        cwd=str(exe_path.parent),
                        input=b"\n",
                        timeout=15,
                    )
                    return proc.returncode, proc.stdout, proc.stderr

                rc, _stdout, stderr = await asyncio.to_thread(_smoke)
                # 0xC0000135 (-1073741515) → STATUS_DLL_NOT_FOUND
                # 0xC0000142 (-1073741502) → STATUS_DLL_INIT_FAILED
                if rc in (-1073741515, -1073741502, 3221225781, 3221225794):
                    smoke_ok = False
                    smoke_message = (
                        f"Installed but the binary can't load — "
                        f"a required DLL is missing (Windows error 0x{rc & 0xFFFFFFFF:08x}). "
                        f"This usually means the Qt6 plugins didn't download correctly. "
                        f"Try uninstalling and reinstalling, or install manually from "
                        f"https://github.com/{_DDC_REPO}/tree/{_DDC_BRANCH}/"
                        f"{_WINDOWS_CONVERTER_DIRS[converter_id]}"
                    )
            except subprocess.TimeoutExpired:
                # Timeout = binary loaded successfully and is waiting for
                # stdin (likely showing a Qt window or sitting in a
                # message loop).  That's exactly what we want to confirm:
                # the loader works.  Treat as healthy and log at INFO
                # instead of WARNING so the line doesn't read as a
                # contradiction next to the "smoke_test_passed: true"
                # response — see BUG-RVT02.
                logger.info(
                    "Smoke test for %s converter timed out — binary loaded "
                    "but waiting for stdin; treating as healthy.",
                    converter_id,
                )
            except Exception as exc:  # noqa: BLE001 — smoke test is best-effort
                # Any other exception (FileNotFoundError, PermissionError,
                # OSError) genuinely means the binary couldn't run.  Don't
                # claim ``smoke_test_passed: true`` in the response.
                smoke_ok = False
                smoke_message = (
                    f"Installed but the smoke test failed: {exc}. "
                    f"Try the Re-check button on the BIM page or reinstall."
                )
                logger.warning(
                    "Smoke test for %s converter failed: %s", converter_id, exc
                )

            size_bytes = exe_path.stat().st_size if exe_path.exists() else 0

            # Drop any stale "not_installed" health entry so the next
            # ``/converters/?verify=true`` poll re-runs the smoke test
            # against the freshly installed binary instead of replaying
            # a cached pre-install ``not_installed`` result.
            from app.modules.boq.cad_import import invalidate_converter_health
            invalidate_converter_health(converter_id)

            # Bust the 6-hour version-check cache so the "Update available"
            # banner re-evaluates against the freshly written blob — without
            # this the badge stays "outdated" for up to 6 h after a Windows
            # install and the user thinks the Update button did nothing.
            try:
                request.app.state._converter_version_cache = None
            except AttributeError:
                pass

            _clear_install_progress(converter_id)
            return {
                "converter_id": converter_id,
                "installed": smoke_ok,
                "path": str(exe_path),
                "already_installed": False,
                "size_bytes": size_bytes,
                "platform": "windows",
                "smoke_test_passed": smoke_ok,
                "message": smoke_message or (
                    f"{meta['name']} installed successfully at {exe_path}"
                ),
            }

        if platform.startswith("linux"):
            # Linux: surface apt instructions instead of auto-installing.
            # We do not write to /etc/apt or sudo from a web handler —
            # that needs root and a privilege-elevation policy we
            # don't ship by default.
            #
            # The apt repo at `pkg.datadrivenconstruction.io` is signed,
            # serves amd64+arm64, and the `.deb` packages drop a single
            # ELF binary into `/usr/bin/{Format}Exporter`. find_converter()
            # picks it up automatically on the next status poll.
            apt_pkg = _LINUX_APT_PACKAGES.get(converter_id, f"ddc-{converter_id}converter")
            linux_binary_name = (meta["exe"] or "").removesuffix(".exe")
            binary_path = f"/usr/bin/{linux_binary_name}" if linux_binary_name else None

            # Detect whether the user has already added the DDC apt
            # source. If yes, we can skip the source-setup lines and
            # surface a one-line install command instead.
            apt_source_path = Path("/etc/apt/sources.list.d/ddc.list")
            source_already_present = apt_source_path.exists()

            if source_already_present:
                instructions = f"sudo apt update && sudo apt install -y {apt_pkg}"
                short_message = (
                    f"DDC apt source already configured. Run "
                    f"`sudo apt install -y {apt_pkg}` to install "
                    f"{meta['name']}. find_converter picks it up "
                    f"automatically on the next status poll — no service "
                    f"restart needed."
                )
            else:
                instructions = (
                    f"# 1. Add the DDC apt source (one-time setup)\n"
                    f"echo 'deb [trusted=yes] https://pkg.datadrivenconstruction.io stable main' "
                    f"| sudo tee /etc/apt/sources.list.d/ddc.list\n"
                    f"sudo apt update\n\n"
                    f"# 2. Install the {meta['name']} (lands at {binary_path or '/usr/bin/'})\n"
                    f"sudo apt install -y {apt_pkg}"
                )
                short_message = (
                    f"One-time apt setup for {meta['name']}. Copy the "
                    f"two-step `instructions` into a root terminal — apt "
                    f"resolves the SDK shared libraries automatically and "
                    f"drops the binary at {binary_path or '/usr/bin/'}. "
                    f"find_converter picks it up on the next status poll, "
                    f"no service restart needed."
                )

            return {
                "converter_id": converter_id,
                "installed": False,
                "platform": "linux",
                # `platform_unsupported` is kept for backwards-compat with
                # frontend toast logic, but it's misleading now — Linux IS
                # supported, just via a one-time apt setup. The frontend
                # banner branches on `platform === 'linux'` to render the
                # softer "One-time apt setup" wording.
                "platform_unsupported": True,
                "apt_package": apt_pkg,
                "apt_source_present": source_already_present,
                "expected_binary_path": binary_path,
                "instructions": instructions,
                "message": short_message,
            }

        # macOS / other — no DDC build available
        return {
            "converter_id": converter_id,
            "installed": False,
            "platform": platform,
            "platform_unsupported": True,
            "message": (
                f"{meta['name']} is not yet available for {platform}. "
                f"Convert to IFC on a Windows machine first, then upload the IFC "
                f"file — IFC has a built-in text fallback parser that works on "
                f"every platform."
            ),
        }
    except HTTPException:
        # Already a structured response — let FastAPI translate it.
        raise
    except Exception as exc:  # noqa: BLE001 — last-ditch error envelope
        # Any uncaught exception (json.JSONDecodeError on a rate-limited
        # GitHub response, OSError on a permission-denied install dir,
        # etc.) used to leak as a generic 500 "Internal server error"
        # which gave the user no actionable signal.  Translate it to a
        # 502 with the real error class + message so the install banner
        # shows something useful — Hans's DWG/DGN failure case (Linux
        # VPS hitting an edge case in find_converter / urlretrieve).
        _clear_install_progress(converter_id)
        logger.exception(
            "Unhandled exception in install_converter for %s: %s",
            converter_id, exc,
        )
        raise HTTPException(
            status_code=502,
            detail=(
                f"Could not install {meta['name']}: "
                f"{type(exc).__name__}: {exc}. "
                f"Check the server logs for the full traceback. You can "
                f"install manually from "
                f"https://github.com/{_DDC_REPO}/tree/{_DDC_BRANCH}/"
                f"{_WINDOWS_CONVERTER_DIRS.get(converter_id, '')}"
            ),
        ) from exc


@router.post(
    "/converters/manifest/install/{component_name}",
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def install_from_manifest(
    component_name: str,
    _user_id: CurrentUserId,
) -> dict[str, Any]:
    """Install a component using the signed manifest (Audit A1).

    Closes the A1 gap by sourcing the download URL + expected SHA-256
    from a signed manifest instead of trusting whatever the GitHub
    Contents API returns at runtime. The signature is verified against
    an Ed25519 public key embedded in ``manifest_verifier.py`` — see
    that module's header comment for the threat model.

    Failure modes (all map to HTTP 502 with a clear ``detail`` so the
    install banner can render the real reason instead of "internal
    error"):

    * Bad signature → manifest was tampered with in transit OR the
      signing key was rotated and this client is stale
    * Platform missing → no build for this OS/arch combination
    * SHA mismatch → the file at the manifest URL is NOT the file the
      publisher signed for (CDN poison, MITM, or upstream replace)
    """
    import asyncio

    cache_dir = _CONVERTER_CACHE_DIR
    cache_dir.mkdir(parents=True, exist_ok=True)
    target = cache_dir / f"{component_name}.manifest-download"

    try:
        manifest = await asyncio.to_thread(fetch_manifest)
        resolved = resolve_install(manifest, component_name)

        # Reuse the hardened download helper from A2/A9/A11 — it gives
        # us host allow-listing, symlink guards, and the streaming
        # size cap for free.
        bytes_written = await asyncio.to_thread(
            _download_one_file, resolved.url, target,
        )

        # SHA verification — this is the new A1 check. Mismatch deletes
        # the partial file so a retry can't pick up a poisoned blob.
        await asyncio.to_thread(
            verify_downloaded_file,
            target, resolved.sha256, resolved.size_bytes,
        )

        return {
            "component": component_name,
            "installed": True,
            "version": resolved.version,
            "platform": resolved.platform_key,
            "path": str(target),
            "size_bytes": bytes_written,
            "sha256": resolved.sha256,
            "manifest_version": manifest.version,
            "manifest_signed_at": manifest.signed_at,
        }
    except ManifestSignatureInvalid as exc:
        logger.error("Manifest signature verification failed: %s", exc)
        raise HTTPException(
            status_code=502,
            detail=(
                f"Manifest signature did not verify. Refusing to install. "
                f"This usually means the signing key was rotated and your "
                f"client is stale, or there is an active MITM between you "
                f"and the package CDN. Details: {exc}"
            ),
        ) from exc
    except InstallNotSupported as exc:
        raise HTTPException(
            status_code=502,
            detail=(
                f"Component {component_name!r} is not available for this "
                f"platform. Please file an issue. Details: {exc}"
            ),
        ) from exc
    except InstallSHAMismatch as exc:
        logger.error("Manifest SHA mismatch installing %s: %s", component_name, exc)
        raise HTTPException(
            status_code=502,
            detail=(
                f"Downloaded file does not match the manifest hash. "
                f"Refusing to install — the file at the manifest URL is "
                f"NOT the file the publisher signed for. Partial file "
                f"has been deleted. Details: {exc}"
            ),
        ) from exc
    except ManifestError as exc:
        raise HTTPException(
            status_code=502,
            detail=f"Manifest install failed: {type(exc).__name__}: {exc}",
        ) from exc


@router.post(
    "/converters/{converter_id}/uninstall/",
    dependencies=[Depends(RequirePermission("takeoff.delete"))],
)
async def uninstall_converter(
    converter_id: str,
    _user_id: CurrentUserId,
) -> dict[str, Any]:
    """Remove an installed DDC CAD/BIM converter."""
    meta = _META_BY_ID.get(converter_id)
    if not meta:
        raise HTTPException(
            status_code=404,
            detail=f"Unknown converter: '{converter_id}'",
        )

    import sys

    exe_name: str = meta["exe"]
    removed = False

    # Linux: the converter was apt-installed under `/usr/bin/`, so we
    # can't delete it from a web handler (no root, and `dpkg --remove`
    # needs to update the package database). Surface `apt remove`
    # instructions and let the user run them.
    platform = sys.platform.lower()
    if platform.startswith("linux"):
        apt_pkg = _LINUX_APT_PACKAGES.get(
            converter_id, f"ddc-{converter_id}converter"
        )
        return {
            "converter_id": converter_id,
            "removed": False,
            "platform": "linux",
            "apt_package": apt_pkg,
            "instructions": f"sudo apt remove -y {apt_pkg}",
            "message": (
                f"Run `sudo apt remove -y {apt_pkg}` to uninstall "
                f"{meta['name']}. The status poll picks up the change "
                f"automatically; no service restart needed."
            ),
        }

    # Windows uninstall: sweep the install dir for the binary + bundled
    # plugin folders. The new installer drops files into a per-format
    # folder ({ext}_windows/) so its bundled Qt6 DLLs don't collide
    # with other converters; older builds may still have files at the
    # install root or in arbitrary subdirs, so we sweep all of those.
    candidates: list[Path] = [_CONVERTER_INSTALL_DIR / exe_name]
    per_format_root = _CONVERTER_INSTALL_DIR / f"{converter_id}_windows"
    if per_format_root.exists():
        candidates.append(per_format_root / exe_name)
    if _CONVERTER_INSTALL_DIR.exists():
        for child in _CONVERTER_INSTALL_DIR.iterdir():
            if child.is_dir():
                candidates.append(child / exe_name)

    # Remove from install dir
    for candidate in candidates:
        if candidate.exists():
            candidate.unlink()
            removed = True
            logger.info("Removed converter executable: %s", candidate)

    # Also wipe the per-format folder entirely if we created one — no
    # point in keeping orphaned Qt DLLs around once the user opted out.
    if per_format_root.exists() and per_format_root.is_dir():
        import shutil as _shutil
        try:
            _shutil.rmtree(per_format_root)
            logger.info("Removed per-format converter folder: %s", per_format_root)
        except OSError as exc:
            logger.warning("Could not remove %s: %s", per_format_root, exc)

    # Drop cached health for this converter so the next status poll
    # re-runs the smoke test against the empty install dir (and reports
    # ``not_installed`` instead of a stale ``ok``).
    from app.modules.boq.cad_import import invalidate_converter_health
    invalidate_converter_health(converter_id)

    return {
        "converter_id": converter_id,
        "removed": removed,
        "message": f"{meta['name']} uninstalled" if removed else f"{meta['name']} was not installed",
    }


# ── CAD quantity extraction (no AI) ──────────────────────────────────────

MAX_CAD_SIZE = 100 * 1024 * 1024  # 100 MB

_SUPPORTED_CAD_EXTS = {"rvt", "ifc", "dwg", "dgn", "rfa", "dxf"}


@router.post(
    "/cad-extract/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_extract(
    file: UploadFile = File(..., description="CAD/BIM file (.rvt, .ifc, .dwg, .dgn)"),
) -> dict[str, Any]:
    """Extract grouped quantity tables from a CAD/BIM file.

    Converts the file using a DDC Community converter, parses the resulting
    Excel output, and groups elements deterministically by category and type.
    **No AI key required** — this is pure file conversion + grouping.

    Returns quantity tables with per-category and grand totals for:
    count, volume (m3), area (m2), and length (m).
    """
    import tempfile
    import time

    from app.modules.boq.cad_import import (
        convert_cad_to_excel,
        find_converter,
        group_cad_elements,
        parse_cad_excel,
    )

    filename = file.filename or "file"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in _SUPPORTED_CAD_EXTS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unsupported file type: .{ext}. Accepted: {', '.join(f'.{e}' for e in sorted(_SUPPORTED_CAD_EXTS))}"
            ),
        )

    converter = find_converter(ext)
    if not converter:
        raise HTTPException(
            status_code=400,
            detail=(
                f"DDC converter for .{ext} files is not installed. "
                f"Install it from the Quantities page (/quantities) or download "
                f"from https://github.com/datadrivenconstruction/ddc-community-toolkit/releases"
            ),
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    start_time = time.monotonic()

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / filename
        input_path.write_bytes(content)

        output_dir = Path(tmpdir) / "output"
        output_dir.mkdir()

        excel_path = await convert_cad_to_excel(input_path, output_dir, ext)
        if not excel_path:
            raise HTTPException(
                status_code=502,
                detail=(
                    f"CAD conversion failed for .{ext} file. "
                    "Ensure the converter is properly installed and the file is valid."
                ),
            )

        elements = parse_cad_excel(excel_path)

    if not elements:
        raise HTTPException(
            status_code=422,
            detail="Converter produced no elements. The file may be empty or unsupported.",
        )

    grouped = group_cad_elements(elements)
    duration_ms = int((time.monotonic() - start_time) * 1000)

    return {
        "filename": filename,
        "format": ext,
        "total_elements": grouped["total_elements"],
        "duration_ms": duration_ms,
        "groups": grouped["groups"],
        "grand_totals": grouped["grand_totals"],
    }


# ── CAD interactive grouping (two-step flow) ──────────────────────────────

# In-memory cache kept as fast fallback; primary storage is the database.
_cad_sessions: dict[str, dict] = {}
_CAD_SESSION_TTL = 86400  # 24 hours (was 5 minutes)


def _cleanup_memory_sessions() -> None:
    """Remove expired CAD extraction sessions from the in-memory cache."""
    now = _time.time()
    expired = [k for k, v in _cad_sessions.items() if now - v["created"] > _CAD_SESSION_TTL]
    for k in expired:
        del _cad_sessions[k]


async def _cleanup_db_sessions(session: Any) -> None:
    """Remove expired CAD extraction sessions from the database (skip permanent)."""
    now = datetime.now(UTC)
    await session.execute(
        delete(CadExtractionSession).where(
            CadExtractionSession.expires_at < now,
            CadExtractionSession.is_permanent == False,  # noqa: E712
        )
    )


async def _save_session_to_db(
    session: Any,
    session_id: str,
    elements: list[dict],
    filename: str,
    file_format: str,
    columns_metadata: dict | None = None,
    user_id: str = "",
    extraction_time: float = 0,
) -> None:
    """Persist a CAD extraction session to the database."""
    expires_at = datetime.now(UTC) + timedelta(seconds=_CAD_SESSION_TTL)
    db_session = CadExtractionSession(
        session_id=session_id,
        user_id=user_id,
        filename=filename,
        file_format=file_format,
        element_count=len(elements),
        extraction_time=extraction_time,
        elements_data=elements,
        columns_metadata=columns_metadata or {},
        expires_at=expires_at,
        created_by=user_id,
    )
    session.add(db_session)
    await session.flush()


async def _get_session_from_db(session: Any, session_id: str) -> dict | None:
    """Retrieve a CAD extraction session from the database.

    Returns a dict matching the old in-memory format, or None if not found / expired.

    Audit B6 — the returned dict now carries ``user_id`` and ``project_id``
    so the ownership helper can gate access without re-querying the row.
    """
    now = datetime.now(UTC)
    result = await session.execute(
        select(CadExtractionSession).where(
            CadExtractionSession.session_id == session_id,
            CadExtractionSession.expires_at > now,
        )
    )
    row = result.scalar_one_or_none()
    if row is None:
        return None
    return {
        "elements": row.elements_data or [],
        "filename": row.filename,
        "format": row.file_format,
        "created": row.created_at.timestamp() if row.created_at else _time.time(),
        "columns_metadata": row.columns_metadata or {},
        # B6 — carry ownership fields for IDOR gate
        "user_id": row.user_id or "",
        "project_id": row.project_id or None,
    }


async def _get_cad_session(session: Any, session_id: str) -> dict | None:
    """Look up a CAD session from memory first, then fall back to database."""
    # Fast path: in-memory
    mem = _cad_sessions.get(session_id)
    if mem is not None:
        return mem
    # Slow path: database
    return await _get_session_from_db(session, session_id)


async def _verify_cad_session_access(
    cad_session: dict,
    user_id: str,
    db_session: Any,
) -> None:
    """Gate access to a CAD extraction session by tenant.

    Audit B6 — used by every endpoint that consumes a session_id
    (``cad_data_elements`` / ``cad_data_aggregate`` /
    ``cad_data_save`` / ``cad_data_list_sessions`` /
    ``cad_data_delete_session``). Two cases mirror the document
    helper:

    1. Session is bound to a project — reuse
       ``verify_project_access`` so admin bypass + ownership work
       identically across the takeoff module.
    2. Standalone session (no project) — only the original uploader
       can touch it. We return 404 on access failure to avoid
       leaking session existence.
    """
    raw_pid = cad_session.get("project_id")
    if raw_pid:
        try:
            pid = _uuid.UUID(str(raw_pid))
        except (ValueError, TypeError):
            pid = None
        if pid is not None:
            await verify_project_access(pid, str(user_id), db_session)
            return

    owner = str(cad_session.get("user_id", "") or "")
    if owner and owner != str(user_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found.",
        )


class CadGroupRequest(BaseModel):
    """Request body for the ``POST /cad-group`` endpoint."""

    session_id: str = Field(..., description="Session ID returned by /cad-columns")
    group_by: list[str] = Field(..., min_length=1, description="Columns to group by")
    sum_columns: list[str] = Field(default_factory=list, description="Numeric columns to sum")


@router.post(
    "/cad-columns/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_columns(
    file: UploadFile = File(..., description="CAD/BIM file (.rvt, .ifc, .dwg, .dgn)"),
    session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Upload a CAD file and analyze its columns for interactive grouping.

    Step 1 of the two-step interactive QTO flow:
    1. Upload file -> get available columns + session_id
    2. POST /cad-group with session_id + chosen columns -> get grouped results

    Returns column classification (grouping / quantity / text), suggested
    defaults, a preview of the first 10 elements, and a ``session_id`` for
    the follow-up grouping request.
    """
    import tempfile
    import time

    from app.modules.boq.cad_import import (
        convert_cad_to_excel,
        find_converter,
        get_available_columns,
        parse_cad_excel,
    )

    filename = file.filename or "file"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in _SUPPORTED_CAD_EXTS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unsupported file type: .{ext}. Accepted: {', '.join(f'.{e}' for e in sorted(_SUPPORTED_CAD_EXTS))}"
            ),
        )

    converter = find_converter(ext)
    if not converter:
        raise HTTPException(
            status_code=400,
            detail=(
                f"DDC converter for .{ext} files is not installed. "
                f"Install it from the Quantities page (/quantities) or download "
                f"from https://github.com/datadrivenconstruction/ddc-community-toolkit/releases"
            ),
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    start_time = time.monotonic()

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / filename
        input_path.write_bytes(content)

        output_dir = Path(tmpdir) / "output"
        output_dir.mkdir()

        excel_path = await convert_cad_to_excel(input_path, output_dir, ext)
        if not excel_path:
            raise HTTPException(
                status_code=502,
                detail=(
                    f"CAD conversion failed for .{ext} file. "
                    "Ensure the converter is properly installed and the file is valid."
                ),
            )

        elements = parse_cad_excel(excel_path)

    if not elements:
        raise HTTPException(
            status_code=422,
            detail="Converter produced no elements. The file may be empty or unsupported.",
        )

    # Filter out empty/system elements (Phases, Patterns, Views, etc.)
    _SKIP_CATEGORIES = {
        "none",
        "",
        "ost_phases",
        "ost_materials",
        "ost_viewports",
        "ost_colorfilllegends",
        "ost_views",
        "ost_grids",
        "ost_levels",
        "ost_sheets",
        "ost_titleblocks",
    }
    real_elements = [el for el in elements if str(el.get("category", "")).strip().lower() not in _SKIP_CATEGORIES]
    # If filtering removed everything, keep originals
    if not real_elements:
        real_elements = elements

    columns = get_available_columns(real_elements, file_format=ext)
    duration_ms = int((time.monotonic() - start_time) * 1000)

    # Cleanup expired sessions, then store new one
    _cleanup_memory_sessions()
    session_id = str(_uuid.uuid4())
    _cad_sessions[session_id] = {
        "elements": real_elements,
        "filename": filename,
        "format": ext,
        "created": _time.time(),
        "columns_metadata": columns,
        # B6 — carry owner so memory-fast-path access checks work too
        "user_id": user_id or "",
        "project_id": None,
    }

    # Persist to database for durability
    if session is not None:
        await _cleanup_db_sessions(session)
        await _save_session_to_db(
            session=session,
            session_id=session_id,
            elements=real_elements,
            filename=filename,
            file_format=ext,
            columns_metadata=columns,
            user_id=user_id or "",
            extraction_time=duration_ms / 1000.0,
        )

    # Preview: pick elements that have actual quantity data (volume > 0 or area > 0)
    def _has_quantity(el: dict) -> bool:
        for key in ("volume", "area", "length", "gross volume", "gross area"):
            val = el.get(key)
            if val is not None:
                try:
                    if float(val) > 0:
                        return True
                except (ValueError, TypeError):
                    pass
        return False

    preview_candidates = [el for el in real_elements if _has_quantity(el)][:10]
    if not preview_candidates:
        # Fallback: elements with type name
        preview_candidates = [el for el in real_elements if el.get("type name")][:10]
    if not preview_candidates:
        preview_candidates = real_elements[:10]

    return {
        "session_id": session_id,
        "filename": filename,
        "format": ext,
        "total_elements": len(real_elements),
        "duration_ms": duration_ms,
        "columns": {
            "grouping": columns.get("grouping", []),
            "quantity": columns.get("quantity", []),
            "text": columns.get("text", []),
        },
        "suggested_grouping": columns.get("suggested_grouping", []),
        "suggested_quantities": columns.get("suggested_quantities", []),
        "presets": columns.get("presets", {}),
        "unit_labels": columns.get("unit_labels", {}),
        "confidence": columns.get("confidence", {}),
        "preview": preview_candidates,
    }


@router.post(
    "/cad-group/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_group(
    body: CadGroupRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Group previously uploaded CAD elements by user-selected columns.

    Step 2 of the two-step interactive QTO flow. Requires a valid
    ``session_id`` from a prior ``POST /cad-columns`` call.

    Sessions are stored in the database and expire after 24 hours.
    If expired, the user must re-upload the file.

    Audit B6 — was IDOR. Same threat as ``cad_data_elements``.
    """
    from app.modules.boq.cad_import import group_cad_elements_dynamic

    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]

    # Validate that requested columns actually exist in the data
    all_columns: set[str] = set()
    for el in elements:
        all_columns.update(el.keys())

    missing_group = [c for c in body.group_by if c not in all_columns]
    if missing_group:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown grouping column(s): {missing_group}. Available: {sorted(all_columns)}",
        )

    # "count" is a virtual column (computed as number of elements per group)
    missing_sum = [c for c in body.sum_columns if c not in all_columns and c != "count"]
    if missing_sum:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown sum column(s): {missing_sum}. Available: {sorted(all_columns)}",
        )

    grouped = group_cad_elements_dynamic(elements, body.group_by, body.sum_columns)

    return {
        "filename": cad_session["filename"],
        "format": cad_session["format"],
        **grouped,
    }


# ── Element detail view for a specific group ──────────────────────────────


class CadGroupElementsRequest(BaseModel):
    """Request body for the ``POST /cad-group/elements`` endpoint."""

    session_id: str = Field(..., description="Session ID returned by /cad-columns")
    group_key: dict[str, str] = Field(
        ...,
        description='Key-value pairs identifying the group (e.g. {"category": "Walls", "type name": "Exterior Wall"})',
    )


@router.post(
    "/cad-group/elements/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def get_group_elements(
    body: CadGroupElementsRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Get individual elements for a specific group.

    Returns all raw elements matching the provided ``group_key`` filter,
    allowing users to inspect what makes up each grouped row.

    Audit B6 — was IDOR. Same threat as ``cad_data_elements``.
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]

    # Filter elements matching all key-value pairs in group_key
    matching: list[dict] = []
    for el in elements:
        match = True
        for col, expected in body.group_key.items():
            raw = el.get(col)
            val = str(raw).strip() if raw is not None else ""
            normalized = val if val and val != "None" else "(empty)"
            if normalized != expected:
                match = False
                break
        if match:
            matching.append(el)

    # Discover all column names across matching elements
    all_cols: list[str] = []
    seen: set[str] = set()
    for el in matching:
        for k in el:
            if k not in seen:
                seen.add(k)
                all_cols.append(k)

    # Compute totals for numeric columns
    totals: dict[str, float] = {}
    for col in all_cols:
        numeric_sum = 0.0
        is_numeric = False
        for el in matching:
            val = el.get(col)
            if val is not None:
                try:
                    numeric_sum += float(val)
                    is_numeric = True
                except (ValueError, TypeError):
                    pass
        if is_numeric:
            totals[col] = round(numeric_sum, 4)

    return {
        "group_key": body.group_key,
        "total_elements": len(matching),
        "columns": all_cols,
        "elements": matching[:500],  # Limit to 500 elements for performance
        "totals": totals,
        "truncated": len(matching) > 500,
    }


# ── Create BOQ from CAD QTO ───────────────────────────────────────────────


class CreateBOQFromCadRequest(BaseModel):
    """Request body for creating a BOQ from grouped CAD QTO data."""

    session_id: str = Field(..., description="Session ID from /cad-columns")
    project_id: str = Field(..., description="Project UUID to create BOQ in")
    boq_name: str = Field(default="CAD Import", description="Name for the new BOQ")
    group_by: list[str] = Field(default_factory=list, description="Columns used for grouping")
    sum_columns: list[str] = Field(default_factory=list, description="Columns used for summing")


@router.post(
    "/cad-group/create-boq/",
    status_code=201,
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def create_boq_from_cad_qto(
    body: CreateBOQFromCadRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Create a BOQ directly from grouped CAD QTO data.

    Retrieves the cached CAD session, runs grouping using the provided
    (or stored) column selections, creates a new BOQ in the specified
    project, and adds positions for each group.

    Audit B6 — was IDOR-on-write on both sides:
    1. Source CAD session ownership (data theft)
    2. Destination project ownership (planting BOQ in foreign project)
    """
    import uuid

    from app.modules.boq.cad_import import group_cad_elements_dynamic
    from app.modules.boq.models import BOQ, Position

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("CAD session not found or expired. Please re-upload the CAD file."),
        )

    # Source-side access check
    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    # Destination-side access check (target project)
    try:
        target_pid = _uuid.UUID(body.project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=400,
            detail="Invalid project_id (must be a UUID)",
        ) from exc
    await verify_project_access(target_pid, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]

    # Use provided grouping or fall back to stored metadata
    group_by = body.group_by
    sum_columns = body.sum_columns
    if not group_by:
        stored = cad_session.get("columns_metadata", {})
        group_by = stored.get("suggested_grouping", [])
    if not sum_columns:
        stored = cad_session.get("columns_metadata", {})
        sum_columns = stored.get("suggested_quantities", [])

    if not group_by:
        raise HTTPException(
            status_code=400,
            detail="No grouping columns specified and no defaults available.",
        )

    grouped = group_cad_elements_dynamic(elements, group_by, sum_columns)
    groups = grouped.get("groups", [])

    # Determine unit labels from stored metadata
    stored_meta = cad_session.get("columns_metadata", {})
    unit_labels: dict[str, str] = stored_meta.get("unit_labels", {})

    # Create BOQ
    project_uuid = uuid.UUID(body.project_id)
    boq = BOQ(
        project_id=project_uuid,
        name=body.boq_name,
        description=f"Auto-generated from CAD file: {cad_session['filename']}",
        status="draft",
        metadata_={"source": "cad_qto", "cad_filename": cad_session["filename"]},
    )
    db_session.add(boq)
    await db_session.flush()

    # Create positions from groups
    position_count = 0
    for idx, group in enumerate(groups):
        # Skip empty groups
        sums = group.get("sums", {})
        count = group.get("count", 0)
        if count == 0 and all(v == 0 for v in sums.values()):
            continue

        # Build description from group key parts
        key_parts = group.get("key_parts", {})
        parts = []
        for col, val in key_parts.items():
            cleaned = str(val or "").strip()
            if col == "category":
                cleaned = cleaned.replace("OST_", "")
            if cleaned:
                parts.append(cleaned)
        description = " — ".join(parts) if parts else group.get("key", f"Group {idx + 1}")

        # Determine best unit and quantity (volume > area > length > count)
        unit = "pcs"
        quantity = float(count)
        for col_name in ["volume", "area", "length"]:
            if col_name in sums and sums[col_name] > 0:
                unit = unit_labels.get(col_name, col_name)
                quantity = round(sums[col_name], 4)
                break

        ordinal = f"{idx + 1:03d}"

        position = Position(
            boq_id=boq.id,
            ordinal=ordinal,
            description=description,
            unit=unit,
            quantity=str(quantity),
            unit_rate="0",
            total="0",
            source="cad_import",
            sort_order=idx,
            metadata_={
                "cad_source": "cad_qto",
                "cad_count": count,
                "cad_sums": sums,
            },
        )
        db_session.add(position)
        position_count += 1

    await db_session.flush()

    logger.info(
        "Created BOQ '%s' with %d positions from CAD QTO session %s",
        body.boq_name,
        position_count,
        body.session_id,
    )

    return {
        "boq_id": str(boq.id),
        "project_id": body.project_id,
        "position_count": position_count,
        "boq_name": body.boq_name,
    }


# ── Export grouped CAD QTO as Excel ───────────────────────────────────────


@router.get(
    "/cad-group/export/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def export_cad_group(
    session_id: str = Query(..., description="Session ID from /cad-columns"),
    group_by: str = Query(default="", description="Comma-separated grouping columns"),
    sum_columns: str = Query(default="", description="Comma-separated sum columns"),
    format: str = Query(default="xlsx", pattern="^(xlsx)$"),
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> StreamingResponse:
    """Export grouped QTO results as an Excel spreadsheet.

    Retrieves the CAD session, runs grouping, and returns an xlsx file
    with headers, data rows, and a bold grand-total row.

    Audit B6 — was IDOR. Anyone with ``takeoff.read`` could download
    another tenant's CAD QTO as XLSX (a one-click data-exfil path).
    """
    import io

    from app.modules.boq.cad_import import group_cad_elements_dynamic

    cad_session = await _get_cad_session(db_session, session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("CAD session not found or expired. Please re-upload the CAD file."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]

    # Parse column lists
    group_by_list = [c.strip() for c in group_by.split(",") if c.strip()] if group_by else []
    sum_columns_list = [c.strip() for c in sum_columns.split(",") if c.strip()] if sum_columns else []

    # Fall back to stored metadata
    if not group_by_list:
        stored = cad_session.get("columns_metadata", {})
        group_by_list = stored.get("suggested_grouping", [])
    if not sum_columns_list:
        stored = cad_session.get("columns_metadata", {})
        sum_columns_list = stored.get("suggested_quantities", [])

    if not group_by_list:
        raise HTTPException(
            status_code=400,
            detail="No grouping columns specified.",
        )

    grouped = group_cad_elements_dynamic(elements, group_by_list, sum_columns_list)
    groups = grouped.get("groups", [])
    grand_totals = grouped.get("grand_totals", {})

    # Build Excel workbook
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Cannot generate Excel export.",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QTO Export"

    # Determine column order
    all_sum_cols = [c for c in sum_columns_list if c != "count"]
    header = list(group_by_list) + all_sum_cols + ["Count"]

    # Write header — column names are user-supplied via the API, so they
    # need formula-injection neutralisation too (BUG-CSV-INJECTION).
    bold_font = Font(bold=True)
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(
            row=1,
            column=col_idx,
            value=neutralise_formula(col_name.replace("_", " ").title()),
        )
        cell.font = bold_font

    # Write data rows
    for row_idx, group in enumerate(groups, 2):
        key_parts = group.get("key_parts", {})
        sums = group.get("sums", {})
        count = group.get("count", 0)

        col_idx = 1
        for gc in group_by_list:
            val = str(key_parts.get(gc, "")).replace("OST_", "")
            ws.cell(row=row_idx, column=col_idx, value=neutralise_formula(val))
            col_idx += 1
        for sc in all_sum_cols:
            ws.cell(row=row_idx, column=col_idx, value=round(sums.get(sc, 0), 4))
            col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=count)

    # Grand total row
    total_row = len(groups) + 2
    col_idx = 1
    total_cell = ws.cell(row=total_row, column=col_idx, value="TOTAL")
    total_cell.font = bold_font
    col_idx = len(group_by_list) + 1
    for sc in all_sum_cols:
        cell = ws.cell(row=total_row, column=col_idx, value=round(grand_totals.get(sc, 0), 4))
        cell.font = bold_font
        col_idx += 1
    total_count = grand_totals.get("count", sum(g.get("count", 0) for g in groups))
    cell = ws.cell(row=total_row, column=col_idx, value=total_count)
    cell.font = bold_font

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter  # type: ignore[union-attr]
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Write to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_base = cad_session.get("filename", "export").rsplit(".", 1)[0]
    download_name = f"{filename_base}_qto.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


# ── CAD Data Explorer ────────────────────────────────────────────────────


class CadDataDescribeRequest(BaseModel):
    """Request body for the ``POST /cad-data/describe`` endpoint."""

    session_id: str = Field(..., description="Session ID returned by /cad-columns")


class CadDataValueCountsRequest(BaseModel):
    """Request body for the ``POST /cad-data/value-counts`` endpoint."""

    session_id: str = Field(..., description="Session ID returned by /cad-columns")
    column: str = Field(..., description="Column name to count values for")
    limit: int = Field(default=50, ge=1, le=500, description="Max number of distinct values")


class CadDataAggregateRequest(BaseModel):
    """Request body for the ``POST /cad-data/aggregate`` endpoint."""

    session_id: str = Field(..., description="Session ID returned by /cad-columns")
    group_by: list[str] = Field(..., min_length=1, description="Columns to group by")
    aggregations: dict[str, str] = Field(
        ...,
        description=("Mapping of column -> aggregation function. Supported: sum, avg, mean, min, max, count."),
    )


def _is_numeric(value: Any) -> bool:
    """Return True if *value* can be converted to a float."""
    if value is None:
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def _to_float(value: Any) -> float:
    """Convert *value* to float, raising on failure."""
    return float(value)


def _collect_column_names(elements: list[dict]) -> list[str]:
    """Return a stable list of all column names across *elements*."""
    seen: set[str] = set()
    columns: list[str] = []
    for el in elements:
        for k in el:
            if k not in seen:
                seen.add(k)
                columns.append(k)
    return columns


@router.post(
    "/cad-data/describe/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_data_describe(
    body: CadDataDescribeRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Return a DataFrame-like describe of the CAD session data.

    For each column, reports dtype, non-null count, unique count, and
    summary statistics (min/max/mean/sum for numbers, top/top_freq for strings).

    Audit B6 — was IDOR.
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]
    all_columns = _collect_column_names(elements)

    columns_info: list[dict[str, Any]] = []
    for col in all_columns:
        values = [el.get(col) for el in elements]
        non_null = [v for v in values if v is not None]
        non_null_count = len(non_null)

        # Determine if column is numeric
        numeric_vals: list[float] = []
        for v in non_null:
            if _is_numeric(v):
                numeric_vals.append(_to_float(v))

        is_numeric_col = len(numeric_vals) > len(non_null) * 0.5 and numeric_vals

        unique_vals = set(str(v) for v in non_null)
        unique_count = len(unique_vals)

        col_info: dict[str, Any] = {
            "name": col,
            "dtype": "number" if is_numeric_col else "string",
            "non_null": non_null_count,
            "unique": unique_count,
        }

        if is_numeric_col:
            col_info["min"] = round(min(numeric_vals), 4)
            col_info["max"] = round(max(numeric_vals), 4)
            col_info["mean"] = round(_mean(numeric_vals), 4)
            col_info["sum"] = round(sum(numeric_vals), 4)
        else:
            # Find the most common value
            freq: dict[str, int] = {}
            for v in non_null:
                key = str(v)
                freq[key] = freq.get(key, 0) + 1
            if freq:
                top_value = max(freq, key=freq.get)  # type: ignore[arg-type]
                col_info["top"] = top_value
                col_info["top_freq"] = freq[top_value]

        columns_info.append(col_info)

    return {
        "filename": cad_session.get("filename", ""),
        "format": cad_session.get("format", ""),
        "total_elements": len(elements),
        "total_columns": len(all_columns),
        "columns": columns_info,
    }


# ── CAD Data Explorer: missingno-style column fill-rate ───────────────────


_MISSINGNESS_SAMPLE_CAP = 1000
_MISSINGNESS_SAMPLE_SEED = 42


def _resolve_filter_column(all_columns: list[str], candidates: list[str]) -> str | None:
    """Return the first column from *all_columns* whose lowercased name is in *candidates*.

    Makes filter resolution case-insensitive and tolerant of naming variants
    (e.g. "Category" vs "category", "type name" vs "Type").
    """
    if not candidates:
        return None
    lower_map = {c.lower(): c for c in all_columns}
    for cand in candidates:
        hit = lower_map.get(cand.lower())
        if hit is not None:
            return hit
    return None


def _infer_dtype(values: list[Any]) -> str:
    """Infer a coarse dtype label from a list of non-null values."""
    if not values:
        return "object"
    numeric = 0
    bool_like = 0
    for v in values:
        if isinstance(v, bool):
            bool_like += 1
        elif _is_numeric(v):
            numeric += 1
    total = len(values)
    if bool_like / total > 0.9:
        return "bool"
    if numeric / total > 0.5:
        return "number"
    if all(isinstance(v, str) for v in values):
        return "string"
    return "object"


@router.get(
    "/cad-data/missingness/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_data_missingness(
    session_id: str = Query(..., description="Session ID returned by /cad-columns"),
    category_filter: str | None = Query(
        default=None,
        description="Value to match against the 'category' column (case-insensitive). Omit for all.",
    ),
    element_type_filter: str | None = Query(
        default=None,
        description=(
            "Value to match against the element-type column (tries 'type name', 'type', 'family')."
        ),
    ),
    sort: str = Query(
        default="fill_desc",
        pattern="^(fill_desc|fill_asc|alpha_asc|alpha_desc)$",
        description="Column ordering for the client",
    ),
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Return per-column fill-rate and per-row completeness for a missingno-style visualisation.

    The row-level snapshot is capped at 1000 rows (random-sampled with a fixed
    seed for determinism) so the frontend never has to render more than that.
    Per-column fill-rates are always computed on the full (filtered) set.

    Audit B6 — was IDOR.
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]
    all_columns = _collect_column_names(elements)

    # --- Apply filters ----------------------------------------------------
    applied_filters: dict[str, str] = {}

    category_col = _resolve_filter_column(all_columns, ["category"])
    if category_filter and category_col:
        needle = category_filter.strip().lower()
        elements = [el for el in elements if str(el.get(category_col, "")).strip().lower() == needle]
        applied_filters[category_col] = category_filter

    type_col = _resolve_filter_column(all_columns, ["type name", "type", "family"])
    if element_type_filter and type_col:
        needle = element_type_filter.strip().lower()
        elements = [el for el in elements if str(el.get(type_col, "")).strip().lower() == needle]
        applied_filters[type_col] = element_type_filter

    total_rows = len(elements)

    # --- Per-column fill-rate on full (filtered) set ----------------------
    columns_info: list[dict[str, Any]] = []
    for col in all_columns:
        non_null_values: list[Any] = []
        for el in elements:
            v = el.get(col)
            # Treat empty strings as missing too — consistent with missingno.
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            non_null_values.append(v)
        non_null_count = len(non_null_values)
        fill_rate = (non_null_count / total_rows) if total_rows else 0.0
        columns_info.append(
            {
                "name": col,
                "non_null_count": non_null_count,
                "fill_rate": round(fill_rate, 6),
                "dtype": _infer_dtype(non_null_values),
            }
        )

    # --- Column ordering --------------------------------------------------
    if sort == "fill_desc":
        columns_info.sort(key=lambda c: (-c["fill_rate"], c["name"].lower()))
    elif sort == "fill_asc":
        columns_info.sort(key=lambda c: (c["fill_rate"], c["name"].lower()))
    elif sort == "alpha_asc":
        columns_info.sort(key=lambda c: c["name"].lower())
    elif sort == "alpha_desc":
        columns_info.sort(key=lambda c: c["name"].lower(), reverse=True)

    # --- Row-sample + per-row completeness (capped at 1000) ---------------
    sampled = total_rows > _MISSINGNESS_SAMPLE_CAP
    if sampled:
        rng = _random.Random(_MISSINGNESS_SAMPLE_SEED)
        sample_indices = sorted(rng.sample(range(total_rows), _MISSINGNESS_SAMPLE_CAP))
        sample = [elements[i] for i in sample_indices]
    else:
        sample = elements

    col_count = len(all_columns) or 1
    row_completeness: list[float] = []
    for el in sample:
        filled = 0
        for col in all_columns:
            v = el.get(col)
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            filled += 1
        row_completeness.append(round(filled / col_count, 6))

    # --- Matrix presence bitmap (ordered to match `columns_info`) ---------
    # 1 = present, 0 = missing. Keeping this compact (int list) keeps the
    # response small — 1000 rows × ~50 columns ≈ 50k ints ≈ ~150kB JSON.
    ordered_col_names = [c["name"] for c in columns_info]
    presence_matrix: list[list[int]] = []
    for el in sample:
        row_bits: list[int] = []
        for col in ordered_col_names:
            v = el.get(col)
            if v is None or (isinstance(v, str) and v.strip() == ""):
                row_bits.append(0)
            else:
                row_bits.append(1)
        presence_matrix.append(row_bits)

    return {
        "total_rows": total_rows,
        "sampled_rows": len(sample),
        "sampled": sampled,
        "columns": columns_info,
        "row_completeness": row_completeness,
        "presence_matrix": presence_matrix,
        "applied_filters": applied_filters,
    }


@router.post(
    "/cad-data/value-counts/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_data_value_counts(
    body: CadDataValueCountsRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Return value counts for a single column, sorted by frequency descending.

    Audit B6 — was IDOR.
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]

    # Validate column exists
    all_columns = _collect_column_names(elements)
    if body.column not in all_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown column: '{body.column}'. Available: {sorted(all_columns)}",
        )

    # Count values
    freq: dict[str, int] = {}
    for el in elements:
        raw = el.get(body.column)
        key = str(raw) if raw is not None else "(null)"
        freq[key] = freq.get(key, 0) + 1

    total = len(elements)
    sorted_values = sorted(freq.items(), key=lambda kv: kv[1], reverse=True)

    values = [
        {
            "value": val,
            "count": cnt,
            "percentage": round(cnt / total * 100, 1) if total else 0,
        }
        for val, cnt in sorted_values[: body.limit]
    ]

    return {
        "column": body.column,
        "total": total,
        "values": values,
    }


@router.get(
    "/cad-data/elements/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_data_elements(
    session_id: str = Query(..., description="Session ID from /cad-columns"),
    offset: int = Query(default=0, ge=0, description="Number of rows to skip"),
    limit: int = Query(default=100, ge=1, le=1000, description="Max rows to return"),
    sort_by: str | None = Query(default=None, description="Column to sort by"),
    sort_order: str = Query(default="asc", pattern="^(asc|desc)$", description="Sort direction"),
    filter_column: str | None = Query(default=None, description="Column to filter on"),
    filter_value: str | None = Query(default=None, description="Value to match (equality)"),
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Return a paginated, sortable, filterable table of CAD elements.

    Audit B6 — was IDOR. Session IDs are UUIDs but anyone with
    ``takeoff.read`` could read another tenant's extracted CAD data
    (potentially confidential pricing or geometry) by guessing or
    scraping the id. We gate via the unified CAD-session helper.
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]
    all_columns = _collect_column_names(elements)

    # --- Filter ---
    if filter_column and filter_value is not None:
        if filter_column not in all_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown filter column: '{filter_column}'. Available: {sorted(all_columns)}",
            )
        elements = [el for el in elements if str(el.get(filter_column, "")) == filter_value]

    # --- Sort ---
    if sort_by is not None:
        if sort_by not in all_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown sort column: '{sort_by}'. Available: {sorted(all_columns)}",
            )
        reverse = sort_order == "desc"

        def _sort_key(el: dict) -> tuple:
            v = el.get(sort_by)
            if v is None:
                # None sorts last regardless of direction
                return (1, "")
            if _is_numeric(v):
                return (0, _to_float(v))
            return (0, str(v).lower())

        elements = sorted(elements, key=_sort_key, reverse=reverse)

    total = len(elements)
    page = elements[offset : offset + limit]

    return {
        "total": total,
        "offset": offset,
        "limit": limit,
        "columns": all_columns,
        "rows": page,
    }


_SUPPORTED_AGG_FUNCS = {"sum", "avg", "mean", "min", "max", "count"}


@router.post(
    "/cad-data/aggregate/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_data_aggregate(
    body: CadDataAggregateRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Aggregate CAD element data by grouping columns.

    Supported aggregation functions: sum, avg (alias: mean), min, max, count.
    ``count`` ignores the column values and counts elements in each group.

    Audit B6 — was IDOR. Same threat as ``cad_data_elements``: gate on
    the session's owning project (or uploader for standalone sessions).
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail=("Session not found or expired. Please re-upload the CAD file via POST /cad-columns."),
        )

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session["elements"]
    all_columns_set: set[str] = set()
    for el in elements:
        all_columns_set.update(el.keys())

    # Validate group_by columns
    missing_group = [c for c in body.group_by if c not in all_columns_set]
    if missing_group:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown grouping column(s): {missing_group}. Available: {sorted(all_columns_set)}",
        )

    # Validate aggregation specs
    for col, func in body.aggregations.items():
        func_lower = func.lower()
        if func_lower not in _SUPPORTED_AGG_FUNCS:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Unsupported aggregation function '{func}' for column '{col}'. "
                    f"Supported: {sorted(_SUPPORTED_AGG_FUNCS)}"
                ),
            )
        if func_lower != "count" and col not in all_columns_set:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown aggregation column: '{col}'. Available: {sorted(all_columns_set)}",
            )

    # --- Build groups ---
    groups_map: dict[tuple, list[dict]] = {}
    for el in elements:
        key = tuple(str(el.get(c, "")) for c in body.group_by)
        groups_map.setdefault(key, []).append(el)

    def _aggregate(vals: list[dict], col: str, func: str) -> float:
        func = func.lower()
        if func == "count":
            return len(vals)
        numeric = []
        for el in vals:
            v = el.get(col)
            if _is_numeric(v):
                numeric.append(_to_float(v))
        if not numeric:
            return 0.0
        if func == "sum":
            return round(sum(numeric), 4)
        if func in ("avg", "mean"):
            return round(_mean(numeric), 4)
        if func == "min":
            return round(min(numeric), 4)
        if func == "max":
            return round(max(numeric), 4)
        return 0.0

    result_groups: list[dict[str, Any]] = []
    for key_tuple, group_elements in groups_map.items():
        key_dict = {c: v for c, v in zip(body.group_by, key_tuple, strict=False)}
        results: dict[str, float] = {}
        for col, func in body.aggregations.items():
            results[col] = _aggregate(group_elements, col, func)
        result_groups.append({"key": key_dict, "results": results, "count": len(group_elements)})

    # Sort groups by first group_by column for stable output
    result_groups.sort(key=lambda g: tuple(g["key"].get(c, "") for c in body.group_by))

    # --- Compute totals across all elements ---
    totals: dict[str, float] = {}
    for col, func in body.aggregations.items():
        totals[col] = _aggregate(elements, col, func)

    return {
        "groups": result_groups,
        "totals": totals,
        "total_count": len(elements),
    }


# ── CAD Data Explorer: Session Management (save, list, delete) ─────────────


class CadDataSaveRequest(BaseModel):
    """Save a CAD session permanently to a project."""

    session_id: str
    project_id: str
    display_name: str


@router.post(
    "/cad-data/save/",
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def cad_data_save(
    body: CadDataSaveRequest,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Mark a CAD session as permanent and link it to a project.

    Audit B6 — IDOR-on-write. Two attack vectors closed:

    1. Source side — caller could pass a foreign tenant's session id
       and rehome it under their own project (data theft).
    2. Destination side — caller could pass another tenant's
       ``project_id`` and dump their session into it.

    We verify BOTH sides before the UPDATE.
    """
    _cleanup_memory_sessions()

    cad_session = await _get_cad_session(db_session, body.session_id)
    if not cad_session:
        raise HTTPException(status_code=404, detail="Session not found or expired.")

    # Source-side check (existing session ownership)
    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    # Destination-side check (target project)
    try:
        target_pid = _uuid.UUID(body.project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=400,
            detail="Invalid project_id (must be a UUID)",
        ) from exc
    await verify_project_access(target_pid, str(user_id) if user_id else "", db_session)

    # Update the database record
    from sqlalchemy import update as sa_update

    stmt = (
        sa_update(CadExtractionSession)
        .where(CadExtractionSession.session_id == body.session_id)
        .values(
            project_id=body.project_id,
            display_name=body.display_name,
            is_permanent=True,
            expires_at=datetime.now(UTC) + timedelta(days=365 * 10),
        )
    )
    await db_session.execute(stmt)
    await db_session.commit()

    return {
        "status": "saved",
        "session_id": body.session_id,
        "project_id": body.project_id,
        "display_name": body.display_name,
    }


@router.get(
    "/cad-data/sessions/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def cad_data_list_sessions(
    project_id: str | None = Query(default=None),
    saved_only: bool = Query(default=False),
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> list[dict[str, Any]]:
    """List CAD sessions. By default shows all non-expired. Use saved_only=true for permanent only.

    Audit B6 — was a global enumeration vulnerability. The endpoint
    returned every tenant's sessions to any authenticated user with
    ``takeoff.read`` (a cross-tenant data leak: filenames, project ids,
    element counts, timestamps). We now scope the query to sessions
    owned by the caller's user_id, OR to sessions in projects the
    caller actually owns. Admins see everything (matching
    ``verify_project_access`` admin-bypass semantics).
    """
    from sqlalchemy import or_, select

    from app.modules.users.repository import UserRepository

    now = datetime.now(UTC)
    stmt = select(CadExtractionSession)
    if saved_only:
        stmt = stmt.where(CadExtractionSession.is_permanent == True)  # noqa: E712
    else:
        # Show permanent + non-expired temporary
        stmt = stmt.where(
            (CadExtractionSession.is_permanent == True)  # noqa: E712
            | (CadExtractionSession.expires_at > now)
        )

    # Admin-aware tenant scoping. Non-admins only see their own
    # sessions or sessions inside projects they own. We resolve
    # the user's owned projects via a sub-query so this stays a
    # single round-trip.
    is_admin = False
    try:
        user_repo = UserRepository(db_session)
        if user_id:
            user_row = await user_repo.get_by_id(_uuid.UUID(str(user_id)))
            if user_row is not None and getattr(user_row, "role", "") == "admin":
                is_admin = True
    except Exception:
        logger.exception("Admin-role lookup failed in cad_data_list_sessions")

    if not is_admin:
        # Collect owned project ids (as strings — session model stores
        # project_id as String, not UUID).
        from app.modules.projects.models import Project as _Project

        owned_proj_stmt = select(_Project.id).where(_Project.owner_id == str(user_id))
        owned_projects = (await db_session.execute(owned_proj_stmt)).scalars().all()
        owned_project_ids = {str(pid) for pid in owned_projects}

        # Match: sessions I uploaded OR sessions in a project I own.
        ownership_filter = CadExtractionSession.user_id == str(user_id)
        if owned_project_ids:
            ownership_filter = or_(
                ownership_filter,
                CadExtractionSession.project_id.in_(owned_project_ids),
            )
        stmt = stmt.where(ownership_filter)

    if project_id:
        # Caller asked for one project — verify access first so the
        # error is symmetric with single-resource endpoints (404 on
        # foreign projects rather than empty list).
        try:
            req_pid = _uuid.UUID(project_id)
        except (ValueError, TypeError) as exc:
            raise HTTPException(
                status_code=400,
                detail="Invalid project_id (must be a UUID)",
            ) from exc
        await verify_project_access(req_pid, str(user_id) if user_id else "", db_session)
        stmt = stmt.where(CadExtractionSession.project_id == project_id)
    stmt = stmt.order_by(CadExtractionSession.created_at.desc())

    result = await db_session.execute(stmt)
    rows = result.scalars().all()

    return [
        {
            "session_id": row.session_id,
            "display_name": row.display_name or row.filename,
            "filename": row.filename,
            "file_format": row.file_format,
            "element_count": row.element_count,
            "extraction_time": row.extraction_time,
            "project_id": row.project_id,
            "is_permanent": bool(row.is_permanent),
            "created_at": row.created_at.isoformat() if row.created_at else None,
        }
        for row in rows
    ]


@router.delete(
    "/cad-data/sessions/{session_id}",
    dependencies=[Depends(RequirePermission("takeoff.delete"))],
    status_code=204,
)
async def cad_data_delete_session(
    session_id: str,
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> None:
    """Delete a saved CAD session.

    Audit B6 — was IDOR-on-write. Any user with ``takeoff.delete``
    could destroy another tenant's permanent extraction sessions
    by UUID. We resolve ownership through the standard helper
    before the DELETE.
    """
    from sqlalchemy import delete as sa_delete

    # Look up first (memory + DB) so we can apply the same ownership
    # check as every other CAD-session endpoint. ``_get_cad_session``
    # is TTL-aware but ``DELETE`` may legitimately target permanent
    # sessions that are still well within their lifetime — they will
    # be returned by the lookup.
    cad_session = await _get_cad_session(db_session, session_id)
    if not cad_session:
        # Match the historic "Session not found." string (some callers
        # branch on it) instead of the more verbose memory message.
        raise HTTPException(status_code=404, detail="Session not found.")

    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)

    stmt = sa_delete(CadExtractionSession).where(CadExtractionSession.session_id == session_id)
    result = await db_session.execute(stmt)
    await db_session.commit()

    if result.rowcount == 0:  # type: ignore[union-attr]
        raise HTTPException(status_code=404, detail="Session not found.")

    # Also remove from memory cache
    _cad_sessions.pop(session_id, None)


# ── Save CAD session to project as BIM model ────────────────────────────


class SaveToProjectRequest(BaseModel):
    """Request body for saving a takeoff session to a project as a BIM model."""

    model_name: str = Field(default="Imported from Takeoff", max_length=255)


@router.post(
    "/sessions/{session_id}/save-to-project/",
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def save_session_to_project(
    session_id: str,
    body: SaveToProjectRequest,
    project_id: str = Query(..., description="Target project UUID"),
    db_session: SessionDep = None,  # type: ignore[assignment]
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Convert a takeoff session's extracted elements into a persistent BIM model.

    Creates a BIMModel + BIMElements from the session's extraction results.
    The session itself remains (not deleted). Requires the bim_hub module
    to be loaded; returns a clear error if it is not available.
    """
    import uuid as _uuid_mod

    # 1. Load the takeoff session
    _cleanup_memory_sessions()
    cad_session = await _get_cad_session(db_session, session_id)
    if not cad_session:
        raise HTTPException(
            status_code=404,
            detail="CAD session not found or expired. Please re-upload the file.",
        )

    # Audit B6 — IDOR-on-write on both sides:
    # - source: prevents stealing a foreign session into your project
    # - destination: prevents writing into a foreign project
    await _verify_cad_session_access(cad_session, str(user_id) if user_id else "", db_session)
    try:
        target_pid = _uuid_mod.UUID(project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=400,
            detail="Invalid project_id (must be a UUID)",
        ) from exc
    await verify_project_access(target_pid, str(user_id) if user_id else "", db_session)

    elements: list[dict] = cad_session.get("elements", [])
    if not elements:
        raise HTTPException(
            status_code=400,
            detail="Session contains no elements to save.",
        )

    # 2. Check that the bim_hub module models are available
    try:
        from app.modules.bim_hub.models import BIMElement, BIMModel
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail=(
                "BIM Hub module is not loaded. Cannot save takeoff session "
                "as a BIM model. Enable the bim_hub module and try again."
            ),
        )

    # 3. Create BIMModel record
    project_uuid = target_pid
    filename = cad_session.get("filename", "unknown")
    file_format = cad_session.get("format", "")

    bim_model = BIMModel(
        project_id=project_uuid,
        name=body.model_name,
        discipline="general",
        model_format=file_format,
        version="1",
        status="ready",
        element_count=len(elements),
        created_by=_uuid_mod.UUID(user_id) if user_id else None,
        metadata_={
            "source": "takeoff_session",
            "takeoff_session_id": session_id,
            "original_filename": filename,
        },
    )
    db_session.add(bim_model)
    await db_session.flush()  # Get the model ID

    # 4. Create BIMElement records from extraction data
    element_count = 0
    for idx, el in enumerate(elements):
        # Build quantities dict from numeric fields
        quantities: dict[str, float] = {}
        for key in ("volume", "area", "length", "gross volume", "gross area", "count"):
            val = el.get(key)
            if val is not None:
                try:
                    quantities[key] = float(val)
                except (ValueError, TypeError):
                    pass

        # Build properties from non-numeric fields
        properties: dict[str, str] = {}
        for key, val in el.items():
            if key in ("id", "volume", "area", "length", "gross volume", "gross area", "count"):
                continue
            if val is not None:
                properties[key] = str(val)

        bim_element = BIMElement(
            model_id=bim_model.id,
            stable_id=str(el.get("id", f"el_{idx}")),
            element_type=str(el.get("category", el.get("type name", ""))),
            name=str(el.get("type name", el.get("family", f"Element {idx + 1}"))),
            storey=str(el.get("level", el.get("storey", ""))) or None,
            discipline="general",
            properties=properties,
            quantities=quantities,
            metadata_={"source_index": idx},
        )
        db_session.add(bim_element)
        element_count += 1

    # 5. Update the CAD session to mark it as persistent and linked
    from sqlalchemy import update as sa_update

    stmt = (
        sa_update(CadExtractionSession)
        .where(CadExtractionSession.session_id == session_id)
        .values(
            is_persistent=True,
            bim_model_id=str(bim_model.id),
            project_id=project_id,
        )
    )
    await db_session.execute(stmt)
    await db_session.flush()

    logger.info(
        "Saved takeoff session %s to project %s as BIM model %s (%d elements)",
        session_id,
        project_id,
        bim_model.id,
        element_count,
    )

    return {
        "model_id": str(bim_model.id),
        "element_count": element_count,
        "model_name": body.model_name,
        "project_id": project_id,
    }


MAX_PDF_SIZE = 50 * 1024 * 1024  # 50 MB


def _get_service(session: SessionDep) -> TakeoffService:
    return TakeoffService(session)


# ── Upload ────────────────────────────────────────────────────────────────


@router.post(
    "/documents/upload/",
    status_code=201,
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def upload_document(
    user_id: CurrentUserId,
    file: UploadFile = File(..., description="PDF file (.pdf)"),
    project_id: str | None = Query(default=None),
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Upload a PDF document for quantity takeoff.

    Audit B5 — was IDOR-on-write. ``project_id`` came in as a free-form
    query string and was persisted verbatim, so anyone with
    ``takeoff.create`` could attach a PDF to another tenant's project
    (and trigger the Documents-hub cross-link). We resolve and verify
    access *before* any disk write or DB insert.
    """
    # Verify access first — fail fast, before reading the upload body
    # into memory or hitting disk. We tolerate ``None`` for legacy
    # standalone uploads but require a valid UUID when present.
    verified_pid: _uuid.UUID | None = None
    if project_id:
        try:
            verified_pid = _uuid.UUID(project_id)
        except (ValueError, TypeError) as exc:
            raise HTTPException(
                status_code=400,
                detail="Invalid project_id (must be a UUID)",
            ) from exc
        await verify_project_access(verified_pid, str(user_id), session)

    allowed, _ = upload_limiter.is_allowed(str(user_id))
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail="Too many uploads. Please wait a moment and try again.",
            headers={"Retry-After": "60"},
        )
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext != "pdf":
        raise HTTPException(
            status_code=400,
            detail=f"Only PDF files are supported, got .{ext}",
        )

    content = await file.read()

    # Magic byte check — every legitimate PDF starts with "%PDF-".
    # Block JPGs/HTML/other files that have been renamed to .pdf to bypass
    # the extension check (security finding from QA report).
    if not content.startswith(b"%PDF-"):
        raise HTTPException(
            status_code=400,
            detail="File does not appear to be a valid PDF (missing %PDF- header)",
        )

    doc = await service.upload_document(
        filename=file.filename,
        content=content,
        size_bytes=len(content),
        owner_id=user_id,
        project_id=project_id,
    )

    # Cross-link: create Document record so takeoff PDFs appear in
    # Documents hub.  Uses the ORM Document model directly so the row
    # picks up timestamps + defaults from the Base mixin and stays in
    # sync with any future schema migration.  Best-effort: failure
    # here MUST NOT break the upload — the takeoff doc is already
    # persisted via service.upload_document().
    if project_id:
        try:
            from app.modules.documents.models import Document

            xlink_doc = Document(
                project_id=_uuid.UUID(project_id),
                name=file.filename,
                description="Takeoff document",
                category="drawing",
                file_size=len(content),
                mime_type="application/pdf",
                file_path=doc.file_path or "",
                version=1,
                uploaded_by=user_id or "",
                tags=["takeoff", "pdf"],
                metadata_={
                    "source_module": "takeoff",
                    "source_id": str(doc.id),
                },
            )
            service.session.add(xlink_doc)
            await service.session.flush()
            logger.info("Cross-linked takeoff doc %s -> document %s", doc.id, xlink_doc.id)
        except Exception:
            logger.exception("Failed to cross-link takeoff document to Documents hub")

    return {
        "id": str(doc.id),
        "filename": doc.filename,
        "pages": doc.pages,
        "size_bytes": doc.size_bytes,
    }


# ── Access helper (Audit B5) ──────────────────────────────────────────────


async def _verify_takeoff_doc_access(
    doc: Any,
    user_id: str,
    session: Any,
) -> None:
    """Gate access to a TakeoffDocument by tenant.

    Audit B5 — used by ``get_document`` / ``extract_tables`` /
    ``download_document`` / ``analyze_document`` to enforce IDOR
    protection. Two cases:

    1. Document is bound to a project (``project_id`` non-empty) —
       reuse ``verify_project_access`` which also handles admin bypass.
    2. Standalone upload (no project) — only the original uploader
       can touch it. We compare by string to tolerate UUID-vs-str
       drift coming from older rows.

    We return 404 (not 403) on access failures to avoid leaking
    document existence to attackers probing UUIDs.
    """
    raw_pid = getattr(doc, "project_id", None)
    if raw_pid:
        try:
            pid = _uuid.UUID(str(raw_pid))
        except (ValueError, TypeError):
            pid = None
        if pid is not None:
            await verify_project_access(pid, str(user_id), session)
            return

    # Standalone — owner-only. Admins still pass via the project-bound
    # path above; here we have no project, so we fall through to a
    # strict owner-match.
    #
    # TakeoffDocument uses ``owner_id`` (a UUID column) while the
    # CadExtractionSession sibling uses ``user_id`` (string). Try
    # both names so a legacy row layout doesn't bypass the gate.
    owner = str(
        getattr(doc, "owner_id", None) or getattr(doc, "user_id", "") or ""
    )
    if owner and owner != str(user_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Document not found",
        )


# ── List documents ────────────────────────────────────────────────────────


@router.get(
    "/documents/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def list_documents(
    user_id: CurrentUserId,
    project_id: str | None = Query(default=None),
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> list[dict[str, Any]]:
    """List uploaded takeoff documents.

    Audit B5 — when filtered by ``project_id`` we additionally verify
    project access so the caller cannot enumerate another tenant's
    documents by guessing the project UUID. When unfiltered the
    underlying ``service.list_documents`` already scopes by
    ``user_id`` so per-tenant isolation holds.
    """
    if project_id:
        try:
            pid = _uuid.UUID(project_id)
        except (ValueError, TypeError) as exc:
            raise HTTPException(
                status_code=400,
                detail="Invalid project_id (must be a UUID)",
            ) from exc
        await verify_project_access(pid, str(user_id), session)
    docs = await service.list_documents(user_id, project_id=project_id)
    return [
        {
            "id": str(d.id),
            "filename": d.filename,
            "pages": d.pages,
            "size_bytes": d.size_bytes,
            "status": d.status,
            "uploaded_at": d.created_at.isoformat() if d.created_at else None,
        }
        for d in docs
    ]


# ── Get single document ──────────────────────────────────────────────────


@router.get(
    "/documents/{doc_id}",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def get_document(
    doc_id: str,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Get a single takeoff document with its data.

    Audit B5 — was IDOR. The endpoint returned the document blindly
    by UUID, exposing extracted text, page data and AI analysis of a
    foreign tenant's PDF. We gate on the owning project (when bound)
    or fall back to the row's ``user_id`` for standalone uploads.
    """
    doc = await service.get_document(doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")

    await _verify_takeoff_doc_access(doc, str(user_id) if user_id else "", session)

    return {
        "id": str(doc.id),
        "filename": doc.filename,
        "pages": doc.pages,
        "size_bytes": doc.size_bytes,
        "status": doc.status,
        "extracted_text": doc.extracted_text[:2000] if doc.extracted_text else "",
        "page_data": doc.page_data,
        "analysis": doc.analysis,
        "uploaded_at": doc.created_at.isoformat() if doc.created_at else None,
    }


# ── Extract tables ────────────────────────────────────────────────────────


@router.post(
    "/documents/{doc_id}/extract-tables/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def extract_tables(
    doc_id: str,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> dict[str, Any]:
    """Extract tabular data from an uploaded document.

    Audit B5 — was IDOR. Anyone with ``takeoff.read`` could trigger
    table extraction on a foreign tenant's PDF and read the result.
    Gate access through the standard helper before any extraction
    work is dispatched.
    """
    doc = await service.get_document(doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")

    await _verify_takeoff_doc_access(doc, str(user_id) if user_id else "", session)

    return await service.extract_tables(doc_id)


# ── Download stored PDF ─────────────────────────────────────────────────


@router.get(
    "/documents/{doc_id}/download/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def download_document(
    doc_id: str,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> FileResponse:
    """Download the stored PDF file for a takeoff document.

    Audit B5 — was IDOR. Anyone with ``takeoff.read`` could fetch the
    raw PDF bytes of a foreign tenant's drawing by guessing the doc
    UUID. Gate access through the standard helper before serving the
    file (which would otherwise stream the whole PDF off disk).
    """
    doc = await service.get_document(doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")

    await _verify_takeoff_doc_access(doc, str(user_id) if user_id else "", session)

    if not doc.file_path:
        raise HTTPException(status_code=404, detail="PDF file not available for this document")

    file_path = Path(doc.file_path).resolve()

    # Security: ensure resolved path is within the takeoff upload directory
    allowed_base = (Path.home() / ".openestimator").resolve()
    if not str(file_path).startswith(str(allowed_base)):
        raise HTTPException(status_code=403, detail="Access denied")

    if not file_path.exists() or file_path.is_symlink():
        raise HTTPException(status_code=404, detail="PDF file not found on disk")

    return FileResponse(
        path=str(file_path),
        media_type="application/pdf",
        filename=doc.filename,
    )


# ── AI Analysis ──────────────────────────────────────────────────────────


@router.post(
    "/documents/{doc_id}/analyze/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def analyze_document(
    doc_id: str,
    user_id: CurrentUserId,
    session: SessionDep,
    service: TakeoffService = Depends(_get_service),
) -> dict[str, Any]:
    """Analyze a takeoff document's extracted text using AI.

    Sends the document's previously extracted text to the configured AI provider
    and returns structured BOQ items parsed from the AI response.
    """
    import time
    import uuid as _uuid

    from app.modules.ai.ai_client import call_ai, extract_json, resolve_provider_and_key
    from app.modules.ai.prompts import (
        SMART_IMPORT_PROMPT,
        SYSTEM_PROMPT,
        USER_FENCE_MAX_LEN,
        fence_user_content,
    )
    from app.modules.ai.repository import AISettingsRepository

    # 1. Get the document
    doc = await service.get_document(doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")

    # Audit B5 — IDOR. AI analysis dispatches the PDF text to a third
    # party LLM and bills tokens; without this check, any user could
    # exfiltrate (and rack up bills against) another tenant's PDFs.
    await _verify_takeoff_doc_access(doc, str(user_id) if user_id else "", session)

    # 2. Check extracted text
    extracted_text = doc.extracted_text or ""
    if not extracted_text.strip():
        raise HTTPException(
            status_code=400,
            detail="Document has no extracted text. Please re-upload the PDF.",
        )

    # 3. Get user AI settings and resolve provider
    settings_repo = AISettingsRepository(session)
    settings = await settings_repo.get_by_user_id(_uuid.UUID(user_id))

    try:
        provider, api_key = resolve_provider_and_key(settings)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=str(exc),
        ) from exc

    # 4. Build prompt
    filename = doc.filename or "document.pdf"
    # Audit AI1 — fence the untrusted document text so prompt-injection
    # attempts embedded in the PDF can't override the system prompt.
    # ``fence_user_content`` also applies the length cap so we no longer
    # need a manual slice here.
    text_for_prompt = fence_user_content(extracted_text, max_len=USER_FENCE_MAX_LEN)
    prompt = SMART_IMPORT_PROMPT.format(filename=filename, text=text_for_prompt)

    # 5. Call AI
    start_time = time.monotonic()
    try:
        raw_response, tokens = await call_ai(
            provider=provider,
            api_key=api_key,
            system=SYSTEM_PROMPT,
            prompt=prompt,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("AI analysis failed for document %s: %s", doc_id, exc)
        raise HTTPException(
            status_code=502,
            detail=f"AI analysis failed: {exc}",
        ) from exc

    duration_ms = int((time.monotonic() - start_time) * 1000)

    # 6. Parse AI response
    parsed = extract_json(raw_response)
    if not isinstance(parsed, list):
        parsed = []

    # 7. Convert to the AnalysisResult format expected by frontend
    elements: list[dict[str, Any]] = []
    categories: dict[str, dict[str, Any]] = {}

    # Audit AI3 — validate numeric fields. The LLM occasionally hallucinates
    # negative rates ("rebate items") or absurd values; we clamp to a
    # plausible band so a malicious / confused response can't pollute the
    # BOQ. Sane upper bounds: quantity 10_000_000 (1M m² building),
    # unit_rate 1_000_000 currency units (luxury fitout per m²).
    _MAX_QUANTITY = 10_000_000.0
    _MAX_UNIT_RATE = 1_000_000.0

    for idx, item in enumerate(parsed):
        if not isinstance(item, dict):
            continue

        description = str(item.get("description", "")).strip()
        if len(description) < 3:
            continue

        try:
            quantity = float(item.get("quantity", 0))
        except (ValueError, TypeError):
            quantity = 0.0
        # Negative quantities are nonsense; outside-band values are
        # almost certainly hallucinations.
        if quantity < 0 or quantity > _MAX_QUANTITY:
            logger.warning(
                "AI returned out-of-band quantity %s for item %d — clamping to 0",
                quantity,
                idx,
            )
            quantity = 0.0

        unit = str(item.get("unit", "pcs")).strip() or "pcs"
        category = str(item.get("category", "General")).strip() or "General"

        try:
            unit_rate = float(item.get("unit_rate", 0))
        except (ValueError, TypeError):
            unit_rate = 0.0
        if unit_rate < 0 or unit_rate > _MAX_UNIT_RATE:
            logger.warning(
                "AI returned out-of-band unit_rate %s for item %d — clamping to 0",
                unit_rate,
                idx,
            )
            unit_rate = 0.0

        element = {
            "id": f"ai_{idx + 1}",
            "category": category,
            "description": description,
            "quantity": round(quantity, 2),
            "unit": unit,
            "confidence": 0.8,
        }
        elements.append(element)

        # Build category summary
        if category not in categories:
            categories[category] = {"count": 0, "total_quantity": 0, "unit": unit}
        categories[category]["count"] += 1
        categories[category]["total_quantity"] += quantity

    logger.info(
        "AI analysis completed: doc=%s, items=%d, tokens=%d, duration=%dms",
        doc_id,
        len(elements),
        tokens,
        duration_ms,
    )

    return {
        "elements": elements,
        "summary": {
            "total_elements": len(elements),
            "categories": categories,
        },
    }


# ── Delete ────────────────────────────────────────────────────────────────


@router.delete(
    "/documents/{doc_id}",
    dependencies=[Depends(RequirePermission("takeoff.delete"))],
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_document(
    doc_id: str,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
) -> None:
    """Delete an uploaded takeoff document.

    Audit B5 — IDOR. Reuses the unified takeoff-doc access helper so
    standalone uploads are owner-locked too (previous code let any
    user with ``takeoff.delete`` blow away orphan rows).
    """
    doc = await service.get_document(doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")

    await _verify_takeoff_doc_access(doc, str(user_id) if user_id else "", session)

    await service.delete_document(doc_id)


# ═══════════════════════════════════════════════════════════════════════════
# Measurement endpoints
# ═══════════════════════════════════════════════════════════════════════════


def _measurement_to_response(item: object) -> TakeoffMeasurementResponse:
    """Build a TakeoffMeasurementResponse from an ORM object."""
    return TakeoffMeasurementResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        document_id=item.document_id,  # type: ignore[attr-defined]
        page=item.page,  # type: ignore[attr-defined]
        type=item.type,  # type: ignore[attr-defined]
        group_name=item.group_name,  # type: ignore[attr-defined]
        group_color=item.group_color,  # type: ignore[attr-defined]
        annotation=item.annotation,  # type: ignore[attr-defined]
        points=item.points or [],  # type: ignore[attr-defined]
        measurement_value=item.measurement_value,  # type: ignore[attr-defined]
        measurement_unit=item.measurement_unit,  # type: ignore[attr-defined]
        depth=item.depth,  # type: ignore[attr-defined]
        volume=item.volume,  # type: ignore[attr-defined]
        perimeter=item.perimeter,  # type: ignore[attr-defined]
        count_value=item.count_value,  # type: ignore[attr-defined]
        scale_pixels_per_unit=item.scale_pixels_per_unit,  # type: ignore[attr-defined]
        linked_boq_position_id=item.linked_boq_position_id,  # type: ignore[attr-defined]
        metadata=getattr(item, "metadata_", {}),  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
    )


# ── Summary (must be before /{measurement_id} to avoid route collision) ──


@router.get(
    "/measurements/summary/",
    response_model=TakeoffMeasurementSummary,
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def measurement_summary(
    project_id: _uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> TakeoffMeasurementSummary:
    """Aggregated measurement stats for a project.

    Audit B4 — was IDOR. Any authenticated user with ``takeoff.read``
    could request another tenant's ``project_id`` and read aggregated
    counts (which leaks both the project's existence and its volume of
    work). Gated by ``verify_project_access`` so foreign projects 404.
    """
    await verify_project_access(project_id, str(user_id), session)
    data = await service.get_measurement_summary(project_id)
    return TakeoffMeasurementSummary(**data)


# ── Export ───────────────────────────────────────────────────────────────


@router.get(
    "/measurements/export/",
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def export_measurements(
    project_id: _uuid.UUID = Query(...),
    format: str = Query(default="csv", pattern="^(csv|json)$"),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> Any:
    """Export measurements for a project.

    Supported formats: csv, json.
    CSV returns a downloadable text response; JSON returns a list of dicts.

    Audit B4 — was IDOR. Without the access check, any authenticated user
    with ``takeoff.read`` could download another tenant's measurements as
    CSV/JSON. Gated by ``verify_project_access`` so foreign projects 404.
    """
    await verify_project_access(project_id, str(user_id), session)
    rows = await service.export_measurements(project_id, fmt=format)

    if format == "csv":
        import csv
        import io

        if not rows:
            return {"csv": "", "count": 0}

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        # Neutralise any string cell that would otherwise be parsed by Excel as
        # a formula (BUG-CSV-INJECTION). Numeric values pass through unchanged.
        safe_rows = [{k: neutralise_formula(v) for k, v in r.items()} for r in rows]
        writer.writerows(safe_rows)
        csv_text = output.getvalue()
        return {"csv": csv_text, "count": len(rows)}

    return {"measurements": rows, "count": len(rows)}


# ── Bulk create ──────────────────────────────────────────────────────────


@router.post(
    "/measurements/bulk/",
    response_model=list[TakeoffMeasurementResponse],
    status_code=201,
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def bulk_create_measurements(
    data: TakeoffMeasurementBulkCreate,
    user_id: CurrentUserId,
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> list[TakeoffMeasurementResponse]:
    """Bulk create measurements (e.g. importing from localStorage).

    Audit B4 — was IDOR-on-write. A user could pass arbitrary
    ``project_id`` values inside each payload and seed measurements
    into a foreign project. We verify access for every distinct
    ``project_id`` present in the batch up-front (one DB lookup per
    project, not per measurement) before any rows are written.
    """
    if not data.measurements:
        return []

    # Collect the unique project IDs touched by this batch. A single
    # bulk import may cover one project (the common case) or several
    # (cross-project paste from another window) — both must be gated.
    project_ids = {m.project_id for m in data.measurements}
    for pid in project_ids:
        await verify_project_access(pid, str(user_id), session)

    try:
        items = await service.bulk_create_measurements(data.measurements, created_by=user_id)
        return [_measurement_to_response(i) for i in items]
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed to bulk create measurements")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to bulk create measurements",
        )


# ── Create ───────────────────────────────────────────────────────────────


@router.post(
    "/measurements/",
    response_model=TakeoffMeasurementResponse,
    status_code=201,
    dependencies=[Depends(RequirePermission("takeoff.create"))],
)
async def create_measurement(
    data: TakeoffMeasurementCreate,
    user_id: CurrentUserId,
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> TakeoffMeasurementResponse:
    """Create a new takeoff measurement.

    Audit B4 — was IDOR-on-write. ``project_id`` in the body was trusted
    blindly, so any user with ``takeoff.create`` could pin a row to
    another tenant's project. We gate writes through
    ``verify_project_access`` first.
    """
    await verify_project_access(data.project_id, str(user_id), session)
    try:
        item = await service.create_measurement(data, created_by=user_id)
        return _measurement_to_response(item)
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed to create measurement")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create measurement",
        )


# ── List ─────────────────────────────────────────────────────────────────


@router.get(
    "/measurements/",
    response_model=list[TakeoffMeasurementResponse],
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def list_measurements(
    project_id: _uuid.UUID = Query(...),
    document_id: str | None = Query(default=None),
    page: int | None = Query(default=None, ge=1),
    group: str | None = Query(default=None),
    type: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=200, ge=1, le=500),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> list[TakeoffMeasurementResponse]:
    """List measurements for a project with optional filters.

    Audit B4 — was IDOR. Any authenticated user with ``takeoff.read``
    could supply another tenant's ``project_id`` and exfiltrate their
    measurements. Gated here through ``verify_project_access`` so
    foreign project ids 404 the same way a missing one does.
    """
    await verify_project_access(project_id, str(user_id), session)
    items = await service.list_measurements(
        project_id,
        document_id=document_id,
        page=page,
        group_name=group,
        measurement_type=type,
        offset=offset,
        limit=limit,
    )
    return [_measurement_to_response(i) for i in items]


# ── Get single ───────────────────────────────────────────────────────────


@router.get(
    "/measurements/{measurement_id}",
    response_model=TakeoffMeasurementResponse,
    dependencies=[Depends(RequirePermission("takeoff.read"))],
)
async def get_measurement(
    measurement_id: _uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> TakeoffMeasurementResponse:
    """Get a single measurement by ID.

    Audit B4 — was IDOR. Anyone with ``takeoff.read`` could guess a
    measurement UUID (or scrape one from a leaked log line) and read
    a foreign tenant's row. We resolve the owning project from the
    measurement itself and gate via ``verify_project_access`` so the
    response is identical to "measurement not found".
    """
    item = await service.get_measurement(measurement_id)
    await verify_project_access(item.project_id, str(user_id), session)
    return _measurement_to_response(item)


# ── Update ───────────────────────────────────────────────────────────────


@router.patch(
    "/measurements/{measurement_id}",
    response_model=TakeoffMeasurementResponse,
    dependencies=[Depends(RequirePermission("takeoff.update"))],
)
async def update_measurement(
    measurement_id: _uuid.UUID,
    data: TakeoffMeasurementUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> TakeoffMeasurementResponse:
    """Update a measurement.

    Audit B4 — was IDOR-on-write. Resolve the owning project from the
    target row and gate via ``verify_project_access`` before any
    mutation. We check ownership *before* calling the update service
    so we don't leak existence via different error codes.
    """
    existing = await service.get_measurement(measurement_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    item = await service.update_measurement(measurement_id, data)
    return _measurement_to_response(item)


# ── Delete ───────────────────────────────────────────────────────────────


@router.delete(
    "/measurements/{measurement_id}",
    status_code=204,
    dependencies=[Depends(RequirePermission("takeoff.delete"))],
)
async def delete_measurement(
    measurement_id: _uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> None:
    """Delete a measurement.

    Audit B4 — was IDOR-on-write. Without the owner check, any user with
    ``takeoff.delete`` could destroy another tenant's measurements by
    UUID. We resolve the owning project from the target row and gate
    via ``verify_project_access``.
    """
    existing = await service.get_measurement(measurement_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    await service.delete_measurement(measurement_id)


# ── Link to BOQ ──────────────────────────────────────────────────────────


@router.post(
    "/measurements/{measurement_id}/link-to-boq/",
    response_model=TakeoffMeasurementResponse,
    dependencies=[Depends(RequirePermission("takeoff.update"))],
)
async def link_measurement_to_boq(
    measurement_id: _uuid.UUID,
    data: LinkToBoqRequest,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TakeoffService = Depends(_get_service),
    session: SessionDep = None,  # type: ignore[assignment]
) -> TakeoffMeasurementResponse:
    """Link a measurement to a BOQ position.

    Audit B4 — was IDOR-on-write. A user could redirect a foreign
    tenant's measurement at their own BOQ position (or vice versa)
    without permission on the measurement side. Gate on the
    measurement's owning project before performing the link.
    """
    existing = await service.get_measurement(measurement_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    item = await service.link_measurement_to_boq(measurement_id, data.boq_position_id)
    return _measurement_to_response(item)
