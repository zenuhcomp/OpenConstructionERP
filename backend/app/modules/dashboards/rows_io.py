"""Tabular Data I/O for snapshots (T06).

Three things live here:

1. :func:`read_rows` — paginated row reader against the snapshot's
   entities Parquet. Used by the ``GET /snapshots/{id}/rows`` endpoint
   and re-used by the export path so a "filtered query → export"
   round-trip can't drift.

2. :func:`export_rows` — streams the same query out as CSV / XLSX /
   Parquet. CSV uses the stdlib ``csv`` writer (zero deps); XLSX uses
   ``openpyxl`` (already in base deps); Parquet uses ``pyarrow``.

3. :class:`SnapshotImportStaging` — in-memory staging area for the
   two-step import path. ``preview()`` parses + validates; ``commit()``
   materialises (today: just returns the row count — actual mutation
   into the snapshot is a T06 follow-up; the ADR-tracked path requires
   a snapshot-level "supplementary data" attachment that doesn't yet
   exist in the schema).
"""

from __future__ import annotations

import csv
import io
import json
import logging
import secrets
import time
import uuid
from dataclasses import dataclass
from typing import Any

logger = logging.getLogger(__name__)

# In-process staging for two-step imports.
#
# A real deployment would back this with Redis + S3 so the preview
# survives a worker restart. For T06 today, in-memory is enough — the
# UI flow (preview → review → commit) happens inside one user session,
# and the tests cover the round-trip.
_STAGING: dict[str, _StagedImport] = {}
_STAGING_TTL_SECONDS = 60 * 60  # one hour


@dataclass
class _StagedImport:
    snapshot_id: uuid.UUID
    columns: list[str]
    rows: list[dict[str, Any]]
    expires_at: float


# ── Errors ─────────────────────────────────────────────────────────────────


class RowsIOError(Exception):
    """Base class for tabular-IO errors."""

    http_status: int = 500
    message_key: str = "common.unknown_error"

    def __init__(self, message: str, *, details: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.details = details or {}


class UnsupportedFormatError(RowsIOError):
    http_status = 422
    message_key = "export.format.unsupported"


class InvalidQueryError(RowsIOError):
    http_status = 422
    message_key = "rows.query.invalid"


class StagingNotFoundError(RowsIOError):
    http_status = 404
    message_key = "import.staging.not_found"


class ImportSchemaError(RowsIOError):
    http_status = 422
    message_key = "import.schema.mismatch"


# ── Helpers shared by read + export ────────────────────────────────────────


def _resolve_filters(filters: str | dict | None) -> dict[str, Any]:
    """Coerce a filter parameter into a dict[str, list[str]]-ish shape."""
    if filters is None or filters == "":
        return {}
    if isinstance(filters, dict):
        return filters
    if isinstance(filters, str):
        try:
            parsed = json.loads(filters)
        except json.JSONDecodeError as exc:
            raise InvalidQueryError(
                f"filters is not valid JSON: {exc}",
                details={"filters": filters[:200]},
            ) from exc
        if not isinstance(parsed, dict):
            raise InvalidQueryError(
                "filters must decode to a JSON object.",
                details={"got_type": type(parsed).__name__},
            )
        return parsed
    raise InvalidQueryError(
        "filters must be a JSON string or dict.",
        details={"got_type": type(filters).__name__},
    )


def _resolve_columns(columns: str | list[str] | None) -> list[str] | None:
    """``"a,b,c"`` → ``["a", "b", "c"]``. Whitespace-trimmed; empties dropped."""
    if columns is None:
        return None
    if isinstance(columns, list):
        out = [str(c).strip() for c in columns if str(c).strip()]
        return out or None
    if isinstance(columns, str):
        out = [c.strip() for c in columns.split(",") if c.strip()]
        return out or None
    raise InvalidQueryError("columns must be a string or list.")


def _resolve_order_by(order_by: str | None) -> tuple[str, str] | None:
    """``"col:asc"`` or ``"col:desc"`` → ``(col, "ASC"|"DESC")``.

    Bare ``"col"`` (no direction) defaults to ASC. Direction is
    case-insensitive. Anything else raises :class:`InvalidQueryError`
    so accidental typos surface immediately.
    """
    if not order_by:
        return None
    parts = order_by.split(":", 1)
    col = parts[0].strip()
    if not col:
        raise InvalidQueryError(
            "order_by column cannot be empty.",
            details={"order_by": order_by},
        )
    direction = (parts[1].strip().lower() if len(parts) == 2 else "asc")
    if direction not in ("asc", "desc"):
        raise InvalidQueryError(
            "order_by direction must be 'asc' or 'desc'.",
            details={"direction": direction},
        )
    return (col, direction.upper())


def _is_safe_identifier(name: str) -> bool:
    """Defensive check — only allow [A-Za-z0-9_.-] in column names so a
    user-supplied column can be quoted into SQL without injection
    surface. DuckDB's identifier quoting (\"col\") still needs the
    underlying value to be reasonable — control chars are out either
    way.
    """
    if not name:
        return False
    return all(ch.isalnum() or ch in "_-." for ch in name)


def _quote_ident(name: str) -> str:
    """DuckDB identifier quoting — embed double quotes by doubling."""
    if not _is_safe_identifier(name):
        raise InvalidQueryError(
            f"column name contains unsupported characters: {name!r}",
            details={"column": name},
        )
    return '"' + name.replace('"', '""') + '"'


# ── Read rows ──────────────────────────────────────────────────────────────


@dataclass
class RowsResult:
    columns: list[str]
    rows: list[dict[str, Any]]
    total: int


async def read_rows(
    *,
    pool: Any,
    snapshot_id: uuid.UUID | str,
    project_id: uuid.UUID | str,
    columns: str | list[str] | None = None,
    filters: str | dict | None = None,
    order_by: str | None = None,
    limit: int = 100,
    offset: int = 0,
) -> RowsResult:
    """Run a paginated, filtered SELECT against a snapshot's entities view.

    Driven by :class:`DuckDBPool` — the pool already registers the
    snapshot's Parquet file as the ``entities`` view, so we just compose
    the SQL on top.
    """
    if limit <= 0:
        raise InvalidQueryError("limit must be positive.")
    if offset < 0:
        raise InvalidQueryError("offset must be >= 0.")
    limit = min(limit, 5000)

    cols = _resolve_columns(columns)
    flt = _resolve_filters(filters)
    order = _resolve_order_by(order_by)

    # Discover the actual column set in the Parquet so we (a) project a
    # known list back to the caller and (b) verify selected columns
    # exist before issuing the data query.
    schema_rows = await pool.execute(
        snapshot_id, project_id, "DESCRIBE entities", parameters=[],
    )
    all_columns = [str(r[0]) for r in schema_rows]
    if not all_columns:
        return RowsResult(columns=[], rows=[], total=0)

    if cols:
        unknown = [c for c in cols if c not in all_columns]
        if unknown:
            raise InvalidQueryError(
                f"unknown columns in selection: {unknown}",
                details={"unknown": unknown, "available": all_columns},
            )
        projected = cols
    else:
        projected = all_columns

    # WHERE clause: each filter is column → list[value]. Empty list ⇒
    # drop the predicate (matches T04 behaviour). Unknown columns get a
    # 422 so silent typos can't return whole-table results.
    where_parts: list[str] = []
    where_params: list[Any] = []
    for col, raw_values in flt.items():
        if col not in all_columns:
            raise InvalidQueryError(
                f"unknown column in filters: {col}",
                details={"column": col, "available": all_columns},
            )
        if not isinstance(raw_values, list):
            raise InvalidQueryError(
                f"filter value for '{col}' must be a list of strings.",
                details={"column": col},
            )
        values = [v for v in raw_values if v is not None and v != ""]
        if not values:
            continue
        placeholders = ",".join(["?"] * len(values))
        where_parts.append(f"{_quote_ident(col)} IN ({placeholders})")
        where_params.extend([str(v) for v in values])

    where_sql = ""
    if where_parts:
        where_sql = " WHERE " + " AND ".join(where_parts)

    order_sql = ""
    if order is not None:
        col_name, direction = order
        if col_name not in all_columns:
            raise InvalidQueryError(
                f"unknown order_by column: {col_name}",
                details={"column": col_name, "available": all_columns},
            )
        order_sql = f" ORDER BY {_quote_ident(col_name)} {direction}"

    select_cols_sql = ", ".join(_quote_ident(c) for c in projected)

    # Total (matches the WHERE clause but ignores limit/offset).
    total_sql = f"SELECT COUNT(*) FROM entities{where_sql}"
    total_rows = await pool.execute(
        snapshot_id, project_id, total_sql, parameters=where_params,
    )
    total = int(total_rows[0][0]) if total_rows else 0

    page_sql = (
        f"SELECT {select_cols_sql} FROM entities"
        f"{where_sql}{order_sql} LIMIT {int(limit)} OFFSET {int(offset)}"
    )
    page = await pool.execute(
        snapshot_id, project_id, page_sql, parameters=where_params,
    )

    rows = [_row_to_dict(projected, r) for r in page]
    return RowsResult(columns=projected, rows=rows, total=total)


def _row_to_dict(columns: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for col, value in zip(columns, row, strict=False):
        out[col] = _coerce_for_json(value)
    return out


def _coerce_for_json(value: Any) -> Any:
    """Make a DuckDB-returned value JSON/CSV-serialisable."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple)):
        return [_coerce_for_json(v) for v in value]
    if isinstance(value, dict):
        return {str(k): _coerce_for_json(v) for k, v in value.items()}
    return str(value)


# ── Export ─────────────────────────────────────────────────────────────────


SUPPORTED_EXPORT_FORMATS = ("csv", "xlsx", "parquet")


def export_to_format(
    *,
    columns: list[str],
    rows: list[dict[str, Any]],
    format: str,
) -> tuple[bytes, str, str]:
    """Serialise ``rows`` into the requested format.

    Returns ``(payload_bytes, content_type, file_extension)``.
    """
    fmt = format.lower()
    if fmt not in SUPPORTED_EXPORT_FORMATS:
        raise UnsupportedFormatError(
            f"unsupported export format '{format}'.",
            details={"format": format, "supported": list(SUPPORTED_EXPORT_FORMATS)},
        )

    if fmt == "csv":
        return _export_csv(columns, rows), "text/csv", "csv"
    if fmt == "xlsx":
        return (
            _export_xlsx(columns, rows),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    if fmt == "parquet":
        return _export_parquet(columns, rows), "application/octet-stream", "parquet"
    # Defensive — list updated above so this branch is unreachable.
    raise UnsupportedFormatError(  # pragma: no cover
        f"unsupported export format '{format}'.",
    )


def _export_csv(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {col: _stringify_for_csv(row.get(col)) for col in columns}
        )
    return buf.getvalue().encode("utf-8")


def _stringify_for_csv(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False, default=str)


def _export_xlsx(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "snapshot"
    ws.append(columns)
    for row in rows:
        ws.append([_stringify_for_xlsx(row.get(col)) for col in columns])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stringify_for_xlsx(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _export_parquet(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    import pyarrow as pa
    import pyarrow.parquet as pq

    table_data = {col: [] for col in columns}
    for row in rows:
        for col in columns:
            table_data[col].append(_coerce_for_parquet(row.get(col)))

    table = pa.table(table_data)
    buf = io.BytesIO()
    pq.write_table(table, buf, compression="snappy")
    return buf.getvalue()


def _coerce_for_parquet(value: Any) -> Any:
    """Round-trip-safe value for pyarrow.

    pyarrow promotes mixed-type columns (e.g. ``[1, "two"]``) to string
    in our use case; we serialise non-scalar values as JSON strings
    rather than letting pyarrow guess.
    """
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


# ── Import (two-step: preview then commit) ─────────────────────────────────


def stage_import(
    *,
    snapshot_id: uuid.UUID,
    snapshot_columns: list[str],
    upload_filename: str,
    upload_bytes: bytes,
    preview_size: int = 50,
) -> dict[str, Any]:
    """Parse + validate the uploaded file. Stash for a follow-up commit.

    Returns the preview payload directly so the router can wrap it in
    its Pydantic out-schema. The caller is responsible for honouring
    :class:`UnsupportedFormatError` / :class:`ImportSchemaError`.
    """
    fmt = (upload_filename or "").rsplit(".", 1)[-1].lower()
    if fmt == "csv":
        columns, rows = _parse_csv(upload_bytes)
    elif fmt == "xlsx":
        columns, rows = _parse_xlsx(upload_bytes)
    else:
        raise UnsupportedFormatError(
            f"unsupported import format '{fmt}'.",
            details={"format": fmt, "supported": ["csv", "xlsx"]},
        )

    matched = [c for c in columns if c in snapshot_columns]
    missing = [c for c in snapshot_columns if c not in columns]
    extra = [c for c in columns if c not in snapshot_columns]

    if not matched:
        # If literally no columns line up the upload is unusable;
        # 422 is more helpful than committing garbage.
        raise ImportSchemaError(
            "uploaded file shares no columns with the snapshot.",
            details={
                "uploaded_columns": columns,
                "snapshot_columns": snapshot_columns,
            },
        )

    staging_id = secrets.token_urlsafe(16)
    _STAGING[staging_id] = _StagedImport(
        snapshot_id=snapshot_id,
        columns=columns,
        rows=rows,
        expires_at=time.monotonic() + _STAGING_TTL_SECONDS,
    )
    _gc_staging()

    return {
        "snapshot_id": snapshot_id,
        "staging_id": staging_id,
        "columns": columns,
        "matched_columns": matched,
        "missing_columns": missing,
        "extra_columns": extra,
        "total_rows": len(rows),
        "preview_rows": rows[:preview_size],
    }


def commit_import(
    *,
    snapshot_id: uuid.UUID,
    staging_id: str,
) -> dict[str, Any]:
    """Finalise a previously staged import.

    Today this just returns the row count — full materialisation into a
    snapshot-side "supplementary data" Parquet is tracked as a T06
    follow-up (the schema for that attachment doesn't exist yet, and
    silently writing onto the entities Parquet would corrupt the
    immutable-snapshot invariant). The two-step protocol is wired
    end-to-end so the UI / API contract is stable.
    """
    _gc_staging()
    staged = _STAGING.get(staging_id)
    if staged is None:
        raise StagingNotFoundError(
            f"staging id '{staging_id}' not found or expired.",
            details={"staging_id": staging_id},
        )
    if staged.snapshot_id != snapshot_id:
        raise StagingNotFoundError(
            f"staging id '{staging_id}' does not belong to snapshot {snapshot_id}.",
            details={"staging_id": staging_id, "snapshot_id": str(snapshot_id)},
        )

    rows_count = len(staged.rows)
    # One-shot consumption — drop after commit so a replay can't double-count.
    _STAGING.pop(staging_id, None)
    return {
        "snapshot_id": snapshot_id,
        "staging_id": staging_id,
        "rows_committed": rows_count,
    }


def _gc_staging() -> None:
    now = time.monotonic()
    expired = [k for k, v in _STAGING.items() if v.expires_at < now]
    for k in expired:
        _STAGING.pop(k, None)


def _parse_csv(payload: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = payload.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ImportSchemaError(
            "uploaded CSV is empty.",
            details={"reason": "no header row"},
        ) from exc
    columns = [c.strip() for c in header]
    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        rows.append({col: raw_row[i] if i < len(raw_row) else None
                     for i, col in enumerate(columns)})
    return columns, rows


def _parse_xlsx(payload: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise ImportSchemaError(
            "uploaded XLSX is empty.",
            details={"reason": "no header row"},
        ) from exc
    columns = [str(c).strip() if c is not None else "" for c in header]
    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        rows.append(
            {col: raw[i] if i < len(raw) else None
             for i, col in enumerate(columns)}
        )
    return columns, rows


__all__ = [
    "ImportSchemaError",
    "InvalidQueryError",
    "RowsIOError",
    "RowsResult",
    "SUPPORTED_EXPORT_FORMATS",
    "StagingNotFoundError",
    "UnsupportedFormatError",
    "commit_import",
    "export_to_format",
    "read_rows",
    "stage_import",
]
