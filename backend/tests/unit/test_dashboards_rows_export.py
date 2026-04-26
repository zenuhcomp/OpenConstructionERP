"""T06 unit tests for rows reader + export round-trip + import staging.

Covers (8+ tests):

* CSV / XLSX / Parquet round-trip — bytes produced are parseable and
  carry the expected column set + row count.
* Filter + pagination correctness against a real DuckDB / Parquet
  fixture (uses :class:`DuckDBPool` so the SQL composition path is
  exercised end-to-end).
* ``order_by`` ASC / DESC correctness.
* Unknown-column / unknown-filter rejection (422 surface).
* Filters with empty value lists are dropped (matches T04 semantics).
* Two-step import: stage → commit returns the row count and frees the
  staging slot; replays raise :class:`StagingNotFoundError`.
* Schema mismatch when an upload shares no columns with the snapshot.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from pathlib import Path

import pandas as pd
import pytest

from app.core.storage import LocalStorageBackend
from app.modules.dashboards.duckdb_pool import DuckDBPool
from app.modules.dashboards.rows_io import (
    ImportSchemaError,
    InvalidQueryError,
    StagingNotFoundError,
    UnsupportedFormatError,
    commit_import,
    export_to_format,
    read_rows,
    stage_import,
)
from app.modules.dashboards.snapshot_storage import write_parquet

# ── Fixtures ────────────────────────────────────────────────────────────────


@pytest.fixture
def small_df() -> pd.DataFrame:
    """50-row DataFrame with two categorical columns + one numeric.

    Layout:
        category: 30 'wall' + 20 'door'
        material: alternating 'concrete' / 'steel'
        thickness_mm: 100..149
    """
    rows = []
    for i in range(50):
        rows.append(
            {
                "entity_guid": f"guid-{i}",
                "category": "wall" if i < 30 else "door",
                "material": "concrete" if i % 2 == 0 else "steel",
                "thickness_mm": 100 + i,
            }
        )
    return pd.DataFrame(rows)


@pytest.fixture
def parquet_fixture(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    small_df: pd.DataFrame,
):
    """Persist ``small_df`` as the entities Parquet for a fresh snapshot."""
    backend = LocalStorageBackend(base_dir=tmp_path)
    monkeypatch.setattr(
        "app.modules.dashboards.snapshot_storage.get_storage_backend",
        lambda: backend,
    )
    project_id = uuid.uuid4()
    snapshot_id = uuid.uuid4()

    import asyncio

    asyncio.run(
        write_parquet(project_id, snapshot_id, "entities", small_df),
    )
    return project_id, snapshot_id


# ── Read rows ──────────────────────────────────────────────────────────────


class TestReadRows:
    @pytest.mark.asyncio
    async def test_pagination_returns_correct_total_and_page(
        self, parquet_fixture,
    ) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                limit=10,
                offset=0,
            )
        finally:
            await pool.close_all()

        assert result.total == 50
        assert len(result.rows) == 10
        assert "category" in result.columns

    @pytest.mark.asyncio
    async def test_filters_narrow_the_result(self, parquet_fixture) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                filters=json.dumps({"category": ["wall"]}),
                limit=100,
            )
        finally:
            await pool.close_all()

        assert result.total == 30
        assert all(r["category"] == "wall" for r in result.rows)

    @pytest.mark.asyncio
    async def test_filter_with_empty_list_is_dropped(self, parquet_fixture) -> None:
        """``{"category": []}`` must not narrow results — matches T04."""
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                filters=json.dumps({"category": []}),
                limit=100,
            )
        finally:
            await pool.close_all()

        assert result.total == 50

    @pytest.mark.asyncio
    async def test_unknown_column_in_filters_raises(
        self, parquet_fixture,
    ) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            with pytest.raises(InvalidQueryError):
                await read_rows(
                    pool=pool,
                    snapshot_id=snapshot_id,
                    project_id=project_id,
                    filters=json.dumps({"made_up_column": ["x"]}),
                )
        finally:
            await pool.close_all()

    @pytest.mark.asyncio
    async def test_order_by_ascending(self, parquet_fixture) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                order_by="thickness_mm:asc",
                limit=5,
            )
        finally:
            await pool.close_all()

        thicknesses = [r["thickness_mm"] for r in result.rows]
        assert thicknesses == sorted(thicknesses)
        assert thicknesses[0] == 100

    @pytest.mark.asyncio
    async def test_order_by_descending(self, parquet_fixture) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                order_by="thickness_mm:desc",
                limit=5,
            )
        finally:
            await pool.close_all()

        thicknesses = [r["thickness_mm"] for r in result.rows]
        assert thicknesses == sorted(thicknesses, reverse=True)
        assert thicknesses[0] == 149

    @pytest.mark.asyncio
    async def test_columns_subset_projection(self, parquet_fixture) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                columns="category,thickness_mm",
                limit=3,
            )
        finally:
            await pool.close_all()

        assert result.columns == ["category", "thickness_mm"]
        for row in result.rows:
            assert set(row.keys()) == {"category", "thickness_mm"}


# ── Export ─────────────────────────────────────────────────────────────────


class TestExportRoundTrips:
    @pytest.mark.asyncio
    async def test_csv_round_trip_matches_filtered_query(
        self, parquet_fixture,
    ) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                filters=json.dumps({"material": ["concrete"]}),
                limit=100,
            )
        finally:
            await pool.close_all()

        payload, content_type, ext = export_to_format(
            columns=result.columns, rows=result.rows, format="csv",
        )
        assert content_type == "text/csv"
        assert ext == "csv"

        decoded = payload.decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        roundtrip = list(reader)
        # Every concrete row should round-trip; row count matches the
        # filtered query result, not the full snapshot.
        assert len(roundtrip) == result.total
        assert all(r["material"] == "concrete" for r in roundtrip)

    @pytest.mark.asyncio
    async def test_xlsx_round_trip_preserves_rows(
        self, parquet_fixture,
    ) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                limit=20,
            )
        finally:
            await pool.close_all()

        payload, content_type, ext = export_to_format(
            columns=result.columns, rows=result.rows, format="xlsx",
        )
        assert ext == "xlsx"
        assert content_type.endswith("spreadsheetml.sheet")

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(payload), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # First row is the header.
        assert list(rows[0]) == result.columns
        assert len(rows) - 1 == 20

    @pytest.mark.asyncio
    async def test_parquet_round_trip_preserves_rows(
        self, parquet_fixture,
    ) -> None:
        project_id, snapshot_id = parquet_fixture
        pool = DuckDBPool()
        try:
            result = await read_rows(
                pool=pool,
                snapshot_id=snapshot_id,
                project_id=project_id,
                limit=15,
            )
        finally:
            await pool.close_all()

        payload, content_type, ext = export_to_format(
            columns=result.columns, rows=result.rows, format="parquet",
        )
        assert ext == "parquet"

        import pyarrow.parquet as pq

        table = pq.read_table(io.BytesIO(payload))
        df = table.to_pandas()
        assert list(df.columns) == result.columns
        assert len(df) == 15

    def test_unsupported_format_raises(self) -> None:
        with pytest.raises(UnsupportedFormatError):
            export_to_format(columns=["a"], rows=[], format="xls")


# ── Import staging ─────────────────────────────────────────────────────────


class TestImportStaging:
    def test_csv_round_trip_stage_then_commit(self) -> None:
        snapshot_id = uuid.uuid4()
        upload = (
            b"category,material,thickness_mm\n"
            b"wall,concrete,200\n"
            b"door,wood,45\n"
            b"wall,concrete,180\n"
        )
        preview = stage_import(
            snapshot_id=snapshot_id,
            snapshot_columns=["category", "material", "thickness_mm"],
            upload_filename="rows.csv",
            upload_bytes=upload,
        )
        assert preview["total_rows"] == 3
        assert preview["matched_columns"] == [
            "category", "material", "thickness_mm",
        ]
        assert preview["missing_columns"] == []
        assert preview["extra_columns"] == []
        assert len(preview["preview_rows"]) == 3

        commit = commit_import(
            snapshot_id=snapshot_id, staging_id=preview["staging_id"],
        )
        assert commit["rows_committed"] == 3

        # Replay must fail — staging is one-shot.
        with pytest.raises(StagingNotFoundError):
            commit_import(
                snapshot_id=snapshot_id, staging_id=preview["staging_id"],
            )

    def test_extra_and_missing_columns_reported(self) -> None:
        snapshot_id = uuid.uuid4()
        upload = b"category,note\nwall,extra-column\n"
        preview = stage_import(
            snapshot_id=snapshot_id,
            snapshot_columns=["category", "material"],
            upload_filename="rows.csv",
            upload_bytes=upload,
        )
        assert preview["matched_columns"] == ["category"]
        assert preview["missing_columns"] == ["material"]
        assert preview["extra_columns"] == ["note"]

    def test_no_overlap_raises_schema_error(self) -> None:
        snapshot_id = uuid.uuid4()
        upload = b"alpha,beta\n1,2\n"
        with pytest.raises(ImportSchemaError):
            stage_import(
                snapshot_id=snapshot_id,
                snapshot_columns=["category", "material"],
                upload_filename="rows.csv",
                upload_bytes=upload,
            )

    def test_unsupported_extension_raises(self) -> None:
        snapshot_id = uuid.uuid4()
        with pytest.raises(UnsupportedFormatError):
            stage_import(
                snapshot_id=snapshot_id,
                snapshot_columns=["category"],
                upload_filename="rows.txt",
                upload_bytes=b"category\nwall\n",
            )

    def test_commit_with_wrong_snapshot_id_rejected(self) -> None:
        right = uuid.uuid4()
        wrong = uuid.uuid4()
        upload = b"category\nwall\n"
        preview = stage_import(
            snapshot_id=right,
            snapshot_columns=["category"],
            upload_filename="r.csv",
            upload_bytes=upload,
        )
        with pytest.raises(StagingNotFoundError):
            commit_import(
                snapshot_id=wrong, staging_id=preview["staging_id"],
            )
