"""Tests for the EAC Excel/CSV import-export module."""

from __future__ import annotations

import openpyxl

from app.modules.requirements.excel_io import (
    COLUMNS,
    build_template_xlsx,
    export_xlsx,
    parse_csv,
    parse_xlsx,
)


def test_template_has_headers_and_sample() -> None:
    payload = build_template_xlsx()
    rows, warnings = parse_xlsx(payload)
    assert warnings == []
    assert len(rows) == 1  # the sample row
    sample = rows[0]
    assert sample["entity"] == "Walls"
    assert sample["attribute"] == "FireRating"
    assert sample["constraint_type"] == "regex"
    assert sample["priority"] == "must"


def test_template_has_operators_legend() -> None:
    import io as _io

    wb = openpyxl.load_workbook(_io.BytesIO(build_template_xlsx()))
    assert "Operators" in wb.sheetnames
    legend = wb["Operators"]
    operators_listed = {row[0] for row in legend.iter_rows(min_row=2, values_only=True) if row[0]}
    expected = {
        "equals", "not_equals", "contains", "not_contains",
        "min", "max", "range", "regex", "exists", "not_exists",
    }
    assert expected.issubset(operators_listed)


def test_csv_roundtrip_preserves_data() -> None:
    csv_text = (
        b"entity,attribute,constraint_type,constraint_value,priority\n"
        b"Doors,Width,min,800,should\n"
        b"Slabs,Volume,range,1.0..50.0,must\n"
    )
    rows, warnings = parse_csv(csv_text)
    assert warnings == []
    assert len(rows) == 2
    assert rows[0]["entity"] == "Doors"
    assert rows[0]["constraint_type"] == "min"
    assert rows[1]["constraint_value"] == "1.0..50.0"


def test_csv_quoted_value_with_comma() -> None:
    # CSV writers quote fields that contain commas — verify our parser
    # respects the standard so exports re-import cleanly.
    csv_text = b'entity,attribute,constraint_type,constraint_value\nWalls,FireRating,regex,"^F\\d{2,3}$"\n'
    rows, _ = parse_csv(csv_text)
    assert len(rows) == 1
    assert rows[0]["constraint_value"] == r"^F\d{2,3}$"


def test_unknown_operator_is_warned_and_defaulted() -> None:
    csv_text = (
        b"entity,attribute,constraint_type,constraint_value\n"
        b"Walls,FireRating,greater_than,50\n"
    )
    rows, warnings = parse_csv(csv_text)
    assert len(rows) == 1
    assert rows[0]["constraint_type"] == "equals"
    assert any("greater_than" in w for w in warnings)


def test_missing_required_column_yields_warning_and_no_rows() -> None:
    csv_text = b"entity,attribute,unit\nWalls,FireRating,F90\n"
    rows, warnings = parse_csv(csv_text)
    assert rows == []
    assert any("constraint_type" in w for w in warnings)


def test_empty_rows_are_skipped() -> None:
    csv_text = (
        b"entity,attribute,constraint_type,constraint_value\n"
        b"Walls,FireRating,equals,F90\n"
        b",,,\n"
        b"Doors,Width,min,800\n"
    )
    rows, _ = parse_csv(csv_text)
    assert len(rows) == 2


def test_export_xlsx_roundtrips_through_parser() -> None:
    rows = [
        {
            "entity": "Walls",
            "attribute": "FireRating",
            "constraint_type": "regex",
            "constraint_value": r"^F\d{2,3}$",
            "unit": "",
            "category": "fire_safety",
            "priority": "must",
            "status": "open",
            "confidence": "",
            "source_ref": "DIN 4102",
            "notes": "",
        },
        {
            "entity": "Doors",
            "attribute": "Width",
            "constraint_type": "min",
            "constraint_value": "800",
            "unit": "mm",
            "category": "structural",
            "priority": "should",
            "status": "open",
            "confidence": "",
            "source_ref": "",
            "notes": "",
        },
    ]
    payload = export_xlsx(rows, title="Roundtrip")
    parsed, warnings = parse_xlsx(payload)
    assert warnings == []
    assert len(parsed) == 2
    assert parsed[0]["entity"] == "Walls"
    assert parsed[0]["constraint_type"] == "regex"
    assert parsed[1]["constraint_value"] == "800"


def test_columns_match_create_payload_keys() -> None:
    # The export contract must match the fields RequirementCreate expects.
    expected = {
        "entity", "attribute", "constraint_type", "constraint_value",
        "unit", "category", "priority", "source_ref", "notes",
    }
    assert expected.issubset(set(COLUMNS))
