"""Unit tests for the COBie UK 2.4 workbook builder.

Structural tests (sheet names + header rows + row counts) instead of a
full byte-for-byte snapshot, because openpyxl embeds build metadata
that makes byte-level reproducibility brittle across versions. We pin
the timestamp via ``CobieOptions.frozen_now`` so VALUES (not bytes)
stay deterministic.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.modules.bim_hub.exporters import build_cobie_workbook
from app.modules.bim_hub.exporters.cobie import (
    COMPONENT_COLUMNS,
    CONTACT_COLUMNS,
    FACILITY_COLUMNS,
    FLOOR_COLUMNS,
    SPACE_COLUMNS,
    SYSTEM_COLUMNS,
    TYPE_COLUMNS,
    CobieOptions,
)


def _fixture_model() -> SimpleNamespace:
    return SimpleNamespace(
        id=uuid.UUID("11111111-1111-1111-1111-111111111111"),
        project_id=uuid.UUID("22222222-2222-2222-2222-222222222222"),
        name="Test Facility",
        discipline="Architectural",
    )


def _fixture_elements() -> list[SimpleNamespace]:
    """Mix of tracked assets, spaces, untracked walls.

    - 2 spaces on 2 different storeys (→ 2 Floor rows + 2 Space rows)
    - 2 AHU units on same make/model (→ 1 Type + 2 Component)
    - 1 pump, different make (→ +1 Type + 1 Component)
    - 1 untracked wall (should NOT appear in Type/Component)
    - 2 components share ``parent_system="HVAC"`` (→ 1 System row)
    """
    return [
        # Space rows
        SimpleNamespace(
            stable_id="room-101",
            element_type="IfcSpace",
            name="Office 101",
            storey="Floor 1",
            discipline="Architectural",
            asset_info={},
            is_tracked_asset=False,
            quantities={"area": 42.5, "height": 3.0},
            properties={},
        ),
        SimpleNamespace(
            stable_id="room-201",
            element_type="Room",
            name="Office 201",
            storey="Floor 2",
            discipline="Architectural",
            asset_info={},
            is_tracked_asset=False,
            quantities={"area": 38.0, "height": 3.0},
            properties={},
        ),
        # Two identical AHU components
        SimpleNamespace(
            stable_id="ahu-01",
            element_type="AirHandlingUnit",
            name="AHU-01",
            storey="Floor 1",
            discipline="MEP",
            asset_info={
                "manufacturer": "Siemens",
                "model": "SV-100",
                "serial_number": "SN-AHU-001",
                "parent_system": "HVAC",
                "asset_tag": "AHU-01",
            },
            is_tracked_asset=True,
            quantities={},
            properties={},
        ),
        SimpleNamespace(
            stable_id="ahu-02",
            element_type="AirHandlingUnit",
            name="AHU-02",
            storey="Floor 2",
            discipline="MEP",
            asset_info={
                "manufacturer": "Siemens",
                "model": "SV-100",
                "serial_number": "SN-AHU-002",
                "parent_system": "HVAC",
                "asset_tag": "AHU-02",
            },
            is_tracked_asset=True,
            quantities={},
            properties={},
        ),
        # Different-make pump (its own Type)
        SimpleNamespace(
            stable_id="pump-01",
            element_type="Pump",
            name="CHW-Pump-01",
            storey="Floor 1",
            discipline="MEP",
            asset_info={
                "manufacturer": "Grundfos",
                "model": "MAGNA3",
                "serial_number": "SN-PMP-001",
                "parent_system": "CHW",
            },
            is_tracked_asset=True,
            quantities={},
            properties={},
        ),
        # Untracked wall — should NOT appear in Type/Component
        SimpleNamespace(
            stable_id="wall-001",
            element_type="Wall",
            name="Wall 001",
            storey="Floor 1",
            discipline="Architectural",
            asset_info={},
            is_tracked_asset=False,
            quantities={"area": 15.0},
            properties={},
        ),
    ]


def _workbook_from_bytes(xlsx: bytes):
    return load_workbook(BytesIO(xlsx), read_only=True, data_only=True)


class TestCobieWorkbookStructure:
    def test_workbook_has_all_seven_sheets(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        expected = ["Contact", "Facility", "Floor", "Space", "Type", "Component", "System"]
        assert wb.sheetnames == expected

    def test_each_sheet_has_correct_header_row(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        header_map = {
            "Contact": CONTACT_COLUMNS,
            "Facility": FACILITY_COLUMNS,
            "Floor": FLOOR_COLUMNS,
            "Space": SPACE_COLUMNS,
            "Type": TYPE_COLUMNS,
            "Component": COMPONENT_COLUMNS,
            "System": SYSTEM_COLUMNS,
        }
        for sheet_name, columns in header_map.items():
            ws = wb[sheet_name]
            actual = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            assert actual == columns, f"{sheet_name} headers mismatch"


class TestCobieRowCounts:
    def test_facility_has_exactly_one_row(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # 1 header + 1 data = 2 total rows.
        assert wb["Facility"].max_row == 2

    def test_floor_has_two_unique_storeys(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # 1 header + 2 storeys.
        assert wb["Floor"].max_row == 3

    def test_space_counts_only_room_like_elements(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # 1 header + 2 spaces.
        assert wb["Space"].max_row == 3

    def test_type_aggregates_by_make_model(self):
        """Two AHUs with same make/model → one Type row. Pump different
        make → its own row. Untracked wall → no row. Total: 2."""
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # 1 header + 2 types.
        assert wb["Type"].max_row == 3

    def test_component_has_one_row_per_tracked_asset(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # 1 header + 3 tracked (2 AHU + 1 pump). Wall excluded.
        assert wb["Component"].max_row == 4

    def test_system_aggregates_by_parent_system(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # 1 header + 2 systems (HVAC + CHW).
        assert wb["System"].max_row == 3


class TestCobieDataCorrectness:
    def test_facility_name_comes_from_model(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        facility_row = next(wb["Facility"].iter_rows(min_row=2, max_row=2))
        values = {c.value for c in facility_row}
        assert "Test Facility" in values

    def test_component_references_type_name(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        # Type column for component must equal the Type sheet's Name.
        types = {
            row[0].value
            for row in wb["Type"].iter_rows(min_row=2)
            if row[0].value
        }
        components_type_names = {
            row[COMPONENT_COLUMNS.index("TypeName")].value
            for row in wb["Component"].iter_rows(min_row=2)
            if row[0].value
        }
        assert components_type_names.issubset(types), (
            "Every Component.TypeName must reference an existing Type.Name"
        )

    def test_system_has_correct_component_names(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        rows = list(wb["System"].iter_rows(min_row=2))
        hvac = next(r for r in rows if r[0].value == "HVAC")
        members = hvac[SYSTEM_COLUMNS.index("ComponentNames")].value
        # Both AHU components are in HVAC (names preserved via asset_tag).
        assert "AHU-01" in members
        assert "AHU-02" in members

    def test_untracked_elements_excluded_from_component(self):
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(xlsx)
        component_names = {
            row[0].value for row in wb["Component"].iter_rows(min_row=2)
        }
        # wall-001 is NOT tracked, so should not be here.
        assert "Wall 001" not in component_names
        assert "wall-001" not in component_names


class TestCobieDeterminism:
    def test_frozen_timestamp_produces_repeatable_timestamps(self):
        """Same input with same frozen_now yields identical CreatedOn
        strings across runs — snapshot-style comparison."""
        opts = CobieOptions(
            frozen_now=datetime(2026, 4, 22, 9, 0, 0, tzinfo=timezone.utc)
        )
        xlsx1 = build_cobie_workbook(_fixture_model(), _fixture_elements(), options=opts)
        xlsx2 = build_cobie_workbook(_fixture_model(), _fixture_elements(), options=opts)
        # Load both and compare the Facility row (which contains CreatedOn).
        wb1 = _workbook_from_bytes(xlsx1)
        wb2 = _workbook_from_bytes(xlsx2)
        r1 = [c.value for c in next(wb1["Facility"].iter_rows(min_row=2, max_row=2))]
        r2 = [c.value for c in next(wb2["Facility"].iter_rows(min_row=2, max_row=2))]
        assert r1 == r2

    def test_created_on_format_is_iso_without_timezone(self):
        opts = CobieOptions(
            frozen_now=datetime(2026, 4, 22, 9, 15, 42, tzinfo=timezone.utc)
        )
        xlsx = build_cobie_workbook(_fixture_model(), _fixture_elements(), options=opts)
        wb = _workbook_from_bytes(xlsx)
        facility_row = next(wb["Facility"].iter_rows(min_row=2, max_row=2))
        row_values = [c.value for c in facility_row]
        # CreatedOn is at index 2 in FACILITY_COLUMNS.
        assert row_values[FACILITY_COLUMNS.index("CreatedOn")] == "2026-04-22T09:15:42"


class TestCobieEdgeCases:
    def test_empty_elements_still_produces_valid_workbook(self):
        xlsx = build_cobie_workbook(_fixture_model(), [])
        wb = _workbook_from_bytes(xlsx)
        # All 7 sheets present; data-less sheets have only the header row.
        assert "Facility" in wb.sheetnames
        assert wb["Component"].max_row == 1  # header only
        assert wb["Type"].max_row == 1  # header only
        # Floor falls back to "Floor 1" placeholder so Space rows have
        # something to reference.
        assert wb["Floor"].max_row == 2

    def test_performance_baseline_5000_elements(self):
        """Synthetic 5k-element stress test — must finish under 5s.

        Not a byte snapshot, just a smoke-test for the scale we expect
        a medium office building to land at. Larger models (50k+) are
        exercised via the integration suite with a real RVT.
        """
        import time

        elements = []
        for i in range(5000):
            elements.append(
                SimpleNamespace(
                    stable_id=f"elem-{i:05d}",
                    element_type="AirHandlingUnit" if i % 50 == 0 else "Wall",
                    name=f"Element {i}",
                    storey=f"Floor {(i // 500) + 1}",
                    discipline="MEP" if i % 50 == 0 else "Arch",
                    asset_info=(
                        {"manufacturer": "ACME", "model": "M-1", "parent_system": "SYS"}
                        if i % 50 == 0
                        else {}
                    ),
                    is_tracked_asset=(i % 50 == 0),
                    quantities={},
                    properties={},
                )
            )
        t0 = time.perf_counter()
        xlsx = build_cobie_workbook(_fixture_model(), elements)
        elapsed = time.perf_counter() - t0
        assert elapsed < 5.0, f"COBie export too slow: {elapsed:.2f}s"
        assert len(xlsx) > 0
