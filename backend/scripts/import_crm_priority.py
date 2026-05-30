"""‌⁠‍Import contacts from a CRM-priority spreadsheet into oe_contacts_contact.

The expected workbook is the e-mail triage export ``crm_priority.xlsx``,
which carries five tier sheets:

    1_paid_customers
    2_hot_leads
    3_active_warm
    4_inbound_leads
    5_cold_outreach

Each sheet has the columns (header row 1):

    email, name, company, country, language, tier, consent_basis,
    last_contact, days_since_last, msg_count, response_rate, topics,
    last_subject, inboxes

The script normalises every row into a ``oe_contacts_contact`` insert,
maps the sheet name to a ``contact_type`` (``customer`` for tier 1,
``lead`` for the rest), folds tier / language / country / consent_basis
/ topics / inboxes into a single ``metadata.tags`` array for the
front-end chip filter, and stores the full per-row CRM payload under
``metadata.crm`` so nothing is lost.

The script is idempotent: it upserts on ``primary_email`` (lower-cased).
If a contact already exists, only metadata.tags / metadata.crm / notes /
country_code / company_name are refreshed; existing first_name /
last_name / phone / vat are preserved.

Usage::

    python -m backend.scripts.import_crm_priority \\
        --xlsx "C:/path/to/crm_priority.xlsx" \\
        [--db backend/openestimate.db] \\
        [--owner-email demo@openconstructionerp.com] \\
        [--dry-run]
"""

from __future__ import annotations

import argparse
import json
import logging
import sqlite3
import sys
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger("import_crm_priority")

SHEET_TO_TYPE: dict[str, tuple[str, str]] = {
    # sheet name → (contact_type, tier_tag)
    "1_paid_customers": ("customer", "paid"),
    "2_hot_leads": ("lead", "hot"),
    "3_active_warm": ("lead", "warm"),
    "4_inbound_leads": ("lead", "inbound"),
    "5_cold_outreach": ("lead", "cold"),
}

COUNTRY_TO_ISO2: dict[str, str] = {
    # The xlsx mixes ISO codes and country names. Map the names that
    # actually appear in crm_priority.xlsx to ISO 3166-1 alpha-2.
    "germany": "DE",
    "deutschland": "DE",
    "uk": "GB",
    "united kingdom": "GB",
    "great britain": "GB",
    "usa": "US",
    "united states": "US",
    "us": "US",
    "russia": "RU",
    "russian federation": "RU",
    "switzerland": "CH",
    "schweiz": "CH",
    "austria": "AT",
    "österreich": "AT",
    "france": "FR",
    "spain": "ES",
    "italy": "IT",
    "netherlands": "NL",
    "belgium": "BE",
    "poland": "PL",
    "czechia": "CZ",
    "czech republic": "CZ",
    "ukraine": "UA",
    "turkey": "TR",
    "türkiye": "TR",
    "australia": "AU",
    "canada": "CA",
    "india": "IN",
    "china": "CN",
    "japan": "JP",
    "brazil": "BR",
    "mexico": "MX",
    "south africa": "ZA",
    "uae": "AE",
    "united arab emirates": "AE",
}


def _normalise_country(value: Any) -> str | None:
    """‌⁠‍Return a 2-char ISO country code or None."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if len(text) == 2 and text.isalpha():
        return text.upper()
    iso = COUNTRY_TO_ISO2.get(text.lower())
    return iso


def _split_name(full: str | None) -> tuple[str | None, str | None]:
    if not full:
        return None, None
    cleaned = " ".join(full.split())
    if not cleaned:
        return None, None
    if " " not in cleaned:
        return cleaned, None
    parts = cleaned.split(" ", 1)
    return parts[0], parts[1]


def _split_csv(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [p.strip().lower() for p in text.replace(";", ",").split(",")]
    return [p for p in parts if p]


def _coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _coerce_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _build_tags(
    *,
    group: str,
    tier: str | None,
    language: str | None,
    country_iso: str | None,
    consent_basis: str | None,
    topics: list[str],
    inboxes: list[str],
) -> list[str]:
    """‌⁠‍Compose the metadata.tags array used by the contacts filter chips."""
    bag: list[str] = [f"group:{group}"]
    if tier:
        bag.append(f"tier:{tier.lower()}")
    if language:
        bag.append(f"lang:{language.lower()}")
    if country_iso:
        bag.append(f"country:{country_iso.lower()}")
    if consent_basis:
        bag.append(f"consent:{consent_basis.lower()}")
    for t in topics:
        bag.append(f"topic:{t}")
    for ib in inboxes:
        bag.append(f"inbox:{ib}")
    # De-duplicate while preserving order.
    seen: set[str] = set()
    out: list[str] = []
    for tag in bag:
        if tag not in seen:
            seen.add(tag)
            out.append(tag)
    return out


def _build_crm_payload(row: dict[str, Any], *, group: str) -> dict[str, Any]:
    return {
        "source": "crm_priority.xlsx",
        "group": group,
        "tier": row.get("tier"),
        "consent_basis": row.get("consent_basis"),
        "language": row.get("language"),
        "last_contact": row.get("last_contact"),
        "days_since_last": _coerce_int(row.get("days_since_last")),
        "msg_count": _coerce_int(row.get("msg_count")),
        "response_rate": _coerce_float(row.get("response_rate")),
        "topics": _split_csv(row.get("topics")),
        "last_subject": row.get("last_subject"),
        "inboxes": _split_csv(row.get("inboxes")),
    }


def _read_workbook(xlsx: Path) -> dict[str, list[dict[str, Any]]]:
    """Return ``{sheet_name: [row_dict, ...]}`` for the five tier sheets."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        logger.error("openpyxl is required. Install with: pip install openpyxl")
        sys.exit(2)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in SHEET_TO_TYPE:
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet not found: %s", sheet_name)
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            continue
        header_map = {(h.strip().lower() if isinstance(h, str) else None): i for i, h in enumerate(headers)}
        rows: list[dict[str, Any]] = []
        for raw in rows_iter:
            if raw is None:
                continue
            row = {key: raw[idx] for key, idx in header_map.items() if key is not None and idx < len(raw)}
            rows.append(row)
        result[sheet_name] = rows
    wb.close()
    return result


def _resolve_owner_id(conn: sqlite3.Connection, email: str | None) -> str:
    """Find the owner UUID. Falls back to the first admin user."""
    cur = conn.cursor()
    if email:
        cur.execute(
            "SELECT id FROM oe_users_user WHERE LOWER(email) = ? LIMIT 1",
            (email.lower(),),
        )
        row = cur.fetchone()
        if row:
            return row[0]
        logger.warning("Owner email not found: %s — falling back to first admin", email)
    cur.execute("SELECT id FROM oe_users_user WHERE role = 'admin' ORDER BY created_at ASC LIMIT 1")
    row = cur.fetchone()
    if not row:
        logger.error("No admin user found in DB. Cannot assign tenant_id.")
        sys.exit(2)
    return row[0]


def _fetch_existing(conn: sqlite3.Connection, emails: list[str]) -> dict[str, dict]:
    """Return {email_lower: {id, metadata, notes, ...}} for existing rows."""
    if not emails:
        return {}
    out: dict[str, dict] = {}
    placeholder = ",".join("?" * len(emails))
    cur = conn.cursor()
    cur.execute(
        f"SELECT id, primary_email, metadata, notes, first_name, last_name, "
        f"company_name, country_code FROM oe_contacts_contact "
        f"WHERE LOWER(primary_email) IN ({placeholder})",
        [e.lower() for e in emails],
    )
    for r in cur.fetchall():
        out[r[1].lower()] = {
            "id": r[0],
            "metadata": r[2],
            "notes": r[3],
            "first_name": r[4],
            "last_name": r[5],
            "company_name": r[6],
            "country_code": r[7],
        }
    return out


def _merge_metadata(existing_raw: str | None, fresh: dict[str, Any]) -> dict[str, Any]:
    """Merge an existing metadata blob with freshly built CRM data.

    Tags are union-merged; the crm sub-dict is replaced with the latest
    payload so re-imports reflect the newest export.
    """
    try:
        existing = json.loads(existing_raw) if existing_raw else {}
    except json.JSONDecodeError:
        existing = {}
    if not isinstance(existing, dict):
        existing = {}

    existing_tags = existing.get("tags") if isinstance(existing.get("tags"), list) else []
    fresh_tags = fresh.get("tags") or []
    seen: set[str] = set()
    merged_tags: list[str] = []
    for tag in (*existing_tags, *fresh_tags):
        if isinstance(tag, str) and tag and tag not in seen:
            seen.add(tag)
            merged_tags.append(tag)

    merged = dict(existing)
    merged["tags"] = merged_tags
    merged["crm"] = fresh.get("crm")
    return merged


def _process_workbook(
    conn: sqlite3.Connection,
    sheets: dict[str, list[dict[str, Any]]],
    *,
    owner_id: str,
    dry_run: bool,
) -> dict[str, Any]:
    counters = {
        "inserted": 0,
        "updated": 0,
        "skipped_no_email": 0,
        "skipped_invalid": 0,
        "by_sheet": {},
    }

    # First pass: collect every email we'll touch so we can fetch
    # existing rows in one round-trip.
    pending: list[tuple[str, dict[str, Any]]] = []
    for sheet_name, rows in sheets.items():
        for row in rows:
            email_raw = row.get("email")
            if not email_raw or "@" not in str(email_raw):
                counters["skipped_no_email"] += 1
                continue
            pending.append((sheet_name, row))

    emails = [str(p[1]["email"]).strip().lower() for p in pending]
    existing = _fetch_existing(conn, emails)

    cur = conn.cursor()
    now = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")

    for sheet_name, row in pending:
        contact_type, tier_tag = SHEET_TO_TYPE[sheet_name]
        email = str(row["email"]).strip().lower()
        first_name, last_name = _split_name(row.get("name"))
        company = (str(row.get("company")).strip() if row.get("company") else None) or None
        country_iso = _normalise_country(row.get("country"))
        last_subject = str(row.get("last_subject")).strip() if row.get("last_subject") else None
        notes = last_subject or None

        topics = _split_csv(row.get("topics"))
        inboxes = _split_csv(row.get("inboxes"))
        language = str(row.get("language")).strip().lower() if row.get("language") else None

        tags = _build_tags(
            group=sheet_name.split("_", 1)[1],  # paid_customers / hot_leads / ...
            tier=str(row.get("tier")) if row.get("tier") else tier_tag,
            language=language,
            country_iso=country_iso,
            consent_basis=str(row.get("consent_basis")) if row.get("consent_basis") else None,
            topics=topics,
            inboxes=inboxes,
        )
        crm = _build_crm_payload(row, group=sheet_name)
        fresh_metadata = {"tags": tags, "crm": crm}

        existing_row = existing.get(email)
        sheet_count = counters["by_sheet"].setdefault(sheet_name, {"inserted": 0, "updated": 0})

        if existing_row:
            merged = _merge_metadata(existing_row.get("metadata"), fresh_metadata)
            updates: dict[str, Any] = {
                "metadata": json.dumps(merged, ensure_ascii=False),
                "notes": notes or existing_row.get("notes"),
                "country_code": country_iso or existing_row.get("country_code"),
                "company_name": company or existing_row.get("company_name"),
                "first_name": existing_row.get("first_name") or first_name,
                "last_name": existing_row.get("last_name") or last_name,
                "is_active": 1,
                "updated_at": now,
            }
            if not dry_run:
                cur.execute(
                    "UPDATE oe_contacts_contact SET "
                    "metadata = :metadata, notes = :notes, "
                    "country_code = :country_code, company_name = :company_name, "
                    "first_name = :first_name, last_name = :last_name, "
                    "is_active = :is_active, updated_at = :updated_at "
                    "WHERE id = :id",
                    {**updates, "id": existing_row["id"]},
                )
            counters["updated"] += 1
            sheet_count["updated"] += 1
            continue

        contact_id = str(uuid.uuid4())
        params = {
            "id": contact_id,
            "contact_type": contact_type,
            "is_platform_user": 0,
            "user_id": None,
            "first_name": first_name,
            "last_name": last_name,
            "company_name": company,
            "legal_name": None,
            "vat_number": None,
            "country_code": country_iso,
            "address": None,
            "primary_email": email,
            "primary_phone": None,
            "website": None,
            "certifications": "[]",
            "insurance": "[]",
            "prequalification_status": None,
            "qualified_until": None,
            "payment_terms_days": None,
            "currency_code": None,
            "name_translations": None,
            "notes": notes,
            "is_active": 1,
            "tenant_id": owner_id,
            "created_by": owner_id,
            "metadata": json.dumps(fresh_metadata, ensure_ascii=False),
            "created_at": now,
            "updated_at": now,
        }
        if not dry_run:
            cur.execute(
                "INSERT INTO oe_contacts_contact ("
                "id, contact_type, is_platform_user, user_id, first_name, "
                "last_name, company_name, legal_name, vat_number, country_code, "
                "address, primary_email, primary_phone, website, certifications, "
                "insurance, prequalification_status, qualified_until, "
                "payment_terms_days, currency_code, name_translations, notes, "
                "is_active, tenant_id, created_by, metadata, created_at, updated_at"
                ") VALUES ("
                ":id, :contact_type, :is_platform_user, :user_id, :first_name, "
                ":last_name, :company_name, :legal_name, :vat_number, :country_code, "
                ":address, :primary_email, :primary_phone, :website, :certifications, "
                ":insurance, :prequalification_status, :qualified_until, "
                ":payment_terms_days, :currency_code, :name_translations, :notes, "
                ":is_active, :tenant_id, :created_by, :metadata, :created_at, :updated_at"
                ")",
                params,
            )
        counters["inserted"] += 1
        sheet_count["inserted"] += 1

    if not dry_run:
        conn.commit()
    return counters


def _find_default_db() -> Path | None:
    candidates = [
        Path.cwd() / "backend" / "openestimate.db",
        Path.cwd() / "openestimate.db",
        Path(__file__).resolve().parent.parent / "openestimate.db",
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="Path to crm_priority.xlsx")
    parser.add_argument("--db", help="Path to openestimate.db (auto-detected by default)")
    parser.add_argument(
        "--owner-email",
        default="demo@openconstructionerp.com",
        help="User email whose tenant the contacts attach to (default: demo admin)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate but do not write to the DB",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        logger.error("xlsx not found: %s", xlsx_path)
        return 2

    db_path = Path(args.db) if args.db else _find_default_db()
    if db_path is None or not db_path.is_file():
        logger.error("openestimate.db not found. Pass --db.")
        return 2
    logger.info("DB: %s", db_path)

    sheets = _read_workbook(xlsx_path)
    total_rows = sum(len(rows) for rows in sheets.values())
    logger.info("Loaded %d rows across %d sheets", total_rows, len(sheets))

    conn = sqlite3.connect(db_path)
    try:
        owner_id = _resolve_owner_id(conn, args.owner_email)
        logger.info("Owner: %s (%s)", args.owner_email, owner_id)

        counters = _process_workbook(conn, sheets, owner_id=owner_id, dry_run=args.dry_run)
    finally:
        conn.close()

    logger.info("")
    logger.info("=== summary ===")
    logger.info("inserted:        %d", counters["inserted"])
    logger.info("updated:         %d", counters["updated"])
    logger.info("skipped_no_email: %d", counters["skipped_no_email"])
    for sheet_name, sub in counters["by_sheet"].items():
        logger.info(
            "  %s: +%d new / %d updated",
            sheet_name,
            sub["inserted"],
            sub["updated"],
        )
    if args.dry_run:
        logger.info("(dry run — no changes written)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
